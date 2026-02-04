"""
Script untuk membuat Model Excel LENGKAP Perhitungan Fuzzy AHP
VERSI 2 - Input bobot kriteria dengan angka 1-10 (lebih sederhana)

Fitur:
- Input data kegiatan & peserta
- Input bobot kriteria dengan skala 1-10 (dikonversi ke TFN)
- Input nilai peserta per kriteria
- Perhitungan skor akhir dengan Fuzzy AHP
- Ranking peserta
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
input_font = Font(size=11, color="0000FF")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

kriteria_list = ["K1", "K2", "K3", "K4", "K5", "K6"]
kriteria_full = [
    "Status Keaktifan di Gugus Depan",
    "Pencapaian SKU", 
    "Pencapaian SPG",
    "Kesehatan Jasmani dan Rohani",
    "Tes Wawancara",
    "Tes Pilihan Ganda"
]
n_kriteria = 6
n_peserta = 10

def header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def input_style(cell):
    cell.font = input_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = input_fill

def section(ws, row, text, col_end=10):
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)

# ============================================
# SHEET 1: BOBOT KRITERIA (Input 1-10)
# ============================================
ws1 = wb.active
ws1.title = "1. Bobot Kriteria"

row = 1
ws1['A1'] = "FUZZY AHP - PERHITUNGAN BOBOT KRITERIA"
ws1['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws1.merge_cells('A1:H1')

ws1['A2'] = "Masukkan bobot kriteria dengan skala 1-10 di kolom INPUT"
ws1['A2'].font = Font(italic=True, color="FF0000")
ws1.merge_cells('A2:H2')

row = 4

# TABEL KONVERSI SKALA KE TFN
section(ws1, row, "TABEL KONVERSI SKALA KE FUZZY (TFN)", col_end=5)
row += 1

ws1.cell(row=row, column=1).value = "Skala"
ws1.cell(row=row, column=2).value = "Keterangan"
ws1.cell(row=row, column=3).value = "l"
ws1.cell(row=row, column=4).value = "m"
ws1.cell(row=row, column=5).value = "u"
for c in range(1, 6):
    header_style(ws1.cell(row=row, column=c))
row += 1

tfn_start = row
# Skala 1-10 dengan TFN
tfn_data = [
    (1, "Sangat Tidak Penting", 0.0, 0.1, 0.2),
    (2, "Tidak Penting", 0.1, 0.2, 0.3),
    (3, "Agak Tidak Penting", 0.2, 0.3, 0.4),
    (4, "Sedikit Tidak Penting", 0.3, 0.4, 0.5),
    (5, "Cukup Penting", 0.4, 0.5, 0.6),
    (6, "Agak Penting", 0.5, 0.6, 0.7),
    (7, "Penting", 0.6, 0.7, 0.8),
    (8, "Sangat Penting", 0.7, 0.8, 0.9),
    (9, "Amat Sangat Penting", 0.8, 0.9, 1.0),
    (10, "Mutlak Penting", 0.9, 1.0, 1.0),
]

for skala, ket, l, m, u in tfn_data:
    ws1.cell(row=row, column=1).value = skala
    ws1.cell(row=row, column=2).value = ket
    ws1.cell(row=row, column=3).value = l
    ws1.cell(row=row, column=4).value = m
    ws1.cell(row=row, column=5).value = u
    for c in range(1, 6):
        cell_style(ws1.cell(row=row, column=c))
    row += 1

tfn_end = row - 1
row += 1

# INPUT BOBOT KRITERIA
section(ws1, row, "INPUT BOBOT KRITERIA (Skala 1-10)", col_end=8)
row += 1

ws1.cell(row=row, column=1).value = "Kode"
ws1.cell(row=row, column=2).value = "Nama Kriteria"
ws1.cell(row=row, column=3).value = "INPUT\n(1-10)"
ws1.cell(row=row, column=4).value = "l"
ws1.cell(row=row, column=5).value = "m"
ws1.cell(row=row, column=6).value = "u"
ws1.cell(row=row, column=7).value = "Defuzzifikasi"
ws1.cell(row=row, column=8).value = "Bobot Normal"
for c in range(1, 9):
    header_style(ws1.cell(row=row, column=c))
header_style(ws1.cell(row=row, column=3), fill=orange_fill)
header_style(ws1.cell(row=row, column=8), fill=green_fill)
row += 1

bobot_start = row

# Default weights
default_weights = [8, 7, 7, 8, 6, 6]

for i in range(n_kriteria):
    ws1.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws1.cell(row=row, column=1))
    ws1.cell(row=row, column=1).fill = light_blue_fill
    
    ws1.cell(row=row, column=2).value = kriteria_full[i]
    cell_style(ws1.cell(row=row, column=2))
    
    # INPUT (1-10)
    ws1.cell(row=row, column=3).value = default_weights[i]
    input_style(ws1.cell(row=row, column=3))
    
    input_ref = f"C{row}"
    
    # l - lookup from TFN table
    # Using simple IF chain instead of VLOOKUP
    l_cell = ws1.cell(row=row, column=4)
    l_cell.value = f"=({input_ref}-1)*0.1"
    cell_style(l_cell)
    l_cell.fill = light_green_fill
    l_cell.number_format = '0.00'
    
    # m
    m_cell = ws1.cell(row=row, column=5)
    m_cell.value = f"={input_ref}*0.1"
    cell_style(m_cell)
    m_cell.fill = light_green_fill
    m_cell.number_format = '0.00'
    
    # u
    u_cell = ws1.cell(row=row, column=6)
    u_cell.value = f"=MIN(1,({input_ref}+1)*0.1)"
    cell_style(u_cell)
    u_cell.fill = light_green_fill
    u_cell.number_format = '0.00'
    
    # Defuzzifikasi (Center of Area): (l + m + u) / 3
    defuzz = ws1.cell(row=row, column=7)
    defuzz.value = f"=(D{row}+E{row}+F{row})/3"
    cell_style(defuzz)
    defuzz.fill = light_green_fill
    defuzz.number_format = '0.0000'
    
    row += 1

bobot_end = row - 1

# Total Defuzzifikasi
ws1.cell(row=row, column=6).value = "Total:"
ws1.cell(row=row, column=6).font = Font(bold=True)
total_defuzz = ws1.cell(row=row, column=7)
total_defuzz.value = f"=SUM(G{bobot_start}:G{bobot_end})"
total_defuzz.font = Font(bold=True)
cell_style(total_defuzz)
total_defuzz.fill = yellow_fill
total_defuzz.number_format = '0.0000'
total_defuzz_ref = f"$G${row}"

# Bobot Normal = Defuzzifikasi / Total
for i in range(n_kriteria):
    r = bobot_start + i
    bobot_norm = ws1.cell(row=r, column=8)
    bobot_norm.value = f"=G{r}/{total_defuzz_ref}"
    cell_style(bobot_norm)
    bobot_norm.fill = light_green_fill
    bobot_norm.number_format = '0.0000'

# Total Bobot Normal (should be 1)
total_bobot = ws1.cell(row=row, column=8)
total_bobot.value = f"=SUM(H{bobot_start}:H{bobot_end})"
total_bobot.font = Font(bold=True)
cell_style(total_bobot)
total_bobot.fill = yellow_fill
total_bobot.number_format = '0.0000'

row += 2

# RINGKASAN BOBOT
section(ws1, row, "RINGKASAN BOBOT KRITERIA", col_end=4)
row += 1

ws1.cell(row=row, column=1).value = "Kriteria"
ws1.cell(row=row, column=2).value = "Bobot"
ws1.cell(row=row, column=3).value = "Persentase"
header_style(ws1.cell(row=row, column=1), fill=green_fill)
header_style(ws1.cell(row=row, column=2), fill=green_fill)
header_style(ws1.cell(row=row, column=3), fill=green_fill)
row += 1

ringkasan_start = row
for i in range(n_kriteria):
    ws1.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws1.cell(row=row, column=1))
    
    ws1.cell(row=row, column=2).value = f"=H{bobot_start + i}"
    cell_style(ws1.cell(row=row, column=2))
    ws1.cell(row=row, column=2).fill = light_green_fill
    ws1.cell(row=row, column=2).number_format = '0.0000'
    
    ws1.cell(row=row, column=3).value = f"=H{bobot_start + i}*100"
    cell_style(ws1.cell(row=row, column=3))
    ws1.cell(row=row, column=3).number_format = '0.00"%"'
    
    row += 1

# Column widths
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 10
ws1.column_dimensions['G'].width = 14
ws1.column_dimensions['H'].width = 14

# ============================================
# SHEET 2: DATA PESERTA
# ============================================
ws2 = wb.create_sheet("2. Data Peserta")

row = 1
ws2['A1'] = "DATA KEGIATAN & PESERTA"
ws2['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws2.merge_cells('A1:E1')

row = 3

# DATA KEGIATAN
section(ws2, row, "DATA KEGIATAN", col_end=4)
row += 1

ws2.cell(row=row, column=1).value = "Nama Kegiatan:"
ws2.cell(row=row, column=1).font = Font(bold=True)
ws2.cell(row=row, column=2).value = "Raimuna Daerah"
input_style(ws2.cell(row=row, column=2))
ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

ws2.cell(row=row, column=1).value = "Tanggal:"
ws2.cell(row=row, column=1).font = Font(bold=True)
ws2.cell(row=row, column=2).value = "25 Januari 2026"
input_style(ws2.cell(row=row, column=2))
ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

ws2.cell(row=row, column=1).value = "Kuota Peserta:"
ws2.cell(row=row, column=1).font = Font(bold=True)
kuota_cell = ws2.cell(row=row, column=2)
kuota_cell.value = 5
input_style(kuota_cell)
kuota_ref = f"B{row}"
row += 2

# DAFTAR PESERTA
section(ws2, row, "DAFTAR PESERTA (Input Nama di Kolom Kuning)", col_end=4)
row += 1

ws2.cell(row=row, column=1).value = "No"
ws2.cell(row=row, column=2).value = "Nama Peserta"
ws2.cell(row=row, column=3).value = "Asal/Kwartir"
header_style(ws2.cell(row=row, column=1))
header_style(ws2.cell(row=row, column=2))
header_style(ws2.cell(row=row, column=3))
row += 1

peserta_start = row
sample_peserta = [
    ("David Kulian", "Merauke"),
    ("Siti Aminah", "Jayapura"),
    ("Budi Santoso", "Sorong"),
    ("Dewi Lestari", "Manokwari"),
    ("Ahmad Fauzi", "Biak"),
    ("", ""),
    ("", ""),
    ("", ""),
    ("", ""),
    ("", "")
]

for i in range(n_peserta):
    ws2.cell(row=row, column=1).value = i + 1
    cell_style(ws2.cell(row=row, column=1))
    
    ws2.cell(row=row, column=2).value = sample_peserta[i][0]
    input_style(ws2.cell(row=row, column=2))
    
    ws2.cell(row=row, column=3).value = sample_peserta[i][1]
    input_style(ws2.cell(row=row, column=3))
    
    row += 1

peserta_end = row - 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 20

# ============================================
# SHEET 3: NILAI PESERTA
# ============================================
ws3 = wb.create_sheet("3. Nilai Peserta")

row = 1
ws3['A1'] = "INPUT NILAI PESERTA PER KRITERIA"
ws3['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws3.merge_cells('A1:I1')

ws3['A2'] = "Masukkan nilai 0-100 untuk setiap kriteria"
ws3['A2'].font = Font(italic=True, color="FF0000")
ws3.merge_cells('A2:I2')

row = 4

section(ws3, row, "NILAI PESERTA", col_end=n_kriteria+2)
row += 1

# Header
ws3.cell(row=row, column=1).value = "No"
ws3.cell(row=row, column=2).value = "Nama Peserta"
header_style(ws3.cell(row=row, column=1))
header_style(ws3.cell(row=row, column=2))

for j in range(n_kriteria):
    cell = ws3.cell(row=row, column=j+3)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

nilai_start = row

sample_nilai = [
    [85, 90, 88, 92, 87, 85],
    [78, 85, 82, 88, 80, 78],
    [92, 88, 90, 85, 88, 90],
    [80, 82, 85, 90, 82, 80],
    [88, 92, 87, 80, 90, 88],
]

for i in range(n_peserta):
    ws3.cell(row=row, column=1).value = i + 1
    cell_style(ws3.cell(row=row, column=1))
    
    # Reference nama dari Sheet 2
    ws3.cell(row=row, column=2).value = f"='2. Data Peserta'!B{peserta_start + i}"
    cell_style(ws3.cell(row=row, column=2))
    ws3.cell(row=row, column=2).fill = light_blue_fill
    
    for j in range(n_kriteria):
        cell = ws3.cell(row=row, column=j+3)
        if i < len(sample_nilai):
            cell.value = sample_nilai[i][j]
        else:
            cell.value = 0
        input_style(cell)
    
    row += 1

nilai_end = row - 1

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 25
for j in range(n_kriteria):
    ws3.column_dimensions[get_column_letter(j+3)].width = 12

# ============================================
# SHEET 4: PERHITUNGAN & RANKING
# ============================================
ws4 = wb.create_sheet("4. Hasil & Ranking")

row = 1
ws4['A1'] = "PERHITUNGAN SKOR AKHIR & RANKING"
ws4['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws4.merge_cells('A1:L1')

ws4['A2'] = "Skor Akhir = Sum(Nilai x Bobot)"
ws4['A2'].font = Font(italic=True, color="666666")
ws4.merge_cells('A2:L2')

row = 4

# BOBOT KRITERIA
section(ws4, row, "BOBOT KRITERIA (dari Sheet 1)", col_end=n_kriteria+1)
row += 1

for j in range(n_kriteria):
    cell = ws4.cell(row=row, column=j+1)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

bobot_ref_row = row
for j in range(n_kriteria):
    cell = ws4.cell(row=row, column=j+1)
    cell.value = f"='1. Bobot Kriteria'!H{bobot_start + j}"
    cell_style(cell)
    cell.fill = light_green_fill
    cell.number_format = '0.0000'
row += 2

# PERHITUNGAN SKOR
section(ws4, row, "PERHITUNGAN SKOR AKHIR", col_end=n_kriteria+4)
row += 1

ws4.cell(row=row, column=1).value = "No"
ws4.cell(row=row, column=2).value = "Nama Peserta"
header_style(ws4.cell(row=row, column=1))
header_style(ws4.cell(row=row, column=2))

for j in range(n_kriteria):
    cell = ws4.cell(row=row, column=j+3)
    cell.value = f"{kriteria_list[j]}xW"
    header_style(cell)

ws4.cell(row=row, column=n_kriteria+3).value = "Skor Akhir"
ws4.cell(row=row, column=n_kriteria+4).value = "Ranking"
header_style(ws4.cell(row=row, column=n_kriteria+3), fill=orange_fill)
header_style(ws4.cell(row=row, column=n_kriteria+4), fill=orange_fill)
row += 1

skor_start = row

for i in range(n_peserta):
    ws4.cell(row=row, column=1).value = i + 1
    cell_style(ws4.cell(row=row, column=1))
    
    ws4.cell(row=row, column=2).value = f"='2. Data Peserta'!B{peserta_start + i}"
    cell_style(ws4.cell(row=row, column=2))
    ws4.cell(row=row, column=2).fill = light_blue_fill
    
    for j in range(n_kriteria):
        cell = ws4.cell(row=row, column=j+3)
        nilai_ref = f"'3. Nilai Peserta'!{get_column_letter(j+3)}{nilai_start + i}"
        bobot_ref = f"${get_column_letter(j+1)}${bobot_ref_row}"
        cell.value = f"={nilai_ref}*{bobot_ref}"
        cell_style(cell)
        cell.fill = light_green_fill
        cell.number_format = '0.00'
    
    # Skor Akhir
    skor = ws4.cell(row=row, column=n_kriteria+3)
    skor.value = f"=SUM(C{row}:{get_column_letter(n_kriteria+2)}{row})"
    skor.font = Font(bold=True)
    cell_style(skor)
    skor.fill = light_yellow_fill
    skor.number_format = '0.00'
    
    # Ranking
    rank = ws4.cell(row=row, column=n_kriteria+4)
    skor_range = f"${get_column_letter(n_kriteria+3)}${skor_start}:${get_column_letter(n_kriteria+3)}${skor_start + n_peserta - 1}"
    rank.value = f'=IF(B{row}="","",RANK({get_column_letter(n_kriteria+3)}{row},{skor_range},0))'
    rank.font = Font(bold=True, size=12)
    cell_style(rank)
    rank.fill = yellow_fill
    
    row += 1

skor_end = row - 1
row += 2

# HASIL RANKING
section(ws4, row, "HASIL RANKING FINAL", col_end=5)
row += 1

ws4.cell(row=row, column=1).value = "Rank"
ws4.cell(row=row, column=2).value = "Nama Peserta"
ws4.cell(row=row, column=3).value = "Skor"
ws4.cell(row=row, column=4).value = "Status"
header_style(ws4.cell(row=row, column=1), fill=orange_fill)
header_style(ws4.cell(row=row, column=2), fill=orange_fill)
header_style(ws4.cell(row=row, column=3), fill=orange_fill)
header_style(ws4.cell(row=row, column=4), fill=orange_fill)
row += 1

for i in range(n_peserta):
    ws4.cell(row=row, column=1).value = i + 1
    cell_style(ws4.cell(row=row, column=1))
    ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    # Nama dengan ranking ini
    rank_col = get_column_letter(n_kriteria+4)
    skor_col = get_column_letter(n_kriteria+3)
    
    nama = ws4.cell(row=row, column=2)
    nama.value = f'=IFERROR(INDEX(B{skor_start}:B{skor_end},MATCH({i+1},{rank_col}{skor_start}:{rank_col}{skor_end},0)),"")'
    cell_style(nama)
    
    skor = ws4.cell(row=row, column=3)
    skor.value = f'=IFERROR(INDEX({skor_col}{skor_start}:{skor_col}{skor_end},MATCH({i+1},{rank_col}{skor_start}:{rank_col}{skor_end},0)),"")'
    cell_style(skor)
    skor.number_format = '0.00'
    
    # Status berdasarkan kuota
    status = ws4.cell(row=row, column=4)
    status.value = f'=IF(A{row}<="\'2. Data Peserta\'!{kuota_ref}","LOLOS","TIDAK LOLOS")'
    cell_style(status)
    
    row += 1

# Column widths
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 25
for j in range(n_kriteria):
    ws4.column_dimensions[get_column_letter(j+3)].width = 12
ws4.column_dimensions[get_column_letter(n_kriteria+3)].width = 14
ws4.column_dimensions[get_column_letter(n_kriteria+4)].width = 10

# ============================================
# SHEET 5: REFERENSI
# ============================================
ws5 = wb.create_sheet("5. Referensi")

row = 1
ws5['A1'] = "TABEL REFERENSI"
ws5['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws5.merge_cells('A1:E1')

row = 3

section(ws5, row, "SKALA BOBOT KRITERIA (1-10)", col_end=6)
row += 1

ws5.cell(row=row, column=1).value = "Skala"
ws5.cell(row=row, column=2).value = "Keterangan"
ws5.cell(row=row, column=3).value = "TFN (l)"
ws5.cell(row=row, column=4).value = "TFN (m)"
ws5.cell(row=row, column=5).value = "TFN (u)"
for c in range(1, 6):
    header_style(ws5.cell(row=row, column=c))
row += 1

for skala, ket, l, m, u in tfn_data:
    ws5.cell(row=row, column=1).value = skala
    ws5.cell(row=row, column=2).value = ket
    ws5.cell(row=row, column=3).value = l
    ws5.cell(row=row, column=4).value = m
    ws5.cell(row=row, column=5).value = u
    for c in range(1, 6):
        cell_style(ws5.cell(row=row, column=c))
    row += 1

row += 1

section(ws5, row, "RUMUS FUZZY AHP", col_end=6)
row += 1

ws5.cell(row=row, column=1).value = "1. Konversi ke TFN"
ws5.cell(row=row, column=2).value = "l = (skala-1)*0.1, m = skala*0.1, u = MIN(1,(skala+1)*0.1)"
ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
row += 1

ws5.cell(row=row, column=1).value = "2. Defuzzifikasi"
ws5.cell(row=row, column=2).value = "D = (l + m + u) / 3 (Center of Area Method)"
ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
row += 1

ws5.cell(row=row, column=1).value = "3. Normalisasi"
ws5.cell(row=row, column=2).value = "W = D / Sum(D)"
ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
row += 1

ws5.cell(row=row, column=1).value = "4. Skor Akhir"
ws5.cell(row=row, column=2).value = "S = Sum(Nilai x Bobot)"
ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
row += 1

ws5.cell(row=row, column=1).value = "5. Ranking"
ws5.cell(row=row, column=2).value = "Berdasarkan Skor Akhir (descending)"
ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 25
for c in ['C', 'D', 'E', 'F']:
    ws5.column_dimensions[c].width = 12

# Save
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_Simple_Input.xlsx"
wb.save(output_file)

print(f"[OK] File berhasil dibuat: {output_file}")
print(f"\n=== CARA PENGGUNAAN ===")
print(f"1. Sheet 1: Input bobot kriteria (1-10)")
print(f"   - Otomatis dikonversi ke TFN")
print(f"   - Otomatis di-defuzzifikasi")
print(f"   - Otomatis dinormalisasi")
print(f"\n2. Sheet 2: Input data kegiatan & peserta")
print(f"\n3. Sheet 3: Input nilai peserta (0-100)")
print(f"\n4. Sheet 4: Lihat hasil skor & ranking")
