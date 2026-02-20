"""
Excel Fuzzy AHP 3 Penilai - LENGKAP Langkah 1-6
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf=Font(bold=True,size=11,color="FFFFFF");nf=Font(size=10)
ca=Alignment(horizontal='center',vertical='center',wrap_text=True)
tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

hdr=PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
s1=PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
s2=PatternFill(start_color="16A34A",end_color="16A34A",fill_type="solid")
s3=PatternFill(start_color="CA8A04",end_color="CA8A04",fill_type="solid")
s4=PatternFill(start_color="9333EA",end_color="9333EA",fill_type="solid")
s5=PatternFill(start_color="D97706",end_color="D97706",fill_type="solid")
s6=PatternFill(start_color="0D9488",end_color="0D9488",fill_type="solid")
ylw=PatternFill(start_color="FFC000",end_color="FFC000",fill_type="solid")
lb=PatternFill(start_color="D6DCE4",end_color="D6DCE4",fill_type="solid")
lg=PatternFill(start_color="E2EFDA",end_color="E2EFDA",fill_type="solid")
ly=PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")
lp=PatternFill(start_color="E9D5FF",end_color="E9D5FF",fill_type="solid")
la=PatternFill(start_color="FEF3C7",end_color="FEF3C7",fill_type="solid")
lt=PatternFill(start_color="CCFBF1",end_color="CCFBF1",fill_type="solid")
inp=PatternFill(start_color="FFFFCC",end_color="FFFFCC",fill_type="solid")
p1f=PatternFill(start_color="DBEAFE",end_color="DBEAFE",fill_type="solid")
p2f=PatternFill(start_color="D1FAE5",end_color="D1FAE5",fill_type="solid")
p3f=PatternFill(start_color="FEE2E2",end_color="FEE2E2",fill_type="solid")
p1h=PatternFill(start_color="3B82F6",end_color="3B82F6",fill_type="solid")
p2h=PatternFill(start_color="10B981",end_color="10B981",fill_type="solid")
p3h=PatternFill(start_color="EF4444",end_color="EF4444",fill_type="solid")

def hs(c,f=hdr):c.font=hf;c.alignment=ca;c.border=tb;c.fill=f
def cs(c):c.font=nf;c.alignment=ca;c.border=tb
def ins(c):c.font=Font(size=11,color="0000FF");c.alignment=ca;c.border=tb;c.fill=inp
def sec(ws,r,num,txt,f,ce=10):
    c=ws.cell(r,1);c.value=f"LANGKAH {num}: {txt}";c.font=Font(bold=True,size=12,color="FFFFFF");c.fill=f;c.alignment=ca
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ce)
    for i in range(1,ce+1):ws.cell(r,i).border=tb;ws.cell(r,i).fill=f

K=["K1","K2","K3","K4","K5","K6"]
KN=["Status Keaktifan","Pencapaian SKU","Pencapaian SPG","Kesehatan Jasmani","Tes Wawancara","Tes Pilihan Ganda"]
n=6;NP=10
W=[1,2,2,3,4,5]

TFN={1:(1,1,1),2:(0.5,1,1.5),3:(1,1.5,2),4:(1.5,2,2.5),5:(2,2.5,3),6:(2.5,3,3.5),7:(3,3.5,4),8:(3.5,4,4.5),9:(4,4.5,4.5)}
TFNR={1:(1,1,1),2:(2/3,1,2),3:(0.5,2/3,1),4:(0.4,0.5,2/3),5:(1/3,0.4,0.5),6:(2/7,1/3,0.4),7:(0.25,2/7,1/3),8:(2/9,0.25,2/7),9:(2/9,2/9,0.25)}

def get_tfn(ratio):
    if ratio>=1:return TFN[max(1,min(9,round(ratio)))]
    else:return TFNR[max(1,min(9,round(1/ratio)))]

FMV=[]
for i in range(n):
    row=[]
    for j in range(n):
        row.append((1,1,1) if i==j else get_tfn(W[i]/W[j]))
    FMV.append(row)

# ======= Sheet 1: Input =======
ws1=wb.active;ws1.title="Input"
ws1['A1']="FUZZY AHP - 3 PENILAI";ws1['A1'].font=Font(bold=True,size=16);ws1.merge_cells('A1:D1')
r=3
ws1.cell(r,1).value="BOBOT KRITERIA";ws1.cell(r,1).font=Font(bold=True,size=12)
r+=1
for c,v in enumerate(["Kode","Kriteria","Bobot"],1):hs(ws1.cell(r,c));ws1.cell(r,c).value=v
r+=1;BS=r
for i in range(n):
    ws1.cell(r,1).value=K[i];cs(ws1.cell(r,1));ws1.cell(r,1).fill=lb
    ws1.cell(r,2).value=KN[i];cs(ws1.cell(r,2))
    ws1.cell(r,3).value=W[i];ins(ws1.cell(r,3));r+=1
r+=1
ws1.cell(r,1).value="PESERTA";ws1.cell(r,1).font=Font(bold=True,size=12)
r+=1
for c,v in enumerate(["No","Nama"],1):hs(ws1.cell(r,c));ws1.cell(r,c).value=v
r+=1;PS=r
PN=["David Kulian","Siti Aminah","Budi Santoso","Dewi Lestari","Ahmad Fauzi","Rina Wati","Joko Susilo","Maya Sari","Andi Pratama","Lina Marlina"]
for i in range(NP):ws1.cell(r,1).value=i+1;cs(ws1.cell(r,1));ws1.cell(r,2).value=PN[i];ins(ws1.cell(r,2));r+=1
ws1.column_dimensions['A'].width=12;ws1.column_dimensions['B'].width=25;ws1.column_dimensions['C'].width=10

# ======= Sheet 2-4: Penilai (masing-masing menilai semua K1-K6) =======
NPV={1:[[85,78,88,92,87,85],[78,82,82,88,80,78],[92,88,90,85,88,90],[80,85,85,90,82,80],
         [88,90,87,80,90,88],[75,80,80,85,85,82],[82,78,88,82,80,85],[90,85,85,88,88,92],
         [85,82,82,85,85,80],[78,80,85,90,82,85]],
     2:[[88,80,90,88,85,87],[82,85,88,85,82,80],[90,82,85,90,90,88],[85,88,90,85,80,82],
         [87,90,80,87,88,90],[80,78,85,80,82,85],[88,82,82,88,85,80],[85,90,88,85,92,88],
         [82,85,85,82,80,85],[85,82,90,85,85,82]],
     3:[[87,82,85,90,88,85],[80,80,78,82,78,80],[88,90,90,88,90,88],[82,85,80,88,80,82],
         [90,88,88,85,90,87],[85,82,82,78,82,80],[80,85,85,80,85,82],[88,88,92,90,88,90],
         [85,80,80,85,80,82],[82,85,85,82,85,80]]}
pfs=[p1f,p2f,p3f];phs=[p1h,p2h,p3h];NRS={}

for p in range(1,4):
    ws=wb.create_sheet(f"Penilai {p}")
    ws['A1']=f"INPUT PENILAI {p} (Semua Kriteria K1-K6)";ws['A1'].font=Font(bold=True,size=14,color="FFFFFF");ws['A1'].fill=phs[p-1]
    ws.merge_cells(f'A1:{get_column_letter(n+2)}1')
    r=3
    headers=["No","Nama"]+K
    for c,v in enumerate(headers,1):hs(ws.cell(r,c),phs[p-1]);ws.cell(r,c).value=v
    r+=1;NRS[p]=r
    for i in range(NP):
        ws.cell(r,1).value=i+1;cs(ws.cell(r,1))
        ws.cell(r,2).value=f"='Input'!B{PS+i}";cs(ws.cell(r,2));ws.cell(r,2).fill=pfs[p-1]
        for j in range(n):
            ws.cell(r,j+3).value=NPV[p][i][j];ins(ws.cell(r,j+3))
        r+=1
    ws.column_dimensions['A'].width=8;ws.column_dimensions['B'].width=20
    for j in range(n):ws.column_dimensions[get_column_letter(j+3)].width=12

# ======= Sheet 5: Rekap (rata-rata 3 penilai) =======
ws5=wb.create_sheet("Rekap Nilai")
ws5['A1']="REKAP NILAI (RATA-RATA 3 PENILAI)";ws5['A1'].font=Font(bold=True,size=14);ws5.merge_cells('A1:H1')
ws5['A2']="Nilai = AVERAGE(Penilai1, Penilai2, Penilai3) per kriteria";ws5['A2'].font=Font(italic=True,size=9);ws5.merge_cells('A2:H2')
r=4
ws5.cell(r,1).value="No";ws5.cell(r,2).value="Nama";hs(ws5.cell(r,1));hs(ws5.cell(r,2))
for j in range(n):c=ws5.cell(r,j+3);c.value=K[j];hs(c,hdr)
r+=1;RNS=r
for i in range(NP):
    ws5.cell(r,1).value=i+1;cs(ws5.cell(r,1))
    ws5.cell(r,2).value=f"='Input'!B{PS+i}";cs(ws5.cell(r,2));ws5.cell(r,2).fill=lb
    for j in range(n):
        col_letter=get_column_letter(j+3)
        p1=f"'Penilai 1'!{col_letter}{NRS[1]+i}"
        p2=f"'Penilai 2'!{col_letter}{NRS[2]+i}"
        p3=f"'Penilai 3'!{col_letter}{NRS[3]+i}"
        c=ws5.cell(r,j+3);c.value=f"=AVERAGE({p1},{p2},{p3})";cs(c);c.fill=lg;c.number_format='0.00'
    r+=1
ws5.column_dimensions['A'].width=8;ws5.column_dimensions['B'].width=20
for j in range(n):ws5.column_dimensions[get_column_letter(j+3)].width=10

# ======= Sheet 6: Langkah 1-3 =======
ws6=wb.create_sheet("Langkah 1-3")
ws6['A1']="LANGKAH 1-3: MATRIKS CRISP, EIGENVECTOR & UJI KONSISTENSI"
ws6['A1'].font=Font(bold=True,size=14);ws6.merge_cells('A1:K1')

r=3
# --- LANGKAH 1 ---
sec(ws6,r,1,"MATRIKS PERBANDINGAN BERPASANGAN (CRISP)",s1,n+1)
r+=1
ws6.cell(r,1).value="";hs(ws6.cell(r,1),s1)
for j in range(n):hs(ws6.cell(r,j+2),s1);ws6.cell(r,j+2).value=K[j]
r+=1;CRS=r
for i in range(n):
    ws6.cell(r,1).value=K[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lb
    for j in range(n):
        c=ws6.cell(r,j+2)
        if i==j:c.value=1
        else:
            bi=f"'Input'!$C${BS+i}";bj=f"'Input'!$C${BS+j}"
            c.value=f"=IF({bi}/{bj}>=1,MIN(9,MAX(1,{bi}/{bj})),MAX(1/9,{bi}/{bj}))"
        cs(c);c.number_format='0.00'
    r+=1
r+=1

# --- LANGKAH 2 ---
sec(ws6,r,2,"PERHITUNGAN VECTOR EIGEN (GEOMETRIC MEAN)",s2,n+1)
r+=1
ws6.cell(r,1).value="Kriteria";ws6.cell(r,2).value="Geometric Mean";ws6.cell(r,3).value="Eigenvector (Wi)"
for c in range(1,4):hs(ws6.cell(r,c),s2)
r+=1;GMS=r
for i in range(n):
    ws6.cell(r,1).value=KN[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lb
    cr=CRS+i
    ws6.cell(r,2).value=f"=(B{cr}*C{cr}*D{cr}*E{cr}*F{cr}*G{cr})^(1/6)"
    cs(ws6.cell(r,2));ws6.cell(r,2).fill=lg;ws6.cell(r,2).number_format='0.0000'
    r+=1
GME=r-1
# Total GM
ws6.cell(r,1).value="Total";ws6.cell(r,1).font=Font(bold=True);cs(ws6.cell(r,1))
GMT=r
ws6.cell(r,2).value=f"=SUM(B{GMS}:B{GME})";ws6.cell(r,2).font=Font(bold=True);cs(ws6.cell(r,2));ws6.cell(r,2).fill=ylw;ws6.cell(r,2).number_format='0.0000'
# Wi
for i in range(n):
    rr=GMS+i
    ws6.cell(rr,3).value=f"=B{rr}/$B${GMT}";cs(ws6.cell(rr,3));ws6.cell(rr,3).fill=lg;ws6.cell(rr,3).number_format='0.0000'
ws6.cell(r,3).value=f"=SUM(C{GMS}:C{GME})";ws6.cell(r,3).font=Font(bold=True);cs(ws6.cell(r,3));ws6.cell(r,3).fill=ylw
r+=2

# --- LANGKAH 3 ---
sec(ws6,r,3,"UJI KONSISTENSI",s3,n+1)
r+=1
ws6.cell(r,1).value="Kriteria";ws6.cell(r,2).value="A*w";ws6.cell(r,3).value="(A*w)/w"
for c in range(1,4):hs(ws6.cell(r,c),s3)
r+=1;AWS=r
for i in range(n):
    ws6.cell(r,1).value=KN[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lb
    cr=CRS+i
    parts=[f"{get_column_letter(j+2)}{cr}*$C${GMS+j}" for j in range(n)]
    ws6.cell(r,2).value="="+"+".join(parts);cs(ws6.cell(r,2));ws6.cell(r,2).fill=ly;ws6.cell(r,2).number_format='0.0000'
    ws6.cell(r,3).value=f"=B{r}/$C${GMS+i}";cs(ws6.cell(r,3));ws6.cell(r,3).fill=ly;ws6.cell(r,3).number_format='0.0000'
    r+=1
AWE=r-1;r+=1

LR=r
ws6.cell(r,1).value="Lambda Max:";ws6.cell(r,1).font=Font(bold=True,color="1F4E79")
ws6.cell(r,2).value=f"=AVERAGE(C{AWS}:C{AWE})";ws6.cell(r,2).font=Font(bold=True,size=14,color="16A34A")
cs(ws6.cell(r,2));ws6.cell(r,2).fill=ylw;ws6.cell(r,2).number_format='0.0000';r+=1

CIR=r
ws6.cell(r,1).value="CI:";ws6.cell(r,1).font=Font(bold=True,color="1F4E79")
ws6.cell(r,2).value=f"=(B{LR}-{n})/({n}-1)";cs(ws6.cell(r,2));ws6.cell(r,2).fill=ly;ws6.cell(r,2).number_format='0.0000';r+=1

RIR=r
ws6.cell(r,1).value="RI (n=6):";ws6.cell(r,1).font=Font(bold=True,color="1F4E79")
ws6.cell(r,2).value=1.24;cs(ws6.cell(r,2));r+=1

CRR=r
ws6.cell(r,1).value="CR:";ws6.cell(r,1).font=Font(bold=True,color="1F4E79")
ws6.cell(r,2).value=f"=B{CIR}/B{RIR}";ws6.cell(r,2).font=Font(bold=True,size=14);cs(ws6.cell(r,2));ws6.cell(r,2).fill=ylw;ws6.cell(r,2).number_format='0.0000';r+=1

ws6.cell(r,1).value="Status:";ws6.cell(r,1).font=Font(bold=True,color="1F4E79")
ws6.cell(r,2).value=f'=IF(B{CRR}<=0.1,"KONSISTEN","TIDAK KONSISTEN")';ws6.cell(r,2).font=Font(bold=True,size=12)
cs(ws6.cell(r,2));ws6.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)

ws6.column_dimensions['A'].width=25
for j in range(n+3):ws6.column_dimensions[get_column_letter(j+2)].width=14

# ======= Sheet 7: Langkah 4-6 =======
ws7=wb.create_sheet("Langkah 4-6")
ws7['A1']="LANGKAH 4-6: FUZZIFIKASI, SYNTHETIC EXTENT & BOBOT GLOBAL"
ws7['A1'].font=Font(bold=True,size=14);ws7.merge_cells('A1:T1')

r=3
# --- Table TFN Reference (Hidden/Side) ---
tfn_rows = []
for k, v in TFN.items():
    tfn_rows.append([k, v[0], v[1], v[2]])

# Write TFN table at X4
ws7.cell(3, 24).value = "Key"
ws7.cell(3, 25).value = "l"
ws7.cell(3, 26).value = "m"
ws7.cell(3, 27).value = "u"
for idx, row_data in enumerate(tfn_rows):
    curr_r = 4 + idx
    for c_idx, val in enumerate(row_data):
        ws7.cell(curr_r, 24 + c_idx).value = val

# --- LANGKAH 4 (FORMULA) ---
sec(ws7,r,4,"FUZZIFIKASI MATRIKS PERBANDINGAN BERPASANGAN",s4,n*3+1)
r+=1
ws7.cell(r,1).value="";hs(ws7.cell(r,1),s4)
col=2
for j in range(n):
    ws7.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+2)
    hs(ws7.cell(r,col),s4);ws7.cell(r,col).value=K[j];col+=3
r+=1
ws7.cell(r,1).value="";hs(ws7.cell(r,1),s4)
col=2
for _ in range(n):
    for t in ["l","m","u"]:c=ws7.cell(r,col);c.value=t;hs(c,lp);c.font=Font(bold=True,size=10,color="000000");col+=1
r+=1;FMS=r

for i in range(n):
    ws7.cell(r,1).value=K[i];cs(ws7.cell(r,1));ws7.cell(r,1).fill=lb
    col=2
    for j in range(n):
        # Referensi nilai crisp dari Langkah 1-3
        crisp_ref = f"'Langkah 1-3'!{get_column_letter(2+j)}${CRS+i}"
        
        # Formula TFN Lookup with Python-style banker's rounding (round half to even)
        # raw = IF(val>=1, val, 1/val)
        # banker_round = IF(MOD(raw,1)=0.5, 2*ROUND(raw/2,0), ROUND(raw,0))
        raw_val = f"IF({crisp_ref}>=1,{crisp_ref},1/{crisp_ref})"
        banker_round = f"IF(MOD({raw_val},1)=0.5,2*ROUND({raw_val}/2,0),ROUND({raw_val},0))"
        key_formula = f"MAX(1,MIN(9,{banker_round}))"
        
        l_look = f"VLOOKUP({key_formula},$X$4:$AA$12,2,FALSE)"
        m_look = f"VLOOKUP({key_formula},$X$4:$AA$12,3,FALSE)"
        u_look = f"VLOOKUP({key_formula},$X$4:$AA$12,4,FALSE)"
        
        # Jika val >= 1: TFN(val)
        # Jika val < 1: Reciprocal TFN(1/val) = (1/u, 1/m, 1/l)
        f_l = f"=IF({crisp_ref}>=1,{l_look},1/{u_look})"
        f_m = f"=IF({crisp_ref}>=1,{m_look},1/{m_look})"
        f_u = f"=IF({crisp_ref}>=1,{u_look},1/{l_look})"
        
        # Kolom l
        c=ws7.cell(r,col);c.value=f_l;cs(c);c.number_format='0.0000'
        if i!=j:c.fill=lp
        # Kolom m
        c=ws7.cell(r,col+1);c.value=f_m;cs(c);c.number_format='0.0000'
        if i!=j:c.fill=lp
        # Kolom u
        c=ws7.cell(r,col+2);c.value=f_u;cs(c);c.number_format='0.0000'
        if i!=j:c.fill=lp
        
        col+=3
    r+=1
r+=1

# --- LANGKAH 5 ---
sec(ws7,r,5,"FUZZY SYNTHETIC EXTENT",s5,n*3+1)
r+=1
ws7.cell(r,1).value="Penjumlahan Baris Matriks Fuzzy";ws7.cell(r,1).font=Font(bold=True,size=11)
r+=1
ws7.cell(r,1).value="Kriteria";ws7.cell(r,2).value="Σl";ws7.cell(r,3).value="Σm";ws7.cell(r,4).value="Σu"
for c in range(1,5):hs(ws7.cell(r,c),s5)
r+=1;RSS=r
for i in range(n):
    ws7.cell(r,1).value=KN[i];cs(ws7.cell(r,1));ws7.cell(r,1).fill=lb
    fr=FMS+i
    lc=[get_column_letter(2+j*3) for j in range(n)]
    mc=[get_column_letter(3+j*3) for j in range(n)]
    uc=[get_column_letter(4+j*3) for j in range(n)]
    ws7.cell(r,2).value=f"={'+'.join([f'{c}{fr}' for c in lc])}";cs(ws7.cell(r,2));ws7.cell(r,2).number_format='0.00';ws7.cell(r,2).fill=la
    ws7.cell(r,3).value=f"={'+'.join([f'{c}{fr}' for c in mc])}";cs(ws7.cell(r,3));ws7.cell(r,3).number_format='0.00';ws7.cell(r,3).fill=la
    ws7.cell(r,4).value=f"={'+'.join([f'{c}{fr}' for c in uc])}";cs(ws7.cell(r,4));ws7.cell(r,4).number_format='0.00';ws7.cell(r,4).fill=la
    r+=1
RSE=r-1
TR=r
ws7.cell(r,1).value="Total";ws7.cell(r,1).font=Font(bold=True);cs(ws7.cell(r,1))
for c in range(2,5):
    ws7.cell(r,c).value=f"=SUM({get_column_letter(c)}{RSS}:{get_column_letter(c)}{RSE})"
    ws7.cell(r,c).font=Font(bold=True);cs(ws7.cell(r,c));ws7.cell(r,c).fill=ylw;ws7.cell(r,c).number_format='0.00'
r+=2

ws7.cell(r,1).value="Fuzzy Synthetic Extent (Si)";ws7.cell(r,1).font=Font(bold=True,size=11)
r+=1
ws7.cell(r,1).value="Kriteria";ws7.cell(r,2).value="Si(l)";ws7.cell(r,3).value="Si(m)";ws7.cell(r,4).value="Si(u)"
for c in range(1,5):hs(ws7.cell(r,c),s5)
r+=1;SIS=r
for i in range(n):
    rr=RSS+i
    ws7.cell(r,1).value=KN[i];cs(ws7.cell(r,1));ws7.cell(r,1).fill=lb
    ws7.cell(r,2).value=f"=B{rr}/$D${TR}";cs(ws7.cell(r,2));ws7.cell(r,2).number_format='0.0000';ws7.cell(r,2).fill=la
    ws7.cell(r,3).value=f"=C{rr}/$C${TR}";cs(ws7.cell(r,3));ws7.cell(r,3).number_format='0.0000';ws7.cell(r,3).fill=la
    ws7.cell(r,4).value=f"=D{rr}/$B${TR}";cs(ws7.cell(r,4));ws7.cell(r,4).number_format='0.0000';ws7.cell(r,4).fill=la
    r+=1
r+=1

# --- LANGKAH 6 ---
sec(ws7,r,6,"PERBANDINGAN PROBABILITAS, NORMALISASI & BOBOT GLOBAL",s6,n*3+1)
r+=1
ws7.cell(r,1).value="V(Si>=Sk) = 1 jika mi>=mk | 0 jika lk>=ui | (lk-ui)/((mi-ui)-(mk-lk))"
ws7.cell(r,1).font=Font(italic=True,size=9)
ws7.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n+1)
r+=1

# --- Matriks V(Si >= Sk) ---
ws7.cell(r,1).value="Matriks V(Si >= Sk)"
ws7.cell(r,1).font=Font(bold=True,size=11,color="0D9488")
r+=1
ws7.cell(r,1).value="V(Si>=Sk)";hs(ws7.cell(r,1),s6)
for j in range(n):hs(ws7.cell(r,j+2),s6);ws7.cell(r,j+2).value=K[j]
r+=1;VMS=r
for i in range(n):
    ws7.cell(r,1).value=K[i];cs(ws7.cell(r,1));ws7.cell(r,1).fill=lb
    si_l=f"$B${SIS+i}";si_m=f"$C${SIS+i}";si_u=f"$D${SIS+i}"
    for j in range(n):
        c=ws7.cell(r,j+2)
        if i==j:
            c.value="-";cs(c);c.fill=lb
        else:
            sk_l=f"$B${SIS+j}";sk_m=f"$C${SIS+j}";sk_u=f"$D${SIS+j}"
            # V(Si >= Sk): m1=sk, m2=si
            # if si_m >= sk_m: 1
            # elif sk_l >= si_u: 0
            # else: (sk_l - si_u) / ((si_m - si_u) - (sk_m - sk_l))
            c.value=f"=IF({si_m}>={sk_m},1,IF({sk_l}>={si_u},0,({sk_l}-{si_u})/(({si_m}-{si_u})-({sk_m}-{sk_l}))))"
            cs(c);c.number_format='0.0000';c.fill=lt
    r+=1
r+=1

# --- d'(Ai) = min V(Si >= Sk) ---
ws7.cell(r,1).value="Nilai Ordinat d'(Ai) = min V(Si >= Sk)"
ws7.cell(r,1).font=Font(bold=True,size=11,color="0D9488")
r+=1
ws7.cell(r,1).value="Kriteria";ws7.cell(r,2).value="d'(Ai)"
hs(ws7.cell(r,1),s6);hs(ws7.cell(r,2),s6)
r+=1;DPS=r
for i in range(n):
    ws7.cell(r,1).value=KN[i];cs(ws7.cell(r,1));ws7.cell(r,1).fill=lb
    vr=VMS+i
    # MIN of all non-diagonal values in V matrix row
    vals=[f"{get_column_letter(j+2)}{vr}" for j in range(n) if j!=i]
    ws7.cell(r,2).value=f"=MIN({','.join(vals)})";cs(ws7.cell(r,2));ws7.cell(r,2).number_format='0.0000';ws7.cell(r,2).fill=lt
    r+=1
DPE=r-1
r+=1

# --- Bobot Global Ternormalisasi ---
ws7.cell(r,1).value="Bobot Global Ternormalisasi (Wi)"
ws7.cell(r,1).font=Font(bold=True,size=11,color="0D9488")
r+=1
ws7.cell(r,1).value="wi = d'(Ai) / Σ d'(Ai)"
ws7.cell(r,1).font=Font(italic=True,size=9)
r+=1
ws7.cell(r,1).value="Kriteria";ws7.cell(r,2).value="Bobot Global (Wi)"
hs(ws7.cell(r,1),s6);hs(ws7.cell(r,2),s6)
r+=1;WIS=r
DTR_formula=f"SUM(B{DPS}:B{DPE})"
for i in range(n):
    ws7.cell(r,1).value=KN[i];cs(ws7.cell(r,1));ws7.cell(r,1).fill=lb
    ws7.cell(r,2).value=f"=B{DPS+i}/{DTR_formula}";cs(ws7.cell(r,2));ws7.cell(r,2).number_format='0.0000';ws7.cell(r,2).fill=lt
    r+=1
WIE=r-1
ws7.cell(r,1).value="Total";ws7.cell(r,1).font=Font(bold=True);cs(ws7.cell(r,1))
ws7.cell(r,3).value="";  # placeholder
ws7.cell(r,2).value=f"=SUM(B{WIS}:B{WIE})";ws7.cell(r,2).font=Font(bold=True);cs(ws7.cell(r,2));ws7.cell(r,2).fill=ylw;ws7.cell(r,2).number_format='0.00'
# Keep DPS reference for Hasil sheet (bobot is now at WIS)
# Override DPS to WIS for bobot reference
DPS=WIS

ws7.column_dimensions['A'].width=25
for c in range(2,n*3+5):ws7.column_dimensions[get_column_letter(c)].width=8

# ======= Sheet 8: Hasil =======
ws8=wb.create_sheet("Hasil & Ranking")
ws8['A1']="SKOR AKHIR (DEFUZZIFIKASI)";ws8['A1'].font=Font(bold=True,size=16);ws8.merge_cells('A1:H1')
ws8['A2']="Skor = (ΣL×w + ΣM×w + ΣU×w) / 3";ws8['A2'].font=Font(italic=True);ws8.merge_cells('A2:H2')
ws8['A3']="Bobot = Wi dari Langkah 6 (Fuzzy AHP). Fuzzy: nilai<=5 → Likert TFN, nilai>5 → (nilai-5, nilai, nilai+5)"
ws8['A3'].font=Font(italic=True,size=9,color="666666");ws8.merge_cells('A3:H3')

r=5
ws8.cell(r,1).value="BOBOT (Wi dari Fuzzy AHP)";ws8.cell(r,1).font=Font(bold=True)
r+=1
for j in range(n):c=ws8.cell(r,j+1);c.value=K[j];hs(c,[p1h,p1h,p2h,p2h,p3h,p3h][j])
r+=1;WR=r
# Bobot = Wi dari Langkah 6 (Bobot Global Ternormalisasi)
for j in range(n):
    c=ws8.cell(r,j+1)
    c.value=f"='Langkah 4-6'!B{DPS+j}"
    cs(c);c.fill=lt;c.number_format='0.0000'
r+=2

ws8.cell(r,1).value="PERHITUNGAN SKOR";ws8.cell(r,1).font=Font(bold=True,size=12)
r+=1
for c,v in enumerate(["No","Nama","ΣL×w","ΣM×w","ΣU×w","Skor","Rank"],1):hs(ws8.cell(r,c));ws8.cell(r,c).value=v
r+=1;SRS=r

for i in range(NP):
    ws8.cell(r,1).value=i+1;cs(ws8.cell(r,1))
    ws8.cell(r,2).value=f"='Input'!B{PS+i}";cs(ws8.cell(r,2));ws8.cell(r,2).fill=lb
    Lp=[];Mp=[];Up=[]
    for j in range(n):
        nr=f"'Rekap Nilai'!{get_column_letter(j+3)}{RNS+i}"
        wr=f"${get_column_letter(j+1)}${WR}"
        # Fuzzifikasi sesuai sistem: auto-detect
        # Jika nilai<=5 (Likert): L=IF(<=1,1,IF(<=2,1,IF(<=3,2,IF(<=4,3,4))))
        #                         M=IF(<=1,1,IF(<=2,2,IF(<=3,3,IF(<=4,4,5)))), U=IF(<=1,2,IF(<=2,3,IF(<=3,4,IF(<=4,5,5))))
        # Jika nilai>5 (Numeric): L=MAX(0,nilai-5), M=nilai, U=MIN(100,nilai+5)
        L_f=f"IF({nr}<=5,IF({nr}<=1,1,IF({nr}<=2,1,IF({nr}<=3,2,IF({nr}<=4,3,4)))),MAX(0,{nr}-5))"
        M_f=f"IF({nr}<=5,IF({nr}<=1,1,IF({nr}<=2,2,IF({nr}<=3,3,IF({nr}<=4,4,5)))),{nr})"
        U_f=f"IF({nr}<=5,IF({nr}<=1,2,IF({nr}<=2,3,IF({nr}<=3,4,IF({nr}<=4,5,5)))),MIN(100,{nr}+5))"
        Lp.append(f"{L_f}*{wr}")
        Mp.append(f"{M_f}*{wr}")
        Up.append(f"{U_f}*{wr}")
    ws8.cell(r,3).value=f"={'+'.join(Lp)}";cs(ws8.cell(r,3));ws8.cell(r,3).number_format='0.00';ws8.cell(r,3).fill=lg
    ws8.cell(r,4).value=f"={'+'.join(Mp)}";cs(ws8.cell(r,4));ws8.cell(r,4).number_format='0.00';ws8.cell(r,4).fill=lg
    ws8.cell(r,5).value=f"={'+'.join(Up)}";cs(ws8.cell(r,5));ws8.cell(r,5).number_format='0.00';ws8.cell(r,5).fill=lg
    ws8.cell(r,6).value=f"=(C{r}+D{r}+E{r})/3";cs(ws8.cell(r,6));ws8.cell(r,6).font=Font(bold=True);ws8.cell(r,6).fill=ly;ws8.cell(r,6).number_format='0.00'
    sr=f"$F${SRS}:$F${SRS+NP-1}"
    ws8.cell(r,7).value=f'=IF(B{r}="","",RANK(F{r},{sr},0))';cs(ws8.cell(r,7));ws8.cell(r,7).font=Font(bold=True,size=12);ws8.cell(r,7).fill=ylw
    r+=1

for c in 'ABCDEFG':ws8.column_dimensions[c].width=14

out="d:/laragon/www/appSaringPramuka/Fuzzy_AHP_3_Penilai_v11.xlsx"
wb.save(out)
print(f"[OK] {out}")
print("Sheet: Input | Penilai 1-3 | Rekap Nilai | Langkah 1-3 | Langkah 4-6 | Hasil & Ranking")
