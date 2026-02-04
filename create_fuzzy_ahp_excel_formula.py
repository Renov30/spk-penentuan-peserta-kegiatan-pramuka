"""
Script untuk membuat Model Tabel Excel Perhitungan Fuzzy AHP dengan RUMUS EXCEL
VERSI 3 - Menggunakan rumus eksplisit (tanpa SUMPRODUCT) untuk kompatibilitas maksimal
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fuzzy AHP Calculator"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="1F4E79")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
input_font = Font(size=11, color="0000FF")
formula_font = Font(size=10, italic=True, color="808080")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Kriteria
kriteria = [
    "Status Keaktifan",
    "Pencapaian SKU", 
    "Pencapaian SPG",
    "Kesehatan Jasmani",
    "Tes Wawancara",
    "Tes Pilihan Ganda"
]
n = len(kriteria)

# Random Index values
ri_values = {1: 0, 2: 0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.46, 10: 1.49}

def apply_header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def apply_cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def apply_input_style(cell):
    cell.font = input_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = input_fill

def apply_section_header(ws, row, text, col_end=10):
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    for c in range(1, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border

current_row = 1

# ============================================
# HEADER
# ============================================
ws['A1'] = "MODEL PERHITUNGAN FUZZY AHP - DENGAN RUMUS EXCEL"
ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws.merge_cells('A1:J1')

ws['A2'] = "Sel berwarna KUNING adalah INPUT yang dapat diubah."
ws['A2'].font = Font(italic=True, size=10, color="FF0000")
ws.merge_cells('A2:J2')

ws['A4'] = "Nama Peserta:"
ws['B4'] = "David Kulian"
ws['B4'].font = Font(bold=True)
ws['B4'].fill = input_fill

current_row = 6

# ============================================
# MATRIKS PERBANDINGAN BERPASANGAN (CRISP) - INPUT
# ============================================
apply_section_header(ws, current_row, "LANGKAH 1: MATRIKS PERBANDINGAN BERPASANGAN (CRISP)")
current_row += 1

ws.cell(row=current_row, column=1).value = "Masukkan nilai perbandingan (1-9) di sel KUNING. Diagonal selalu 1."
ws.cell(row=current_row, column=1).font = Font(italic=True, size=9, color="666666")
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
current_row += 1

# Header row untuk matriks crisp
crisp_header_row = current_row
ws.cell(row=current_row, column=1).value = "Kriteria"
apply_header_style(ws.cell(row=current_row, column=1))

# Kolom kriteria: B, C, D, E, F, G (kolom 2-7)
krit_cols = []
for j in range(n):
    col = j + 2  # Kolom 2 sampai 7 (B-G)
    krit_cols.append(get_column_letter(col))
    cell = ws.cell(row=current_row, column=col)
    cell.value = kriteria[j]
    apply_header_style(cell)

# Kolom tambahan: GM (H), Wi (I)
gm_col = get_column_letter(n+2)  # H
wi_col = get_column_letter(n+3)  # I

ws.cell(row=current_row, column=n+2).value = "GM"
ws.cell(row=current_row, column=n+3).value = "Wi (Bobot)"
apply_header_style(ws.cell(row=current_row, column=n+2), fill=green_fill)
apply_header_style(ws.cell(row=current_row, column=n+3), fill=green_fill)
current_row += 1

crisp_data_start = current_row  # Row 9
print(f"crisp_data_start = {crisp_data_start}")

# Data rows matriks crisp
for i in range(n):
    cell = ws.cell(row=current_row, column=1)
    cell.value = kriteria[i]
    cell.font = Font(bold=True, size=10)
    cell.border = thin_border
    cell.fill = light_blue_fill
    cell.alignment = left_align
    
    for j in range(n):
        col = j + 2
        cell = ws.cell(row=current_row, column=col)
        
        if i == j:
            cell.value = 1
            apply_cell_style(cell)
        elif i < j:
            cell.value = 1
            apply_input_style(cell)
        else:
            upper_row = crisp_data_start + j
            upper_col = krit_cols[i]  # Kolom untuk kriteria i
            cell.value = f"=1/{upper_col}{upper_row}"
            apply_cell_style(cell)
            cell.fill = light_green_fill
    
    # GM = (B*C*D*E*F*G)^(1/6) - explicit multiplication
    gm_parts = []
    for j in range(n):
        gm_parts.append(f"{krit_cols[j]}{current_row}")
    gm_formula = f"=({'+'.join([f'{p}' for p in gm_parts]).replace('+', '*')})^(1/{n})"
    # Simpler: just multiply all
    gm_formula = f"=({krit_cols[0]}{current_row}*{krit_cols[1]}{current_row}*{krit_cols[2]}{current_row}*{krit_cols[3]}{current_row}*{krit_cols[4]}{current_row}*{krit_cols[5]}{current_row})^(1/6)"
    
    gm_cell = ws.cell(row=current_row, column=n+2)
    gm_cell.value = gm_formula
    apply_cell_style(gm_cell)
    gm_cell.fill = light_green_fill
    gm_cell.number_format = '0.0000'
    
    current_row += 1

crisp_data_end = current_row - 1  # Row 14
print(f"crisp_data_end = {crisp_data_end}")

# Total GM
ws.cell(row=current_row, column=n+1).value = "Total:"
ws.cell(row=current_row, column=n+1).font = Font(bold=True)

gm_sum_row = current_row
gm_sum_cell = ws.cell(row=current_row, column=n+2)
gm_sum_cell.value = f"=SUM({gm_col}{crisp_data_start}:{gm_col}{crisp_data_end})"
gm_sum_cell.font = Font(bold=True)
apply_cell_style(gm_sum_cell)
gm_sum_cell.fill = yellow_fill
gm_sum_cell.number_format = '0.0000'

gm_sum_ref = f"${gm_col}${gm_sum_row}"
print(f"gm_sum_ref = {gm_sum_ref}")

# Wi (Eigenvector) = GM / Total GM
for i in range(n):
    row = crisp_data_start + i
    wi_cell = ws.cell(row=row, column=n+3)
    wi_cell.value = f"={gm_col}{row}/{gm_sum_ref}"
    apply_cell_style(wi_cell)
    wi_cell.fill = light_green_fill
    wi_cell.number_format = '0.0000'

# Total Wi (should be 1)
wi_sum_cell = ws.cell(row=current_row, column=n+3)
wi_sum_cell.value = f"=SUM({wi_col}{crisp_data_start}:{wi_col}{crisp_data_end})"
wi_sum_cell.font = Font(bold=True)
apply_cell_style(wi_sum_cell)
wi_sum_cell.fill = yellow_fill
wi_sum_cell.number_format = '0.0000'

current_row += 2

# ============================================
# PERHITUNGAN LAMBDA MAX & KONSISTENSI
# ============================================
apply_section_header(ws, current_row, "LANGKAH 2 & 3: PERHITUNGAN LAMBDA MAX & UJI KONSISTENSI")
current_row += 1

ws.cell(row=current_row, column=1).value = "A*w = Matriks x Bobot (untuk setiap baris)"
ws.cell(row=current_row, column=1).font = subtitle_font
current_row += 1

# Header
ws.cell(row=current_row, column=1).value = "Kriteria"
ws.cell(row=current_row, column=2).value = "A*w"
ws.cell(row=current_row, column=3).value = "(A*w)/w"
apply_header_style(ws.cell(row=current_row, column=1))
apply_header_style(ws.cell(row=current_row, column=2))
apply_header_style(ws.cell(row=current_row, column=3))
current_row += 1

aw_start = current_row
print(f"aw_start = {aw_start}")

for i in range(n):
    ws.cell(row=current_row, column=1).value = kriteria[i]
    apply_cell_style(ws.cell(row=current_row, column=1))
    ws.cell(row=current_row, column=1).fill = light_blue_fill
    
    crisp_row = crisp_data_start + i
    
    # A*w = B*$I$9 + C*$I$10 + D*$I$11 + E*$I$12 + F*$I$13 + G*$I$14 (explicit)
    aw_parts = []
    for j in range(n):
        crisp_cell = f"{krit_cols[j]}{crisp_row}"
        wi_ref = f"${wi_col}${crisp_data_start + j}"
        aw_parts.append(f"{crisp_cell}*{wi_ref}")
    aw_formula = "=" + "+".join(aw_parts)
    
    aw_cell = ws.cell(row=current_row, column=2)
    aw_cell.value = aw_formula
    apply_cell_style(aw_cell)
    aw_cell.fill = light_green_fill
    aw_cell.number_format = '0.0000'
    
    # (A*w)/w
    wi_ref = f"${wi_col}${crisp_data_start + i}"
    ratio_cell = ws.cell(row=current_row, column=3)
    ratio_cell.value = f"=B{current_row}/{wi_ref}"
    apply_cell_style(ratio_cell)
    ratio_cell.fill = light_green_fill
    ratio_cell.number_format = '0.0000'
    
    current_row += 1

aw_end = current_row - 1
print(f"aw_end = {aw_end}")

current_row += 1

# Lambda Max
lambda_row = current_row
ws.cell(row=current_row, column=1).value = "Lambda Max (lmax):"
ws.cell(row=current_row, column=1).font = subtitle_font
lambda_cell = ws.cell(row=current_row, column=2)
lambda_cell.value = f"=AVERAGE(C{aw_start}:C{aw_end})"
lambda_cell.font = Font(bold=True, size=14, color="ED7D31")
lambda_cell.fill = yellow_fill
apply_cell_style(lambda_cell)
lambda_cell.number_format = '0.0000'
lambda_ref = f"$B${lambda_row}"

ws.cell(row=current_row, column=4).value = "= rata-rata (A*w)/w"
ws.cell(row=current_row, column=4).font = formula_font
current_row += 1

# CI
ci_row = current_row
ws.cell(row=current_row, column=1).value = "Consistency Index (CI):"
ws.cell(row=current_row, column=1).font = subtitle_font
ci_cell = ws.cell(row=current_row, column=2)
ci_cell.value = f"=({lambda_ref}-{n})/({n}-1)"
ci_cell.font = Font(bold=True, size=12)
ci_cell.fill = light_green_fill
apply_cell_style(ci_cell)
ci_cell.number_format = '0.0000'
ci_ref = f"$B${ci_row}"

ws.cell(row=current_row, column=4).value = f"= (lmax - {n}) / ({n} - 1)"
ws.cell(row=current_row, column=4).font = formula_font
current_row += 1

# RI
ri_row = current_row
ws.cell(row=current_row, column=1).value = "Random Index (RI):"
ws.cell(row=current_row, column=1).font = subtitle_font
ri_cell = ws.cell(row=current_row, column=2)
ri_cell.value = ri_values[n]
ri_cell.font = Font(bold=True, size=12)
apply_cell_style(ri_cell)
ri_ref = f"$B${ri_row}"

ws.cell(row=current_row, column=4).value = f"Untuk n={n}, RI = {ri_values[n]}"
ws.cell(row=current_row, column=4).font = formula_font
current_row += 1

# CR
cr_row = current_row
ws.cell(row=current_row, column=1).value = "Consistency Ratio (CR):"
ws.cell(row=current_row, column=1).font = subtitle_font
cr_cell = ws.cell(row=current_row, column=2)
cr_cell.value = f"=IF({ri_ref}=0,0,{ci_ref}/{ri_ref})"
cr_cell.font = Font(bold=True, size=14)
cr_cell.fill = yellow_fill
apply_cell_style(cr_cell)
cr_cell.number_format = '0.0000'
cr_ref = f"$B${cr_row}"

ws.cell(row=current_row, column=4).value = "= CI / RI"
ws.cell(row=current_row, column=4).font = formula_font
current_row += 1

# Status
ws.cell(row=current_row, column=1).value = "Status Konsistensi:"
ws.cell(row=current_row, column=1).font = subtitle_font
status_cell = ws.cell(row=current_row, column=2)
status_cell.value = f'=IF({cr_ref}<=0.1,"KONSISTEN","TIDAK KONSISTEN")'
status_cell.font = Font(bold=True, size=11)
apply_cell_style(status_cell)
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)

current_row += 2

# ============================================
# TABEL TFN REFERENSI
# ============================================
apply_section_header(ws, current_row, "TABEL REFERENSI SKALA FUZZY (TFN)")
current_row += 1

headers = ["Intensitas", "l", "m", "u", "", "Kebalikan", "1/u", "1/m", "1/l"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=i)
    cell.value = h
    apply_header_style(cell)
current_row += 1

tfn_start = current_row
tfn_data = [
    (1, 1, 1, 1),
    (2, 0.5, 1, 1.5),
    (3, 1, 1.5, 2),
    (4, 1.5, 2, 2.5),
    (5, 2, 2.5, 3),
    (6, 2.5, 3, 3.5),
    (7, 3, 3.5, 4),
    (8, 3.5, 4, 4.5),
    (9, 4, 4.5, 4.5)
]

for intensity, l, m, u in tfn_data:
    ws.cell(row=current_row, column=1).value = intensity
    ws.cell(row=current_row, column=2).value = l
    ws.cell(row=current_row, column=3).value = m
    ws.cell(row=current_row, column=4).value = u
    
    ws.cell(row=current_row, column=6).value = intensity
    ws.cell(row=current_row, column=7).value = round(1/u, 4)
    ws.cell(row=current_row, column=8).value = round(1/m, 4)
    ws.cell(row=current_row, column=9).value = round(1/l, 4)
    
    for c in range(1, 10):
        apply_cell_style(ws.cell(row=current_row, column=c))
    current_row += 1

tfn_end = current_row - 1
print(f"tfn_start = {tfn_start}, tfn_end = {tfn_end}")

current_row += 1

# ============================================
# MATRIKS FUZZY (TFN)
# ============================================
apply_section_header(ws, current_row, "LANGKAH 4: FUZZIFIKASI MATRIKS (TFN)", col_end=20)
current_row += 1

ws.cell(row=current_row, column=1).value = "Konversi otomatis berdasarkan nilai crisp"
ws.cell(row=current_row, column=1).font = Font(italic=True, size=9, color="666666")
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
current_row += 1

# Header
ws.cell(row=current_row, column=1).value = "Kriteria"
apply_header_style(ws.cell(row=current_row, column=1))

col_idx = 2
for krit in kriteria:
    ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx+2)
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = krit[:10]
    apply_header_style(cell)
    col_idx += 3
current_row += 1

# Sub-header
ws.cell(row=current_row, column=1).value = ""
apply_header_style(ws.cell(row=current_row, column=1), fill=yellow_fill)
col_idx = 2
for _ in kriteria:
    for label in ["l", "m", "u"]:
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = label
        apply_header_style(cell, fill=yellow_fill)
        col_idx += 1
current_row += 1

fuzzy_start = current_row

# Data rows dengan formula INDEX/MATCH
for i in range(n):
    cell = ws.cell(row=current_row, column=1)
    cell.value = kriteria[i]
    cell.font = Font(bold=True, size=9)
    cell.border = thin_border
    cell.fill = light_blue_fill
    
    col_idx = 2
    for j in range(n):
        crisp_row = crisp_data_start + i
        crisp_col = krit_cols[j]
        crisp_ref = f"{crisp_col}{crisp_row}"
        
        if i <= j:
            # Upper triangle - TFN langsung
            l_formula = f"=INDEX($B${tfn_start}:$B${tfn_end},MATCH(MAX(1,MIN(9,ROUND(ABS({crisp_ref}),0))),$A${tfn_start}:$A${tfn_end},0))"
            m_formula = f"=INDEX($C${tfn_start}:$C${tfn_end},MATCH(MAX(1,MIN(9,ROUND(ABS({crisp_ref}),0))),$A${tfn_start}:$A${tfn_end},0))"
            u_formula = f"=INDEX($D${tfn_start}:$D${tfn_end},MATCH(MAX(1,MIN(9,ROUND(ABS({crisp_ref}),0))),$A${tfn_start}:$A${tfn_end},0))"
        else:
            # Lower triangle - kebalikan
            upper_crisp_row = crisp_data_start + j
            upper_crisp_col = krit_cols[i]
            upper_ref = f"{upper_crisp_col}{upper_crisp_row}"
            
            l_formula = f"=1/INDEX($D${tfn_start}:$D${tfn_end},MATCH(MAX(1,MIN(9,ROUND(ABS({upper_ref}),0))),$A${tfn_start}:$A${tfn_end},0))"
            m_formula = f"=1/INDEX($C${tfn_start}:$C${tfn_end},MATCH(MAX(1,MIN(9,ROUND(ABS({upper_ref}),0))),$A${tfn_start}:$A${tfn_end},0))"
            u_formula = f"=1/INDEX($B${tfn_start}:$B${tfn_end},MATCH(MAX(1,MIN(9,ROUND(ABS({upper_ref}),0))),$A${tfn_start}:$A${tfn_end},0))"
        
        ws.cell(row=current_row, column=col_idx).value = l_formula
        ws.cell(row=current_row, column=col_idx+1).value = m_formula
        ws.cell(row=current_row, column=col_idx+2).value = u_formula
        
        for c in range(col_idx, col_idx+3):
            apply_cell_style(ws.cell(row=current_row, column=c))
            ws.cell(row=current_row, column=c).number_format = '0.00'
            if i > j:
                ws.cell(row=current_row, column=c).fill = light_green_fill
        
        col_idx += 3
    
    current_row += 1

fuzzy_end = current_row - 1
current_row += 2

# ============================================
# RINGKASAN BOBOT
# ============================================
apply_section_header(ws, current_row, "RINGKASAN: BOBOT KRITERIA")
current_row += 1

ws.cell(row=current_row, column=1).value = "No"
ws.cell(row=current_row, column=2).value = "Kriteria"
ws.cell(row=current_row, column=3).value = "Bobot"
ws.cell(row=current_row, column=4).value = "Persentase"
apply_header_style(ws.cell(row=current_row, column=1), fill=green_fill)
apply_header_style(ws.cell(row=current_row, column=2), fill=green_fill)
apply_header_style(ws.cell(row=current_row, column=3), fill=green_fill)
apply_header_style(ws.cell(row=current_row, column=4), fill=green_fill)
current_row += 1

bobot_start = current_row
for i in range(n):
    ws.cell(row=current_row, column=1).value = i + 1
    apply_cell_style(ws.cell(row=current_row, column=1))
    
    ws.cell(row=current_row, column=2).value = kriteria[i]
    apply_cell_style(ws.cell(row=current_row, column=2))
    
    wi_ref = f"{wi_col}{crisp_data_start + i}"
    ws.cell(row=current_row, column=3).value = f"={wi_ref}"
    apply_cell_style(ws.cell(row=current_row, column=3))
    ws.cell(row=current_row, column=3).fill = light_green_fill
    ws.cell(row=current_row, column=3).number_format = '0.0000'
    
    ws.cell(row=current_row, column=4).value = f"={wi_ref}*100"
    apply_cell_style(ws.cell(row=current_row, column=4))
    ws.cell(row=current_row, column=4).fill = light_yellow_fill
    ws.cell(row=current_row, column=4).number_format = '0.00"%"'
    
    current_row += 1

bobot_end = current_row - 1

ws.cell(row=current_row, column=2).value = "TOTAL"
ws.cell(row=current_row, column=2).font = Font(bold=True)
apply_cell_style(ws.cell(row=current_row, column=2))

ws.cell(row=current_row, column=3).value = f"=SUM(C{bobot_start}:C{bobot_end})"
ws.cell(row=current_row, column=3).font = Font(bold=True)
apply_cell_style(ws.cell(row=current_row, column=3))
ws.cell(row=current_row, column=3).fill = yellow_fill

ws.cell(row=current_row, column=4).value = f"=SUM(D{bobot_start}:D{bobot_end})"
ws.cell(row=current_row, column=4).font = Font(bold=True)
apply_cell_style(ws.cell(row=current_row, column=4))
ws.cell(row=current_row, column=4).fill = yellow_fill

current_row += 2

# Status
ws.cell(row=current_row, column=1).value = "STATUS:"
ws.cell(row=current_row, column=1).font = subtitle_font
ws.cell(row=current_row, column=2).value = f'=IF({cr_ref}<=0.1,"VALID - Dapat digunakan","TIDAK VALID")'
ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

# ============================================
# Set column widths
# ============================================
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 14
for j in range(10, 25):
    ws.column_dimensions[get_column_letter(j)].width = 7

# Print summary
print("\n=== POSISI SEL PENTING ===")
print(f"Matriks Crisp: B{crisp_data_start}:G{crisp_data_end}")
print(f"GM Column: {gm_col}{crisp_data_start}:{gm_col}{crisp_data_end}")
print(f"Wi Column: {wi_col}{crisp_data_start}:{wi_col}{crisp_data_end}")
print(f"GM Sum: {gm_sum_ref}")
print(f"A*w: B{aw_start}:B{aw_end}")
print(f"(A*w)/w: C{aw_start}:C{aw_end}")
print(f"Lambda Max: {lambda_ref}")
print(f"CI: {ci_ref}")
print(f"RI: {ri_ref}")
print(f"CR: {cr_ref}")
print(f"TFN Table: A{tfn_start}:D{tfn_end}")

# Save
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_Calculator_v3.xlsx"
wb.save(output_file)
print(f"\n[OK] File berhasil dibuat: {output_file}")
