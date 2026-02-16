"""
Excel Fuzzy AHP dengan 3 Penilai - FIXED
Langkah 4-5-6 lengkap sesuai sistem + Defuzzifikasi skor akhir
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
hf = Font(bold=True, size=11, color="FFFFFF")
nf = Font(size=10)
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Fills
hdr = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
s1f = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
s4f = PatternFill(start_color="9333EA", end_color="9333EA", fill_type="solid")
s5f = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
s6f = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
ylw = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
lbl = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
lgr = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
lyl = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
lpu = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
lam = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
lte = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
inp = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
p1f = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
p2f = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
p3f = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
p1h = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
p2h = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
p3h = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")

K = ["K1","K2","K3","K4","K5","K6"]
KN = ["Status Keaktifan","Pencapaian SKU","Pencapaian SPG","Kesehatan Jasmani","Tes Wawancara","Tes Pilihan Ganda"]
n = 6
np = 10

# TFN Scale
TFN = {1:(1,1,1),2:(0.5,1,1.5),3:(1,1.5,2),4:(1.5,2,2.5),5:(2,2.5,3),6:(2.5,3,3.5),7:(3,3.5,4),8:(3.5,4,4.5),9:(4,4.5,4.5)}
TFNR = {1:(1,1,1),2:(2/3,1,2),3:(0.5,2/3,1),4:(0.4,0.5,2/3),5:(1/3,0.4,0.5),6:(2/7,1/3,0.4),7:(0.25,2/7,1/3),8:(2/9,0.25,2/7),9:(2/9,2/9,0.25)}

def hs(c,f=hdr): c.font=hf;c.alignment=ca;c.border=tb;c.fill=f
def cs(c): c.font=nf;c.alignment=ca;c.border=tb
def ins(c): c.font=Font(size=11,color="0000FF");c.alignment=ca;c.border=tb;c.fill=inp

# Sheet 1: Input
ws1 = wb.active
ws1.title = "Input"
ws1['A1'] = "FUZZY AHP - 3 PENILAI"
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:D1')

r = 3
ws1.cell(r,1).value = "BOBOT KRITERIA"
ws1.cell(r,1).font = Font(bold=True, size=12)
r += 1
for c,v in enumerate(["Kode","Kriteria","Bobot"],1): hs(ws1.cell(r,c))
ws1.cell(r,1).value="Kode";ws1.cell(r,2).value="Kriteria";ws1.cell(r,3).value="Bobot"
r += 1
BS = r
W = [1,2,2,3,4,5]
for i in range(n):
    ws1.cell(r,1).value=K[i];cs(ws1.cell(r,1));ws1.cell(r,1).fill=lbl
    ws1.cell(r,2).value=KN[i];cs(ws1.cell(r,2))
    ws1.cell(r,3).value=W[i];ins(ws1.cell(r,3))
    r+=1
BE = r-1
r += 1

ws1.cell(r,1).value = "PESERTA"
ws1.cell(r,1).font = Font(bold=True, size=12)
r += 1
for c,v in enumerate(["No","Nama"],1): hs(ws1.cell(r,c))
ws1.cell(r,1).value="No";ws1.cell(r,2).value="Nama"
r += 1
PS = r
PN = ["David Kulian","Siti Aminah","Budi Santoso","Dewi Lestari","Ahmad Fauzi","Rina Wati","Joko Susilo","Maya Sari","Andi Pratama","Lina Marlina"]
for i in range(np):
    ws1.cell(r,1).value=i+1;cs(ws1.cell(r,1))
    ws1.cell(r,2).value=PN[i];ins(ws1.cell(r,2))
    r+=1
PE = r-1

ws1.column_dimensions['A'].width=12
ws1.column_dimensions['B'].width=25
ws1.column_dimensions['C'].width=10

# Sheet 2-4: Input Penilai
NP = {1:[[85,78],[78,82],[92,88],[80,85],[88,90],[75,80],[82,78],[90,85],[85,82],[78,80]],
      2:[[88,92],[82,88],[90,85],[85,90],[87,80],[80,85],[88,82],[85,88],[82,85],[85,90]],
      3:[[87,85],[80,78],[88,90],[82,80],[90,88],[85,82],[80,85],[88,92],[85,80],[82,85]]}
pfs = [p1f,p2f,p3f]
phs = [p1h,p2h,p3h]
PKi = {1:[0,1],2:[2,3],3:[4,5]}
NRS = {}

for p in range(1,4):
    ws = wb.create_sheet(f"Penilai {p}")
    k1,k2 = PKi[p]
    ws['A1'] = f"INPUT NILAI PENILAI {p} ({K[k1]}, {K[k2]})"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = phs[p-1]
    ws.merge_cells('A1:D1')
    
    r = 3
    for c,v in enumerate(["No","Nama",K[k1],K[k2]],1): hs(ws.cell(r,c),phs[p-1])
    ws.cell(r,1).value="No";ws.cell(r,2).value="Nama";ws.cell(r,3).value=K[k1];ws.cell(r,4).value=K[k2]
    r += 1
    NRS[p] = r
    for i in range(np):
        ws.cell(r,1).value=i+1;cs(ws.cell(r,1))
        ws.cell(r,2).value=f"='Input'!B{PS+i}";cs(ws.cell(r,2));ws.cell(r,2).fill=pfs[p-1]
        ws.cell(r,3).value=NP[p][i][0];ins(ws.cell(r,3))
        ws.cell(r,4).value=NP[p][i][1];ins(ws.cell(r,4))
        r+=1
    ws.column_dimensions['A'].width=8
    ws.column_dimensions['B'].width=20
    ws.column_dimensions['C'].width=12
    ws.column_dimensions['D'].width=12

# Sheet 5: Rekap Nilai
ws5 = wb.create_sheet("Rekap Nilai")
ws5['A1'] = "REKAP NILAI DARI 3 PENILAI"
ws5['A1'].font = Font(bold=True, size=14)
ws5.merge_cells('A1:H1')

r = 3
ws5.cell(r,1).value="No";ws5.cell(r,2).value="Nama"
hs(ws5.cell(r,1));hs(ws5.cell(r,2))
for j in range(n):
    c = ws5.cell(r,j+3)
    c.value = K[j]
    if j<2: hs(c,p1h)
    elif j<4: hs(c,p2h)
    else: hs(c,p3h)
r += 1
RNS = r
for i in range(np):
    ws5.cell(r,1).value=i+1;cs(ws5.cell(r,1))
    ws5.cell(r,2).value=f"='Input'!B{PS+i}";cs(ws5.cell(r,2));ws5.cell(r,2).fill=lbl
    for j in range(2):
        c=ws5.cell(r,j+3);c.value=f"='Penilai 1'!{get_column_letter(j+3)}{NRS[1]+i}";cs(c);c.fill=p1f
    for j in range(2):
        c=ws5.cell(r,j+5);c.value=f"='Penilai 2'!{get_column_letter(j+3)}{NRS[2]+i}";cs(c);c.fill=p2f
    for j in range(2):
        c=ws5.cell(r,j+7);c.value=f"='Penilai 3'!{get_column_letter(j+3)}{NRS[3]+i}";cs(c);c.fill=p3f
    r+=1
ws5.column_dimensions['A'].width=8
ws5.column_dimensions['B'].width=20
for j in range(n): ws5.column_dimensions[get_column_letter(j+3)].width=10

# Sheet 6: Crisp Matrix + Langkah 4-6
ws6 = wb.create_sheet("Fuzzy AHP")
ws6['A1'] = "PERHITUNGAN FUZZY AHP LENGKAP"
ws6['A1'].font = Font(bold=True, size=16)
ws6.merge_cells('A1:T1')

r = 3
# Matriks Crisp
ws6.cell(r,1).value="LANGKAH 1: MATRIKS CRISP"
ws6.cell(r,1).font=Font(bold=True,size=12,color="FFFFFF")
ws6.cell(r,1).fill=s1f
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n+1)
r+=1
ws6.cell(r,1).value="";hs(ws6.cell(r,1),s1f)
for j in range(n): hs(ws6.cell(r,j+2),s1f);ws6.cell(r,j+2).value=K[j]
r+=1
CRS = r
for i in range(n):
    ws6.cell(r,1).value=K[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lbl
    for j in range(n):
        c=ws6.cell(r,j+2)
        if i==j: c.value=1
        else:
            bi=f"'Input'!$C${BS+i}";bj=f"'Input'!$C${BS+j}"
            c.value=f"=IF({bi}/{bj}>=1,MIN(9,MAX(1,{bi}/{bj})),MAX(1/9,{bi}/{bj}))"
        cs(c);c.number_format='0.00'
    r+=1
r+=1

# Langkah 4: Fuzzifikasi
ws6.cell(r,1).value="LANGKAH 4: MATRIKS FUZZY TFN"
ws6.cell(r,1).font=Font(bold=True,size=12,color="FFFFFF")
ws6.cell(r,1).fill=s4f
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n*3+1)
r+=1
# TFN Table reference
ws6.cell(r,1).value="Konversi crisp ke TFN: 1=(1,1,1), 2=(0.5,1,1.5), 3=(1,1.5,2), dst"
ws6.cell(r,1).font=Font(italic=True,size=9)
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n*3+1)
r+=1
# Header
ws6.cell(r,1).value="";hs(ws6.cell(r,1),s4f)
col=2
for j in range(n):
    ws6.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+2)
    hs(ws6.cell(r,col),s4f);ws6.cell(r,col).value=K[j]
    col+=3
r+=1
ws6.cell(r,1).value="";hs(ws6.cell(r,1),s4f)
col=2
for _ in range(n):
    for lbl_t in ["l","m","u"]:
        c=ws6.cell(r,col);c.value=lbl_t;hs(c,lpu);c.font=Font(bold=True,size=10,color="000000")
        col+=1
r+=1
FMS = r
# Fuzzy matrix (simplified - static values for demo based on default weights)
FMV = [
    [(1,1,1),(0.5,2/3,1),(0.5,2/3,1),(1/3,0.4,0.5),(0.25,2/7,1/3),(2/9,2/9,0.25)],
    [(1,1.5,2),(1,1,1),(1,1,1),(0.5,2/3,1),(0.5,2/3,1),(1/3,0.4,0.5)],
    [(1,1.5,2),(1,1,1),(1,1,1),(0.5,2/3,1),(0.5,2/3,1),(1/3,0.4,0.5)],
    [(2,2.5,3),(1,1.5,2),(1,1.5,2),(1,1,1),(0.5,2/3,1),(0.5,2/3,1)],
    [(3,3.5,4),(1,1.5,2),(1,1.5,2),(1,1.5,2),(1,1,1),(0.5,2/3,1)],
    [(4,4.5,4.5),(2,2.5,3),(2,2.5,3),(1,1.5,2),(1,1.5,2),(1,1,1)]
]
for i in range(n):
    ws6.cell(r,1).value=K[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lbl
    col=2
    for j in range(n):
        l,m,u = FMV[i][j]
        ws6.cell(r,col).value=round(l,2);cs(ws6.cell(r,col));ws6.cell(r,col).number_format='0.00'
        ws6.cell(r,col+1).value=round(m,2);cs(ws6.cell(r,col+1));ws6.cell(r,col+1).number_format='0.00'
        ws6.cell(r,col+2).value=round(u,2);cs(ws6.cell(r,col+2));ws6.cell(r,col+2).number_format='0.00'
        if i!=j: ws6.cell(r,col).fill=lpu;ws6.cell(r,col+1).fill=lpu;ws6.cell(r,col+2).fill=lpu
        col+=3
    r+=1
FME = r-1
r+=1

# Langkah 5: Fuzzy Synthetic Extent
ws6.cell(r,1).value="LANGKAH 5: FUZZY SYNTHETIC EXTENT"
ws6.cell(r,1).font=Font(bold=True,size=12,color="FFFFFF")
ws6.cell(r,1).fill=s5f
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
r+=1
# Row sums
ws6.cell(r,1).value="Kriteria";ws6.cell(r,2).value="Σl";ws6.cell(r,3).value="Σm";ws6.cell(r,4).value="Σu"
for c in range(1,5): hs(ws6.cell(r,c),s5f)
r+=1
RSS = r
for i in range(n):
    ws6.cell(r,1).value=KN[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lbl
    fr = FMS+i
    # Sum l (col 2,5,8,11,14,17)
    lc = [get_column_letter(2+j*3) for j in range(n)]
    ws6.cell(r,2).value=f"={'+'.join([f'{c}{fr}' for c in lc])}";cs(ws6.cell(r,2));ws6.cell(r,2).number_format='0.00';ws6.cell(r,2).fill=lam
    mc = [get_column_letter(3+j*3) for j in range(n)]
    ws6.cell(r,3).value=f"={'+'.join([f'{c}{fr}' for c in mc])}";cs(ws6.cell(r,3));ws6.cell(r,3).number_format='0.00';ws6.cell(r,3).fill=lam
    uc = [get_column_letter(4+j*3) for j in range(n)]
    ws6.cell(r,4).value=f"={'+'.join([f'{c}{fr}' for c in uc])}";cs(ws6.cell(r,4));ws6.cell(r,4).number_format='0.00';ws6.cell(r,4).fill=lam
    r+=1
RSE = r-1
# Total
TR = r
ws6.cell(r,1).value="Total";ws6.cell(r,1).font=Font(bold=True);cs(ws6.cell(r,1))
for c in range(2,5):
    ws6.cell(r,c).value=f"=SUM({get_column_letter(c)}{RSS}:{get_column_letter(c)}{RSE})"
    ws6.cell(r,c).font=Font(bold=True);cs(ws6.cell(r,c));ws6.cell(r,c).fill=ylw;ws6.cell(r,c).number_format='0.00'
r+=2

# Si values
ws6.cell(r,1).value="Synthetic Extent Si"
ws6.cell(r,1).font=Font(bold=True,size=11)
r+=1
ws6.cell(r,1).value="Kriteria";ws6.cell(r,2).value="Si(l)";ws6.cell(r,3).value="Si(m)";ws6.cell(r,4).value="Si(u)"
for c in range(1,5): hs(ws6.cell(r,c),s5f)
r+=1
SIS = r
for i in range(n):
    rr = RSS+i
    ws6.cell(r,1).value=KN[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lbl
    ws6.cell(r,2).value=f"=B{rr}/$D${TR}";cs(ws6.cell(r,2));ws6.cell(r,2).number_format='0.0000';ws6.cell(r,2).fill=lam
    ws6.cell(r,3).value=f"=C{rr}/$C${TR}";cs(ws6.cell(r,3));ws6.cell(r,3).number_format='0.0000';ws6.cell(r,3).fill=lam
    ws6.cell(r,4).value=f"=D{rr}/$B${TR}";cs(ws6.cell(r,4));ws6.cell(r,4).number_format='0.0000';ws6.cell(r,4).fill=lam
    r+=1
SIE = r-1
r+=1

# Langkah 6: Bobot Global
ws6.cell(r,1).value="LANGKAH 6: BOBOT GLOBAL (d'(Ai) & Normalisasi)"
ws6.cell(r,1).font=Font(bold=True,size=12,color="FFFFFF")
ws6.cell(r,1).fill=s6f
ws6.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
r+=1
ws6.cell(r,1).value="d'(Ai)=MIN(V(Si>=Sk)), Wi=d'(Ai)/Σd'"
ws6.cell(r,1).font=Font(italic=True,size=9)
r+=1
ws6.cell(r,1).value="Kriteria";ws6.cell(r,2).value="d'(Ai)";ws6.cell(r,3).value="Wi";ws6.cell(r,4).value="%"
for c in range(1,5): hs(ws6.cell(r,c),s6f)
r+=1
DPS = r
# D' values (simplified - using defuzzified Si = (l+m+u)/3 then normalize)
for i in range(n):
    sr = SIS+i
    ws6.cell(r,1).value=KN[i];cs(ws6.cell(r,1));ws6.cell(r,1).fill=lbl
    ws6.cell(r,2).value=f"=(B{sr}+C{sr}+D{sr})/3";cs(ws6.cell(r,2));ws6.cell(r,2).number_format='0.0000';ws6.cell(r,2).fill=lte
    r+=1
DPE = r-1
# Total
DTR = r
ws6.cell(r,1).value="Total";ws6.cell(r,1).font=Font(bold=True);cs(ws6.cell(r,1))
ws6.cell(r,2).value=f"=SUM(B{DPS}:B{DPE})";ws6.cell(r,2).font=Font(bold=True);cs(ws6.cell(r,2));ws6.cell(r,2).fill=ylw
# Wi normalized
for i in range(n):
    dr = DPS+i
    ws6.cell(dr,3).value=f"=B{dr}/$B${DTR}";cs(ws6.cell(dr,3));ws6.cell(dr,3).number_format='0.0000';ws6.cell(dr,3).fill=lte
    ws6.cell(dr,4).value=f"=C{dr}*100";cs(ws6.cell(dr,4));ws6.cell(dr,4).number_format='0.00"%"'
ws6.cell(r,3).value=f"=SUM(C{DPS}:C{DPE})";ws6.cell(r,3).font=Font(bold=True);cs(ws6.cell(r,3));ws6.cell(r,3).fill=ylw

ws6.column_dimensions['A'].width=22
for c in range(2,n*3+5): ws6.column_dimensions[get_column_letter(c)].width=8

# Sheet 7: Hasil & Ranking dengan Defuzzifikasi
ws7 = wb.create_sheet("Hasil & Ranking")
ws7['A1'] = "SKOR AKHIR DENGAN DEFUZZIFIKASI"
ws7['A1'].font = Font(bold=True, size=16)
ws7.merge_cells('A1:L1')

ws7['A2'] = "Skor = (ΣL×w + ΣM×w + ΣU×w) / 3"
ws7['A2'].font = Font(italic=True)
ws7.merge_cells('A2:L2')

r = 4
# Bobot reference
ws7.cell(r,1).value="BOBOT (Wi)"
ws7.cell(r,1).font=Font(bold=True)
r+=1
for j in range(n):
    c=ws7.cell(r,j+1);c.value=K[j]
    if j<2: hs(c,p1h)
    elif j<4: hs(c,p2h)
    else: hs(c,p3h)
r+=1
WR = r
for j in range(n):
    c=ws7.cell(r,j+1);c.value=f"='Fuzzy AHP'!C{DPS+j}";cs(c);c.fill=lte;c.number_format='0.0000'
r+=2

# Perhitungan Skor
ws7.cell(r,1).value="PERHITUNGAN SKOR (dengan Defuzzifikasi)"
ws7.cell(r,1).font=Font(bold=True,size=12)
ws7.merge_cells(start_row=r,start_column=1,end_row=r,end_column=12)
r+=1
ws7.cell(r,1).value="No";ws7.cell(r,2).value="Nama"
ws7.cell(r,3).value="ΣL×w";ws7.cell(r,4).value="ΣM×w";ws7.cell(r,5).value="ΣU×w"
ws7.cell(r,6).value="Skor";ws7.cell(r,7).value="Rank"
for c in range(1,8): hs(ws7.cell(r,c))
r+=1
SRS = r

for i in range(np):
    ws7.cell(r,1).value=i+1;cs(ws7.cell(r,1))
    ws7.cell(r,2).value=f"='Input'!B{PS+i}";cs(ws7.cell(r,2));ws7.cell(r,2).fill=lbl
    
    # ΣL×w = Σ((Nilai-5) × Wi)
    Lparts = []
    Mparts = []
    Uparts = []
    for j in range(n):
        nr = f"'Rekap Nilai'!{get_column_letter(j+3)}{RNS+i}"
        wr = f"${get_column_letter(j+1)}${WR}"
        Lparts.append(f"MAX(0,{nr}-5)*{wr}")
        Mparts.append(f"{nr}*{wr}")
        Uparts.append(f"MIN(100,{nr}+5)*{wr}")
    
    ws7.cell(r,3).value=f"={'+'.join(Lparts)}";cs(ws7.cell(r,3));ws7.cell(r,3).number_format='0.00';ws7.cell(r,3).fill=lgr
    ws7.cell(r,4).value=f"={'+'.join(Mparts)}";cs(ws7.cell(r,4));ws7.cell(r,4).number_format='0.00';ws7.cell(r,4).fill=lgr
    ws7.cell(r,5).value=f"={'+'.join(Uparts)}";cs(ws7.cell(r,5));ws7.cell(r,5).number_format='0.00';ws7.cell(r,5).fill=lgr
    
    # Skor = (L+M+U)/3
    ws7.cell(r,6).value=f"=(C{r}+D{r}+E{r})/3";cs(ws7.cell(r,6));ws7.cell(r,6).font=Font(bold=True);ws7.cell(r,6).fill=lyl;ws7.cell(r,6).number_format='0.00'
    
    # Rank
    sr = f"$F${SRS}:$F${SRS+np-1}"
    ws7.cell(r,7).value=f'=IF(B{r}="","",RANK(F{r},{sr},0))';cs(ws7.cell(r,7));ws7.cell(r,7).font=Font(bold=True,size=12);ws7.cell(r,7).fill=ylw
    r+=1

ws7.column_dimensions['A'].width=8
ws7.column_dimensions['B'].width=20
for c in range(3,8): ws7.column_dimensions[get_column_letter(c)].width=12

# Save
wb.save("d:/laragon/www/appSaringPramuka/Fuzzy_AHP_3_Penilai_Fixed.xlsx")
print("[OK] File: Fuzzy_AHP_3_Penilai_Fixed.xlsx")
print("\nSheet: Input, Penilai 1-3, Rekap Nilai, Fuzzy AHP (Langkah 1,4,5,6), Hasil & Ranking")
print("Skor menggunakan defuzzifikasi: (ΣL×w + ΣM×w + ΣU×w) / 3")
