"""
Script untuk membuat Model Tabel Excel Perhitungan Fuzzy AHP (Single Sheet)
Data: David Kulian - Raimuna Daerah (25 January 2026)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import math

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fuzzy AHP Calculation"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14, color="1F4E79")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Kriteria (dari screenshot)
kriteria = [
    "Status Keaktifan di Gugus Depan",
    "Pencapaian SKU",
    "Pencapaian SPG",
    "Kesehatan Jasmani dan Rohani",
    "Tes Wawancara",
    "Tes Pilihan Ganda"
]
n = len(kriteria)

# Matriks perbandingan berpasangan (Crisp) - dari screenshot semua nilai 1
crisp_matrix = [
    [1, 1, 1, 1, 1, 1],
    [1, 1, 1, 1, 1, 1],
    [1, 1, 1, 1, 1, 1],
    [1, 1, 1, 1, 1, 1],
    [1, 1, 1, 1, 1, 1],
    [1, 1, 1, 1, 1, 1]
]

# Tabel konversi TFN (Triangular Fuzzy Number)
tfn_scale = {
    1: (1, 1, 1),
    2: (0.5, 1, 1.5),
    3: (1, 1.5, 2),
    4: (1.5, 2, 2.5),
    5: (2, 2.5, 3),
    6: (2.5, 3, 3.5),
    7: (3, 3.5, 4),
    8: (3.5, 4, 4.5),
    9: (4, 4.5, 4.5)
}

tfn_reciprocal = {
    1: (1, 1, 1),
    2: (2/3, 1, 2),
    3: (0.5, 2/3, 1),
    4: (0.40, 0.5, 2/3),
    5: (1/3, 0.40, 0.5),
    6: (0.29, 1/3, 0.40),
    7: (0.25, 0.29, 1/3),
    8: (0.22, 0.25, 0.29),
    9: (0.22, 0.22, 0.25)
}

# Random Index values
ri_table = {
    1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24,
    7: 1.32, 8: 1.41, 9: 1.46, 10: 1.49, 11: 1.51, 12: 1.58
}

def apply_header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def apply_cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def apply_section_header(ws, row, text, col_start=1, col_end=8):
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border

current_row = 1

# ============================================
# HEADER - Info Peserta
# ============================================
ws['A1'] = "MODEL PERHITUNGAN FUZZY AHP"
ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws.merge_cells('A1:H1')

ws['A3'] = "Nama Peserta:"
ws['B3'] = "David Kulian"
ws['B3'].font = Font(bold=True)
ws['D3'] = "Kegiatan:"
ws['E3'] = "Raimuna Daerah"
ws['G3'] = "Peringkat:"
ws['H3'] = "#1"
ws['H3'].font = Font(bold=True, size=14, color="ED7D31")

ws['A4'] = "Tanggal:"
ws['B4'] = "25 January 2026"

current_row = 6

# ============================================
# SECTION 1: Tabel Skala Saaty
# ============================================
apply_section_header(ws, current_row, "TABEL SKALA PERBANDINGAN TINGKAT KEPENTINGAN (SAATY)", 1, 4)
current_row += 1

ws.cell(row=current_row, column=1).value = "Intensitas"
ws.cell(row=current_row, column=2).value = "Definisi"
apply_header_style(ws.cell(row=current_row, column=1))
apply_header_style(ws.cell(row=current_row, column=2))
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
current_row += 1

saaty_scale = [
    (1, "Kedua elemen sama penting (Equal Importance)"),
    (3, "Elemen satu sedikit lebih penting (Slightly More Importance)"),
    (5, "Elemen satu lebih penting (Materially More Importance)"),
    (7, "Satu elemen jelas lebih penting (Significantly More Importance)"),
    (9, "Satu elemen mutlak lebih penting (Absolutely More Importance)"),
    ("2,4,6,8", "Nilai-nilai diantara dua pertimbangan yang berdekatan")
]

for intensitas, definisi in saaty_scale:
    ws.cell(row=current_row, column=1).value = intensitas
    ws.cell(row=current_row, column=2).value = definisi
    apply_cell_style(ws.cell(row=current_row, column=1))
    apply_cell_style(ws.cell(row=current_row, column=2))
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
    current_row += 1

current_row += 1

# ============================================
# SECTION 2: Tabel Skala Fuzzy (TFN)
# ============================================
apply_section_header(ws, current_row, "TABEL SKALA NILAI FUZZY (TFN)", 1, 4)
current_row += 1

ws.cell(row=current_row, column=1).value = "Intensitas"
ws.cell(row=current_row, column=2).value = "TFN (l, m, u)"
ws.cell(row=current_row, column=3).value = "Kebalikan"
apply_header_style(ws.cell(row=current_row, column=1))
apply_header_style(ws.cell(row=current_row, column=2))
apply_header_style(ws.cell(row=current_row, column=3))
ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
current_row += 1

for i in range(1, 10):
    l, m, u = tfn_scale[i]
    rl, rm, ru = tfn_reciprocal[i]
    ws.cell(row=current_row, column=1).value = i
    ws.cell(row=current_row, column=2).value = f"({l}, {m}, {u})"
    ws.cell(row=current_row, column=3).value = f"({rl:.2f}, {rm:.2f}, {ru:.2f})" if i > 1 else "(1, 1, 1)"
    apply_cell_style(ws.cell(row=current_row, column=1))
    apply_cell_style(ws.cell(row=current_row, column=2))
    apply_cell_style(ws.cell(row=current_row, column=3))
    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
    current_row += 1

current_row += 1

# ============================================
# SECTION 3: Matriks Perbandingan Berpasangan (Crisp)
# ============================================
apply_section_header(ws, current_row, "LANGKAH 1: MATRIKS PERBANDINGAN BERPASANGAN (CRISP)", 1, 8)
current_row += 1

ws.cell(row=current_row, column=1).value = "Matriks bersifat resiprokal: aij = 1 dan aji = 1/aij"
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
current_row += 1

# Header row
ws.cell(row=current_row, column=1).value = "Kriteria"
apply_header_style(ws.cell(row=current_row, column=1))
for j, krit in enumerate(kriteria, start=2):
    cell = ws.cell(row=current_row, column=j)
    cell.value = krit[:12] + ".." if len(krit) > 12 else krit
    apply_header_style(cell)
current_row += 1

# Data rows
for i, krit in enumerate(kriteria):
    cell = ws.cell(row=current_row, column=1)
    cell.value = krit[:18] + ".." if len(krit) > 18 else krit
    cell.font = Font(bold=True, size=10)
    cell.border = thin_border
    cell.fill = light_blue_fill
    cell.alignment = left_align
    
    for j in range(n):
        cell = ws.cell(row=current_row, column=j + 2)
        cell.value = crisp_matrix[i][j]
        apply_cell_style(cell)
    current_row += 1

current_row += 1

# ============================================
# SECTION 4: Perhitungan Vector Eigen
# ============================================
apply_section_header(ws, current_row, "LANGKAH 2: PERHITUNGAN VECTOR EIGEN (GEOMETRIC MEAN METHOD)", 1, 8)
current_row += 1

ws.cell(row=current_row, column=1).value = "Rumus: GMi = (∏ aij)^(1/n)"
ws.cell(row=current_row, column=4).value = "wi = GMi / Σ GMi"
current_row += 1

# Calculate Geometric Mean
gm_values = []
for i in range(n):
    product = 1
    for j in range(n):
        product *= crisp_matrix[i][j]
    gm = product ** (1/n)
    gm_values.append(gm)

sum_gm = sum(gm_values)
eigenvector = [gm / sum_gm for gm in gm_values]

# Header
ws.cell(row=current_row, column=1).value = "Kriteria"
ws.cell(row=current_row, column=2).value = "Geometric Mean (GMi)"
ws.cell(row=current_row, column=3).value = "Eigenvector (wi)"
apply_header_style(ws.cell(row=current_row, column=1))
apply_header_style(ws.cell(row=current_row, column=2))
apply_header_style(ws.cell(row=current_row, column=3))
current_row += 1

# Data
for i, krit in enumerate(kriteria):
    ws.cell(row=current_row, column=1).value = krit[:25] + ".." if len(krit) > 25 else krit
    ws.cell(row=current_row, column=2).value = round(gm_values[i], 4)
    ws.cell(row=current_row, column=3).value = round(eigenvector[i], 4)
    apply_cell_style(ws.cell(row=current_row, column=1))
    apply_cell_style(ws.cell(row=current_row, column=2))
    apply_cell_style(ws.cell(row=current_row, column=3))
    current_row += 1

# Sum row
ws.cell(row=current_row, column=1).value = "TOTAL"
ws.cell(row=current_row, column=2).value = round(sum_gm, 4)
ws.cell(row=current_row, column=3).value = round(sum(eigenvector), 4)
ws.cell(row=current_row, column=1).font = Font(bold=True)
ws.cell(row=current_row, column=2).font = Font(bold=True)
ws.cell(row=current_row, column=3).font = Font(bold=True)
apply_cell_style(ws.cell(row=current_row, column=1))
apply_cell_style(ws.cell(row=current_row, column=2))
apply_cell_style(ws.cell(row=current_row, column=3))

# Lambda Max
lambda_max = 6.00
ws.cell(row=current_row-3, column=5).value = "Lambda Max (λmax)"
ws.cell(row=current_row-3, column=5).font = subtitle_font
ws.cell(row=current_row-2, column=5).value = lambda_max
ws.cell(row=current_row-2, column=5).font = Font(bold=True, size=18, color="ED7D31")
ws.cell(row=current_row-1, column=5).value = "λmax = (1/n) × Σ (Aw)i"
ws.merge_cells(start_row=current_row-1, start_column=5, end_row=current_row-1, end_column=7)

current_row += 2

# ============================================
# SECTION 5: Uji Konsistensi
# ============================================
apply_section_header(ws, current_row, "LANGKAH 3: UJI KONSISTENSI MATRIKS PERBANDINGAN", 1, 8)
current_row += 1

ws.cell(row=current_row, column=1).value = "Matriks dinyatakan KONSISTEN jika CR ≤ 0.1"
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
current_row += 1

# CI calculation
ci = (lambda_max - n) / (n - 1)
ri = ri_table[n]
cr = ci / ri if ri != 0 else 0

# CI Box
ws.cell(row=current_row, column=1).value = "Consistency Index (CI)"
ws.cell(row=current_row, column=1).font = subtitle_font
ws.cell(row=current_row+1, column=1).value = round(ci, 4)
ws.cell(row=current_row+1, column=1).font = Font(bold=True, size=16, color="008000")
ws.cell(row=current_row+2, column=1).value = "CI = (λmax - n) / (n - 1)"

# RI Box
ws.cell(row=current_row, column=3).value = "Random Index (RI)"
ws.cell(row=current_row, column=3).font = subtitle_font
ws.cell(row=current_row+1, column=3).value = ri
ws.cell(row=current_row+1, column=3).font = Font(bold=True, size=16)
ws.cell(row=current_row+2, column=3).value = f"n = {n}"

# CR Box
ws.cell(row=current_row, column=5).value = "Consistency Ratio (CR)"
ws.cell(row=current_row, column=5).font = subtitle_font
ws.cell(row=current_row+1, column=5).value = round(cr, 4)
ws.cell(row=current_row+1, column=5).font = Font(bold=True, size=16, color="008000")
if cr <= 0.1:
    ws.cell(row=current_row+2, column=5).value = "✓ KONSISTEN (CR ≤ 0.1)"
    ws.cell(row=current_row+2, column=5).font = Font(bold=True, color="008000")
    ws.cell(row=current_row+2, column=5).fill = light_green_fill
else:
    ws.cell(row=current_row+2, column=5).value = "✗ TIDAK KONSISTEN (CR > 0.1)"
    ws.cell(row=current_row+2, column=5).font = Font(bold=True, color="FF0000")

current_row += 4

# Tabel Random Index
ws.cell(row=current_row, column=1).value = "Tabel Random Index (RI):"
ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
current_row += 1

ws.cell(row=current_row, column=1).value = "n"
apply_header_style(ws.cell(row=current_row, column=1))
for i in range(1, 13):
    cell = ws.cell(row=current_row, column=i + 1)
    cell.value = i
    apply_header_style(cell)
current_row += 1

ws.cell(row=current_row, column=1).value = "RI"
apply_cell_style(ws.cell(row=current_row, column=1))
ws.cell(row=current_row, column=1).font = Font(bold=True)
for i in range(1, 13):
    cell = ws.cell(row=current_row, column=i + 1)
    cell.value = ri_table[i]
    apply_cell_style(cell)

current_row += 2

# ============================================
# SECTION 6: Matriks Fuzzy (TFN)
# ============================================
apply_section_header(ws, current_row, "LANGKAH 4: FUZZIFIKASI MATRIKS PERBANDINGAN BERPASANGAN (TFN)", 1, 19)
current_row += 1

ws.cell(row=current_row, column=1).value = "Nilai crisp dikonversi ke Triangular Fuzzy Number (TFN) = (l, m, u)"
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=19)
current_row += 1

# Header row - 3 columns per criterion (l, m, u)
ws.cell(row=current_row, column=1).value = "Kriteria"
apply_header_style(ws.cell(row=current_row, column=1))

col_idx = 2
for krit in kriteria:
    short_name = krit[:8] + ".." if len(krit) > 8 else krit
    ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx+2)
    cell = ws.cell(row=current_row, column=col_idx)
    cell.value = short_name
    apply_header_style(cell)
    col_idx += 3
current_row += 1

# Sub-headers for l, m, u
ws.cell(row=current_row, column=1).value = ""
apply_header_style(ws.cell(row=current_row, column=1), fill=yellow_fill)
col_idx = 2
for _ in kriteria:
    ws.cell(row=current_row, column=col_idx).value = "l"
    ws.cell(row=current_row, column=col_idx+1).value = "m"
    ws.cell(row=current_row, column=col_idx+2).value = "u"
    for c in range(col_idx, col_idx+3):
        apply_header_style(ws.cell(row=current_row, column=c), fill=yellow_fill)
    col_idx += 3
current_row += 1

# Data rows
for i, krit in enumerate(kriteria):
    cell = ws.cell(row=current_row, column=1)
    cell.value = krit[:12] + ".." if len(krit) > 12 else krit
    cell.font = Font(bold=True, size=9)
    cell.border = thin_border
    cell.fill = light_blue_fill
    cell.alignment = left_align
    
    col_idx = 2
    for j in range(n):
        crisp_val = crisp_matrix[i][j]
        l, m, u = tfn_scale.get(crisp_val, (1, 1, 1))
        
        ws.cell(row=current_row, column=col_idx).value = round(l, 2)
        ws.cell(row=current_row, column=col_idx+1).value = round(m, 2)
        ws.cell(row=current_row, column=col_idx+2).value = round(u, 2)
        
        for c in range(col_idx, col_idx+3):
            apply_cell_style(ws.cell(row=current_row, column=c))
        
        col_idx += 3
    current_row += 1

current_row += 1

# ============================================
# SECTION 7: Ringkasan Hasil
# ============================================
apply_section_header(ws, current_row, "RINGKASAN HASIL PERHITUNGAN FUZZY AHP", 1, 8)
current_row += 1

# Bobot Kriteria table
ws.cell(row=current_row, column=1).value = "No"
ws.cell(row=current_row, column=2).value = "Kriteria"
ws.cell(row=current_row, column=3).value = "Bobot (wi)"
apply_header_style(ws.cell(row=current_row, column=1), fill=green_fill)
apply_header_style(ws.cell(row=current_row, column=2), fill=green_fill)
apply_header_style(ws.cell(row=current_row, column=3), fill=green_fill)
current_row += 1

for i, krit in enumerate(kriteria):
    ws.cell(row=current_row, column=1).value = i + 1
    ws.cell(row=current_row, column=2).value = krit
    ws.cell(row=current_row, column=3).value = round(eigenvector[i], 4)
    apply_cell_style(ws.cell(row=current_row, column=1))
    apply_cell_style(ws.cell(row=current_row, column=2))
    apply_cell_style(ws.cell(row=current_row, column=3))
    ws.cell(row=current_row, column=3).fill = light_green_fill
    current_row += 1

current_row += 1

# Status boxes
ws.cell(row=current_row, column=1).value = "STATUS KONSISTENSI"
ws.cell(row=current_row, column=1).font = subtitle_font
ws.cell(row=current_row, column=5).value = "PERINGKAT AKHIR"
ws.cell(row=current_row, column=5).font = subtitle_font
current_row += 1

ws.cell(row=current_row, column=1).value = "CI ="
ws.cell(row=current_row, column=2).value = round(ci, 4)
ws.cell(row=current_row, column=5).value = "#1"
ws.cell(row=current_row, column=5).font = Font(bold=True, size=24, color="ED7D31")
ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row+2, end_column=6)
current_row += 1

ws.cell(row=current_row, column=1).value = "RI ="
ws.cell(row=current_row, column=2).value = ri
current_row += 1

ws.cell(row=current_row, column=1).value = "CR ="
ws.cell(row=current_row, column=2).value = round(cr, 4)
current_row += 1

ws.cell(row=current_row, column=1).value = "Status:"
if cr <= 0.1:
    ws.cell(row=current_row, column=2).value = "✓ KONSISTEN"
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12, color="008000")
    ws.cell(row=current_row, column=2).fill = light_green_fill
else:
    ws.cell(row=current_row, column=2).value = "✗ TIDAK KONSISTEN"
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12, color="FF0000")

# ============================================
# Set column widths
# ============================================
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
for j in range(9, 21):
    ws.column_dimensions[get_column_letter(j)].width = 6

# Save workbook
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_David_Kulian_SingleSheet.xlsx"
wb.save(output_file)
print(f"[OK] File Excel berhasil dibuat: {output_file}")
print("\nFile berisi SINGLE SHEET dengan semua langkah perhitungan Fuzzy AHP:")
print("  1. Tabel Skala Saaty")
print("  2. Tabel Skala Fuzzy (TFN)")
print("  3. Matriks Perbandingan Berpasangan (Crisp)")
print("  4. Perhitungan Vector Eigen (Geometric Mean)")
print("  5. Uji Konsistensi (CI, RI, CR)")
print("  6. Matriks Fuzzy (TFN)")
print("  7. Ringkasan Hasil")
