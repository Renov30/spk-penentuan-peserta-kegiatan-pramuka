"""
Script untuk membuat Model Tabel Excel Perhitungan Fuzzy AHP
VERSI 5 - TANPA LOOKUP, 100% formula dasar (arithmetic only)
Semua TFN menggunakan nilai langsung berdasarkan input crisp = 1
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fuzzy AHP"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
input_font = Font(size=11, color="0000FF")
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

kriteria = ["K1", "K2", "K3", "K4", "K5", "K6"]
kriteria_full = ["Status Keaktifan", "Pencapaian SKU", "Pencapaian SPG",
                 "Kesehatan Jasmani", "Tes Wawancara", "Tes Pilihan Ganda"]
n = 6

def header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def input_style(cell):
    cell.font = input_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = input_fill

def section(row, text, col_end=9):
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)

row = 1

# Title
ws['A1'] = "FUZZY AHP CALCULATOR"
ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws.merge_cells('A1:I1')
ws['A2'] = "Ubah nilai 1-9 di sel KUNING. Hasil otomatis terhitung."
ws['A2'].font = Font(italic=True, color="FF0000")
ws.merge_cells('A2:I2')

row = 4

# ============================================
# MATRIKS PERBANDINGAN BERPASANGAN
# ============================================
section(row, "MATRIKS PERBANDINGAN BERPASANGAN")
row += 1

# Header
ws.cell(row=row, column=1).value = ""
header_style(ws.cell(row=row, column=1))
for j in range(n):
    cell = ws.cell(row=row, column=j+2)
    cell.value = kriteria[j]
    header_style(cell)

ws.cell(row=row, column=8).value = "GM"
ws.cell(row=row, column=9).value = "Wi"
header_style(ws.cell(row=row, column=8), fill=green_fill)
header_style(ws.cell(row=row, column=9), fill=green_fill)
row += 1

data_start = row

# Crisp matrix
for i in range(n):
    cell = ws.cell(row=row, column=1)
    cell.value = kriteria[i]
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = light_blue_fill
    
    for j in range(n):
        cell = ws.cell(row=row, column=j+2)
        if i == j:
            cell.value = 1
            cell_style(cell)
        elif i < j:
            cell.value = 1  # Default input
            input_style(cell)
        else:
            # Reciprocal
            upper_row = data_start + j
            upper_col = get_column_letter(i + 2)
            cell.value = f"=1/{upper_col}{upper_row}"
            cell_style(cell)
            cell.fill = light_green_fill
    
    # GM = (B*C*D*E*F*G)^(1/6)
    gm = ws.cell(row=row, column=8)
    gm.value = f"=(B{row}*C{row}*D{row}*E{row}*F{row}*G{row})^(1/6)"
    cell_style(gm)
    gm.fill = light_green_fill
    gm.number_format = '0.0000'
    
    row += 1

data_end = row - 1

# Total
ws.cell(row=row, column=7).value = "Total:"
ws.cell(row=row, column=7).font = Font(bold=True)
gm_sum = ws.cell(row=row, column=8)
gm_sum.value = f"=SUM(H{data_start}:H{data_end})"
gm_sum.font = Font(bold=True)
cell_style(gm_sum)
gm_sum.fill = yellow_fill
gm_sum_ref = f"$H${row}"

# Wi
for i in range(n):
    r = data_start + i
    wi = ws.cell(row=r, column=9)
    wi.value = f"=H{r}/{gm_sum_ref}"
    cell_style(wi)
    wi.fill = light_green_fill
    wi.number_format = '0.0000'

wi_sum = ws.cell(row=row, column=9)
wi_sum.value = f"=SUM(I{data_start}:I{data_end})"
wi_sum.font = Font(bold=True)
cell_style(wi_sum)
wi_sum.fill = yellow_fill

row += 2

# ============================================
# PERHITUNGAN KONSISTENSI
# ============================================
section(row, "PERHITUNGAN KONSISTENSI")
row += 1

ws.cell(row=row, column=1).value = "Kriteria"
ws.cell(row=row, column=2).value = "A*w"
ws.cell(row=row, column=3).value = "(A*w)/w"
header_style(ws.cell(row=row, column=1))
header_style(ws.cell(row=row, column=2))
header_style(ws.cell(row=row, column=3))
row += 1

aw_start = row

for i in range(n):
    ws.cell(row=row, column=1).value = kriteria[i]
    cell_style(ws.cell(row=row, column=1))
    ws.cell(row=row, column=1).fill = light_blue_fill
    
    crisp_row = data_start + i
    
    # A*w explicit sum
    aw = ws.cell(row=row, column=2)
    parts = []
    for j in range(n):
        col = get_column_letter(j + 2)
        parts.append(f"{col}{crisp_row}*$I${data_start + j}")
    aw.value = "=" + "+".join(parts)
    cell_style(aw)
    aw.fill = light_green_fill
    aw.number_format = '0.0000'
    
    # Ratio
    ratio = ws.cell(row=row, column=3)
    ratio.value = f"=B{row}/$I${data_start + i}"
    cell_style(ratio)
    ratio.fill = light_green_fill
    ratio.number_format = '0.0000'
    
    row += 1

aw_end = row - 1
row += 1

# Lambda Max
lambda_row = row
ws.cell(row=row, column=1).value = "Lambda Max:"
ws.cell(row=row, column=1).font = subtitle_font
lmax = ws.cell(row=row, column=2)
lmax.value = f"=AVERAGE(C{aw_start}:C{aw_end})"
lmax.font = Font(bold=True, size=14, color="ED7D31")
lmax.fill = yellow_fill
cell_style(lmax)
row += 1

# CI
ci_row = row
ws.cell(row=row, column=1).value = "CI:"
ws.cell(row=row, column=1).font = subtitle_font
ci = ws.cell(row=row, column=2)
ci.value = f"=(B{lambda_row}-6)/5"
ci.font = Font(bold=True)
ci.fill = light_green_fill
cell_style(ci)
row += 1

# RI
ri_row = row
ws.cell(row=row, column=1).value = "RI:"
ws.cell(row=row, column=1).font = subtitle_font
ri = ws.cell(row=row, column=2)
ri.value = 1.24
ri.font = Font(bold=True)
cell_style(ri)
row += 1

# CR
cr_row = row
ws.cell(row=row, column=1).value = "CR:"
ws.cell(row=row, column=1).font = subtitle_font
cr = ws.cell(row=row, column=2)
cr.value = f"=B{ci_row}/B{ri_row}"
cr.font = Font(bold=True, size=14)
cr.fill = yellow_fill
cell_style(cr)
row += 1

# Status
ws.cell(row=row, column=1).value = "Status:"
ws.cell(row=row, column=1).font = subtitle_font
status = ws.cell(row=row, column=2)
status.value = f'=IF(B{cr_row}<=0.1,"KONSISTEN","TIDAK KONSISTEN")'
status.font = Font(bold=True, size=12)
cell_style(status)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

row += 2

# ============================================
# TABEL TFN
# ============================================
section(row, "TABEL SKALA FUZZY (TFN)")
row += 1

ws.cell(row=row, column=1).value = "Nilai"
ws.cell(row=row, column=2).value = "l"
ws.cell(row=row, column=3).value = "m"
ws.cell(row=row, column=4).value = "u"
header_style(ws.cell(row=row, column=1))
header_style(ws.cell(row=row, column=2))
header_style(ws.cell(row=row, column=3))
header_style(ws.cell(row=row, column=4))
row += 1

tfn = [(1,1,1,1), (2,0.5,1,1.5), (3,1,1.5,2), (4,1.5,2,2.5), 
       (5,2,2.5,3), (6,2.5,3,3.5), (7,3,3.5,4), (8,3.5,4,4.5), (9,4,4.5,4.5)]

for v, l, m, u in tfn:
    ws.cell(row=row, column=1).value = v
    ws.cell(row=row, column=2).value = l
    ws.cell(row=row, column=3).value = m
    ws.cell(row=row, column=4).value = u
    for c in range(1, 5):
        cell_style(ws.cell(row=row, column=c))
    row += 1

row += 1

# ============================================
# RINGKASAN BOBOT
# ============================================
section(row, "RINGKASAN BOBOT KRITERIA")
row += 1

ws.cell(row=row, column=1).value = "No"
ws.cell(row=row, column=2).value = "Kriteria"
ws.cell(row=row, column=3).value = "Bobot"
ws.cell(row=row, column=4).value = "%"
header_style(ws.cell(row=row, column=1), fill=green_fill)
header_style(ws.cell(row=row, column=2), fill=green_fill)
header_style(ws.cell(row=row, column=3), fill=green_fill)
header_style(ws.cell(row=row, column=4), fill=green_fill)
row += 1

for i in range(n):
    ws.cell(row=row, column=1).value = i + 1
    cell_style(ws.cell(row=row, column=1))
    
    ws.cell(row=row, column=2).value = kriteria_full[i]
    cell_style(ws.cell(row=row, column=2))
    
    ws.cell(row=row, column=3).value = f"=I{data_start + i}"
    cell_style(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).fill = light_green_fill
    ws.cell(row=row, column=3).number_format = '0.0000'
    
    ws.cell(row=row, column=4).value = f"=I{data_start + i}*100"
    cell_style(ws.cell(row=row, column=4))
    ws.cell(row=row, column=4).number_format = '0.00"%"'
    row += 1

# Column widths
for c in ['A','B','C','D','E','F','G','H','I']:
    ws.column_dimensions[c].width = 14

# Save
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_Calculator_v5.xlsx"
wb.save(output_file)
print(f"[OK] File berhasil dibuat: {output_file}")
print(f"\nINFO:")
print(f"  - Matriks input: B{data_start}:G{data_end}")
print(f"  - Bobot (Wi): I{data_start}:I{data_end}")
print(f"  - CR: B{cr_row}")
print(f"\nCatatan: File ini HANYA menggunakan formula arithmetic dasar")
print(f"         Tidak ada VLOOKUP, INDEX, MATCH, atau fungsi kompleks lainnya")
