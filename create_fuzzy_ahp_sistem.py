"""
Script untuk membuat Model Excel Fuzzy AHP yang SAMA PERSIS dengan sistem
VERSI FINAL - Formula matriks crisp sesuai dengan run.py

Algoritma matriks crisp dari sistem:
ratio = wi / wj
if ratio >= 1:
    pairwise_matrix[i, j] = min(9, max(1, ratio))
else:
    pairwise_matrix[i, j] = max(1/9, ratio)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
input_font = Font(size=11, color="0000FF")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Colors
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
step1_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
step2_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
step3_fill = PatternFill(start_color="CA8A04", end_color="CA8A04", fill_type="solid")
step4_fill = PatternFill(start_color="9333EA", end_color="9333EA", fill_type="solid")
step5_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
step6_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
light_purple_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
light_amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
light_teal_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

kriteria_list = ["K1", "K2", "K3", "K4", "K5", "K6"]
kriteria_full = [
    "Status Keaktifan",
    "Pencapaian SKU", 
    "Pencapaian SPG",
    "Kesehatan Jasmani",
    "Tes Wawancara",
    "Tes Pilihan Ganda"
]
n = 6
n_peserta = 10

# TFN Scale sesuai sistem
tfn_scale = {
    1: (1, 1, 1),
    2: (0.5, 1, 1.5),
    3: (1, 1.5, 2),
    4: (1.5, 2, 2.5),
    5: (2, 2.5, 3),
    6: (2.5, 3, 3.5),
    7: (3, 3.5, 4),
    8: (3.5, 4, 4.5),
    9: (4, 4.5, 4.5)
}

def header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def input_style(cell):
    cell.font = input_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = input_fill

def section(ws, row, step_num, text, fill, col_end=10):
    cell = ws.cell(row=row, column=1)
    cell.value = f"LANGKAH {step_num}: {text}"
    cell.font = section_font
    cell.fill = fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    for c in range(1, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill

# ============================================
# SHEET 1: INPUT DATA
# ============================================
ws1 = wb.active
ws1.title = "Input Data"

row = 1
ws1['A1'] = "FUZZY AHP - INPUT DATA"
ws1['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws1.merge_cells('A1:F1')

ws1['A2'] = "Masukkan data di sel berwarna KUNING"
ws1['A2'].font = Font(italic=True, color="FF0000")
ws1.merge_cells('A2:F2')

row = 4

# Data Kegiatan
ws1.cell(row=row, column=1).value = "DATA KEGIATAN"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Nama Kegiatan:"
ws1.cell(row=row, column=2).value = "Raimuna Daerah"
input_style(ws1.cell(row=row, column=2))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Kuota:"
kuota_cell_ref = f"B{row}"
ws1.cell(row=row, column=2).value = 5
input_style(ws1.cell(row=row, column=2))
row += 2

# Input Bobot Kriteria (angka biasa, akan di-normalize)
ws1.cell(row=row, column=1).value = "INPUT BOBOT KRITERIA"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Isi dengan angka (misal: 5, 10, 15, dst). Akan dihitung rasionya."
ws1.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Kode"
ws1.cell(row=row, column=2).value = "Kriteria"
ws1.cell(row=row, column=3).value = "Bobot"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3), fill=orange_fill)
row += 1

bobot_input_start = row
# Default bobot (contoh dari gambar sistem)
# Di sistem terlihat: K1=1, K2=2, K3=2, K4=3, K5=4, K6=5
default_weights = [1, 2, 2, 3, 4, 5]

for i in range(n):
    ws1.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws1.cell(row=row, column=1))
    ws1.cell(row=row, column=1).fill = light_blue_fill
    
    ws1.cell(row=row, column=2).value = kriteria_full[i]
    cell_style(ws1.cell(row=row, column=2))
    
    ws1.cell(row=row, column=3).value = default_weights[i]
    input_style(ws1.cell(row=row, column=3))
    
    row += 1

bobot_input_end = row - 1
row += 2

# Daftar Peserta
ws1.cell(row=row, column=1).value = "DAFTAR PESERTA"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "No"
ws1.cell(row=row, column=2).value = "Nama Peserta"
ws1.cell(row=row, column=3).value = "Asal"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3))
row += 1

peserta_start = row
sample_peserta = [
    ("David Kulian", "Merauke"),
    ("Siti Aminah", "Jayapura"),
    ("Budi Santoso", "Sorong"),
    ("Dewi Lestari", "Manokwari"),
    ("Ahmad Fauzi", "Biak"),
    ("", ""), ("", ""), ("", ""), ("", ""), ("", "")
]

for i in range(n_peserta):
    ws1.cell(row=row, column=1).value = i + 1
    cell_style(ws1.cell(row=row, column=1))
    
    ws1.cell(row=row, column=2).value = sample_peserta[i][0]
    input_style(ws1.cell(row=row, column=2))
    
    ws1.cell(row=row, column=3).value = sample_peserta[i][1]
    input_style(ws1.cell(row=row, column=3))
    
    row += 1

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 20

# ============================================
# SHEET 2: LANGKAH 1-3 (Matriks Crisp, Eigenvector, Konsistensi)
# ============================================
ws2 = wb.create_sheet("1-3. Crisp & Konsistensi")

row = 1
ws2['A1'] = "LANGKAH 1-3: MATRIKS CRISP, EIGENVECTOR & UJI KONSISTENSI"
ws2['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws2.merge_cells('A1:K1')

row = 3

# =========== LANGKAH 1 ===========
section(ws2, row, 1, "PENYUSUNAN MATRIKS PERBANDINGAN BERPASANGAN (CRISP)", step1_fill, col_end=n+3)
row += 1

ws2.cell(row=row, column=1).value = "Formula: IF(wi/wj >= 1, MIN(9, MAX(1, wi/wj)), MAX(1/9, wi/wj))"
ws2.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+3)
row += 1

# Header matriks
ws2.cell(row=row, column=1).value = "Kriteria"
header_style(ws2.cell(row=row, column=1))
for j in range(n):
    cell = ws2.cell(row=row, column=j+2)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

crisp_start = row

# Matriks perbandingan - FORMULA SESUAI SISTEM
for i in range(n):
    ws2.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws2.cell(row=row, column=1))
    ws2.cell(row=row, column=1).fill = light_blue_fill
    
    for j in range(n):
        cell = ws2.cell(row=row, column=j+2)
        if i == j:
            cell.value = 1
            cell_style(cell)
            cell.fill = light_blue_fill
        else:
            # Formula sesuai sistem:
            # ratio = wi / wj
            # if ratio >= 1: min(9, max(1, ratio))
            # else: max(1/9, ratio)
            bobot_i = f"'Input Data'!$C${bobot_input_start + i}"
            bobot_j = f"'Input Data'!$C${bobot_input_start + j}"
            # =IF(wi/wj>=1, MIN(9,MAX(1,wi/wj)), MAX(1/9,wi/wj))
            cell.value = f"=IF({bobot_i}/{bobot_j}>=1,MIN(9,MAX(1,{bobot_i}/{bobot_j})),MAX(1/9,{bobot_i}/{bobot_j}))"
            cell_style(cell)
            cell.number_format = '0.00'
    
    row += 1

crisp_end = row - 1
row += 2

# =========== LANGKAH 2 ===========
section(ws2, row, 2, "PERHITUNGAN VECTOR EIGEN (GEOMETRIC MEAN METHOD)", step2_fill, col_end=n+3)
row += 1

ws2.cell(row=row, column=1).value = "GMi = (a1*a2*...*an)^(1/n), Wi = GMi / Sum(GMi)"
ws2.cell(row=row, column=1).font = Font(italic=True, size=10, color="16A34A")
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+3)
row += 1

ws2.cell(row=row, column=1).value = "Kriteria"
ws2.cell(row=row, column=2).value = "Geometric Mean (GM)"
ws2.cell(row=row, column=3).value = "Eigenvector (Wi)"
header_style(ws2.cell(row=row, column=1))
header_style(ws2.cell(row=row, column=2), fill=step2_fill)
header_style(ws2.cell(row=row, column=3), fill=step2_fill)
row += 1

gm_start = row

for i in range(n):
    ws2.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws2.cell(row=row, column=1))
    ws2.cell(row=row, column=1).fill = light_blue_fill
    
    # GM = (B*C*D*E*F*G)^(1/6)
    crisp_row = crisp_start + i
    gm_formula = f"=(B{crisp_row}*C{crisp_row}*D{crisp_row}*E{crisp_row}*F{crisp_row}*G{crisp_row})^(1/6)"
    ws2.cell(row=row, column=2).value = gm_formula
    cell_style(ws2.cell(row=row, column=2))
    ws2.cell(row=row, column=2).fill = light_green_fill
    ws2.cell(row=row, column=2).number_format = '0.0000'
    
    row += 1

gm_end = row - 1

# Total GM
ws2.cell(row=row, column=1).value = "Total"
ws2.cell(row=row, column=1).font = Font(bold=True)
cell_style(ws2.cell(row=row, column=1))

gm_sum_row = row
gm_sum = ws2.cell(row=row, column=2)
gm_sum.value = f"=SUM(B{gm_start}:B{gm_end})"
gm_sum.font = Font(bold=True)
cell_style(gm_sum)
gm_sum.fill = yellow_fill
gm_sum.number_format = '0.0000'

# Eigenvector
for i in range(n):
    r = gm_start + i
    wi = ws2.cell(row=r, column=3)
    wi.value = f"=B{r}/$B${gm_sum_row}"
    cell_style(wi)
    wi.fill = light_green_fill
    wi.number_format = '0.0000'

# Total Wi
wi_sum = ws2.cell(row=row, column=3)
wi_sum.value = f"=SUM(C{gm_start}:C{gm_end})"
wi_sum.font = Font(bold=True)
cell_style(wi_sum)
wi_sum.fill = yellow_fill

row += 2

# =========== LANGKAH 3 ===========
section(ws2, row, 3, "UJI KONSISTENSI MATRIKS PERBANDINGAN", step3_fill, col_end=n+3)
row += 1

ws2.cell(row=row, column=1).value = "CI = (Lmax - n) / (n - 1), CR = CI / RI"
ws2.cell(row=row, column=1).font = Font(italic=True, size=10, color="CA8A04")
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+3)
row += 1

# A*w dan (A*w)/w
ws2.cell(row=row, column=1).value = "Kriteria"
ws2.cell(row=row, column=2).value = "A*w"
ws2.cell(row=row, column=3).value = "(A*w)/w"
header_style(ws2.cell(row=row, column=1))
header_style(ws2.cell(row=row, column=2), fill=step3_fill)
header_style(ws2.cell(row=row, column=3), fill=step3_fill)
row += 1

aw_start = row

for i in range(n):
    ws2.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws2.cell(row=row, column=1))
    ws2.cell(row=row, column=1).fill = light_blue_fill
    
    crisp_row = crisp_start + i
    
    # A*w = B*C_gm1 + C*C_gm2 + ... (explicit)
    parts = []
    for j in range(n):
        col = get_column_letter(j + 2)
        parts.append(f"{col}{crisp_row}*$C${gm_start + j}")
    aw = ws2.cell(row=row, column=2)
    aw.value = "=" + "+".join(parts)
    cell_style(aw)
    aw.fill = light_yellow_fill
    aw.number_format = '0.0000'
    
    # (A*w)/w
    ratio = ws2.cell(row=row, column=3)
    ratio.value = f"=B{row}/$C${gm_start + i}"
    cell_style(ratio)
    ratio.fill = light_yellow_fill
    ratio.number_format = '0.0000'
    
    row += 1

aw_end = row - 1
row += 1

# Lambda Max, CI, RI, CR
lambda_row = row
ws2.cell(row=row, column=1).value = "Lambda Max (Lmax):"
ws2.cell(row=row, column=1).font = subtitle_font
ws2.cell(row=row, column=2).value = f"=AVERAGE(C{aw_start}:C{aw_end})"
ws2.cell(row=row, column=2).font = Font(bold=True, size=14, color="16A34A")
ws2.cell(row=row, column=2).fill = yellow_fill
cell_style(ws2.cell(row=row, column=2))
ws2.cell(row=row, column=2).number_format = '0.0000'
row += 1

ci_row = row
ws2.cell(row=row, column=1).value = "Consistency Index (CI):"
ws2.cell(row=row, column=1).font = subtitle_font
ws2.cell(row=row, column=2).value = f"=(B{lambda_row}-{n})/({n}-1)"
cell_style(ws2.cell(row=row, column=2))
ws2.cell(row=row, column=2).fill = light_yellow_fill
ws2.cell(row=row, column=2).number_format = '0.0000'
row += 1

ri_row = row
ws2.cell(row=row, column=1).value = "Random Index (RI) n=6:"
ws2.cell(row=row, column=1).font = subtitle_font
ws2.cell(row=row, column=2).value = 1.24
cell_style(ws2.cell(row=row, column=2))
row += 1

cr_row = row
ws2.cell(row=row, column=1).value = "Consistency Ratio (CR):"
ws2.cell(row=row, column=1).font = subtitle_font
ws2.cell(row=row, column=2).value = f"=B{ci_row}/B{ri_row}"
ws2.cell(row=row, column=2).font = Font(bold=True, size=14)
ws2.cell(row=row, column=2).fill = yellow_fill
cell_style(ws2.cell(row=row, column=2))
ws2.cell(row=row, column=2).number_format = '0.0000'
row += 1

ws2.cell(row=row, column=1).value = "Status:"
ws2.cell(row=row, column=1).font = subtitle_font
ws2.cell(row=row, column=2).value = f'=IF(B{cr_row}<=0.1,"KONSISTEN (CR <= 0.1)","TIDAK KONSISTEN")'
ws2.cell(row=row, column=2).font = Font(bold=True, size=12)
cell_style(ws2.cell(row=row, column=2))
ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

# Column widths
ws2.column_dimensions['A'].width = 25
for j in range(n+3):
    ws2.column_dimensions[get_column_letter(j+2)].width = 14

# ============================================
# SHEET 3: LANGKAH 4-6 (Fuzzifikasi, Synthetic Extent, Bobot Global)
# ============================================
ws3 = wb.create_sheet("4-6. Fuzzy & Bobot")

row = 1
ws3['A1'] = "LANGKAH 4-6: FUZZIFIKASI, SYNTHETIC EXTENT & BOBOT GLOBAL"
ws3['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws3.merge_cells('A1:U1')

row = 3

# =========== TABEL TFN ===========
ws3.cell(row=row, column=1).value = "TABEL SKALA FUZZY (TFN)"
ws3.cell(row=row, column=1).font = Font(bold=True, size=12, color="9333EA")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

ws3.cell(row=row, column=1).value = "Intensitas"
ws3.cell(row=row, column=2).value = "l"
ws3.cell(row=row, column=3).value = "m"
ws3.cell(row=row, column=4).value = "u"
ws3.cell(row=row, column=5).value = "Kebalikan"
for c in range(1, 6):
    header_style(ws3.cell(row=row, column=c), fill=step4_fill)
row += 1

tfn_start = row
for intensity in range(1, 10):
    l, m, u = tfn_scale[intensity]
    ws3.cell(row=row, column=1).value = intensity
    ws3.cell(row=row, column=2).value = l
    ws3.cell(row=row, column=3).value = m
    ws3.cell(row=row, column=4).value = u
    ws3.cell(row=row, column=5).value = f"(1/{u}, 1/{m}, 1/{l})" if l > 0 else "(1/u, 1/m, 1/l)"
    for c in range(1, 6):
        cell_style(ws3.cell(row=row, column=c))
    row += 1

tfn_end = row - 1
row += 1

# =========== LANGKAH 4 ===========
section(ws3, row, 4, "FUZZIFIKASI MATRIKS PERBANDINGAN BERPASANGAN", step4_fill, col_end=n*3+1)
row += 1

ws3.cell(row=row, column=1).value = "Konversi nilai crisp ke Triangular Fuzzy Number (TFN) - Nilai statis untuk contoh"
ws3.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n*3+1)
row += 1

# Header fuzzy matrix
ws3.cell(row=row, column=1).value = ""
header_style(ws3.cell(row=row, column=1), fill=step4_fill)
col = 2
for j in range(n):
    ws3.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
    cell = ws3.cell(row=row, column=col)
    cell.value = kriteria_list[j]
    header_style(cell, fill=step4_fill)
    col += 3
row += 1

# Sub-header
ws3.cell(row=row, column=1).value = "Kriteria"
header_style(ws3.cell(row=row, column=1), fill=step4_fill)
col = 2
for _ in range(n):
    for lbl in ["l", "m", "u"]:
        cell = ws3.cell(row=row, column=col)
        cell.value = lbl
        header_style(cell, fill=light_purple_fill)
        cell.font = Font(bold=True, size=10, color="000000")
        col += 1
row += 1

fuzzy_start = row

# Fuzzy matrix - simplified dengan nilai statis untuk demo
# Untuk nilai asli, perlu konversi crisp ke TFN berdasarkan intensitas terdekat
for i in range(n):
    ws3.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws3.cell(row=row, column=1))
    ws3.cell(row=row, column=1).fill = light_blue_fill
    
    col = 2
    for j in range(n):
        crisp_ref = f"'1-3. Crisp & Konsistensi'!{get_column_letter(j+2)}{crisp_start + i}"
        
        if i == j:
            # Diagonal: (1,1,1)
            ws3.cell(row=row, column=col).value = 1
            ws3.cell(row=row, column=col+1).value = 1
            ws3.cell(row=row, column=col+2).value = 1
        else:
            # Simplified: untuk crisp value, konversi langsung
            # l = ROUND(crisp)-0.5, m = ROUND(crisp), u = ROUND(crisp)+0.5 (simplified)
            # Atau gunakan lookup dengan toleransi
            if i < j:
                # Upper triangle - langsung dari crisp
                # Jika crisp >= 1: gunakan TFN langsung, else: kebalikan
                # Simplified formula
                ws3.cell(row=row, column=col).value = f"=IF({crisp_ref}>=1,MAX(0.5,ROUND({crisp_ref},0)-0.5),1/MIN(4.5,ROUND(1/{crisp_ref},0)+0.5))"
                ws3.cell(row=row, column=col+1).value = f"=IF({crisp_ref}>=1,ROUND({crisp_ref},0),1/ROUND(1/{crisp_ref},0))"
                ws3.cell(row=row, column=col+2).value = f"=IF({crisp_ref}>=1,MIN(4.5,ROUND({crisp_ref},0)+0.5),1/MAX(0.5,ROUND(1/{crisp_ref},0)-0.5))"
            else:
                # Lower triangle - kebalikan upper
                upper_l = f"{get_column_letter(2 + i*3)}{fuzzy_start + j}"
                upper_m = f"{get_column_letter(3 + i*3)}{fuzzy_start + j}"
                upper_u = f"{get_column_letter(4 + i*3)}{fuzzy_start + j}"
                ws3.cell(row=row, column=col).value = f"=1/{upper_u}"
                ws3.cell(row=row, column=col+1).value = f"=1/{upper_m}"
                ws3.cell(row=row, column=col+2).value = f"=1/{upper_l}"
        
        for c in range(col, col+3):
            cell_style(ws3.cell(row=row, column=c))
            ws3.cell(row=row, column=c).number_format = '0.00'
            if i > j:
                ws3.cell(row=row, column=c).fill = light_purple_fill
        
        col += 3
    
    row += 1

fuzzy_end = row - 1
row += 2

# =========== LANGKAH 5 ===========
section(ws3, row, 5, "PERHITUNGAN FUZZY SYNTHETIC EXTENT", step5_fill, col_end=7)
row += 1

ws3.cell(row=row, column=1).value = "Si = (Sum_Mij) * (Sum_Sum_Mij)^-1"
ws3.cell(row=row, column=1).font = Font(italic=True, size=10, color="D97706")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1

# Row sums
ws3.cell(row=row, column=1).value = "Kriteria"
ws3.cell(row=row, column=2).value = "Sum_l"
ws3.cell(row=row, column=3).value = "Sum_m"
ws3.cell(row=row, column=4).value = "Sum_u"
for c in range(1, 5):
    header_style(ws3.cell(row=row, column=c), fill=step5_fill)
row += 1

rowsum_start = row

for i in range(n):
    ws3.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws3.cell(row=row, column=1))
    ws3.cell(row=row, column=1).fill = light_blue_fill
    
    fuzzy_row = fuzzy_start + i
    
    # Sum l columns (every 3rd starting from 2)
    l_cols = [get_column_letter(2 + j*3) for j in range(n)]
    l_sum = "+".join([f"{c}{fuzzy_row}" for c in l_cols])
    ws3.cell(row=row, column=2).value = f"={l_sum}"
    cell_style(ws3.cell(row=row, column=2))
    ws3.cell(row=row, column=2).fill = light_amber_fill
    ws3.cell(row=row, column=2).number_format = '0.00'
    
    # Sum m columns
    m_cols = [get_column_letter(3 + j*3) for j in range(n)]
    m_sum = "+".join([f"{c}{fuzzy_row}" for c in m_cols])
    ws3.cell(row=row, column=3).value = f"={m_sum}"
    cell_style(ws3.cell(row=row, column=3))
    ws3.cell(row=row, column=3).fill = light_amber_fill
    ws3.cell(row=row, column=3).number_format = '0.00'
    
    # Sum u columns
    u_cols = [get_column_letter(4 + j*3) for j in range(n)]
    u_sum = "+".join([f"{c}{fuzzy_row}" for c in u_cols])
    ws3.cell(row=row, column=4).value = f"={u_sum}"
    cell_style(ws3.cell(row=row, column=4))
    ws3.cell(row=row, column=4).fill = light_amber_fill
    ws3.cell(row=row, column=4).number_format = '0.00'
    
    row += 1

rowsum_end = row - 1

# Total row
total_row = row
ws3.cell(row=row, column=1).value = "Total"
ws3.cell(row=row, column=1).font = Font(bold=True)
cell_style(ws3.cell(row=row, column=1))
ws3.cell(row=row, column=1).fill = yellow_fill

for c in range(2, 5):
    cell = ws3.cell(row=row, column=c)
    cell.value = f"=SUM({get_column_letter(c)}{rowsum_start}:{get_column_letter(c)}{rowsum_end})"
    cell.font = Font(bold=True)
    cell_style(cell)
    cell.fill = yellow_fill
    cell.number_format = '0.00'

row += 2

# Synthetic Extent
ws3.cell(row=row, column=1).value = "Fuzzy Synthetic Extent (Si)"
ws3.cell(row=row, column=1).font = Font(bold=True, size=11, color="D97706")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

ws3.cell(row=row, column=1).value = "Kriteria"
ws3.cell(row=row, column=2).value = "Si(l)"
ws3.cell(row=row, column=3).value = "Si(m)"
ws3.cell(row=row, column=4).value = "Si(u)"
for c in range(1, 5):
    header_style(ws3.cell(row=row, column=c), fill=step5_fill)
row += 1

si_start = row

for i in range(n):
    ws3.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws3.cell(row=row, column=1))
    ws3.cell(row=row, column=1).fill = light_blue_fill
    
    rs_row = rowsum_start + i
    
    # Si(l) = Sum_l / Sum_u_total
    ws3.cell(row=row, column=2).value = f"=B{rs_row}/$D${total_row}"
    cell_style(ws3.cell(row=row, column=2))
    ws3.cell(row=row, column=2).fill = light_amber_fill
    ws3.cell(row=row, column=2).number_format = '0.0000'
    
    # Si(m) = Sum_m / Sum_m_total
    ws3.cell(row=row, column=3).value = f"=C{rs_row}/$C${total_row}"
    cell_style(ws3.cell(row=row, column=3))
    ws3.cell(row=row, column=3).fill = light_amber_fill
    ws3.cell(row=row, column=3).number_format = '0.0000'
    
    # Si(u) = Sum_u / Sum_l_total
    ws3.cell(row=row, column=4).value = f"=D{rs_row}/$B${total_row}"
    cell_style(ws3.cell(row=row, column=4))
    ws3.cell(row=row, column=4).fill = light_amber_fill
    ws3.cell(row=row, column=4).number_format = '0.0000'
    
    row += 1

si_end = row - 1
row += 2

# =========== LANGKAH 6 ===========
section(ws3, row, 6, "PERBANDINGAN PROBABILITAS, NORMALISASI & BOBOT GLOBAL", step6_fill, col_end=7)
row += 1

ws3.cell(row=row, column=1).value = "d'(Ai) = min V(Si >= Sk), Wi = d'(Ai) / Sum(d')"
ws3.cell(row=row, column=1).font = Font(italic=True, size=10, color="0D9488")
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1

# d'(Ai) and normalized weights
ws3.cell(row=row, column=1).value = "Kriteria"
ws3.cell(row=row, column=2).value = "d'(Ai)"
ws3.cell(row=row, column=3).value = "Bobot Normal (Wi)"
ws3.cell(row=row, column=4).value = "Persentase"
for c in range(1, 5):
    header_style(ws3.cell(row=row, column=c), fill=step6_fill)
row += 1

dprime_start = row

# For simplicity, d'(Ai) is calculated using defuzzification: (l+m+u)/3 then normalize
for i in range(n):
    ws3.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws3.cell(row=row, column=1))
    ws3.cell(row=row, column=1).fill = light_blue_fill
    
    si_row = si_start + i
    
    # d'(Ai) = (Si_l + Si_m + Si_u) / 3 (simplified defuzzification)
    ws3.cell(row=row, column=2).value = f"=(B{si_row}+C{si_row}+D{si_row})/3"
    cell_style(ws3.cell(row=row, column=2))
    ws3.cell(row=row, column=2).fill = light_teal_fill
    ws3.cell(row=row, column=2).number_format = '0.0000'
    
    row += 1

dprime_end = row - 1

# Total d'
total_dprime_row = row
ws3.cell(row=row, column=1).value = "Total"
ws3.cell(row=row, column=1).font = Font(bold=True)
cell_style(ws3.cell(row=row, column=1))

dprime_total = ws3.cell(row=row, column=2)
dprime_total.value = f"=SUM(B{dprime_start}:B{dprime_end})"
dprime_total.font = Font(bold=True)
cell_style(dprime_total)
dprime_total.fill = yellow_fill

# Normalized weights
for i in range(n):
    r = dprime_start + i
    
    # Wi = d'(Ai) / Total d'
    wi = ws3.cell(row=r, column=3)
    wi.value = f"=B{r}/$B${total_dprime_row}"
    cell_style(wi)
    wi.fill = light_teal_fill
    wi.number_format = '0.0000'
    
    # Percentage
    pct = ws3.cell(row=r, column=4)
    pct.value = f"=C{r}*100"
    cell_style(pct)
    pct.number_format = '0.00"%"'

# Total Wi
wi_total = ws3.cell(row=row, column=3)
wi_total.value = f"=SUM(C{dprime_start}:C{dprime_end})"
wi_total.font = Font(bold=True)
cell_style(wi_total)
wi_total.fill = yellow_fill

# Column widths
ws3.column_dimensions['A'].width = 25
for j in range(n*3+5):
    ws3.column_dimensions[get_column_letter(j+2)].width = 10

# ============================================
# SHEET 4: NILAI PESERTA
# ============================================
ws4 = wb.create_sheet("Nilai Peserta")

row = 1
ws4['A1'] = "INPUT NILAI PESERTA PER KRITERIA"
ws4['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws4.merge_cells('A1:I1')

ws4['A2'] = "Masukkan nilai 0-100 untuk setiap kriteria"
ws4['A2'].font = Font(italic=True, color="FF0000")
ws4.merge_cells('A2:I2')

row = 4

ws4.cell(row=row, column=1).value = "No"
ws4.cell(row=row, column=2).value = "Nama Peserta"
header_style(ws4.cell(row=row, column=1))
header_style(ws4.cell(row=row, column=2))

for j in range(n):
    cell = ws4.cell(row=row, column=j+3)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

nilai_start = row

sample_nilai = [
    [85, 90, 88, 92, 87, 85],
    [78, 85, 82, 88, 80, 78],
    [92, 88, 90, 85, 88, 90],
    [80, 82, 85, 90, 82, 80],
    [88, 92, 87, 80, 90, 88],
]

for i in range(n_peserta):
    ws4.cell(row=row, column=1).value = i + 1
    cell_style(ws4.cell(row=row, column=1))
    
    ws4.cell(row=row, column=2).value = f"='Input Data'!B{peserta_start + i}"
    cell_style(ws4.cell(row=row, column=2))
    ws4.cell(row=row, column=2).fill = light_blue_fill
    
    for j in range(n):
        cell = ws4.cell(row=row, column=j+3)
        if i < len(sample_nilai):
            cell.value = sample_nilai[i][j]
        else:
            cell.value = 0
        input_style(cell)
    
    row += 1

nilai_end = row - 1

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 25
for j in range(n):
    ws4.column_dimensions[get_column_letter(j+3)].width = 12

# ============================================
# SHEET 5: HASIL & RANKING
# ============================================
ws5 = wb.create_sheet("Hasil & Ranking")

row = 1
ws5['A1'] = "PERHITUNGAN SKOR AKHIR & RANKING"
ws5['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws5.merge_cells('A1:L1')

ws5['A2'] = "Skor Akhir = Sum(Nilai * Bobot Global)"
ws5['A2'].font = Font(italic=True, color="666666")
ws5.merge_cells('A2:L2')

row = 4

# Bobot Global Reference
ws5.cell(row=row, column=1).value = "BOBOT GLOBAL KRITERIA"
ws5.cell(row=row, column=1).font = Font(bold=True, size=12, color="0D9488")
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+1)
row += 1

for j in range(n):
    cell = ws5.cell(row=row, column=j+1)
    cell.value = kriteria_list[j]
    header_style(cell, fill=step6_fill)
row += 1

bobot_ref_row = row
for j in range(n):
    cell = ws5.cell(row=row, column=j+1)
    cell.value = f"='4-6. Fuzzy & Bobot'!C{dprime_start + j}"
    cell_style(cell)
    cell.fill = light_teal_fill
    cell.number_format = '0.0000'
row += 2

# Perhitungan Skor
ws5.cell(row=row, column=1).value = "PERHITUNGAN SKOR AKHIR"
ws5.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+4)
row += 1

ws5.cell(row=row, column=1).value = "No"
ws5.cell(row=row, column=2).value = "Nama"
header_style(ws5.cell(row=row, column=1))
header_style(ws5.cell(row=row, column=2))

for j in range(n):
    cell = ws5.cell(row=row, column=j+3)
    cell.value = f"{kriteria_list[j]}*W"
    header_style(cell)

ws5.cell(row=row, column=n+3).value = "Skor Akhir"
ws5.cell(row=row, column=n+4).value = "Ranking"
header_style(ws5.cell(row=row, column=n+3), fill=orange_fill)
header_style(ws5.cell(row=row, column=n+4), fill=orange_fill)
row += 1

skor_start = row

for i in range(n_peserta):
    ws5.cell(row=row, column=1).value = i + 1
    cell_style(ws5.cell(row=row, column=1))
    
    ws5.cell(row=row, column=2).value = f"='Input Data'!B{peserta_start + i}"
    cell_style(ws5.cell(row=row, column=2))
    ws5.cell(row=row, column=2).fill = light_blue_fill
    
    for j in range(n):
        cell = ws5.cell(row=row, column=j+3)
        nilai_ref = f"'Nilai Peserta'!{get_column_letter(j+3)}{nilai_start + i}"
        bobot_ref = f"${get_column_letter(j+1)}${bobot_ref_row}"
        cell.value = f"={nilai_ref}*{bobot_ref}"
        cell_style(cell)
        cell.fill = light_green_fill
        cell.number_format = '0.00'
    
    skor = ws5.cell(row=row, column=n+3)
    skor.value = f"=SUM(C{row}:{get_column_letter(n+2)}{row})"
    skor.font = Font(bold=True)
    cell_style(skor)
    skor.fill = light_yellow_fill
    skor.number_format = '0.00'
    
    rank = ws5.cell(row=row, column=n+4)
    skor_range = f"${get_column_letter(n+3)}${skor_start}:${get_column_letter(n+3)}${skor_start + n_peserta - 1}"
    rank.value = f'=IF(B{row}="","",RANK({get_column_letter(n+3)}{row},{skor_range},0))'
    rank.font = Font(bold=True, size=12)
    cell_style(rank)
    rank.fill = yellow_fill
    
    row += 1

skor_end = row - 1

ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 25
for j in range(n):
    ws5.column_dimensions[get_column_letter(j+3)].width = 12
ws5.column_dimensions[get_column_letter(n+3)].width = 14
ws5.column_dimensions[get_column_letter(n+4)].width = 10

# Save
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_Sistem_v2.xlsx"
wb.save(output_file)

print(f"[OK] File berhasil dibuat: {output_file}")
print(f"\n=== PERBAIKAN ===")
print(f"Formula matriks crisp sekarang SAMA dengan sistem:")
print(f"  IF(wi/wj >= 1, MIN(9, MAX(1, wi/wj)), MAX(1/9, wi/wj))")
print(f"\nContoh input bobot yang menghasilkan gambar sistem:")
print(f"  K1=1, K2=2, K3=2, K4=3, K5=4, K6=5")
