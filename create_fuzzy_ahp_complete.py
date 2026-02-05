"""
Script untuk membuat Model Excel LENGKAP Perhitungan Fuzzy AHP
Fitur:
- Input data kegiatan
- Input bobot kriteria (pairwise comparison)
- Input nilai peserta per kriteria
- Perhitungan skor akhir dengan Fuzzy AHP
- Ranking peserta
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
input_font = Font(size=11, color="0000FF")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

kriteria_list = ["K1", "K2", "K3", "K4", "K5", "K6"]
kriteria_full = [
    "Status Keaktifan di Gugus Depan",
    "Pencapaian SKU", 
    "Pencapaian SPG",
    "Kesehatan Jasmani dan Rohani",
    "Tes Wawancara",
    "Tes Pilihan Ganda"
]
n_kriteria = 6
n_peserta = 10  # Support up to 10 participants

def header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def input_style(cell):
    cell.font = input_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = input_fill

def section(ws, row, text, col_end=10):
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)

# ============================================
# SHEET 1: DATA KEGIATAN & KRITERIA
# ============================================
ws1 = wb.active
ws1.title = "1. Data Input"

row = 1
ws1['A1'] = "SISTEM PENDUKUNG KEPUTUSAN - FUZZY AHP"
ws1['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws1.merge_cells('A1:J1')

ws1['A2'] = "Spreadsheet Interaktif untuk Validasi Perhitungan"
ws1['A2'].font = Font(italic=True, size=11, color="666666")
ws1.merge_cells('A2:J2')

row = 4

# DATA KEGIATAN
section(ws1, row, "DATA KEGIATAN")
row += 1

ws1.cell(row=row, column=1).value = "Nama Kegiatan:"
ws1.cell(row=row, column=1).font = Font(bold=True)
ws1.cell(row=row, column=2).value = "Raimuna Daerah"
input_style(ws1.cell(row=row, column=2))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Tanggal:"
ws1.cell(row=row, column=1).font = Font(bold=True)
ws1.cell(row=row, column=2).value = "25 Januari 2026"
input_style(ws1.cell(row=row, column=2))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Kuota Peserta:"
ws1.cell(row=row, column=1).font = Font(bold=True)
ws1.cell(row=row, column=2).value = 5
input_style(ws1.cell(row=row, column=2))
row += 2

# DAFTAR KRITERIA
section(ws1, row, "DAFTAR KRITERIA PENILAIAN")
row += 1

ws1.cell(row=row, column=1).value = "Kode"
ws1.cell(row=row, column=2).value = "Nama Kriteria"
ws1.cell(row=row, column=3).value = "Bobot (dari Sheet 2)"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
row += 1

kriteria_start_row = row
for i in range(n_kriteria):
    ws1.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws1.cell(row=row, column=1))
    ws1.cell(row=row, column=1).fill = light_blue_fill
    
    ws1.cell(row=row, column=2).value = kriteria_full[i]
    cell_style(ws1.cell(row=row, column=2))
    
    # Reference to weight from Sheet 2
    ws1.cell(row=row, column=3).value = f"='2. Bobot Kriteria'!I{6+i}"
    cell_style(ws1.cell(row=row, column=3))
    ws1.cell(row=row, column=3).fill = light_green_fill
    ws1.cell(row=row, column=3).number_format = '0.0000'
    
    row += 1

row += 1

# DAFTAR PESERTA
section(ws1, row, "DAFTAR PESERTA")
row += 1

ws1.cell(row=row, column=1).value = "No"
ws1.cell(row=row, column=2).value = "Nama Peserta"
ws1.cell(row=row, column=3).value = "Asal"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
row += 1

peserta_start_row = row
sample_peserta = [
    ("David Kulian", "Merauke"),
    ("Siti Aminah", "Jayapura"),
    ("Budi Santoso", "Sorong"),
    ("Dewi Lestari", "Manokwari"),
    ("Ahmad Fauzi", "Biak"),
    ("", ""),
    ("", ""),
    ("", ""),
    ("", ""),
    ("", "")
]

for i in range(n_peserta):
    ws1.cell(row=row, column=1).value = i + 1
    cell_style(ws1.cell(row=row, column=1))
    
    ws1.cell(row=row, column=2).value = sample_peserta[i][0]
    input_style(ws1.cell(row=row, column=2))
    
    ws1.cell(row=row, column=3).value = sample_peserta[i][1]
    input_style(ws1.cell(row=row, column=3))
    
    row += 1

# Column widths
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 20

# ============================================
# SHEET 2: BOBOT KRITERIA (Pairwise Comparison)
# ============================================
ws2 = wb.create_sheet("2. Bobot Kriteria")

row = 1
ws2['A1'] = "PERHITUNGAN BOBOT KRITERIA - FUZZY AHP"
ws2['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws2.merge_cells('A1:J1')

ws2['A2'] = "Masukkan nilai perbandingan berpasangan (1-9) di sel KUNING"
ws2['A2'].font = Font(italic=True, color="FF0000")
ws2.merge_cells('A2:J2')

row = 4

# MATRIKS PERBANDINGAN
section(ws2, row, "MATRIKS PERBANDINGAN BERPASANGAN (CRISP)")
row += 1

# Header
ws2.cell(row=row, column=1).value = "Kriteria"
header_style(ws2.cell(row=row, column=1))
for j in range(n_kriteria):
    cell = ws2.cell(row=row, column=j+2)
    cell.value = kriteria_list[j]
    header_style(cell)

ws2.cell(row=row, column=n_kriteria+2).value = "GM"
ws2.cell(row=row, column=n_kriteria+3).value = "Bobot (Wi)"
header_style(ws2.cell(row=row, column=n_kriteria+2), fill=green_fill)
header_style(ws2.cell(row=row, column=n_kriteria+3), fill=green_fill)
row += 1

crisp_start = row

for i in range(n_kriteria):
    cell = ws2.cell(row=row, column=1)
    cell.value = kriteria_list[i]
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.fill = light_blue_fill
    
    for j in range(n_kriteria):
        cell = ws2.cell(row=row, column=j+2)
        if i == j:
            cell.value = 1
            cell_style(cell)
        elif i < j:
            cell.value = 1
            input_style(cell)
        else:
            upper_row = crisp_start + j
            upper_col = get_column_letter(i + 2)
            cell.value = f"=1/{upper_col}{upper_row}"
            cell_style(cell)
            cell.fill = light_green_fill
    
    # GM
    gm = ws2.cell(row=row, column=n_kriteria+2)
    gm.value = f"=(B{row}*C{row}*D{row}*E{row}*F{row}*G{row})^(1/6)"
    cell_style(gm)
    gm.fill = light_green_fill
    gm.number_format = '0.0000'
    
    row += 1

crisp_end = row - 1

# Total GM
ws2.cell(row=row, column=n_kriteria+1).value = "Total:"
ws2.cell(row=row, column=n_kriteria+1).font = Font(bold=True)
gm_sum = ws2.cell(row=row, column=n_kriteria+2)
gm_sum.value = f"=SUM(H{crisp_start}:H{crisp_end})"
gm_sum.font = Font(bold=True)
cell_style(gm_sum)
gm_sum.fill = yellow_fill
gm_sum_ref = f"$H${row}"

# Bobot (Wi)
for i in range(n_kriteria):
    r = crisp_start + i
    wi = ws2.cell(row=r, column=n_kriteria+3)
    wi.value = f"=H{r}/{gm_sum_ref}"
    cell_style(wi)
    wi.fill = light_green_fill
    wi.number_format = '0.0000'

wi_sum = ws2.cell(row=row, column=n_kriteria+3)
wi_sum.value = f"=SUM(I{crisp_start}:I{crisp_end})"
wi_sum.font = Font(bold=True)
cell_style(wi_sum)
wi_sum.fill = yellow_fill

row += 2

# UJI KONSISTENSI
section(ws2, row, "UJI KONSISTENSI")
row += 1

ws2.cell(row=row, column=1).value = "Kriteria"
ws2.cell(row=row, column=2).value = "A*w"
ws2.cell(row=row, column=3).value = "(A*w)/w"
header_style(ws2.cell(row=row, column=1))
header_style(ws2.cell(row=row, column=2))
header_style(ws2.cell(row=row, column=3))
row += 1

aw_start = row

for i in range(n_kriteria):
    ws2.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws2.cell(row=row, column=1))
    ws2.cell(row=row, column=1).fill = light_blue_fill
    
    crisp_row = crisp_start + i
    
    # A*w
    aw = ws2.cell(row=row, column=2)
    parts = []
    for j in range(n_kriteria):
        col = get_column_letter(j + 2)
        parts.append(f"{col}{crisp_row}*$I${crisp_start + j}")
    aw.value = "=" + "+".join(parts)
    cell_style(aw)
    aw.fill = light_green_fill
    aw.number_format = '0.0000'
    
    # Ratio
    ratio = ws2.cell(row=row, column=3)
    ratio.value = f"=B{row}/$I${crisp_start + i}"
    cell_style(ratio)
    ratio.fill = light_green_fill
    ratio.number_format = '0.0000'
    
    row += 1

aw_end = row - 1
row += 1

# Lambda, CI, CR
lambda_row = row
ws2.cell(row=row, column=1).value = "Lambda Max:"
ws2.cell(row=row, column=1).font = subtitle_font
lmax = ws2.cell(row=row, column=2)
lmax.value = f"=AVERAGE(C{aw_start}:C{aw_end})"
lmax.font = Font(bold=True, size=14, color="ED7D31")
lmax.fill = yellow_fill
cell_style(lmax)
lmax.number_format = '0.0000'
row += 1

ci_row = row
ws2.cell(row=row, column=1).value = "CI:"
ws2.cell(row=row, column=1).font = subtitle_font
ci = ws2.cell(row=row, column=2)
ci.value = f"=(B{lambda_row}-6)/5"
ci.font = Font(bold=True)
ci.fill = light_green_fill
cell_style(ci)
ci.number_format = '0.0000'
row += 1

ri_row = row
ws2.cell(row=row, column=1).value = "RI (n=6):"
ws2.cell(row=row, column=1).font = subtitle_font
ri = ws2.cell(row=row, column=2)
ri.value = 1.24
ri.font = Font(bold=True)
cell_style(ri)
row += 1

cr_row = row
ws2.cell(row=row, column=1).value = "CR:"
ws2.cell(row=row, column=1).font = subtitle_font
cr = ws2.cell(row=row, column=2)
cr.value = f"=B{ci_row}/B{ri_row}"
cr.font = Font(bold=True, size=14)
cr.fill = yellow_fill
cell_style(cr)
cr.number_format = '0.0000'
row += 1

ws2.cell(row=row, column=1).value = "Status:"
ws2.cell(row=row, column=1).font = subtitle_font
status = ws2.cell(row=row, column=2)
status.value = f'=IF(B{cr_row}<=0.1,"KONSISTEN","TIDAK KONSISTEN")'
status.font = Font(bold=True, size=12)
cell_style(status)
ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

# Column widths
for c in ['A','B','C','D','E','F','G','H','I']:
    ws2.column_dimensions[c].width = 14

# ============================================
# SHEET 3: NILAI PESERTA
# ============================================
ws3 = wb.create_sheet("3. Nilai Peserta")

row = 1
ws3['A1'] = "INPUT NILAI PESERTA PER KRITERIA"
ws3['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws3.merge_cells('A1:J1')

ws3['A2'] = "Masukkan nilai 0-100 untuk setiap kriteria di sel KUNING"
ws3['A2'].font = Font(italic=True, color="FF0000")
ws3.merge_cells('A2:J2')

row = 4

section(ws3, row, "NILAI PESERTA", col_end=n_kriteria+2)
row += 1

# Header
ws3.cell(row=row, column=1).value = "No"
ws3.cell(row=row, column=2).value = "Nama Peserta"
header_style(ws3.cell(row=row, column=1))
header_style(ws3.cell(row=row, column=2))

for j in range(n_kriteria):
    cell = ws3.cell(row=row, column=j+3)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

nilai_start = row

# Sample values
sample_nilai = [
    [85, 90, 88, 92, 87, 85],
    [78, 85, 82, 88, 80, 78],
    [92, 88, 90, 85, 88, 90],
    [80, 82, 85, 90, 82, 80],
    [88, 92, 87, 80, 90, 88],
]

for i in range(n_peserta):
    ws3.cell(row=row, column=1).value = i + 1
    cell_style(ws3.cell(row=row, column=1))
    
    # Reference to name from Sheet 1
    ws3.cell(row=row, column=2).value = f"='1. Data Input'!B{peserta_start_row + i}"
    cell_style(ws3.cell(row=row, column=2))
    ws3.cell(row=row, column=2).fill = light_blue_fill
    
    for j in range(n_kriteria):
        cell = ws3.cell(row=row, column=j+3)
        if i < len(sample_nilai):
            cell.value = sample_nilai[i][j]
        else:
            cell.value = 0
        input_style(cell)
    
    row += 1

nilai_end = row - 1

# Column widths
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 25
for j in range(n_kriteria):
    ws3.column_dimensions[get_column_letter(j+3)].width = 12

# ============================================
# SHEET 4: PERHITUNGAN & RANKING
# ============================================
ws4 = wb.create_sheet("4. Hasil & Ranking")

row = 1
ws4['A1'] = "PERHITUNGAN SKOR AKHIR & RANKING"
ws4['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws4.merge_cells('A1:L1')

ws4['A2'] = "Skor Akhir = Σ (Nilai × Bobot)"
ws4['A2'].font = Font(italic=True, color="666666")
ws4.merge_cells('A2:L2')

row = 4

# BOBOT KRITERIA (Reference)
section(ws4, row, "BOBOT KRITERIA (dari Sheet 2)", col_end=n_kriteria+1)
row += 1

for j in range(n_kriteria):
    cell = ws4.cell(row=row, column=j+1)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

bobot_ref_row = row
for j in range(n_kriteria):
    cell = ws4.cell(row=row, column=j+1)
    cell.value = f"='2. Bobot Kriteria'!I{crisp_start + j}"
    cell_style(cell)
    cell.fill = light_green_fill
    cell.number_format = '0.0000'
row += 2

# PERHITUNGAN SKOR
section(ws4, row, "PERHITUNGAN SKOR AKHIR", col_end=n_kriteria+4)
row += 1

# Header
ws4.cell(row=row, column=1).value = "No"
ws4.cell(row=row, column=2).value = "Nama Peserta"
header_style(ws4.cell(row=row, column=1))
header_style(ws4.cell(row=row, column=2))

for j in range(n_kriteria):
    cell = ws4.cell(row=row, column=j+3)
    cell.value = f"{kriteria_list[j]}×W{j+1}"
    header_style(cell)

ws4.cell(row=row, column=n_kriteria+3).value = "Skor Akhir"
ws4.cell(row=row, column=n_kriteria+4).value = "Ranking"
header_style(ws4.cell(row=row, column=n_kriteria+3), fill=orange_fill)
header_style(ws4.cell(row=row, column=n_kriteria+4), fill=orange_fill)
row += 1

skor_start = row

for i in range(n_peserta):
    ws4.cell(row=row, column=1).value = i + 1
    cell_style(ws4.cell(row=row, column=1))
    
    # Reference to name
    ws4.cell(row=row, column=2).value = f"='1. Data Input'!B{peserta_start_row + i}"
    cell_style(ws4.cell(row=row, column=2))
    ws4.cell(row=row, column=2).fill = light_blue_fill
    
    # Nilai × Bobot for each kriteria
    for j in range(n_kriteria):
        cell = ws4.cell(row=row, column=j+3)
        nilai_ref = f"'3. Nilai Peserta'!{get_column_letter(j+3)}{nilai_start + i}"
        bobot_ref = f"${get_column_letter(j+1)}${bobot_ref_row}"
        cell.value = f"={nilai_ref}*{bobot_ref}"
        cell_style(cell)
        cell.fill = light_green_fill
        cell.number_format = '0.00'
    
    # Skor Akhir = SUM
    skor = ws4.cell(row=row, column=n_kriteria+3)
    skor.value = f"=SUM(C{row}:{get_column_letter(n_kriteria+2)}{row})"
    skor.font = Font(bold=True)
    cell_style(skor)
    skor.fill = light_yellow_fill
    skor.number_format = '0.00'
    
    # Ranking (only if name exists)
    rank = ws4.cell(row=row, column=n_kriteria+4)
    skor_range = f"${get_column_letter(n_kriteria+3)}${skor_start}:${get_column_letter(n_kriteria+3)}${skor_start + n_peserta - 1}"
    rank.value = f'=IF(B{row}="","",RANK({get_column_letter(n_kriteria+3)}{row},{skor_range},0))'
    rank.font = Font(bold=True, size=12)
    cell_style(rank)
    rank.fill = yellow_fill
    
    row += 1

skor_end = row - 1
row += 2

# HASIL RANKING FINAL
section(ws4, row, "HASIL RANKING FINAL", col_end=4)
row += 1

ws4.cell(row=row, column=1).value = "Ranking"
ws4.cell(row=row, column=2).value = "Nama Peserta"
ws4.cell(row=row, column=3).value = "Skor Akhir"
ws4.cell(row=row, column=4).value = "Status"
header_style(ws4.cell(row=row, column=1), fill=orange_fill)
header_style(ws4.cell(row=row, column=2), fill=orange_fill)
header_style(ws4.cell(row=row, column=3), fill=orange_fill)
header_style(ws4.cell(row=row, column=4), fill=orange_fill)
row += 1

ranking_start = row
for i in range(n_peserta):
    ws4.cell(row=row, column=1).value = i + 1
    cell_style(ws4.cell(row=row, column=1))
    ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
    
    # Find peserta with this rank using INDEX/MATCH alternative with simple lookup
    # Using simpler approach - reference directly
    ws4.cell(row=row, column=2).value = f'=IFERROR(INDEX(B{skor_start}:B{skor_end},MATCH({i+1},{get_column_letter(n_kriteria+4)}{skor_start}:{get_column_letter(n_kriteria+4)}{skor_end},0)),"")'
    cell_style(ws4.cell(row=row, column=2))
    
    ws4.cell(row=row, column=3).value = f'=IFERROR(INDEX({get_column_letter(n_kriteria+3)}{skor_start}:{get_column_letter(n_kriteria+3)}{skor_end},MATCH({i+1},{get_column_letter(n_kriteria+4)}{skor_start}:{get_column_letter(n_kriteria+4)}{skor_end},0)),"")'
    cell_style(ws4.cell(row=row, column=3))
    ws4.cell(row=row, column=3).number_format = '0.00'
    
    # Status: Lolos/Tidak Lolos based on kuota
    ws4.cell(row=row, column=4).value = f'=IF(A{row}<="\'1. Data Input\'!B8","LOLOS","TIDAK LOLOS")'
    cell_style(ws4.cell(row=row, column=4))
    
    row += 1

# Column widths
ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 25
for j in range(n_kriteria):
    ws4.column_dimensions[get_column_letter(j+3)].width = 12
ws4.column_dimensions[get_column_letter(n_kriteria+3)].width = 14
ws4.column_dimensions[get_column_letter(n_kriteria+4)].width = 10

# ============================================
# SHEET 5: TABEL REFERENSI
# ============================================
ws5 = wb.create_sheet("5. Referensi")

row = 1
ws5['A1'] = "TABEL REFERENSI"
ws5['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws5.merge_cells('A1:E1')

row = 3

# Skala Saaty
section(ws5, row, "SKALA PERBANDINGAN SAATY", col_end=5)
row += 1

ws5.cell(row=row, column=1).value = "Nilai"
ws5.cell(row=row, column=2).value = "Definisi"
header_style(ws5.cell(row=row, column=1))
header_style(ws5.cell(row=row, column=2))
ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
row += 1

saaty = [
    (1, "Kedua elemen sama penting"),
    (3, "Elemen satu sedikit lebih penting"),
    (5, "Elemen satu lebih penting"),
    (7, "Satu elemen jelas lebih penting"),
    (9, "Satu elemen mutlak lebih penting"),
    ("2,4,6,8", "Nilai antara dua penilaian berdekatan")
]

for val, desc in saaty:
    ws5.cell(row=row, column=1).value = val
    ws5.cell(row=row, column=2).value = desc
    cell_style(ws5.cell(row=row, column=1))
    cell_style(ws5.cell(row=row, column=2))
    ws5.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

row += 1

# TFN
section(ws5, row, "SKALA FUZZY (TFN)", col_end=5)
row += 1

ws5.cell(row=row, column=1).value = "Nilai"
ws5.cell(row=row, column=2).value = "l"
ws5.cell(row=row, column=3).value = "m"
ws5.cell(row=row, column=4).value = "u"
header_style(ws5.cell(row=row, column=1))
header_style(ws5.cell(row=row, column=2))
header_style(ws5.cell(row=row, column=3))
header_style(ws5.cell(row=row, column=4))
row += 1

tfn = [(1,1,1,1), (2,0.5,1,1.5), (3,1,1.5,2), (4,1.5,2,2.5), 
       (5,2,2.5,3), (6,2.5,3,3.5), (7,3,3.5,4), (8,3.5,4,4.5), (9,4,4.5,4.5)]

for v, l, m, u in tfn:
    ws5.cell(row=row, column=1).value = v
    ws5.cell(row=row, column=2).value = l
    ws5.cell(row=row, column=3).value = m
    ws5.cell(row=row, column=4).value = u
    for c in range(1, 5):
        cell_style(ws5.cell(row=row, column=c))
    row += 1

row += 1

# Random Index
section(ws5, row, "RANDOM INDEX (RI)", col_end=5)
row += 1

ws5.cell(row=row, column=1).value = "n"
for i in range(1, 11):
    ws5.cell(row=row, column=i+1).value = i
    header_style(ws5.cell(row=row, column=i+1))
header_style(ws5.cell(row=row, column=1))
row += 1

ri_vals = [0, 0, 0.58, 0.90, 1.12, 1.24, 1.32, 1.41, 1.46, 1.49]
ws5.cell(row=row, column=1).value = "RI"
cell_style(ws5.cell(row=row, column=1))
for i, ri in enumerate(ri_vals):
    ws5.cell(row=row, column=i+2).value = ri
    cell_style(ws5.cell(row=row, column=i+2))

# Column widths
ws5.column_dimensions['A'].width = 12
ws5.column_dimensions['B'].width = 40
for c in ['C', 'D', 'E']:
    ws5.column_dimensions[c].width = 10

# Save
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_Complete.xlsx"
wb.save(output_file)

print(f"[OK] File berhasil dibuat: {output_file}")
print(f"\n=== STRUKTUR FILE ===")
print(f"Sheet 1: Data Input")
print(f"  - Data kegiatan (nama, tanggal, kuota)")
print(f"  - Daftar kriteria dengan bobot")
print(f"  - Daftar peserta (max 10)")
print(f"\nSheet 2: Bobot Kriteria")
print(f"  - Matriks perbandingan berpasangan (INPUT)")
print(f"  - Perhitungan GM dan Bobot")
print(f"  - Uji konsistensi (CI, RI, CR)")
print(f"\nSheet 3: Nilai Peserta")
print(f"  - Input nilai 0-100 per kriteria per peserta")
print(f"\nSheet 4: Hasil & Ranking")
print(f"  - Perhitungan skor akhir")
print(f"  - Ranking otomatis")
print(f"\nSheet 5: Referensi")
print(f"  - Skala Saaty")
print(f"  - Skala TFN")
print(f"  - Random Index")
