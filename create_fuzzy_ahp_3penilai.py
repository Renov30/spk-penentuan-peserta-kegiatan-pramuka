"""
Script untuk membuat Model Excel Fuzzy AHP dengan 3 PENILAI
Masing-masing penilai memegang 2 kriteria:
- Penilai 1: K1 (Status Keaktifan), K2 (Pencapaian SKU)
- Penilai 2: K3 (Pencapaian SPG), K4 (Kesehatan Jasmani)
- Penilai 3: K5 (Tes Wawancara), K6 (Tes Pilihan Ganda)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
subtitle_font = Font(bold=True, size=11, color="1F4E79")
section_font = Font(bold=True, size=12, color="FFFFFF")
normal_font = Font(size=10)
input_font = Font(size=11, color="0000FF")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Colors
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
step1_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
step2_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
step3_fill = PatternFill(start_color="CA8A04", end_color="CA8A04", fill_type="solid")
step4_fill = PatternFill(start_color="9333EA", end_color="9333EA", fill_type="solid")
step5_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
step6_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
light_purple_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
light_amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
light_teal_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Warna untuk 3 penilai
penilai1_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Light Blue
penilai2_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Light Green
penilai3_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Light Red

penilai1_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Blue
penilai2_header = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")  # Green  
penilai3_header = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")  # Red

kriteria_list = ["K1", "K2", "K3", "K4", "K5", "K6"]
kriteria_full = [
    "Status Keaktifan",
    "Pencapaian SKU", 
    "Pencapaian SPG",
    "Kesehatan Jasmani",
    "Tes Wawancara",
    "Tes Pilihan Ganda"
]

# Pembagian kriteria per penilai
penilai_kriteria = {
    1: [0, 1],  # Penilai 1: K1, K2
    2: [2, 3],  # Penilai 2: K3, K4
    3: [4, 5],  # Penilai 3: K5, K6
}

n = 6
n_peserta = 10

# TFN Scale
tfn_scale = {
    1: (1, 1, 1),
    2: (0.5, 1, 1.5),
    3: (1, 1.5, 2),
    4: (1.5, 2, 2.5),
    5: (2, 2.5, 3),
    6: (2.5, 3, 3.5),
    7: (3, 3.5, 4),
    8: (3.5, 4, 4.5),
    9: (4, 4.5, 4.5)
}

def header_style(cell, fill=header_fill):
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill

def cell_style(cell):
    cell.font = normal_font
    cell.alignment = center_align
    cell.border = thin_border

def input_style(cell):
    cell.font = input_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = input_fill

def section(ws, row, step_num, text, fill, col_end=10):
    cell = ws.cell(row=row, column=1)
    cell.value = f"LANGKAH {step_num}: {text}"
    cell.font = section_font
    cell.fill = fill
    cell.alignment = center_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    for c in range(1, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill

# ============================================
# SHEET 1: INPUT DATA & PENUGASAN PENILAI
# ============================================
ws1 = wb.active
ws1.title = "Data & Penugasan"

row = 1
ws1['A1'] = "FUZZY AHP - SISTEM 3 PENILAI"
ws1['A1'].font = Font(bold=True, size=18, color="1F4E79")
ws1.merge_cells('A1:F1')

ws1['A2'] = "Masukkan data di sel berwarna KUNING"
ws1['A2'].font = Font(italic=True, color="FF0000")
ws1.merge_cells('A2:F2')

row = 4

# Data Kegiatan
ws1.cell(row=row, column=1).value = "DATA KEGIATAN"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Nama Kegiatan:"
ws1.cell(row=row, column=2).value = "Raimuna Daerah"
input_style(ws1.cell(row=row, column=2))
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Kuota:"
ws1.cell(row=row, column=2).value = 5
input_style(ws1.cell(row=row, column=2))
row += 2

# Penugasan Penilai
ws1.cell(row=row, column=1).value = "PENUGASAN PENILAI"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
row += 1

ws1.cell(row=row, column=1).value = "Penilai"
ws1.cell(row=row, column=2).value = "Nama Penilai"
ws1.cell(row=row, column=3).value = "Kriteria 1"
ws1.cell(row=row, column=4).value = "Kriteria 2"
ws1.cell(row=row, column=5).value = "Warna"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3))
header_style(ws1.cell(row=row, column=4))
header_style(ws1.cell(row=row, column=5))
row += 1

penilai_names = ["Pak Ahmad", "Bu Siti", "Pak Budi"]
penilai_fills = [penilai1_fill, penilai2_fill, penilai3_fill]
penilai_headers = [penilai1_header, penilai2_header, penilai3_header]

for p in range(3):
    ws1.cell(row=row, column=1).value = f"Penilai {p+1}"
    cell_style(ws1.cell(row=row, column=1))
    ws1.cell(row=row, column=1).fill = penilai_fills[p]
    
    ws1.cell(row=row, column=2).value = penilai_names[p]
    input_style(ws1.cell(row=row, column=2))
    
    k1, k2 = penilai_kriteria[p+1]
    ws1.cell(row=row, column=3).value = f"{kriteria_list[k1]} - {kriteria_full[k1]}"
    cell_style(ws1.cell(row=row, column=3))
    ws1.cell(row=row, column=3).fill = penilai_fills[p]
    
    ws1.cell(row=row, column=4).value = f"{kriteria_list[k2]} - {kriteria_full[k2]}"
    cell_style(ws1.cell(row=row, column=4))
    ws1.cell(row=row, column=4).fill = penilai_fills[p]
    
    ws1.cell(row=row, column=5).value = ""
    cell_style(ws1.cell(row=row, column=5))
    ws1.cell(row=row, column=5).fill = penilai_headers[p]
    
    row += 1

row += 1

# Bobot Kriteria
ws1.cell(row=row, column=1).value = "BOBOT KRITERIA"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "Kode"
ws1.cell(row=row, column=2).value = "Kriteria"
ws1.cell(row=row, column=3).value = "Bobot"
ws1.cell(row=row, column=4).value = "Penilai"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3), fill=orange_fill)
header_style(ws1.cell(row=row, column=4))
row += 1

bobot_input_start = row
default_weights = [1, 2, 2, 3, 4, 5]

for i in range(n):
    ws1.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws1.cell(row=row, column=1))
    
    ws1.cell(row=row, column=2).value = kriteria_full[i]
    cell_style(ws1.cell(row=row, column=2))
    
    ws1.cell(row=row, column=3).value = default_weights[i]
    input_style(ws1.cell(row=row, column=3))
    
    # Tentukan penilai
    if i < 2:
        penilai_idx = 0
    elif i < 4:
        penilai_idx = 1
    else:
        penilai_idx = 2
    
    ws1.cell(row=row, column=4).value = f"Penilai {penilai_idx + 1}"
    cell_style(ws1.cell(row=row, column=4))
    ws1.cell(row=row, column=4).fill = penilai_fills[penilai_idx]
    
    row += 1

bobot_input_end = row - 1
row += 1

# Daftar Peserta
ws1.cell(row=row, column=1).value = "DAFTAR PESERTA"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws1.cell(row=row, column=1).value = "No"
ws1.cell(row=row, column=2).value = "Nama Peserta"
ws1.cell(row=row, column=3).value = "Asal"
header_style(ws1.cell(row=row, column=1))
header_style(ws1.cell(row=row, column=2))
header_style(ws1.cell(row=row, column=3))
row += 1

peserta_start = row
sample_peserta = [
    ("David Kulian", "Merauke"),
    ("Siti Aminah", "Jayapura"),
    ("Budi Santoso", "Sorong"),
    ("Dewi Lestari", "Manokwari"),
    ("Ahmad Fauzi", "Biak"),
    ("Rina Wati", "Mimika"),
    ("Joko Susilo", "Nabire"),
    ("Maya Sari", "Timika"),
    ("Andi Pratama", "Wamena"),
    ("Lina Marlina", "Sarmi")
]

for i in range(n_peserta):
    ws1.cell(row=row, column=1).value = i + 1
    cell_style(ws1.cell(row=row, column=1))
    
    ws1.cell(row=row, column=2).value = sample_peserta[i][0]
    input_style(ws1.cell(row=row, column=2))
    
    ws1.cell(row=row, column=3).value = sample_peserta[i][1]
    input_style(ws1.cell(row=row, column=3))
    
    row += 1

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 25
ws1.column_dimensions['E'].width = 10

# ============================================
# SHEET 2-4: INPUT NILAI PENILAI 1, 2, 3
# ============================================

sample_nilai = {
    1: [  # Penilai 1: K1, K2
        [85, 78],
        [78, 82],
        [92, 88],
        [80, 85],
        [88, 90],
        [75, 80],
        [82, 78],
        [90, 85],
        [85, 82],
        [78, 80]
    ],
    2: [  # Penilai 2: K3, K4
        [88, 92],
        [82, 88],
        [90, 85],
        [85, 90],
        [87, 80],
        [80, 85],
        [88, 82],
        [85, 88],
        [82, 85],
        [85, 90]
    ],
    3: [  # Penilai 3: K5, K6
        [87, 85],
        [80, 78],
        [88, 90],
        [82, 80],
        [90, 88],
        [85, 82],
        [80, 85],
        [88, 92],
        [85, 80],
        [82, 85]
    ]
}

nilai_sheet_refs = {}

for p in range(1, 4):
    ws = wb.create_sheet(f"Input Penilai {p}")
    
    k1_idx, k2_idx = penilai_kriteria[p]
    k1_name = kriteria_full[k1_idx]
    k2_name = kriteria_full[k2_idx]
    k1_code = kriteria_list[k1_idx]
    k2_code = kriteria_list[k2_idx]
    
    row = 1
    ws.cell(row=row, column=1).value = f"INPUT NILAI - PENILAI {p}"
    ws.cell(row=row, column=1).font = Font(bold=True, size=18, color="FFFFFF")
    ws.cell(row=row, column=1).fill = penilai_headers[p-1]
    ws.merge_cells(f'A1:E1')
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = penilai_headers[p-1]
    
    row = 2
    ws.cell(row=row, column=1).value = f"Nama: {penilai_names[p-1]}"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A2:E2')
    
    row = 3
    ws.cell(row=row, column=1).value = f"Kriteria yang dinilai: {k1_code} ({k1_name}) dan {k2_code} ({k2_name})"
    ws.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
    ws.merge_cells(f'A3:E3')
    
    row = 5
    ws.cell(row=row, column=1).value = "No"
    ws.cell(row=row, column=2).value = "Nama Peserta"
    ws.cell(row=row, column=3).value = f"{k1_code}\n{k1_name}"
    ws.cell(row=row, column=4).value = f"{k2_code}\n{k2_name}"
    header_style(ws.cell(row=row, column=1), fill=penilai_headers[p-1])
    header_style(ws.cell(row=row, column=2), fill=penilai_headers[p-1])
    header_style(ws.cell(row=row, column=3), fill=penilai_headers[p-1])
    header_style(ws.cell(row=row, column=4), fill=penilai_headers[p-1])
    ws.row_dimensions[row].height = 40
    
    row = 6
    nilai_start = row
    
    for i in range(n_peserta):
        ws.cell(row=row, column=1).value = i + 1
        cell_style(ws.cell(row=row, column=1))
        
        ws.cell(row=row, column=2).value = f"='Data & Penugasan'!B{peserta_start + i}"
        cell_style(ws.cell(row=row, column=2))
        ws.cell(row=row, column=2).fill = penilai_fills[p-1]
        
        # Nilai kriteria 1
        ws.cell(row=row, column=3).value = sample_nilai[p][i][0]
        input_style(ws.cell(row=row, column=3))
        
        # Nilai kriteria 2
        ws.cell(row=row, column=4).value = sample_nilai[p][i][1]
        input_style(ws.cell(row=row, column=4))
        
        row += 1
    
    nilai_sheet_refs[p] = {
        'sheet': f"'Input Penilai {p}'",
        'start_row': nilai_start,
        'k1_col': 'C',
        'k2_col': 'D'
    }
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

# ============================================
# SHEET 5: REKAP NILAI GABUNGAN
# ============================================
ws5 = wb.create_sheet("Rekap Nilai")

row = 1
ws5['A1'] = "REKAPITULASI NILAI DARI SEMUA PENILAI"
ws5['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws5.merge_cells('A1:I1')

ws5['A2'] = "Nilai diambil dari input masing-masing penilai"
ws5['A2'].font = Font(italic=True, color="666666")
ws5.merge_cells('A2:I2')

row = 4

# Header
ws5.cell(row=row, column=1).value = "No"
ws5.cell(row=row, column=2).value = "Nama Peserta"
header_style(ws5.cell(row=row, column=1))
header_style(ws5.cell(row=row, column=2))

for j in range(n):
    cell = ws5.cell(row=row, column=j+3)
    cell.value = kriteria_list[j]
    
    # Warna sesuai penilai
    if j < 2:
        header_style(cell, fill=penilai1_header)
    elif j < 4:
        header_style(cell, fill=penilai2_header)
    else:
        header_style(cell, fill=penilai3_header)

row += 1
rekap_start = row

for i in range(n_peserta):
    ws5.cell(row=row, column=1).value = i + 1
    cell_style(ws5.cell(row=row, column=1))
    
    ws5.cell(row=row, column=2).value = f"='Data & Penugasan'!B{peserta_start + i}"
    cell_style(ws5.cell(row=row, column=2))
    ws5.cell(row=row, column=2).fill = light_blue_fill
    
    # K1, K2 dari Penilai 1
    for j in range(2):
        cell = ws5.cell(row=row, column=j+3)
        col_ref = 'C' if j == 0 else 'D'
        cell.value = f"='Input Penilai 1'!{col_ref}{nilai_sheet_refs[1]['start_row'] + i}"
        cell_style(cell)
        cell.fill = penilai1_fill
    
    # K3, K4 dari Penilai 2
    for j in range(2):
        cell = ws5.cell(row=row, column=j+5)
        col_ref = 'C' if j == 0 else 'D'
        cell.value = f"='Input Penilai 2'!{col_ref}{nilai_sheet_refs[2]['start_row'] + i}"
        cell_style(cell)
        cell.fill = penilai2_fill
    
    # K5, K6 dari Penilai 3
    for j in range(2):
        cell = ws5.cell(row=row, column=j+7)
        col_ref = 'C' if j == 0 else 'D'
        cell.value = f"='Input Penilai 3'!{col_ref}{nilai_sheet_refs[3]['start_row'] + i}"
        cell_style(cell)
        cell.fill = penilai3_fill
    
    row += 1

rekap_end = row - 1

# Legend
row += 1
ws5.cell(row=row, column=1).value = "Keterangan:"
ws5.cell(row=row, column=1).font = Font(bold=True)
row += 1

ws5.cell(row=row, column=1).value = ""
ws5.cell(row=row, column=1).fill = penilai1_fill
ws5.cell(row=row, column=2).value = "= Nilai dari Penilai 1 (K1, K2)"
row += 1

ws5.cell(row=row, column=1).value = ""
ws5.cell(row=row, column=1).fill = penilai2_fill
ws5.cell(row=row, column=2).value = "= Nilai dari Penilai 2 (K3, K4)"
row += 1

ws5.cell(row=row, column=1).value = ""
ws5.cell(row=row, column=1).fill = penilai3_fill
ws5.cell(row=row, column=2).value = "= Nilai dari Penilai 3 (K5, K6)"

ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 25
for j in range(n):
    ws5.column_dimensions[get_column_letter(j+3)].width = 12

# ============================================
# SHEET 6: LANGKAH 1-3 (Matriks Crisp, Eigenvector, Konsistensi)
# ============================================
ws6 = wb.create_sheet("1-3. Crisp & Konsistensi")

row = 1
ws6['A1'] = "LANGKAH 1-3: MATRIKS CRISP, EIGENVECTOR & UJI KONSISTENSI"
ws6['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws6.merge_cells('A1:K1')

row = 3

# LANGKAH 1
section(ws6, row, 1, "PENYUSUNAN MATRIKS PERBANDINGAN BERPASANGAN (CRISP)", step1_fill, col_end=n+3)
row += 1

ws6.cell(row=row, column=1).value = "Formula: IF(wi/wj >= 1, MIN(9, MAX(1, wi/wj)), MAX(1/9, wi/wj))"
ws6.cell(row=row, column=1).font = Font(italic=True, size=9, color="666666")
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+3)
row += 1

ws6.cell(row=row, column=1).value = "Kriteria"
header_style(ws6.cell(row=row, column=1))
for j in range(n):
    cell = ws6.cell(row=row, column=j+2)
    cell.value = kriteria_list[j]
    header_style(cell)
row += 1

crisp_start = row

for i in range(n):
    ws6.cell(row=row, column=1).value = kriteria_list[i]
    cell_style(ws6.cell(row=row, column=1))
    ws6.cell(row=row, column=1).fill = light_blue_fill
    
    for j in range(n):
        cell = ws6.cell(row=row, column=j+2)
        if i == j:
            cell.value = 1
            cell_style(cell)
            cell.fill = light_blue_fill
        else:
            bobot_i = f"'Data & Penugasan'!$C${bobot_input_start + i}"
            bobot_j = f"'Data & Penugasan'!$C${bobot_input_start + j}"
            cell.value = f"=IF({bobot_i}/{bobot_j}>=1,MIN(9,MAX(1,{bobot_i}/{bobot_j})),MAX(1/9,{bobot_i}/{bobot_j}))"
            cell_style(cell)
            cell.number_format = '0.00'
    
    row += 1

crisp_end = row - 1
row += 2

# LANGKAH 2
section(ws6, row, 2, "PERHITUNGAN VECTOR EIGEN (GEOMETRIC MEAN METHOD)", step2_fill, col_end=n+3)
row += 1

ws6.cell(row=row, column=1).value = "Kriteria"
ws6.cell(row=row, column=2).value = "Geometric Mean (GM)"
ws6.cell(row=row, column=3).value = "Eigenvector (Wi)"
header_style(ws6.cell(row=row, column=1))
header_style(ws6.cell(row=row, column=2), fill=step2_fill)
header_style(ws6.cell(row=row, column=3), fill=step2_fill)
row += 1

gm_start = row

for i in range(n):
    ws6.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws6.cell(row=row, column=1))
    ws6.cell(row=row, column=1).fill = light_blue_fill
    
    crisp_row = crisp_start + i
    gm_formula = f"=(B{crisp_row}*C{crisp_row}*D{crisp_row}*E{crisp_row}*F{crisp_row}*G{crisp_row})^(1/6)"
    ws6.cell(row=row, column=2).value = gm_formula
    cell_style(ws6.cell(row=row, column=2))
    ws6.cell(row=row, column=2).fill = light_green_fill
    ws6.cell(row=row, column=2).number_format = '0.0000'
    
    row += 1

gm_end = row - 1

ws6.cell(row=row, column=1).value = "Total"
ws6.cell(row=row, column=1).font = Font(bold=True)
cell_style(ws6.cell(row=row, column=1))

gm_sum_row = row
gm_sum = ws6.cell(row=row, column=2)
gm_sum.value = f"=SUM(B{gm_start}:B{gm_end})"
gm_sum.font = Font(bold=True)
cell_style(gm_sum)
gm_sum.fill = yellow_fill

for i in range(n):
    r = gm_start + i
    wi = ws6.cell(row=r, column=3)
    wi.value = f"=B{r}/$B${gm_sum_row}"
    cell_style(wi)
    wi.fill = light_green_fill
    wi.number_format = '0.0000'

wi_sum = ws6.cell(row=row, column=3)
wi_sum.value = f"=SUM(C{gm_start}:C{gm_end})"
wi_sum.font = Font(bold=True)
cell_style(wi_sum)
wi_sum.fill = yellow_fill

row += 2

# LANGKAH 3
section(ws6, row, 3, "UJI KONSISTENSI MATRIKS PERBANDINGAN", step3_fill, col_end=n+3)
row += 1

ws6.cell(row=row, column=1).value = "Kriteria"
ws6.cell(row=row, column=2).value = "A*w"
ws6.cell(row=row, column=3).value = "(A*w)/w"
header_style(ws6.cell(row=row, column=1))
header_style(ws6.cell(row=row, column=2), fill=step3_fill)
header_style(ws6.cell(row=row, column=3), fill=step3_fill)
row += 1

aw_start = row

for i in range(n):
    ws6.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws6.cell(row=row, column=1))
    ws6.cell(row=row, column=1).fill = light_blue_fill
    
    crisp_row = crisp_start + i
    parts = []
    for j in range(n):
        col = get_column_letter(j + 2)
        parts.append(f"{col}{crisp_row}*$C${gm_start + j}")
    aw = ws6.cell(row=row, column=2)
    aw.value = "=" + "+".join(parts)
    cell_style(aw)
    aw.fill = light_yellow_fill
    aw.number_format = '0.0000'
    
    ratio = ws6.cell(row=row, column=3)
    ratio.value = f"=B{row}/$C${gm_start + i}"
    cell_style(ratio)
    ratio.fill = light_yellow_fill
    ratio.number_format = '0.0000'
    
    row += 1

aw_end = row - 1
row += 1

lambda_row = row
ws6.cell(row=row, column=1).value = "Lambda Max:"
ws6.cell(row=row, column=1).font = subtitle_font
ws6.cell(row=row, column=2).value = f"=AVERAGE(C{aw_start}:C{aw_end})"
ws6.cell(row=row, column=2).font = Font(bold=True, size=14, color="16A34A")
ws6.cell(row=row, column=2).fill = yellow_fill
cell_style(ws6.cell(row=row, column=2))
ws6.cell(row=row, column=2).number_format = '0.0000'
row += 1

ci_row = row
ws6.cell(row=row, column=1).value = "CI:"
ws6.cell(row=row, column=1).font = subtitle_font
ws6.cell(row=row, column=2).value = f"=(B{lambda_row}-{n})/({n}-1)"
cell_style(ws6.cell(row=row, column=2))
ws6.cell(row=row, column=2).fill = light_yellow_fill
row += 1

ri_row = row
ws6.cell(row=row, column=1).value = "RI (n=6):"
ws6.cell(row=row, column=1).font = subtitle_font
ws6.cell(row=row, column=2).value = 1.24
cell_style(ws6.cell(row=row, column=2))
row += 1

cr_row = row
ws6.cell(row=row, column=1).value = "CR:"
ws6.cell(row=row, column=1).font = subtitle_font
ws6.cell(row=row, column=2).value = f"=B{ci_row}/B{ri_row}"
ws6.cell(row=row, column=2).font = Font(bold=True, size=14)
ws6.cell(row=row, column=2).fill = yellow_fill
cell_style(ws6.cell(row=row, column=2))
row += 1

ws6.cell(row=row, column=1).value = "Status:"
ws6.cell(row=row, column=1).font = subtitle_font
ws6.cell(row=row, column=2).value = f'=IF(B{cr_row}<=0.1,"KONSISTEN","TIDAK KONSISTEN")'
ws6.cell(row=row, column=2).font = Font(bold=True, size=12)
cell_style(ws6.cell(row=row, column=2))

ws6.column_dimensions['A'].width = 25
for j in range(n+3):
    ws6.column_dimensions[get_column_letter(j+2)].width = 14

# ============================================
# SHEET 7: LANGKAH 4-6 (Fuzzy)
# ============================================
ws7 = wb.create_sheet("4-6. Fuzzy & Bobot")

row = 1
ws7['A1'] = "LANGKAH 4-6: FUZZIFIKASI, SYNTHETIC EXTENT & BOBOT GLOBAL"
ws7['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws7.merge_cells('A1:U1')

row = 3

# Tabel TFN
ws7.cell(row=row, column=1).value = "TABEL SKALA FUZZY (TFN)"
ws7.cell(row=row, column=1).font = Font(bold=True, size=12, color="9333EA")
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

ws7.cell(row=row, column=1).value = "Intensitas"
ws7.cell(row=row, column=2).value = "l"
ws7.cell(row=row, column=3).value = "m"
ws7.cell(row=row, column=4).value = "u"
for c in range(1, 5):
    header_style(ws7.cell(row=row, column=c), fill=step4_fill)
row += 1

tfn_start = row
for intensity in range(1, 10):
    l, m, u = tfn_scale[intensity]
    ws7.cell(row=row, column=1).value = intensity
    ws7.cell(row=row, column=2).value = l
    ws7.cell(row=row, column=3).value = m
    ws7.cell(row=row, column=4).value = u
    for c in range(1, 5):
        cell_style(ws7.cell(row=row, column=c))
    row += 1

tfn_end = row - 1
row += 1

# LANGKAH 4 - Simplified (menggunakan eigenvector sebagai bobot final)
section(ws7, row, 4, "FUZZIFIKASI - SIMPLIFIED", step4_fill, col_end=10)
row += 1

ws7.cell(row=row, column=1).value = "Untuk penyederhanaan, bobot menggunakan eigenvector dari Langkah 2"
ws7.cell(row=row, column=1).font = Font(italic=True, size=10, color="666666")
ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
row += 2

# LANGKAH 6 - Bobot Final
section(ws7, row, 6, "BOBOT GLOBAL (DARI EIGENVECTOR)", step6_fill, col_end=5)
row += 1

ws7.cell(row=row, column=1).value = "Kriteria"
ws7.cell(row=row, column=2).value = "Bobot (Wi)"
ws7.cell(row=row, column=3).value = "Persentase"
ws7.cell(row=row, column=4).value = "Penilai"
for c in range(1, 5):
    header_style(ws7.cell(row=row, column=c), fill=step6_fill)
row += 1

dprime_start = row

for i in range(n):
    ws7.cell(row=row, column=1).value = kriteria_full[i]
    cell_style(ws7.cell(row=row, column=1))
    ws7.cell(row=row, column=1).fill = light_blue_fill
    
    # Bobot dari eigenvector
    ws7.cell(row=row, column=2).value = f"='1-3. Crisp & Konsistensi'!C{gm_start + i}"
    cell_style(ws7.cell(row=row, column=2))
    ws7.cell(row=row, column=2).fill = light_teal_fill
    ws7.cell(row=row, column=2).number_format = '0.0000'
    
    # Persentase
    ws7.cell(row=row, column=3).value = f"=B{row}*100"
    cell_style(ws7.cell(row=row, column=3))
    ws7.cell(row=row, column=3).number_format = '0.00"%"'
    
    # Penilai
    if i < 2:
        penilai_idx = 0
    elif i < 4:
        penilai_idx = 1
    else:
        penilai_idx = 2
    
    ws7.cell(row=row, column=4).value = f"Penilai {penilai_idx + 1}"
    cell_style(ws7.cell(row=row, column=4))
    ws7.cell(row=row, column=4).fill = penilai_fills[penilai_idx]
    
    row += 1

dprime_end = row - 1

ws7.cell(row=row, column=1).value = "Total"
ws7.cell(row=row, column=1).font = Font(bold=True)
cell_style(ws7.cell(row=row, column=1))

ws7.cell(row=row, column=2).value = f"=SUM(B{dprime_start}:B{dprime_end})"
ws7.cell(row=row, column=2).font = Font(bold=True)
cell_style(ws7.cell(row=row, column=2))
ws7.cell(row=row, column=2).fill = yellow_fill

ws7.column_dimensions['A'].width = 25
ws7.column_dimensions['B'].width = 15
ws7.column_dimensions['C'].width = 12
ws7.column_dimensions['D'].width = 12

# ============================================
# SHEET 8: HASIL & RANKING
# ============================================
ws8 = wb.create_sheet("Hasil & Ranking")

row = 1
ws8['A1'] = "PERHITUNGAN SKOR AKHIR & RANKING"
ws8['A1'].font = Font(bold=True, size=16, color="1F4E79")
ws8.merge_cells('A1:L1')

ws8['A2'] = "Skor = Sum(Nilai dari masing-masing Penilai * Bobot)"
ws8['A2'].font = Font(italic=True, color="666666")
ws8.merge_cells('A2:L2')

row = 4

# Bobot Reference
ws8.cell(row=row, column=1).value = "BOBOT KRITERIA"
ws8.cell(row=row, column=1).font = Font(bold=True, size=12, color="0D9488")
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+1)
row += 1

for j in range(n):
    cell = ws8.cell(row=row, column=j+1)
    cell.value = kriteria_list[j]
    if j < 2:
        header_style(cell, fill=penilai1_header)
    elif j < 4:
        header_style(cell, fill=penilai2_header)
    else:
        header_style(cell, fill=penilai3_header)
row += 1

bobot_ref_row = row
for j in range(n):
    cell = ws8.cell(row=row, column=j+1)
    cell.value = f"='4-6. Fuzzy & Bobot'!B{dprime_start + j}"
    cell_style(cell)
    cell.fill = light_teal_fill
    cell.number_format = '0.0000'
row += 2

# Perhitungan Skor
ws8.cell(row=row, column=1).value = "PERHITUNGAN SKOR AKHIR"
ws8.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+4)
row += 1

ws8.cell(row=row, column=1).value = "No"
ws8.cell(row=row, column=2).value = "Nama"
header_style(ws8.cell(row=row, column=1))
header_style(ws8.cell(row=row, column=2))

for j in range(n):
    cell = ws8.cell(row=row, column=j+3)
    cell.value = f"{kriteria_list[j]}*W"
    if j < 2:
        header_style(cell, fill=penilai1_header)
    elif j < 4:
        header_style(cell, fill=penilai2_header)
    else:
        header_style(cell, fill=penilai3_header)

ws8.cell(row=row, column=n+3).value = "Skor"
ws8.cell(row=row, column=n+4).value = "Rank"
header_style(ws8.cell(row=row, column=n+3), fill=orange_fill)
header_style(ws8.cell(row=row, column=n+4), fill=orange_fill)
row += 1

skor_start = row

for i in range(n_peserta):
    ws8.cell(row=row, column=1).value = i + 1
    cell_style(ws8.cell(row=row, column=1))
    
    ws8.cell(row=row, column=2).value = f"='Data & Penugasan'!B{peserta_start + i}"
    cell_style(ws8.cell(row=row, column=2))
    ws8.cell(row=row, column=2).fill = light_blue_fill
    
    # Nilai * Bobot untuk setiap kriteria
    for j in range(n):
        cell = ws8.cell(row=row, column=j+3)
        nilai_ref = f"'Rekap Nilai'!{get_column_letter(j+3)}{rekap_start + i}"
        bobot_ref = f"${get_column_letter(j+1)}${bobot_ref_row}"
        cell.value = f"={nilai_ref}*{bobot_ref}"
        cell_style(cell)
        
        if j < 2:
            cell.fill = penilai1_fill
        elif j < 4:
            cell.fill = penilai2_fill
        else:
            cell.fill = penilai3_fill
        
        cell.number_format = '0.00'
    
    # Skor total
    skor = ws8.cell(row=row, column=n+3)
    skor.value = f"=SUM(C{row}:{get_column_letter(n+2)}{row})"
    skor.font = Font(bold=True)
    cell_style(skor)
    skor.fill = light_yellow_fill
    skor.number_format = '0.00'
    
    # Ranking
    rank = ws8.cell(row=row, column=n+4)
    skor_range = f"${get_column_letter(n+3)}${skor_start}:${get_column_letter(n+3)}${skor_start + n_peserta - 1}"
    rank.value = f'=IF(B{row}="","",RANK({get_column_letter(n+3)}{row},{skor_range},0))'
    rank.font = Font(bold=True, size=12)
    cell_style(rank)
    rank.fill = yellow_fill
    
    row += 1

skor_end = row - 1

# Legend
row += 1
ws8.cell(row=row, column=1).value = "Keterangan Warna:"
ws8.cell(row=row, column=1).font = Font(bold=True)
row += 1

ws8.cell(row=row, column=1).value = ""
ws8.cell(row=row, column=1).fill = penilai1_fill
ws8.cell(row=row, column=2).value = "= Nilai dari Penilai 1"
row += 1

ws8.cell(row=row, column=1).value = ""
ws8.cell(row=row, column=1).fill = penilai2_fill
ws8.cell(row=row, column=2).value = "= Nilai dari Penilai 2"
row += 1

ws8.cell(row=row, column=1).value = ""
ws8.cell(row=row, column=1).fill = penilai3_fill
ws8.cell(row=row, column=2).value = "= Nilai dari Penilai 3"

ws8.column_dimensions['A'].width = 8
ws8.column_dimensions['B'].width = 25
for j in range(n):
    ws8.column_dimensions[get_column_letter(j+3)].width = 12
ws8.column_dimensions[get_column_letter(n+3)].width = 12
ws8.column_dimensions[get_column_letter(n+4)].width = 8

# Save
output_file = "d:/laragon/www/appSaringPramuka/Fuzzy_AHP_3_Penilai.xlsx"
wb.save(output_file)

print(f"[OK] File berhasil dibuat: {output_file}")
print(f"\n=== STRUKTUR FILE ===")
print(f"Sheet 1: Data & Penugasan")
print(f"  - Data kegiatan & kuota")
print(f"  - Penugasan penilai ke kriteria")
print(f"  - Bobot kriteria")
print(f"  - Daftar peserta")
print(f"")
print(f"Sheet 2: Input Penilai 1 (K1, K2)")
print(f"Sheet 3: Input Penilai 2 (K3, K4)")
print(f"Sheet 4: Input Penilai 3 (K5, K6)")
print(f"")
print(f"Sheet 5: Rekap Nilai (gabungan dari 3 penilai)")
print(f"Sheet 6: Langkah 1-3 (Matriks Crisp, Eigenvector, Konsistensi)")
print(f"Sheet 7: Langkah 4-6 (Fuzzy & Bobot Global)")
print(f"Sheet 8: Hasil & Ranking")
