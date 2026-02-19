from flask import Flask, Response, abort, request, render_template, request as flask_request, redirect, url_for, flash, session, jsonify, current_app, send_file, make_response
from flask_session import Session
from flask_session import FileSystemSessionInterface
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from app import create_app, db
from app.models import Users, Participants, Notification, Event, Kuota, Criteria, HasilSeleksi, Penilaian, HimpunanKriteria, News, Comment, CommentLike, tb_participant_kegiatan, PairwiseComparison, AHPResults, ArsipSeleksi, LogAktivitas, Settings
from flask_mail import Mail, Message
from twilio.rest import Client
from authlib.integrations.flask_client import OAuth
from markupsafe import escape
from datetime import datetime, timedelta, date
from flask_wtf import CSRFProtect
from flask_wtf.csrf import CSRFProtect, CSRFError
from forms import LoginForm, RegisterForm
from config import Config
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import current_user, LoginManager, login_user, login_required
from functools import wraps
from app.utils.utils import log_activity
from sqlalchemy.exc import IntegrityError
from slugify import slugify
from urllib.parse import urlparse, urljoin
import io
import base64
from app.translations import TRANSLATIONS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import random, string
import logging
import os
import secrets
import time
import os
import re
import pandas as pd
from dotenv import load_dotenv

load_dotenv()

app = create_app()
app.config['SESSION_FILE_PATH'] = os.path.join(app.root_path, 'flask_session')
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
secret_key = os.getenv("APP_SECRET_KEY")
# Ensure secret_key is a string, not bytes
if isinstance(secret_key, bytes):
    secret_key = secret_key.decode('utf-8')
app.secret_key = secret_key or secrets.token_hex(32)

# Custom Session Interface to fix bytes/string issue
class FixedFileSystemSessionInterface(FileSystemSessionInterface):
    def generate_sid(self):
        """Generate session ID and ensure it's always a string"""
        sid = super().generate_sid()
        if isinstance(sid, bytes):
            try:
                sid = sid.decode('utf-8')
            except (UnicodeDecodeError, AttributeError):
                import base64
                sid = base64.urlsafe_b64encode(sid).decode('utf-8').rstrip('=')
        return sid
    
    def save_session(self, app, session, response):
        """Override save_session to ensure session_id is always a string"""
        original_set_cookie = response.set_cookie
        
        def patched_set_cookie(key, value='', *args, **kwargs):
            if isinstance(value, bytes):
                try:
                    value = value.decode('utf-8')
                except (UnicodeDecodeError, AttributeError):
                    import base64
                    value = base64.urlsafe_b64encode(value).decode('utf-8').rstrip('=')
            return original_set_cookie(key, value, *args, **kwargs)
        response.set_cookie = patched_set_cookie
        try:
            super().save_session(app, session, response)
        finally:
            response.set_cookie = original_set_cookie

# Initialize Flask-Session first to set up configuration
Session(app)
existing_interface = app.session_interface
existing_interface.__class__ = FixedFileSystemSessionInterface

csrf = CSRFProtect(app)
app.config.from_object(Config)
limiter = Limiter(get_remote_address, app=app)
logging.basicConfig(filename='login.log', level=logging.INFO,
                    format='%(asctime)s %(levelname)s:%(message)s')

# Helper function untuk membuat notifikasi
def create_notification(user_id, message):
    """
    Membuat notifikasi untuk user tertentu
    
    Args:
        user_id: ID user yang akan menerima notifikasi
        message: Pesan notifikasi (maks 255 karakter)
    
    Returns:
        Notification object atau None jika gagal
    """
    try:
        # Validasi user_id
        if not user_id:
            logging.error("create_notification: user_id is None or empty")
            return None
        
        # Validasi message
        if not message or not message.strip():
            logging.error("create_notification: message is None or empty")
            return None
        
        # Potong message jika terlalu panjang
        if len(message) > 255:
            message = message[:252] + "..."
        
        notification = Notification(
            user_id=user_id,
            message=message.strip(),
            is_read=False
        )
        db.session.add(notification)
        db.session.flush()  # Flush untuk memastikan ID tersedia
        db.session.commit()
        logging.info(f"Notification created successfully for user_id: {user_id}, message: {message[:50]}...")
        return notification
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating notification for user_id {user_id}: {e}")
        if hasattr(current_app, 'logger'):
            current_app.logger.exception('Error in create_notification:')
        return None

# Helper function untuk membuat notifikasi ke semua admin
def create_notification_to_all_admins(message):
    """
    Membuat notifikasi untuk semua admin
    
    Args:
        message: Pesan notifikasi
    """
    try:
        if not message or not message.strip():
            logging.error("create_notification_to_all_admins: message is None or empty")
            return
        
        admins = Users.query.filter_by(level='admin').all()
        if not admins:
            logging.warning("No admin users found to send notification to")
            return
        
        logging.info(f"[NOTIFICATION] Creating notifications for {len(admins)} admin(s): {message[:50]}...")
        success_count = 0
        for admin in admins:
            logging.info(f"[NOTIFICATION] Sending to admin: {admin.username} (ID: {admin.id}, Email: {admin.email})")
            result = create_notification(admin.id, message)
            if result:
                success_count += 1
                logging.info(f"[NOTIFICATION] ✓ Notification sent to admin: {admin.username} (ID: {admin.id})")
            else:
                logging.warning(f"[NOTIFICATION] ✗ Failed to send notification to admin: {admin.username} (ID: {admin.id})")
        
        logging.info(f"[NOTIFICATION] Summary: Successfully created {success_count}/{len(admins)} notifications")
    except Exception as e:
        logging.error(f"Error creating notification to admins: {e}")
        if hasattr(current_app, 'logger'):
            current_app.logger.exception('Error in create_notification_to_all_admins:')

# Inisialisasi Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)

login_manager.login_view = 'login'  # nama fungsi view untuk login
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return Users.query.get(int(user_id))

# Tentukan folder upload (path absolut dari root aplikasi)
UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Extension yang diizinkan untuk file
ALLOWED_EXTENSIONS_IMAGE = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_EXTENSIONS_DOC = {'csv', 'xls', 'xlsx'}

def allowed_file(filename, file_type='image'):
    """Validasi extension file"""
    if file_type == 'image':
        allowed = ALLOWED_EXTENSIONS_IMAGE
    else:
        allowed = ALLOWED_EXTENSIONS_DOC
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed

# Helper function untuk mendapatkan pengaturan dari database
def get_setting(key, default=None):
    """Mendapatkan pengaturan dari database, fallback ke default jika tidak ada"""
    try:
        setting = Settings.query.filter_by(key=key).first()
        if setting and setting.value:
            return setting.value
    except:
        pass
    return default

def get_email_settings():
    """Mendapatkan pengaturan email dari database atau env/config"""
    return {
        'mail_server': get_setting('mail_server') or os.getenv("MAIL_SERVER", "smtp.gmail.com"),
        'mail_port': int(get_setting('mail_port') or os.getenv("MAIL_PORT", "465")),
        'mail_use_tls': get_setting('mail_use_tls', 'false') == 'true',
        'mail_use_ssl': get_setting('mail_use_ssl', 'true') == 'true',
        'mail_username': get_setting('mail_username') or os.getenv("MAIL_USERNAME", ""),
        'mail_password': get_setting('mail_password') or os.getenv("MAIL_PASSWORD", ""),
        'mail_enabled': get_setting('mail_enabled', 'true') == 'true'
    }

def get_sms_settings():
    """Mendapatkan pengaturan SMS/WhatsApp dari database atau env/config"""
    return {
        'twilio_account_sid': get_setting('twilio_account_sid') or os.getenv("TWILIO_ACCOUNT_SID", ""),
        'twilio_auth_token': get_setting('twilio_auth_token') or os.getenv("TWILIO_AUTH_TOKEN", ""),
        'twilio_whatsapp_from': get_setting('twilio_whatsapp_from') or os.getenv("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886"),
        'sms_enabled': get_setting('sms_enabled', 'true') == 'true'
    }

# Configure Flask-Mail OTP - akan diupdate dari database saat runtime
email_config = get_email_settings()
app.config['MAIL_SERVER'] = email_config['mail_server']
app.config['MAIL_PORT'] = email_config['mail_port']
app.config['MAIL_USE_TLS'] = email_config['mail_use_tls']
app.config['MAIL_USE_SSL'] = email_config['mail_use_ssl']
app.config['MAIL_USERNAME'] = email_config['mail_username']
app.config['MAIL_PASSWORD'] = email_config['mail_password']
mail = Mail(app)

def send_whatsapp_code(phone, code):
    """Mengirim kode verifikasi via WhatsApp menggunakan pengaturan dari database"""
    sms_config = get_sms_settings()
    
    if not sms_config['sms_enabled']:
        print("SMS/WhatsApp tidak diaktifkan")
        return False
    
    if not sms_config['twilio_account_sid'] or not sms_config['twilio_auth_token']:
        print("Twilio credentials tidak ditemukan")
        return False
    
    try:
        client = Client(sms_config['twilio_account_sid'], sms_config['twilio_auth_token'])
        message_body = f"Kode verifikasi Anda adalah: *{code}*.\nJangan bagikan kode ini kepada siapa pun."
        message = client.messages.create(
            body=message_body,
            from_=sms_config['twilio_whatsapp_from'],
            to=f'whatsapp:{phone}'
        )
        print("Pesan berhasil dikirim:", message.sid)
        return True
    except Exception as e:
        print("Gagal mengirim pesan:", e)
        return False

def send_email_message(subject, recipients, html_body, sender=None):
    """Mengirim email menggunakan pengaturan dari database"""
    email_config = get_email_settings()
    
    if not email_config['mail_enabled']:
        print("Email tidak diaktifkan")
        return False
    
    if not email_config['mail_username'] or not email_config['mail_password']:
        print("Email credentials tidak ditemukan")
        return False
    
    try:
        # Update config sementara
        original_config = {
            'MAIL_SERVER': app.config.get('MAIL_SERVER'),
            'MAIL_PORT': app.config.get('MAIL_PORT'),
            'MAIL_USE_SSL': app.config.get('MAIL_USE_SSL'),
            'MAIL_USE_TLS': app.config.get('MAIL_USE_TLS'),
            'MAIL_USERNAME': app.config.get('MAIL_USERNAME'),
            'MAIL_PASSWORD': app.config.get('MAIL_PASSWORD')
        }
        
        app.config['MAIL_SERVER'] = email_config['mail_server']
        app.config['MAIL_PORT'] = email_config['mail_port']
        app.config['MAIL_USE_SSL'] = email_config['mail_use_ssl']
        app.config['MAIL_USE_TLS'] = email_config['mail_use_tls']
        app.config['MAIL_USERNAME'] = email_config['mail_username']
        app.config['MAIL_PASSWORD'] = email_config['mail_password']
        
        # Reinitialize mail dengan config baru
        mail.init_app(app)
        
        msg = Message(
            subject=subject,
            recipients=recipients if isinstance(recipients, list) else [recipients],
            html=html_body,
            sender=sender or email_config['mail_username']
        )
        mail.send(msg)
        
        # Restore original config
        for key, value in original_config.items():
            app.config[key] = value
        mail.init_app(app)
        
        return True
    except Exception as e:
        print(f"Gagal mengirim email: {e}")
        # Restore original config
        try:
            for key, value in original_config.items():
                app.config[key] = value
            mail.init_app(app)
        except:
            pass
        return False

def normalize_phone_number(phone):
    phone = phone.strip()
    if phone.startswith('0'):
        return '+62' + phone[1:]
    elif phone.startswith('+62'):
        return phone
    elif phone.startswith('62'):
        return '+' + phone
    else:
        return phone
    
def generate_username(email):
    name_part = email.split('@')[0]
    random_suffix = ''.join(random.choices(string.digits, k=4))
    return f"{name_part}_{random_suffix}"

# Google OAuth Config
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.getenv('GOOGLE_CLIENT_ID'),
    client_secret=os.getenv('GOOGLE_CLIENT_SECRET'),
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    access_token_url='https://oauth2.googleapis.com/token',
    access_token_params=None,
    authorize_url='https://accounts.google.com/o/oauth2/auth',
    authorize_params=None,
    api_base_url='https://www.googleapis.com/oauth2/v1/',
    userinfo_endpoint='https://www.googleapis.com/oauth2/v1/userinfo',
    client_kwargs={'scope': 'openid email profile'},
)

# Fungsi untuk mengecek keberadaan username, nomor hp, email dan password serta untuk menghasilkan kode verifikasi 6 digit
def check_username_in_db(username):
    user = Users.query.filter_by(username=username).first()
    return user is not None

def check_phone_in_db(phone):
    user = Users.query.filter_by(nomor_hp=phone).first()
    return user is not None


def check_email_in_db(email):
    user = Users.query.filter_by(email=email).first()
    return user is not None

def check_password_in_db(username, password):
    user = Users.query.filter_by(username=username).first()
    if user:
        return check_password_hash(user.password, password)
    return False

def generate_verification_code():
    return random.randint(100000, 999999)

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    try:
        return render_template('csrf_error.html', reason=e.description), 400
    except:
        # Fallback jika template tidak ditemukan
        flash("Sesi Anda mungkin telah kedaluwarsa. Silakan refresh halaman dan coba lagi.", "danger")
        if current_user.is_authenticated:
            if current_user.level == 'admin':
                return redirect(url_for('admin_users'))
            elif current_user.level == 'penilai':
                return redirect(url_for('penilai_dashboard'))
            elif current_user.level == 'peserta':
                return redirect(url_for('peserta_dashboard'))
        return redirect(url_for('login'))

@app.errorhandler(429)
def ratelimit_handler(e):
    return render_template("429.html", message="Terlalu banyak percobaan login. Silakan coba lagi nanti."), 429

@app.context_processor
def inject_notifications():
    user = None
    unread_count = 0
    notifications = []
    if session.get('username'):
        user = Users.query.filter_by(username=session.get('username')).first()
    elif session.get('user'):
        user = Users.query.filter_by(username=session['user'].get('username')).first()
    if user:
        unread_count = Notification.query.filter_by(user_id=user.id, is_read=False).count()
        # Get recent notifications (last 10) for the notification tray
        notifications = Notification.query.filter_by(
            user_id=user.id
        ).order_by(
            Notification.id.desc()
        ).limit(10).all()
    return dict(notification_count=unread_count, notifications=notifications)

# --- Middleware untuk cek login dan role ---
def my_decorator(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # logika tambahan
        return f(*args, **kwargs)
    return decorated_function

def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and ref_url.netloc == test_url.netloc

# Endpoint login
@app.route('/login/', methods=['GET', 'POST'])
@limiter.limit("20 per minute")
def login():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    next_url = request.form.get('next') or request.args.get('next')
    form = LoginForm()
    if current_user.is_authenticated:
        if next_url and is_safe_url(next_url):
            logging.info(f"Redirect after login to: {next_url}")
            return redirect(next_url)
        # Redirect ke dashboard sesuai role, bukan ke landing page
        if current_user.level == "admin":
            return redirect(url_for('admin_dashboard'))
        elif current_user.level == "penilai":
            return redirect(url_for('penilai_dashboard'))
        elif current_user.level == "peserta":
            return redirect(url_for('peserta_dashboard'))
        else:
            return redirect(url_for('index'))

    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        
        # Query user dari database
        user = Users.query.filter_by(username=username).first()

        if not user:
            logging.warning(f"Login gagal: username '{username}' tidak ditemukan.")
            flash(t['username_invalid'], 'danger')
        elif not check_password_hash(user.password, password):
            logging.warning(f"Login gagal: password salah untuk user '{username}'.")
            flash(t['password_invalid'], 'danger')
        else:
            login_user(user)
            session['username'] = username  
            session['role'] = user.level
            safe_username = escape(username)
            logging.info(f"User '{username}' berhasil login sebagai {user.level}.")
            flash(f"{t['login_success']}, {safe_username}.", 'success')
            session['first_time_login'] = True
            
            if next_url and is_safe_url(next_url):
                return redirect(next_url)
            
            # Redirect sesuai role
            if user.level == "admin":
                return redirect(url_for('admin_dashboard'))
            elif user.level == "penilai":
                return redirect(url_for('penilai_dashboard'))
            elif user.level == "peserta":
                return redirect(url_for('peserta_dashboard'))
            else:
                # Default jika role tidak dikenali
                return redirect(url_for('login')) 
    return render_template('login.html', form=form)

# Endpoint login with Google
@app.route('/login/google/', methods=['GET', 'POST'])
@limiter.limit("20 per minute")
def login_google():
    next_url = request.args.get('next')

    if next_url and is_safe_url(next_url):
        session['oauth_next'] = next_url
        
    redirect_uri = url_for('login_google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)

# Endpoint Callback Login With Google
@app.route('/login/google/callback/')
def login_google_callback():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    next_url = session.pop('oauth_next', None)

    try:
        token = google.authorize_access_token()
        resp = google.get('userinfo')
        resp.raise_for_status()
        user_info = resp.json()
    except Exception as e:
        logging.warning(f"Login Google gagal: {e}")
        flash(t['google_login_failed'], "danger")
        return redirect(url_for('login'))

    email = user_info.get('email')
    picture = user_info.get('picture') or "img/default-user.png"

    if not email:
        flash(t['google_email_not_found'], "danger")
        return redirect(url_for('login'))

    if not email.endswith('@gmail.com'):
        flash(t['google_only_gmail'], "danger")
        return redirect(url_for('login'))

    user = Users.query.filter_by(email=email).first()
    if not user:
        session['pending_user'] = user_info
        logging.warning(f"Percobaan login Google dari email '{email}' belum terdaftar.")
        flash(f"{t['google_not_registered']}", "warning")
        return redirect(url_for('confirm_register'))

    if not user.foto or user.foto == "img/default-user.png":
        user.foto = picture
        db.session.commit()

    login_user(user)
    session['user'] = {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'nama_lengkap': user.nama_lengkap,
        'foto': user.foto,
        'level': user.level
    }
    session['first_time_login'] = True
    session.modified = True

    logging.info(f"User '{user.username}' berhasil login via Google.")
    flash(f"{t['login_success']}, {escape(user.nama_lengkap)}.", "success")

    if next_url and is_safe_url(next_url):
        return redirect(next_url)

    user_level = (user.level or "").strip().lower()
    role_redirect = {
        "admin": "admin_dashboard",
        "penilai": "penilai_dashboard",
        "peserta": "peserta_dashboard"
    }
    endpoint = role_redirect.get(user_level)

    if endpoint:
        return redirect(url_for(endpoint))
    return redirect(url_for("index"))
    
# Endpoint Fisrt Confirm Register With Google
@app.route('/confirm-register/')
def confirm_register():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    user_info = session.get('pending_user')
    if not user_info:
        flash(f"{t['user_data_not_found']}", "danger")
        return redirect(url_for('login'))
    form = RegisterForm()
    return render_template("confirm_register.html", user=user_info, form=form)

# Endpoint Confirm Register With Google
@app.route('/confirm-register', methods=['POST'])
def do_register():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    user_info = session.get('pending_user')
    if not user_info:
        flash(f"{t['user_data_not_found']}", "danger")
        return redirect(url_for('login'))
    
    form = RegisterForm()

    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        email = user_info['email']
        
    # Validasi apakah username atau email sudah digunakan
        if Users.query.filter_by(email=email).first():
            flash(f"{t['email_already_used']}", "warning")
            return redirect(url_for('login'))
        if Users.query.filter_by(username=username).first():
            flash(f"{t['username_already_used']}", "warning")
            return redirect(url_for('confirm_register'))
    
        # Simpan ke database
        new_user = Users(
                username=username,
                password=generate_password_hash(secrets.token_urlsafe(12), method='pbkdf2:sha256'),
                nama_lengkap=user_info['name'],
                email=user_info['email'],
                jenis_kelamin='laki-laki',
                usia='0',
                foto=user_info.get('picture', 'img/default-user.png'),
                nomor_hp='',
                level='peserta',
                reset_token="",
                login_method="google",
                sidebar_state="expanded",
                status='aktif'
        )
        try:
            db.session.add(new_user)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logging.error(f"Gagal menyimpan user Google baru: {e}")
            flash(f"{t['register_failed']}", "danger")
            return redirect(url_for('confirm_register'))

        # Set session
        session['user'] = {
            'id': new_user.id,
            'username': new_user.username,
            'email': new_user.email,
            'nama_lengkap': new_user.nama_lengkap,
            'foto': new_user.foto,
            'level': new_user.level
        }
        print("Session setelah login Google:", dict(session))
        logging.info(f"User baru '{username}' berhasil registrasi dan login via Google.")
        flash(f"{t['register_success']}", "welcome")
        session['username'] = new_user.username
        session['first_time_login'] = True
        return redirect(url_for('index'))
    return render_template("confirm_register.html", user=user_info, form=form)

# Endpoint register
@app.route('/register/', methods=['GET', 'POST'])
def register():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if request.method == 'POST':
        full_name = request.form.get('fullName', '').strip()
        email = request.form.get('email', '').strip().lower()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirmPassword', '')
        level = 'peserta'  
        
        # Apakah ada kolom yang kosong?
        if not all([full_name, email, username, password, confirm_password]):
            flash(f"{t['all_fields_required']}", "danger")
            return redirect(url_for('register'))
        
        # Validasi password dengan regex
        password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
        if not re.match(password_pattern, password):
            flash(f"{t['password_invalid']}", "danger")
            return redirect(url_for('register'))
        
        # Validasi apakah password dan confirmPassword cocok
        if password != confirm_password:
            flash(f"{t['password_not_match']}")
            return redirect(url_for('register'))
        
        # Cek keberadaan username dan email di database
        user_exists = check_username_in_db(username)
        email_exists = check_email_in_db(email)
        
        if not user_exists and not email_exists:
           # Enkripsi password
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16) 
            # Masukkan data ke database
            try:
                # Ambil field tambahan dari form atau gunakan default
                jenis_kelamin = request.form.get('jenis_kelamin', 'laki-laki')
                usia = request.form.get('usia', '0')
                nomor_hp = request.form.get('nomor_hp', '')
                foto = request.form.get('foto', 'img/default-user.png')
                
                new_user = Users(
                    username=username,
                    password=hashed_password,
                    nama_lengkap=full_name,
                    email=email,
                    jenis_kelamin=jenis_kelamin,
                    usia=usia,
                    foto=foto,
                    nomor_hp=nomor_hp,
                    level=level,
                    reset_token="",
                    login_method="manual",
                    sidebar_state="expanded",
                    status='aktif'
                )
                db.session.add(new_user)
                db.session.commit()
                flash(f"{t['register_success_manual']}", "welcome")
                return redirect(url_for('login'))
            except IntegrityError as e:
                db.session.rollback()
                logging.error(f"Error during registration: {e}")
                flash(f"{t['register_failed']}", "danger")
                print(e)
        elif user_exists and email_exists:
            flash(f"{t['username_email_exists']}", "danger")
        elif user_exists:
            flash(f"{t['username_exists']}", "danger")
        elif email_exists:
            flash(f"{t['email_exists']}", "danger")
        return redirect(url_for('register'))
    return render_template('register.html')

# Endpoint Register With Google
@app.route('/register/google/', methods=['GET', 'POST'])
def register_google():
    redirect_uri = url_for('register_google_callback', _external=True)
    return google.authorize_redirect(redirect_uri)

# Endpoint Callback Register With Google
@app.route('/register/google/callback/')
def register_google_callback():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    try:
        token = google.authorize_access_token()
        resp = google.get('userinfo')
        resp.raise_for_status() 
        user_info = resp.json()
    except Exception as e:
        print(f"Google registrasi error: {e}") 
        flash(f"{t['google_register_failed']}", "danger")
        return redirect(url_for('register'))
    
    email = user_info['email']
    username = generate_username(email)
    
    # Cek email sudah digunakan
    if Users.query.filter_by(email=email).first():
        flash(f"{t['email_already_used']}", "warning")
        return redirect(url_for('login'))
    
    # Simpan ke database
    new_user = Users(
            username=username,
            password=generate_password_hash(secrets.token_urlsafe(12), method='pbkdf2:sha256'),
            nama_lengkap=user_info['name'],
            email=user_info['email'],
            jenis_kelamin='laki-laki',
            usia='0',
            foto=user_info.get('picture', 'img/default-user.png'),
            nomor_hp='',
            level='peserta',
            reset_token="",
            login_method="google",
            sidebar_state="expanded",
            status='aktif'
    )
    db.session.add(new_user)
    db.session.commit()
    
    # Login langsung setelah registrasi
    login_user(new_user)
    session['username'] = new_user.username
    session['user'] = {
        'id': new_user.id,
        'username': new_user.username,
        'email': new_user.email,
        'nama_lengkap': new_user.nama_lengkap,
        'foto': new_user.foto,
        'level': new_user.level
    }
    flash(f"{t['register_google_success']}", "welcome")
    return redirect(url_for('admin_dashboard'))

# Endpoint Find Account
@app.route('/find_account/', methods=['GET', 'POST'])
def find_account():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        phone = request.form.get('no-hp')
        # Pastikan username diisi
        if not username:
            flash(f"{t['username_required']}", 'danger')
            return redirect(url_for('find_account'))

        # ===== Kondisi 1: Username + Email =====
        if email and not phone:
            # Validasi format email
            email_pattern = r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"
            if not re.match(email_pattern, email):
                flash(f"{t['invalid_email_format']}", 'danger')
                return redirect(url_for('find_account'))
            user_exists = check_username_in_db(username)
            email_exists = check_email_in_db(email)

            if user_exists and email_exists:
                verification_code = generate_verification_code()
                session['verification_code'] = verification_code
                session['verification_code_expiry'] = time.time() + 180
                session['username'] = username
                # Kirim kode via email menggunakan helper function
                safe_code = escape(verification_code)
                html_body = f"""
                <p>Halo,</p>
                <p>Berikut adalah kode verifikasi 6 digit untuk mengakses akun Anda:</p>
                <h2>{safe_code}</h2>
                <p>Atau klik <a href="{url_for('verify_code', _external=True)}" style="color: blue;">link ini</a> untuk melanjutkan.</p>
                """
                send_email_message('Verifikasi Akun Anda', email, html_body)
                flash(f"{t['email_sent_code']} {escape(email)}", 'success')
                return redirect(url_for('verify_code'))
            elif not user_exists and not email_exists:
                flash(f"{t['username_email_not_found']}", 'danger')
            elif not user_exists:
                flash(f"{t['username_not_found']}", 'danger')
            elif not email_exists:
                flash(f"{t['email_not_found']}", 'danger')
    
        # ===== Kondisi 2: Username + Nomor HP =====
        elif phone and not email:
            normalized_phone = normalize_phone_number(phone)
            phone_pattern = r'^\+628\d{7,12}$'
            if not re.match(phone_pattern, normalized_phone):
                flash(f"{t['invalid_phone_format']}", 'danger')
                return redirect(url_for('find_account'))
            # Cek keberadaan username dan nomor HP di database
            user_exists = check_username_in_db(username)
            hp_exists = check_phone_in_db(normalized_phone)
            if user_exists and hp_exists:
                verification_code = generate_verification_code()
                session['verification_code'] = verification_code
                session['verification_code_expiry'] = time.time() + 180  # berlaku 3 menit
                session['username'] = username
                session['phone'] = normalized_phone
                send_whatsapp_code(normalized_phone, verification_code)
                flash(f"{t['whatsapp_sent_code']} {normalized_phone}", 'success')
                return redirect(url_for('verify_code'))
            # Penanganan error spesifik
            if not user_exists and not hp_exists:
                flash(f"{t['username_phone_not_found']}", 'danger')
            elif not user_exists:
                flash(f"{t['username_not_found']}", 'danger')
            elif not hp_exists:
                flash(f"{t['phone_not_found']}", 'danger')
            return redirect(url_for('find_account'))
        else:
            flash(f"{t['email_or_phone_required']}", 'danger')
        return redirect(url_for('find_account'))
    return render_template('find_account.html')

# Endpoint untuk verify_code
@app.route('/verify_code/', methods=['GET', 'POST'])
def verify_code():
    if request.method == 'GET':
        expiry_time = session.get('verification_code_expiry', 0)
        return render_template('verify_code.html', expiry_time=int(expiry_time))
    # Metode untuk memproses data JSON
    expiry_time = session.get('verification_code_expiry', 0)
    data = request.get_json()
    if not data or 'verification_code' not in data:
        return jsonify({'message': 'Verification code is required.'}), 400
    code = data['verification_code']
    if time.time() > expiry_time:
        return jsonify({'message': 'Verification code has expired.'}), 400
    if 'verification_code' in session and session['verification_code'] == int(code):
        # Generate reset token dan set waktu kedaluwarsa
        reset_token = secrets.token_hex(16)
        expiry_time = datetime.now() + timedelta(minutes=10)
        username = session.get('username')
        user = Users.query.filter_by(username=username).first()
        if not user:
            return jsonify({'message': 'User not found.'}), 404
        # Simpan token dan waktu kedaluwarsa di database lalu sertakan URL reset password dengan token
        user.reset_token = reset_token
        user.token_exp = expiry_time
        db.session.commit()
        reset_password_url = escape(url_for('reset_password', token=reset_token, _external=True))
        return jsonify({
            'message': 'Verification successful.',
            'redirect_url': reset_password_url
        }), 200
    else:
        return jsonify({'message': 'Incorrect verification code.'}), 400
    
# Endpoint Reset Password
@app.route('/reset_password/', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'GET':
        reset_token = request.args.get('token')
        if not reset_token:
            return jsonify({'error': "Token tidak ditemukan."}), 400
        # Validasi token
        user = Users.query.filter_by(reset_token=reset_token).first()
        if not user or datetime.now() > user.token_exp:
            if user:
                user.reset_token = None
                user.token_exp = None
                db.session.commit()
            return jsonify({'error': "Token tidak valid atau telah kedaluwarsa."}), 400
        # Jika token valid, arahkan ke halaman reset password
        return render_template('reset_password.html', token=escape(reset_token))
    if request.method == 'POST':
        data = request.get_json()
        reset_token = data.get('reset_token')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')
        # Debug input
        print(f"Reset token: {reset_token}, Password baru: {new_password}")
        # Validasi input
        if not reset_token or not new_password or not confirm_password:
            return jsonify({'error': "Semua data harus diisi."}), 400
        # Validasi pola password
        password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
        if not re.match(password_pattern, new_password):
            return jsonify({'error': "Password harus terdiri dari minimal 8 karakter, termasuk huruf besar, kecil, angka, dan simbol."}), 400
        if new_password != confirm_password:
            return jsonify({'error': "Password dan konfirmasi password tidak cocok."}), 400
        # Validasi token
        user = Users.query.filter_by(reset_token=reset_token).first()
        if not user:
            return jsonify({'error': "Token reset password tidak valid."}), 400
        if datetime.now() > user.token_exp:
            user.reset_token = None
            user.token_exp = None
            db.session.commit()
            return jsonify({'error': "Token reset password telah kedaluwarsa."}), 400
        # Update password
        hashed_password = generate_password_hash(new_password, method='pbkdf2:sha256', salt_length=16)
        user.password = hashed_password
        user.reset_token = None
        user.token_exp = None
        db.session.commit()
        print("Password berhasil diubah.")
        return jsonify({'message': "Password Anda telah berhasil diubah!"}), 200
    
# Route Index
@app.route("/")
@app.route("/index/")
def index():
    username = None
    user_data = None
    notification_count = 0
    profile_picture = None
    
    print("session keys:", session.keys())
    print("first_time_login:", session.get("first_time_login"))
    print("Session di /index/:", dict(session))
    first_time = session.get('first_time_login', None)

    # Cek login via session username (manual)
    if 'username' in session:
        username = session['username']
        user = Users.query.filter_by(username=username).first()
        if user:
            notification_count = Notification.query.filter_by(user_id=user.id, is_read=False).count()
            profile_picture = user.foto
        else:
            session.pop('username', None)

    # Cek login via session user (Google OAuth)
    elif 'user' in session:
        user_data = session['user'] 
        username = user_data.get('username')
        profile_picture = user_data.get('foto')
        user_id = user_data.get('id')
        if user_id:
            notification_count = Notification.query.filter_by(user_id=user_id, is_read=False).count()
            
    # Fix broken or missing profile picture - default to profil-default.png
    if not profile_picture or profile_picture in ['img/default-user.png', '']:
        profile_picture = 'images/profil-default.png'
        
    latest_news = (
        News.query
        .filter(News.status == 'published')
        .order_by(News.published_at.desc())
        .limit(3)
        .all()
    ) 
    return render_template('index.html', username=username, profile_picture=profile_picture, notification_count=notification_count, user_data=user_data, first_time_login=first_time, debug_theme=session.get("theme"), latest_news=latest_news, TRANSLATIONS=TRANSLATIONS)

@app.route("/clear-first-login-flag", methods=["POST"])
@csrf.exempt
def clear_first_login_flag():
    session.pop('first_time_login', None)
    return '', 204

@app.before_request
def check_session():
    print("Session sekarang:", dict(session))
    
@app.route('/set-language/<lang_code>')
def set_language(lang_code):
    if lang_code in ['id', 'en']:
        session['lang'] = lang_code
    return redirect(request.referrer or url_for('index'))

@app.context_processor
def inject_current_lang():
    lang = session.get('lang', 'id')
    return {
        'current_lang': lang,
        't': TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    }

@app.route('/set-theme/<theme>', methods=['POST'])
def set_theme(theme):
    if theme in ['light', 'dark']:
        session['theme'] = theme
        session.modified = True
        return '', 204
    return 'Invalid theme', 400

@app.context_processor
def inject_theme():
    return dict(current_theme=session.get('theme', 'light'))

@app.route('/save_sidebar_state', methods=['POST'])
@login_required
def save_sidebar_state():
    data = request.get_json()
    state = data.get('state')

    if state not in ['expanded', 'collapsed']:
        return jsonify({'status': 'error', 'message': 'Invalid state'}), 400

    current_user.sidebar_state = state
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Sidebar state saved'})

# --- Route Dashboard Admin ---
@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    sidebar_state = current_user.sidebar_state or 'expanded'
    user = current_user
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])

    # Validasi role admin
    if user.level == 'penilai':
        return redirect(url_for('penilai_dashboard'))
    elif user.level == 'peserta':
        return redirect(url_for('peserta_dashboard'))
    elif user.level != 'admin':
        flash(f"{t['access_denied_not_admin']}", "danger")
        return redirect(url_for('index'))

    total_users = Users.query.count()
    total_participants = Participants.query.count() if db.inspect(db.engine).has_table("participants") else 0
    total_criteria = Criteria.query.count() if db.inspect(db.engine).has_table("criteria") else 0
    total_events = Event.query.count() if db.inspect(db.engine).has_table("tb_kegiatan") else 0
    total_notifications = Notification.query.count()
    
    return render_template(
        'dashboard_admin.html',
        total_users=total_users,
        total_participants=total_participants,
        total_criteria=total_criteria,
        total_events=total_events,
        total_notifications=total_notifications,
        user=user,
        sidebar_state=sidebar_state
    )

# API untuk data chart dashboard admin
@app.route('/api/admin/dashboard/charts')
@login_required
def api_admin_dashboard_charts():
    """API untuk mendapatkan data chart dashboard admin"""
    # Cek apakah user adalah admin
    if current_user.level != 'admin':
        return jsonify({
            'success': False,
            'message': 'Akses ditolak. Hanya admin yang bisa mengakses data ini.'
        }), 403
    
    try:
        # 1. Distribusi User (Admin, Penilai, Peserta)
        admin_count = Users.query.filter_by(level='admin').count()
        penilai_count = Users.query.filter_by(level='penilai').count()
        peserta_count = Users.query.filter_by(level='peserta').count()
        
        # 2. Statistik Peserta (Status)
        peserta_aktif = Users.query.filter_by(level='peserta', status='aktif').count()
        peserta_nonaktif = Users.query.filter_by(level='peserta', status='non-aktif').count()
        
        # 3. Statistik Peserta (Jenis Kelamin)
        peserta_laki = db.session.query(Users).filter_by(level='peserta').filter(
            (Users.jenis_kelamin == 'laki-laki') | (Users.jenis_kelamin == 'Laki-laki')
        ).count()
        peserta_perempuan = db.session.query(Users).filter_by(level='peserta').filter(
            (Users.jenis_kelamin == 'perempuan') | (Users.jenis_kelamin == 'Perempuan')
        ).count()
        
        # 4. Statistik Notifikasi (Read vs Unread)
        notifications_read = Notification.query.filter_by(is_read=True).count()
        notifications_unread = Notification.query.filter_by(is_read=False).count()
        
        # 5. Statistik Kegiatan/Event
        total_events = Event.query.count()
        events_aktif = Event.query.filter(
            Event.waktu_pelaksanaan_dimulai <= date.today(),
            Event.waktu_pelaksanaan_selesai >= date.today()
        ).count()
        events_selesai = Event.query.filter(
            Event.waktu_pelaksanaan_selesai < date.today()
        ).count()
        events_mendatang = Event.query.filter(
            Event.waktu_pelaksanaan_dimulai > date.today()
        ).count()
        
        # 6. Statistik Hasil Seleksi (jika ada)
        total_hasil_seleksi = HasilSeleksi.query.count()
        
        return jsonify({
            'success': True,
            'data': {
                'user_distribution': {
                    'admin': admin_count,
                    'penilai': penilai_count,
                    'peserta': peserta_count
                },
                'peserta_status': {
                    'aktif': peserta_aktif,
                    'nonaktif': peserta_nonaktif
                },
                'peserta_gender': {
                    'laki_laki': peserta_laki,
                    'perempuan': peserta_perempuan
                },
                'notifications': {
                    'read': notifications_read,
                    'unread': notifications_unread
                },
                'events': {
                    'total': total_events,
                    'aktif': events_aktif,
                    'selesai': events_selesai,
                    'mendatang': events_mendatang
                },
                'hasil_seleksi': {
                    'total': total_hasil_seleksi
                }
            }
        })
    except Exception as e:
        logging.error(f"Error in api_admin_dashboard_charts: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# Route untuk melihat data penugasan penilai
@app.route('/admin/view_penugasan_penilai')
@login_required
def admin_view_penugasan_penilai():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    if 'username' not in session:
        flash(f"{t['login_required']}", "warning")
        return redirect(url_for('login'))
    user = current_user
    if not user or user.level != 'admin':
        flash(f"{t['access_denied']}", "danger")
        return redirect(url_for('index'))
    
    # Ambil semua event beserta kriteria dan evaluator yang ditugaskan
    events = Event.query.all()
    
    # Siapkan data untuk ditampilkan
    assignment_data = []
    for event in events:
        event_info = {
            'event': event,
            'criteria_assignments': []
        }
        
        # Ambil semua kriteria untuk event ini
        criteria_list = Criteria.query.filter_by(event_id=event.id_kegiatan).all()
        for criteria in criteria_list:
            # Ambil evaluator yang ditugaskan untuk kriteria ini
            evaluators = criteria.evaluators  # Menggunakan relationship yang sudah didefinisikan
            event_info['criteria_assignments'].append({
                'criteria': criteria,
                'evaluators': evaluators
            })
        
        assignment_data.append(event_info)
    return render_template(
        'penugasan_penilai_view.html',
        assignment_data=assignment_data,
        user=user,
        sidebar_state=sidebar_state
    )

# Middleware untuk membatasi akses hanya admin
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        lang = session.get('lang', 'id')
        t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
        
        if not current_user.is_authenticated:
            flash(f"{t['access_denied']}", "warning")
            return redirect(url_for('login'))

        if current_user.level != 'admin':
            flash(f"{t['admin_only_access']}", "danger")
            return redirect(url_for('index'))

        return f(*args, **kwargs)
    return decorated_function

# API Get Penilaian Peserta per Kegiatan
@app.route('/api/penilaian/peserta/<int:kegiatan_id>')
@login_required
@admin_required
def get_penilaian_peserta(kegiatan_id):
    try:
        # Ambil semua peserta yang terdaftar di kegiatan ini melalui many-to-many relationship
        peserta_list = Participants.query.join(
            tb_participant_kegiatan,
            Participants.id == tb_participant_kegiatan.c.participant_id
        ).filter(
            tb_participant_kegiatan.c.kegiatan_id == kegiatan_id
        ).all()
        
        data = []
        for p in peserta_list:
            # Cari user yang associated dengan peserta ini (berdasarkan email atau nama)
            # Note: Idealnya Participants punya relasi ke Users, tapi jika tidak ada kita cari manual
            # Asumsi: Participants.email match dengan Users.email
            user = Users.query.filter_by(email=p.email).first()
            
            status_validasi = "Belum Dinilai"
            total_nilai = 0
            has_nilai = False
            
            if user:
                # Cek penilaian
                penilaian = Penilaian.query.filter_by(id_users=user.id, id_kriteria=Criteria.query.filter_by(event_id=kegiatan_id).first().id_kriteria if Criteria.query.filter_by(event_id=kegiatan_id).first() else 0).all()
                
                # Hitung total nilai (simplifikasi, bisa disesuaikan dengan logika penilaian yang kompleks)
                nilai_records = Penilaian.query.filter(
                    Penilaian.id_users == user.id,
                    Penilaian.id_kriteria.in_([c.id_kriteria for c in Criteria.query.filter_by(event_id=kegiatan_id).all()])
                ).all()
                
                if nilai_records:
                    has_nilai = True
                    total_nilai = sum([n.nilai for n in nilai_records])
                    status_validasi = "Sudah Dinilai" # Bisa dikembangkan lagi logikanya
            
            data.append({
                'id': user.id if user else None, # User ID untuk keperluan hapus penilaian
                'participant_id': p.id,
                'nama': p.nama_lengkap,
                'golongan': p.golongan,
                'nilai': total_nilai if has_nilai else '-',
                'status': status_validasi,
                'has_nilai': has_nilai
            })
            
        return jsonify({'status': 'success', 'data': data}), 200
    except Exception as e:
        current_app.logger.exception('Error in /api/penilaian/peserta:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API Get Detail Penilaian per Kriteria
@app.route('/api/penilaian/detail/<int:user_id>/<int:kegiatan_id>')
@login_required
@admin_required
def get_detail_penilaian(user_id, kegiatan_id):
    try:
        # Ambil semua kriteria untuk kegiatan ini
        kriteria_list = Criteria.query.filter_by(event_id=kegiatan_id).all()
        
        data = []
        for kriteria in kriteria_list:
            # Ambil nilai untuk kriteria ini
            penilaian = Penilaian.query.filter_by(
                id_users=user_id,
                id_kriteria=kriteria.id_kriteria
            ).first()
            
            # Ambil nama penilai jika ada
            penilai_nama = None
            if penilaian and penilaian.evaluator_id:
                evaluator = Users.query.get(penilaian.evaluator_id)
                if evaluator:
                    penilai_nama = evaluator.nama_lengkap
            
            data.append({
                'kriteria': kriteria.nama_kriteria,
                'bobot': kriteria.bobot,
                'nilai': penilaian.nilai if penilaian else 0,
                'penilai': penilai_nama
            })
        
        return jsonify({'status': 'success', 'data': data}), 200
    except Exception as e:
        current_app.logger.exception('Error in /api/penilaian/detail:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Route View Detail Penilaian Peserta (Halaman)
@app.route('/admin/penilaian/detail_view/<int:user_id>/<int:kegiatan_id>')
@login_required
@admin_required
def admin_penilaian_detail_view(user_id, kegiatan_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
        
    try:
        user_peserta = Users.query.get_or_404(user_id)
        participant = Participants.query.filter_by(email=user_peserta.email).first()
        event = Event.query.get_or_404(kegiatan_id)
        
        # Ambil semua kriteria untuk kegiatan ini
        kriteria_list = Criteria.query.filter_by(event_id=kegiatan_id).order_by(Criteria.id_kriteria).all()
        
        # Ambil SEMUA penilaian untuk peserta ini di kegiatan ini
        all_penilaian = Penilaian.query.filter(
            Penilaian.id_users == user_id,
            Penilaian.id_kriteria.in_([k.id_kriteria for k in kriteria_list])
        ).all()
        
        # Kumpulkan semua evaluator yang memberikan nilai
        evaluator_ids = set()
        for p in all_penilaian:
            if p.evaluator_id:
                evaluator_ids.add(p.evaluator_id)
        
        # Ambil data evaluator
        evaluators = []
        if evaluator_ids:
            evaluator_users = Users.query.filter(Users.id.in_(evaluator_ids)).all()
            evaluators = sorted(evaluator_users, key=lambda x: x.id)
        
        # Buat mapping: (evaluator_id, kriteria_id) -> nilai  
        score_map = {}
        for p in all_penilaian:
            score_map[(p.evaluator_id, p.id_kriteria)] = p.nilai
        
        # Susun data detail per kriteria dengan semua penilai
        detail_scores = []
        
        # Calculate Fuzzy AHP weights for correct bobot values
        from app.ahp_calculator import FuzzyAHPCalculator
        from app.fuzzy_ahp import get_pairwise_matrix_from_db
        
        criteria_names_list = [k.nama_kriteria for k in kriteria_list]
        criteria_ids_list = [k.id_kriteria for k in kriteria_list]
        pairwise_matrix = get_pairwise_matrix_from_db(kegiatan_id, criteria_ids_list)
        
        fuzzy_ahp_weight_map = {}
        if pairwise_matrix is not None and len(kriteria_list) > 1:
            fuzzy_ahp_calc = FuzzyAHPCalculator(criteria_names_list)
            fuzzy_ahp_calc.set_fuzzy_pairwise_matrix(pairwise_matrix)
            fuzzy_ahp_calc.calculate_fuzzy_synthetic_extent()
            fuzzy_weights = fuzzy_ahp_calc.calculate_fuzzy_weights()
            for name in criteria_names_list:
                fuzzy_ahp_weight_map[name] = round(fuzzy_weights.get(name, 0), 4)
        
        # Fallback to database weights if Fuzzy AHP weights not available
        if not fuzzy_ahp_weight_map:
            total_bobot = sum(k.bobot for k in kriteria_list)
            for k in kriteria_list:
                fuzzy_ahp_weight_map[k.nama_kriteria] = (k.bobot / total_bobot) if total_bobot > 0 else 0
        
        # Data untuk perhitungan Fuzzy AHP (seperti di halaman hasil)
        fuzzy_total_l = 0
        fuzzy_total_m = 0
        fuzzy_total_u = 0
        calculation_details = []
        
        for kriteria in kriteria_list:
            # Nilai per evaluator
            evaluator_scores = []
            nilai_list = []
            for ev in evaluators:
                nilai = score_map.get((ev.id, kriteria.id_kriteria), None)
                evaluator_scores.append({
                    'evaluator_id': ev.id,
                    'evaluator_nama': ev.nama_lengkap,
                    'nilai': nilai
                })
                if nilai is not None:
                    nilai_list.append(nilai)
            
            # Hitung rata-rata
            avg_nilai = sum(nilai_list) / len(nilai_list) if nilai_list else 0
            
            # Use Fuzzy AHP weight (consistent with Bobot Global)
            weight = fuzzy_ahp_weight_map.get(kriteria.nama_kriteria, 0)
            
            # Fuzzifikasi (sama dengan logika di fuzzy_ahp.py dan halaman hasil)
            if avg_nilai > 0:
                if avg_nilai <= 5:  # Likert scale
                    if avg_nilai <= 1:
                        l, m, u = 1, 1, 2
                    elif avg_nilai <= 2:
                        l, m, u = 1, 2, 3
                    elif avg_nilai <= 3:
                        l, m, u = 2, 3, 4
                    elif avg_nilai <= 4:
                        l, m, u = 3, 4, 5
                    else:
                        l, m, u = 4, 5, 5
                else:  # 0-100 scale
                    l = max(0, avg_nilai - 5)
                    m = avg_nilai
                    u = min(100, avg_nilai + 5)
                
                weighted_l = l * weight
                weighted_m = m * weight
                weighted_u = u * weight
                
                fuzzy_total_l += weighted_l
                fuzzy_total_m += weighted_m
                fuzzy_total_u += weighted_u
            else:
                l, m, u = 0, 0, 0
                weighted_l, weighted_m, weighted_u = 0, 0, 0
            
            detail_scores.append({
                'kriteria': kriteria.nama_kriteria,
                'kriteria_id': kriteria.id_kriteria,
                'bobot': kriteria.bobot,
                'weight': round(weight, 2),
                'evaluator_scores': evaluator_scores,
                'avg_nilai': round(avg_nilai, 2),
                'fuzzy_l': round(l, 2),
                'fuzzy_m': round(m, 2),
                'fuzzy_u': round(u, 2),
                'weighted_l': round(weighted_l, 2),
                'weighted_m': round(weighted_m, 2),
                'weighted_u': round(weighted_u, 2)
            })
        
        # Skor akhir defuzzifikasi
        final_score = round((fuzzy_total_l + fuzzy_total_m + fuzzy_total_u) / 3, 2) if detail_scores else 0
        
        # Ambil hasil seleksi (ranking)
        hasil_seleksi = HasilSeleksi.query.filter_by(
            id_users=user_id,
            event_id=kegiatan_id
        ).first()
            
        sidebar_state = current_user.sidebar_state or 'expanded'
            
        return render_template(
            'penilaian_peserta_detail.html',
            user=current_user,
            participant=participant,
            user_peserta=user_peserta,
            event=event,
            detail_scores=detail_scores,
            evaluators=evaluators,
            fuzzy_total_l=round(fuzzy_total_l, 2),
            fuzzy_total_m=round(fuzzy_total_m, 2),
            fuzzy_total_u=round(fuzzy_total_u, 2),
            final_score=final_score,
            hasil_seleksi=hasil_seleksi,
            sidebar_state=sidebar_state
        )
    except Exception as e:
        current_app.logger.exception('Error in admin_penilaian_detail_view:')
        flash(f"{t['error_occurred']}: {str(e)}", "danger")
        return redirect(url_for('admin_manajemen_seleksi'))

# API Hapus Penilaian Peserta
@app.route('/api/penilaian/hapus', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def delete_penilaian():
    try:
        data = request.get_json(force=True)
        user_id = data.get('user_id')
        kegiatan_id = data.get('kegiatan_id')
        
        if not user_id or not kegiatan_id:
            return jsonify({'status': 'error', 'message': 'Parameter tidak lengkap'}), 400
            
        # Cari kriteria yang berhubungan dengan kegiatan ini
        criteria_ids = [c.id_kriteria for c in Criteria.query.filter_by(event_id=kegiatan_id).all()]
        
        if not criteria_ids:
             return jsonify({'status': 'error', 'message': 'Tidak ada kriteria untuk kegiatan ini'}), 404

        # Hapus penilaian
        deleted_count = Penilaian.query.filter(
            Penilaian.id_users == user_id,
            Penilaian.id_kriteria.in_(criteria_ids)
        ).delete(synchronize_session=False)
        
        # Hapus juga hasil seleksi jika ada
        HasilSeleksi.query.filter_by(id_users=user_id).delete()
        
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': f'Berhasil menghapus {deleted_count} data penilaian'}), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/penilaian/hapus:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.all()
    users_data = []
    for u in users:
        user_dict = u.to_dict()
        # Cari biodata berdasarkan email
        biodata = Participants.query.filter_by(email=u.email).first()
        if biodata:
            # Tambahkan data biodata ke user_dict
            user_dict['biodata'] = {
                'nama_lengkap': biodata.nama_lengkap or '',
                'tanggal_lahir': biodata.tanggal_lahir.strftime('%Y-%m-%d') if biodata.tanggal_lahir else '',
                'alamat_tinggal': biodata.alamat_tinggal or '',
                'golongan': biodata.golongan or '',
                'tingkatan': biodata.tingkatan or '',
                'asal_gudep': biodata.asal_gudep or '',
                'asal_kwarran': biodata.asal_kwarran or '',
                'asal_kwarcab': biodata.asal_kwarcab or '',
                'asal_kwarda': biodata.asal_kwarda or '',
                'usia': biodata.usia or '',
                'jenis_kelamin': biodata.jenis_kelamin or '',
                'email': biodata.email or '',
                'nomor_hp': biodata.nomor_hp or '',
                'foto': biodata.foto or ''
            }
        else:
            user_dict['biodata'] = None
        users_data.append(user_dict)
    return render_template('manajemen_pengguna.html', sidebar_state=sidebar_state, users=users_data, time=time)

@app.route('/admin/add_user', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_user():
    sidebar_state = current_user.sidebar_state or 'expanded'
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])

    if request.method == 'POST':
        # Ambil data dari form
        nama_lengkap = request.form.get('fullName')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form['confirmPassword']
        level = request.form.get('level')  

        # Validasi password dengan regex
        password_pattern = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
        if not re.match(password_pattern, password):
            flash(t['invalid_password_format'], "danger")
            return redirect(url_for('admin_users'))
        # Validasi apakah password dan confirmPassword cocok
        if password != confirm_password:
            flash(t['password_mismatch'], "danger")
            return redirect(url_for('admin_users'))
        # Cek keberadaan username dan email di database
        user_exists = check_username_in_db(username)
        email_exists = check_email_in_db(email)
        if not user_exists and not email_exists:
            # Enkripsi password
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16) 
            # Masukkan data ke database
            try:
                # Ambil field tambahan dari form atau gunakan default
                jenis_kelamin = request.form.get('jenis_kelamin', 'laki-laki')
                usia = request.form.get('usia', '0')
                nomor_hp = request.form.get('nomor_hp', '')
                foto = request.form.get('foto', 'img/default-user.png')
                
                new_user = Users(
                    username=username,
                    password=hashed_password,
                    nama_lengkap=nama_lengkap,
                    email=email,
                    jenis_kelamin=jenis_kelamin,
                    usia=usia,
                    foto=foto,
                    nomor_hp=nomor_hp,
                    level=level,
                    reset_token="",
                    login_method="manual",
                    sidebar_state="expanded",
                    status='aktif'
                )
                db.session.add(new_user)
                db.session.commit()
                
                # Buat notifikasi untuk admin jika user yang dibuat adalah peserta
                if level == 'peserta':
                    create_notification_to_all_admins(
                        f"Peserta baru terdaftar: {nama_lengkap} ({email})"
                    )
                
                flash(t['user_created'], "success")
                return redirect(url_for('admin_users', page='kelola'))
            except Exception as e:
                db.session.rollback()
                logging.error(f"Error during registration: {e}")
                flash(t['registration_failed'], "danger")
                print(e)
        elif user_exists and email_exists:
            flash(t['username_email_exists'], "danger")
        elif user_exists:
            flash(t['username_exists'], "danger")
        elif email_exists:
            flash(t['email_exists'], "danger")
        return redirect(url_for('admin_users'))
    return render_template('manajemen_pengguna.html', sidebar_state=sidebar_state, users=Users.query.all(), time=time)

# Admin/Import Data User
@app.route('/admin/import_users', methods=['POST'])
@login_required
@admin_required
def admin_import_users():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if 'file' not in request.files:
        flash(t['no_file_uploaded'], 'danger')
        return redirect(url_for('admin_users'))
    
    file = request.files['file']
    if not file.filename:
        flash(t['invalid_file_name'], 'danger')
        return redirect(url_for('admin_users'))

    if file.mimetype not in [
        'text/csv',
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]:
        flash(t['unsupported_file_type'], 'danger')
        return redirect(url_for('admin_users'))

    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if not allowed_file(file.filename, 'doc'):
        flash(t['file_not_allowed'], 'danger')
        return redirect(url_for('admin_users'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        df = pd.read_csv(filepath) if ext == 'csv' else pd.read_excel(filepath)
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_").str.replace("-", "_")

        required_cols = ['nama_lengkap', 'username', 'email', 'level']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            flash(f"{t['missing_columns']}: {', '.join(missing)}", 'danger')
            return redirect(url_for('admin_users'))

        existing_usernames = {u[0] for u in db.session.query(Users.username).all()}
        valid_levels = {'admin', 'penilai', 'peserta'}
        count_added, count_skipped = 0, 0
        new_users = []

        for _, row in df.iterrows():
            # Validasi
            if pd.isna(row['username']) or pd.isna(row['email']):
                count_skipped += 1
                continue
            if row['username'] in existing_usernames:
                count_skipped += 1
                continue
            if row['level'] not in valid_levels:
                count_skipped += 1
                continue
            if '@' not in str(row['email']):
                count_skipped += 1
                continue

            password = row['password'] if 'password' in df.columns and pd.notna(row['password']) else '12345678'
            new_users.append(Users(
                nama_lengkap=row['nama_lengkap'],
                username=row['username'],
                email=row['email'],
                password=generate_password_hash(str(password)),
                level=row['level'],
                jenis_kelamin=row.get('jenis_kelamin'),
                usia=row.get('usia', 0),
                nomor_hp=row.get('nomor_hp', "")
            ))
            count_added += 1

        if new_users:
            db.session.bulk_save_objects(new_users)
            db.session.commit()
            flash(f"{t['import_success']}: {count_added} {t['users_added']}, {count_skipped} {t['users_skipped']}.", 'success')
        else:
            flash(t['no_new_users'], 'warning')
    except Exception as e:
        db.session.rollback()
        app.logger.exception(f"Import gagal: {e}")
        flash(f"{t['import_failed']}: {e}", 'danger')
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
    return redirect(url_for('admin_users'))

# Download Data User
@app.route('/download_users')
def download_users():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    users = Users.query.order_by(Users.nama_lengkap.asc()).all()
    
    # Jika tidak ada data pengguna
    if not users:
        flash(t['no_user_data'], 'warning')
        return redirect(url_for('admin_users'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = t['user_data_sheet']
    headers = [t['no'], t['full_name'], t['username'], t['email'], t['level'], t['status']]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for i, u in enumerate(users, start=1):
        ws.append([
            i,
            u.nama_lengkap or '',
            u.username or '',
            u.email or '',
            u.level or '',
            u.status or ''
        ])
        
    # Auto width untuk setiap kolom
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"data_pengguna_{timestamp}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

# Admin/Delete User
@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    # Hanya admin boleh hapus
    if current_user.level != "admin":
        flash(f"{t['no_permission_delete_user']}", "danger")
        return redirect(url_for('admin_users'))
    
    user = Users.query.get(user_id)
    if not user:
        flash(f"{t['user_not_found']}", "danger")
        return redirect(url_for('admin_users'))
    
    # Proteksi: admin tidak bisa menghapus dirinya sendiri
    if user.id == current_user.id:
        flash(f"{t['cannot_delete_self']}", "warning")
        return redirect(url_for('admin_users'))
    try:
        db.session.delete(user)
        db.session.commit()
        flash(f"{t['user_deleted'].format(username=user.username)}", "success")
        logging.info(f"User '{user.username}' berhasil dihapus oleh admin.")
    except Exception as e:
        db.session.rollback()
        flash(f"{t['error_delete_user']}", "danger")
        logging.error(f"Gagal menghapus user_id {user_id}: {e}")
    return redirect(url_for('admin_users'))

# Admin/Edit User
@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    user = Users.query.get(user_id)
    if not user:
        flash(f"{t['user_not_found']}", "danger")
        return redirect(url_for('admin_users'))
    
    # Jika GET request, hanya redirect ke halaman manajemen pengguna
    if request.method == 'GET':
        return redirect(url_for('admin_users'))
    
    # POST request - proses update data
    try:
        # Field required
        nama_lengkap = request.form.get('nama_lengkap', '').strip()
        if not nama_lengkap:
            flash(f"{t['full_name_required']}", "danger")
            return redirect(url_for('admin_users'))
        user.nama_lengkap = nama_lengkap
        email = request.form.get('email', '').strip().lower()
        if not email:
            flash(f"{t['email_required']}", "danger")
            return redirect(url_for('admin_users'))
        # Cek apakah email sudah digunakan oleh user lain
        existing_user = Users.query.filter(Users.email == email, Users.id != user_id).first()
        if existing_user:
            flash(f"{t['email_used_by_other']}", "danger")
            return redirect(url_for('admin_users'))
        user.email = email
        username = request.form.get('username', '').strip()
        if not username:
            flash(f"{t['username_required']}", "danger")
            return redirect(url_for('admin_users'))
        # Cek apakah username sudah digunakan oleh user lain
        existing_user = Users.query.filter(Users.username == username, Users.id != user_id).first()
        if existing_user:
            flash(f"{t['username_used_by_other']}", "danger")
            return redirect(url_for('admin_users'))
        user.username = username
        
        # Field optional dengan validasi
        level = request.form.get('level', '').strip()
        if level and level in ['admin', 'penilai', 'peserta']:
            user.level = level
        elif not user.level:  # Pastikan level selalu ada
            user.level = 'peserta'
        
        status = request.form.get('status', '').strip()
        if status:
            # Normalisasi status: "nonaktif" -> "non-aktif"
            if status == 'nonaktif':
                status = 'non-aktif'
            if status in ['aktif', 'non-aktif']:
                user.status = status
        elif not user.status:  # Pastikan status selalu ada
            user.status = 'aktif'
        
        jenis_kelamin = request.form.get('jenis_kelamin', '').strip()
        if jenis_kelamin:
            # Normalisasi jenis kelamin
            if jenis_kelamin.lower() in ['laki-laki', 'laki laki']:
                user.jenis_kelamin = 'laki-laki'
            elif jenis_kelamin.lower() == 'perempuan':
                user.jenis_kelamin = 'perempuan'
        elif not user.jenis_kelamin:  # Pastikan jenis_kelamin selalu ada
            user.jenis_kelamin = 'laki-laki'
        
        usia = request.form.get('usia', '').strip()
        if usia:
            user.usia = str(usia)  # Pastikan string
        elif not user.usia:  # Jika kosong dan belum ada nilai sebelumnya
            user.usia = '0'
        
        nomor_hp = request.form.get('nomor_hp', '').strip()
        if nomor_hp:
            user.nomor_hp = nomor_hp
        elif not user.nomor_hp:  # Jika kosong dan belum ada nilai sebelumnya
            user.nomor_hp = ''

        # Handle upload foto jika ada
        foto_file = request.files.get('foto')
        if foto_file and foto_file.filename:
            # Validasi file extension
            if not allowed_file(foto_file.filename, 'image'):
                flash("Format file tidak didukung! Gunakan file gambar (png, jpg, jpeg, gif).", "danger")
                return redirect(url_for('admin_users'))
            
            try:
                # Generate unique filename
                filename = secure_filename(foto_file.filename)
                if not filename:
                    flash(f"{t['invalid_image_format']}", "danger")
                    return redirect(url_for('admin_users'))
                    
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_filename = f"{timestamp}_{filename}"
                
                # Buat folder users jika belum ada
                users_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'users')
                os.makedirs(users_upload_dir, exist_ok=True)
                
                # Simpan file
                foto_path = os.path.join(users_upload_dir, unique_filename)
                foto_file.save(foto_path)
                
                # Verifikasi file berhasil disimpan
                if os.path.exists(foto_path) and os.path.getsize(foto_path) > 0:
                    # Update path foto (relative dari static folder untuk url_for)
                    user.foto = f"uploads/users/{unique_filename}"
                    logging.info(f"Foto berhasil diupload untuk user_id {user_id}: {foto_path}")
                else:
                    flash(f"{t['failed_save_photo']}", "danger")
                    logging.error(f"File tidak ditemukan atau kosong setelah save: {foto_path}")
                    return redirect(url_for('admin_users'))
            except Exception as e:
                flash(f"{t['error_upload_photo'].format(error=str(e))}", "danger")
                logging.error(f"Error uploading foto for user_id {user_id}: {e}")
                current_app.logger.exception('Error uploading foto:')
                return redirect(url_for('admin_users'))

        db.session.commit()
        flash(f"{t['user_updated'].format(username=user.username)}", "success")
        log_activity(
            current_user.id,
            f'Mengupdate data pengguna: {user.username}'
        )
    except IntegrityError as e:
        db.session.rollback()
        flash(f"{t['email_or_username_used']}", "danger")
        logging.error(f"Integrity error saat update user_id {user_id}: {e}")
        current_app.logger.exception('Integrity error in edit_user:')
    except ValueError as e:
        db.session.rollback()
        flash(f"Data tidak valid: {str(e)}", "danger")
        logging.error(f"Value error saat update user_id {user_id}: {e}")
        current_app.logger.exception('Value error in edit_user:')
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        flash(f"{t['error_update_data'].format(error=error_msg)}", "danger")
        logging.error(f"Gagal memperbarui user_id {user_id}: {error_msg}")
        current_app.logger.exception('Error in edit_user:')
        # Print error untuk debugging
        print(f"ERROR in edit_user: {error_msg}")
        import traceback
        traceback.print_exc()
    return redirect(url_for('admin_users'))

@app.route('/api/user/update_status/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def api_update_user_status(user_id):
    """API endpoint untuk update status akun user"""
    try:
        user = Users.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User tidak ditemukan'}), 404
        
        data = request.get_json()
        new_status = data.get('status', '').strip()
        
        # Normalisasi status
        if new_status == 'nonaktif':
            new_status = 'non-aktif'
        elif new_status not in ['aktif', 'non-aktif']:
            return jsonify({'success': False, 'message': 'Status tidak valid'}), 400
        
        # Update status
        old_status = user.status
        user.status = new_status
        db.session.commit()
        
        # Log aktivitas
        status_text = 'mengaktifkan' if new_status == 'aktif' else 'menonaktifkan'
        log_activity(
            current_user.id,
            f'{status_text.capitalize()} akun pengguna: {user.username}'
        )
        
        return jsonify({
            'success': True,
            'message': f'Status akun {user.username} berhasil diubah menjadi {new_status}',
            'status': new_status
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in api_update_user_status:')
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'}), 500

# Manajemen Seleksi   
@app.route('/admin/manajemen_seleksi')
@login_required
@admin_required
def admin_manajemen_seleksi():
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    # Get all events
    kegiatan_list = Event.query.all()
    
    # Convert to serializable format if needed
    kegiatan_data = []
    for kegiatan in kegiatan_list:
        # Hitung jumlah peserta
        jumlah_peserta = kegiatan.registered_participants.count()
        
        kegiatan_data.append({
            'id': kegiatan.id_kegiatan,
            'nama': kegiatan.nama_kegiatan,
            'jenis': kegiatan.jenis_kegiatan,
            'waktu_mulai': kegiatan.waktu_pelaksanaan_dimulai.strftime('%Y-%m-%d') if kegiatan.waktu_pelaksanaan_dimulai else None,
            'waktu_selesai': kegiatan.waktu_pelaksanaan_selesai.strftime('%Y-%m-%d') if kegiatan.waktu_pelaksanaan_selesai else None,
            'jumlah_peserta': jumlah_peserta
        })
    
    # Ambil semua arsip untuk ditampilkan
    arsip_list = ArsipSeleksi.query.order_by(ArsipSeleksi.tanggal_arsip.desc()).all()
    arsip_data = []
    for arsip in arsip_list:
        arsip_data.append({
            'id': arsip.id_arsip,
            'event_id': arsip.event_id,
            'nama_kegiatan': arsip.event.nama_kegiatan if arsip.event else 'N/A',
            'nama_arsip': arsip.nama_arsip,
            'deskripsi': arsip.deskripsi,
            'file_path': arsip.file_path,
            'file_type': arsip.file_type,
            'tanggal_arsip': arsip.tanggal_arsip.strftime('%d %b %Y') if arsip.tanggal_arsip else '',
            'pembuat': arsip.pembuat.nama_lengkap if arsip.pembuat else 'System',
            'status': arsip.status
        })
    
    return render_template(
        "manajemen_seleksi.html", 
        kegiatan_list=kegiatan_list, 
        kegiatan_data=kegiatan_data,
        arsip_list=arsip_data,
        sidebar_state=sidebar_state
    )

# Route untuk halaman penugasan penilai
@app.route('/admin/penugasan_penilai')
@login_required
@admin_required
def admin_penugasan_penilai():
    sidebar_state = current_user.sidebar_state or 'expanded'
    events = Event.query.all()
    evaluators = Users.query.filter_by(level='penilai').all()
    
    # Build assignment matrix
    assignments = {}
    for evaluator in evaluators:
        evaluator_assignments = {}
        
        # Get assigned criteria grouped by event
        # We assume evaluator.assigned_criteria exists due to the backref in Criteria model
        if hasattr(evaluator, 'assigned_criteria'):
            for criterion in evaluator.assigned_criteria:
                if criterion.event_id not in evaluator_assignments:
                    evaluator_assignments[criterion.event_id] = []
                evaluator_assignments[criterion.event_id].append(criterion.id_kriteria)
        
        assignments[evaluator.id] = evaluator_assignments
    
    # Prepare criteria data for frontend
    events_criteria = {}
    for event in events:    
        events_criteria[event.id_kegiatan] = [
            {'id': c.id_kriteria, 'nama': c.nama_kriteria} 
            for c in event.kriteria
        ]

    return render_template(
        "manajemen-seleksi/penugasan_penilai.html", 
        events=events,
        evaluators=evaluators,
        assignments=assignments,
        events_criteria=events_criteria,
        sidebar_state=sidebar_state
    )

# API untuk update penugasan kriteria penilai
@app.route('/api/update_evaluator_criteria', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def update_evaluator_criteria():
    try:
        data = request.get_json()
        event_id = data.get('event_id')
        evaluator_id = data.get('evaluator_id')
        criteria_ids = data.get('criteria_ids', []) # List of selected criteria IDs
        
        if not event_id or not evaluator_id:
            return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400
        
        event = Event.query.get_or_404(event_id)
        evaluator = Users.query.get_or_404(evaluator_id)
        
        # 1. Update Criteria Assignment
        # Get all criteria for this event
        event_criteria = Criteria.query.filter_by(event_id=event_id).all()
        
        # Remove evaluator from all event criteria first
        for criterion in event_criteria:
            if evaluator in criterion.evaluators:
                criterion.evaluators.remove(evaluator)
        
        # Add evaluator to selected criteria
        for c_id in criteria_ids:
            criterion = Criteria.query.get(c_id)
            if criterion and criterion.event_id == int(event_id):
                criterion.evaluators.append(evaluator)
        
        # 2. Sync with Event Assignment (tb_event_evaluator)
        # If evaluator has ANY criteria assigned, they should be in event.evaluators
        # If they have NO criteria assigned, they should be removed from event.evaluators (subject to min 3 rule)
        
        has_criteria = len(criteria_ids) > 0
        
        if has_criteria:
            if evaluator not in event.evaluators:
                event.evaluators.append(evaluator)
        else:
            # Trying to remove evaluator from event completely
            if evaluator in event.evaluators:
                # Check min 3 rule
                # We need to count how many evaluators are assigned to this event
                # excluding the current one if we are about to remove them
                current_evaluator_count = len(event.evaluators)
                
                if current_evaluator_count <= 3:
                     # Revert criteria changes? 
                     # Or just block the whole operation if it results in < 3 evaluators?
                     # But wait, maybe they just wanted to change criteria, not remove the user?
                     # If criteria_ids is empty, it means they are being unassigned.
                     
                     db.session.rollback()
                     return jsonify({
                         'status': 'error', 
                         'message': 'Minimal 3 penilai harus ditugaskan! Tidak dapat menghapus penilai ini.'
                     }), 400
                
                event.evaluators.remove(evaluator)
        
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Penugasan berhasil diperbarui'}), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error updating evaluator criteria:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Legacy API (kept for reference or fallback, but logic moved to update_evaluator_criteria)
@app.route('/api/assign_evaluator', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def assign_evaluator():
    return jsonify({'status': 'error', 'message': 'Please use criteria assignment'}), 400

@app.route('/api/unassign_evaluator', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def unassign_evaluator():
     return jsonify({'status': 'error', 'message': 'Please use criteria assignment'}), 400

# API untuk get assignments
@app.route('/api/get_assignments')
@login_required
@admin_required
def get_assignments():
    try:
        events = Event.query.all()
        result = []
        
        for event in events:
            event_data = {
                'id': event.id_kegiatan,
                'nama': event.nama_kegiatan,
                'evaluators': [{'id': e.id, 'nama': e.nama_lengkap} for e in event.evaluators]
            }
            result.append(event_data)
        
        return jsonify({'status': 'success', 'data': result}), 200
    except Exception as e:
        current_app.logger.exception('Error getting assignments:')
        return jsonify({'status': 'error', 'message': str(e)}), 500


# Konfigurasi Seleksi
@app.route('/api/save_config', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def save_config():
    try:
        data = request.get_json(force=True)
        activities = data.get('activities', [])
        criteria_list = data.get('criteria', [])

        if not activities and not criteria_list:
            return jsonify({'status': 'error', 'message': 'No data provided'}), 400
        created_events = []

        # Buat Event & Kuota
        for act in activities:
            nama = (act.get('nama') or '').strip()
            if not nama:
                continue
            
            # Parse tanggal
            mulai_date = None
            selesai_date = None
            waktu_pelaksanaan_dimulai_date = None
            waktu_pelaksanaan_selesai_date = None
            try:
                if act.get('mulai'):
                    # Handle datetime-local format (YYYY-MM-DDTHH:mm)
                    mulai_str = act['mulai']
                    if 'T' in mulai_str:
                        mulai_date = datetime.strptime(mulai_str.split('T')[0], '%Y-%m-%d').date()
                    else:
                        mulai_date = datetime.strptime(mulai_str, '%Y-%m-%d').date()
                if act.get('selesai'):
                    # Handle datetime-local format (YYYY-MM-DDTHH:mm)
                    selesai_str = act['selesai']
                    if 'T' in selesai_str:
                        selesai_date = datetime.strptime(selesai_str.split('T')[0], '%Y-%m-%d').date()
                    else:
                        selesai_date = datetime.strptime(selesai_str, '%Y-%m-%d').date()
                # Parse waktu pelaksanaan dimulai dan selesai
                if act.get('waktuMulai'):
                    try:
                        # Format datetime-local: "YYYY-MM-DDTHH:mm" atau "YYYY-MM-DD"
                        waktu_str = act['waktuMulai']
                        if 'T' in waktu_str:
                            waktu_pelaksanaan_dimulai_date = datetime.strptime(waktu_str.split('T')[0], '%Y-%m-%d').date()
                        else:
                            waktu_pelaksanaan_dimulai_date = datetime.strptime(waktu_str, '%Y-%m-%d').date()
                    except Exception:
                        waktu_pelaksanaan_dimulai_date = mulai_date if mulai_date else None
                else:
                    waktu_pelaksanaan_dimulai_date = mulai_date if mulai_date else None
                
                if act.get('waktuSelesai'):
                    try:
                        # Format datetime-local: "YYYY-MM-DDTHH:mm" atau "YYYY-MM-DD"
                        waktu_str = act['waktuSelesai']
                        if 'T' in waktu_str:
                            waktu_pelaksanaan_selesai_date = datetime.strptime(waktu_str.split('T')[0], '%Y-%m-%d').date()
                        else:
                            waktu_pelaksanaan_selesai_date = datetime.strptime(waktu_str, '%Y-%m-%d').date()
                    except Exception:
                        waktu_pelaksanaan_selesai_date = selesai_date if selesai_date else waktu_pelaksanaan_dimulai_date
                else:
                    waktu_pelaksanaan_selesai_date = selesai_date if selesai_date else waktu_pelaksanaan_dimulai_date
            except Exception:
                pass
            
            # Normalisasi jenis_kegiatan (ENUM case-sensitive)
            jenis_kegiatan_map = {
                'siaga': 'Siaga',
                'penggalang': 'Penggalang',
                'penegak': 'Penegak',
                'pandega': 'Pandega',
                'penegak dan pandega': 'Penegak dan Pandega'
            }
            jenis_raw = (act.get('jenis') or '').strip().lower()
            jenis_kegiatan = jenis_kegiatan_map.get(jenis_raw, 'Siaga')  # Default ke Siaga
            
            # Normalisasi skala_kegiatan (ENUM case-sensitive)
            skala_kegiatan_map = {
                'ranting': 'Ranting',
                'cabang': 'Cabang',
                'daerah': 'Daerah',
                'nasional': 'Nasional',
                'internasional': 'Internasional'
            }
            skala_raw = (act.get('skala') or '').strip().lower()
            skala_kegiatan = skala_kegiatan_map.get(skala_raw, 'Ranting')  # Default ke Ranting
            
            # Validasi tempat_pelaksanaan
            tempat = (act.get('tempat') or '').strip()
            if not tempat:
                tempat = '-'
            
            # Validasi kwartir_penyelenggara
            kwartir = (act.get('kwartir') or '').strip()
            if not kwartir:
                kwartir = 'Kwartir Ranting'
            
            # Pastikan semua tanggal ada
            if not mulai_date:
                mulai_date = datetime.utcnow().date()
            if not selesai_date:
                selesai_date = mulai_date
            if not waktu_pelaksanaan_dimulai_date:
                waktu_pelaksanaan_dimulai_date = mulai_date
            if not waktu_pelaksanaan_selesai_date:
                waktu_pelaksanaan_selesai_date = waktu_pelaksanaan_dimulai_date
            
            # Validasi: Periode Seleksi harus selesai sebelum Waktu Pelaksanaan dimulai
            # Cek apakah mulai periode seleksi >= waktu pelaksanaan
            if mulai_date >= waktu_pelaksanaan_dimulai_date:
                return jsonify({
                    'status': 'error', 
                    'message': f'Periode Seleksi (mulai) untuk kegiatan "{nama}" harus sebelum Waktu Pelaksanaan dimulai'
                }), 400
            
            # Cek apakah selesai periode seleksi >= waktu pelaksanaan
            if selesai_date >= waktu_pelaksanaan_dimulai_date:
                return jsonify({
                    'status': 'error', 
                    'message': f'Periode Seleksi (selesai) untuk kegiatan "{nama}" harus sebelum Waktu Pelaksanaan dimulai'
                }), 400
            
            # Validasi: Waktu Pelaksanaan tidak boleh dalam kurun waktu Periode Seleksi
            # Cek apakah waktu pelaksanaan mulai dalam periode seleksi
            if waktu_pelaksanaan_dimulai_date >= mulai_date and waktu_pelaksanaan_dimulai_date <= selesai_date:
                return jsonify({
                    'status': 'error', 
                    'message': f'Waktu Pelaksanaan (mulai) untuk kegiatan "{nama}" tidak boleh dalam kurun waktu Periode Seleksi'
                }), 400
            
            # Cek apakah waktu pelaksanaan selesai dalam periode seleksi
            if waktu_pelaksanaan_selesai_date >= mulai_date and waktu_pelaksanaan_selesai_date <= selesai_date:
                return jsonify({
                    'status': 'error', 
                    'message': f'Waktu Pelaksanaan (selesai) untuk kegiatan "{nama}" tidak boleh dalam kurun waktu Periode Seleksi'
                }), 400
            
            # Cek apakah waktu pelaksanaan overlap dengan periode seleksi (waktu pelaksanaan mencakup seluruh periode seleksi)
            if waktu_pelaksanaan_dimulai_date <= mulai_date and waktu_pelaksanaan_selesai_date >= selesai_date:
                return jsonify({
                    'status': 'error', 
                    'message': f'Waktu Pelaksanaan untuk kegiatan "{nama}" tidak boleh mencakup seluruh Periode Seleksi'
                }), 400
            
            # Parse jadwal tes (now as text)
            tanggal_tes = (act.get('tanggalTes') or '').strip()
            
            # Parse tempat tes
            tempat_tes = (act.get('tempatTes') or '').strip()
            
            event = Event(
                jenis_kegiatan=jenis_kegiatan,
                nama_kegiatan=nama,
                waktu_pelaksanaan_dimulai=waktu_pelaksanaan_dimulai_date,
                waktu_pelaksanaan_selesai=waktu_pelaksanaan_selesai_date,
                tempat_pelaksanaan=tempat,
                skala_kegiatan=skala_kegiatan,
                kwartir_penyelenggara=kwartir,
                mulai=mulai_date,
                selesai=selesai_date,
                tanggal_tes=tanggal_tes if tanggal_tes else None,
                tempat_tes=tempat_tes if tempat_tes else None
            )
            db.session.add(event)
            db.session.flush()
            
            # Ambil data kuota dari act (sudah disinkronkan dari contingents di frontend)
            putra = int(act.get('putra') or act.get('umpiPutra') or 0)
            putri = int(act.get('putri') or act.get('umpiPutri') or 0)
            
            kuota = Kuota(
                event_id=event.id_kegiatan,
                putra=putra,
                putri=putri
            )
            db.session.add(kuota)
            created_events.append(event)
            
            # Buat Criteria untuk event ini dari activities[index].criteria
            criteria_list = act.get('criteria', [])
            for c in criteria_list:
                nama_kriteria = (c.get('nama') or '').strip()
                if not nama_kriteria:
                    continue
                
                # Ambil skala (bobot) dari kriteria
                skala = c.get('skala') or c.get('bobot') or 0
                bobot = float(skala) if skala else 0.0
                
                # Ambil jenis kriteria
                jenis_kriteria_raw = c.get('jenis', 'Kualitatif')
                if isinstance(jenis_kriteria_raw, list):
                    # Jika jenis adalah array (untuk Tes Wawancara), gabungkan
                    aspek_str = ', '.join(jenis_kriteria_raw) if jenis_kriteria_raw else ''
                    jenis_kriteria = 'Kualitatif'  # Default untuk wawancara
                else:
                    # Jika jenis adalah string
                    jenis_kriteria = jenis_kriteria_raw if jenis_kriteria_raw else 'Kualitatif'
                    aspek_str = ''
                
                # Ambil jumlah soal jika ada
                jumlah_soal = c.get('jumlah_soal') or c.get('jumlahSoal') or None
                
                # Ambil deskripsi jika ada
                deskripsi = c.get('deskripsi', '')
                
                crit = Criteria(
                    event_id=event.id_kegiatan,
                    nama_kriteria=nama_kriteria,
                    aspek=aspek_str,
                    bobot=bobot,
                    deskripsi=deskripsi,
                    jenis_kriteria=jenis_kriteria,
                    jumlah_soal=int(jumlah_soal) if jumlah_soal else None
                )
                db.session.add(crit)
        
        # Fallback: jika criteria_list dikirim terpisah (backward compatibility)
        # Hanya diproses jika tidak ada criteria di dalam activities
        if criteria_list and not any(act.get('criteria') for act in activities):
            target_event_id = created_events[0].id_kegiatan if created_events else None
            for c in criteria_list:
                nama_kriteria = (c.get('nama') or '').strip()
                if not nama_kriteria:
                    continue
                    
                bobot = float(c.get('bobot') or c.get('skala') or 0)
                aspek = c.get('aspek', [])
                aspek_str = ', '.join(aspek) if isinstance(aspek, list) else (aspek or '')
                jumlah_soal = c.get('jumlah_soal') or c.get('jumlahSoal') or None
                deskripsi = c.get('deskripsi', '')
                jenis_kriteria = c.get('jenis_kriteria', 'Kualitatif')
                if isinstance(c.get('jenis'), list):
                    aspek_str = ', '.join(c.get('jenis', []))
                    jenis_kriteria = 'Kualitatif'
                elif c.get('jenis'):
                    jenis_kriteria = c.get('jenis')
                
                if target_event_id is None:
                    today = datetime.utcnow().date()
                    placeholder = Event(
                        jenis_kegiatan='Siaga',  # Default ENUM value
                        nama_kegiatan='(Default) Konfigurasi Seleksi',
                        waktu_pelaksanaan_dimulai=today,
                        waktu_pelaksanaan_selesai=today,
                        tempat_pelaksanaan='-',
                        skala_kegiatan='Ranting',  # Default ENUM value
                        kwartir_penyelenggara='-',
                        mulai=today,
                        selesai=today
                    )
                    db.session.add(placeholder)
                    db.session.flush()
                    target_event_id = placeholder.id_kegiatan
                crit = Criteria(
                    event_id=target_event_id,
                    nama_kriteria=nama_kriteria,
                    aspek=aspek_str,
                    bobot=bobot,
                    deskripsi=deskripsi,
                    jenis_kriteria=jenis_kriteria,
                    jumlah_soal=int(jumlah_soal) if jumlah_soal else None
                )
                db.session.add(crit)
        db.session.commit()
        
        # Buat notifikasi untuk admin setelah commit berhasil
        if created_events:
            for event in created_events:
                notification_message = f"Kegiatan baru dibuat: {event.nama_kegiatan} ({event.jenis_kegiatan})"
                logging.info(f"[NOTIFICATION] Attempting to create notification: {notification_message}")
                
                try:
                    create_notification_to_all_admins(notification_message)
                    logging.info(f"[NOTIFICATION] Successfully created notifications for event: {event.nama_kegiatan}")
                except Exception as e:
                    # Log error tapi jangan gagalkan proses utama
                    logging.error(f"[NOTIFICATION] Failed to create notification for new event '{event.nama_kegiatan}': {e}")
                    import traceback
                    logging.error(f"[NOTIFICATION] Traceback: {traceback.format_exc()}")
                    if hasattr(current_app, 'logger'):
                        current_app.logger.exception('Error creating notification for new event:')
        
        return jsonify({'status': 'success', 'message': 'Konfigurasi berhasil disimpan'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/save_config:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API Get Konfigurasi Seleksi
@app.route('/api/get_config/<int:event_id>')
@login_required
@admin_required
def get_config(event_id):
    try:
        event = Event.query.get_or_404(event_id)
        kuota = Kuota.query.filter_by(event_id=event_id).first()
        criteria_list = Criteria.query.filter_by(event_id=event_id).all()
        
        config_data = {
            'event': {
                'id': event.id_kegiatan,
                'nama_kegiatan': event.nama_kegiatan,
                'jenis_kegiatan': event.jenis_kegiatan,
                'skala_kegiatan': event.skala_kegiatan,
                'kwartir_penyelenggara': event.kwartir_penyelenggara,
                'tempat_pelaksanaan': event.tempat_pelaksanaan,
                'waktu_pelaksanaan_dimulai': event.waktu_pelaksanaan_dimulai.isoformat() if event.waktu_pelaksanaan_dimulai else None,
                'waktu_pelaksanaan_selesai': event.waktu_pelaksanaan_selesai.isoformat() if event.waktu_pelaksanaan_selesai else None,
                'mulai': event.mulai.isoformat() if event.mulai else None,
                'selesai': event.selesai.isoformat() if event.selesai else None,
                'tanggal_tes': event.tanggal_tes,
                'tempat_tes': event.tempat_tes
            },
            'kuota': {
                'putra': kuota.putra if kuota else 0,
                'putri': kuota.putri if kuota else 0
            },
            'criteria': [{
                'id': c.id_kriteria,
                'nama_kriteria': c.nama_kriteria,
                'bobot': c.bobot,
                'jenis_kriteria': c.jenis_kriteria,
                'aspek': c.aspek,
                'deskripsi': c.deskripsi,
                'jumlah_soal': c.jumlah_soal
            } for c in criteria_list]
        }
        return jsonify({'status': 'success', 'data': config_data}), 200
    except Exception as e:
        current_app.logger.exception('Error in /api/get_config:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Halaman View Konfigurasi Seleksi
@app.route('/admin/view_config')
@login_required
@admin_required
def view_config():
    sidebar_state = current_user.sidebar_state or 'expanded'
    events = Event.query.order_by(Event.id_kegiatan.desc()).all()
    return render_template("manajemen-seleksi/view_konfigurasi.html", 
                         events=events, 
                         sidebar_state=sidebar_state, 
                         user=current_user)

# API Update Konfigurasi Seleksi
@app.route('/api/update_config/<int:event_id>', methods=['PUT', 'POST'])
@login_required
@admin_required
@csrf.exempt
def update_config(event_id):
    try:
        event = Event.query.get_or_404(event_id)
        data = request.get_json(force=True)
        
        # Update Event
        if 'event' in data:
            evt_data = data['event']
            if 'nama_kegiatan' in evt_data:
                event.nama_kegiatan = evt_data['nama_kegiatan'].strip()
            if 'jenis_kegiatan' in evt_data:
                jenis_kegiatan_map = {
                    'siaga': 'Siaga', 'penggalang': 'Penggalang', 'penegak': 'Penegak',
                    'pandega': 'Pandega', 'penegak dan pandega': 'Penegak dan Pandega'
                }
                jenis_raw = evt_data['jenis_kegiatan'].strip().lower()
                event.jenis_kegiatan = jenis_kegiatan_map.get(jenis_raw, event.jenis_kegiatan)
            if 'skala_kegiatan' in evt_data:
                skala_kegiatan_map = {
                    'ranting': 'Ranting', 'cabang': 'Cabang', 'daerah': 'Daerah',
                    'nasional': 'Nasional', 'internasional': 'Internasional'
                }
                skala_raw = evt_data['skala_kegiatan'].strip().lower()
                event.skala_kegiatan = skala_kegiatan_map.get(skala_raw, event.skala_kegiatan)
            if 'kwartir_penyelenggara' in evt_data:
                event.kwartir_penyelenggara = evt_data['kwartir_penyelenggara'].strip()
            if 'tempat_pelaksanaan' in evt_data:
                event.tempat_pelaksanaan = evt_data['tempat_pelaksanaan'].strip()
            if 'waktu_pelaksanaan_dimulai' in evt_data and evt_data['waktu_pelaksanaan_dimulai']:
                try:
                    event.waktu_pelaksanaan_dimulai = datetime.strptime(evt_data['waktu_pelaksanaan_dimulai'], '%Y-%m-%d').date()
                except:
                    pass
            if 'waktu_pelaksanaan_selesai' in evt_data and evt_data['waktu_pelaksanaan_selesai']:
                try:
                    event.waktu_pelaksanaan_selesai = datetime.strptime(evt_data['waktu_pelaksanaan_selesai'], '%Y-%m-%d').date()
                except:
                    pass
            if 'mulai' in evt_data and evt_data['mulai']:
                try:
                    event.mulai = datetime.strptime(evt_data['mulai'], '%Y-%m-%d').date()
                except:
                    pass
            if 'selesai' in evt_data and evt_data['selesai']:
                try:
                    event.selesai = datetime.strptime(evt_data['selesai'], '%Y-%m-%d').date()
                except:
                    pass
            if 'tanggal_tes' in evt_data:
                event.tanggal_tes = evt_data['tanggal_tes'].strip() if evt_data['tanggal_tes'] else None
            if 'tempat_tes' in evt_data:
                event.tempat_tes = evt_data['tempat_tes'].strip() if evt_data['tempat_tes'] else None
        
        # Update Kuota
        if 'kuota' in data:
            kuota = Kuota.query.filter_by(event_id=event_id).first()
            if kuota:
                if 'putra' in data['kuota']:
                    kuota.putra = int(data['kuota']['putra'] or 0)
                if 'putri' in data['kuota']:
                    kuota.putri = int(data['kuota']['putri'] or 0)
            else:
                kuota = Kuota(
                    event_id=event_id,
                    putra=int(data['kuota'].get('putra', 0)),
                    putri=int(data['kuota'].get('putri', 0))
                )
                db.session.add(kuota)
        
        # Update Criteria
        if 'criteria' in data:
            # Get existing criteria map {id: object}
            existing_criteria = {c.id_kriteria: c for c in Criteria.query.filter_by(event_id=event_id).all()}
            
            # Process incoming criteria
            incoming_ids = []
            for c in data['criteria']:
                crit_id = c.get('id')
                
                if crit_id and crit_id in existing_criteria:
                    # Update existing
                    crit = existing_criteria[crit_id]
                    crit.nama_kriteria = c.get('nama_kriteria', '').strip() or 'Unnamed Criteria'
                    crit.bobot = float(c.get('bobot', 0))
                    crit.aspek = ', '.join(c.get('aspek', [])) if isinstance(c.get('aspek'), list) else (c.get('aspek', '') or '')
                    crit.deskripsi = c.get('deskripsi', '')
                    crit.jenis_kriteria = c.get('jenis_kriteria', 'Kualitatif')
                    crit.jumlah_soal = int(c.get('jumlah_soal')) if c.get('jumlah_soal') else None
                    incoming_ids.append(crit_id)
                else:
                    # Create new
                    new_crit = Criteria(
                        event_id=event_id,
                        nama_kriteria=c.get('nama_kriteria', '').strip() or 'Unnamed Criteria',
                        bobot=float(c.get('bobot', 0)),
                        aspek=', '.join(c.get('aspek', [])) if isinstance(c.get('aspek'), list) else (c.get('aspek', '') or ''),
                        deskripsi=c.get('deskripsi', ''),
                        jenis_kriteria=c.get('jenis_kriteria', 'Kualitatif'),
                        jumlah_soal=int(c.get('jumlah_soal')) if c.get('jumlah_soal') else None
                    )
                    db.session.add(new_crit)
            
            # Delete removed criteria (only if not referenced)
            for crit_id, crit in existing_criteria.items():
                if crit_id not in incoming_ids:
                    try:
                        db.session.delete(crit)
                        db.session.flush() # Check for integrity error immediately
                    except IntegrityError:
                        db.session.rollback()
                        # If referenced, just skip deletion or log warning
                        current_app.logger.warning(f"Cannot delete criteria {crit_id} because it is referenced.")
                        pass
        
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Konfigurasi berhasil diperbarui'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/update_config:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API Delete Konfigurasi Seleksi
@app.route('/api/delete_config/<int:event_id>', methods=['DELETE', 'POST'])
@login_required
@admin_required
@csrf.exempt
def delete_config(event_id):
    try:
        event = Event.query.get_or_404(event_id)
        event_name = event.nama_kegiatan
        
        # Ambil semua kriteria ID dari kegiatan
        criteria = Criteria.query.filter_by(event_id=event_id).all()
        criteria_ids = [c.id_kriteria for c in criteria]
        
        # Hapus penilaian yang terkait dengan kriteria tersebut
        if criteria_ids:
            Penilaian.query.filter(Penilaian.id_kriteria.in_(criteria_ids)).delete(synchronize_session=False)
            
            # Hapus himpunan kriteria yang terkait dengan kriteria tersebut
            HimpunanKriteria.query.filter(HimpunanKriteria.id_kriteria.in_(criteria_ids)).delete(synchronize_session=False)
        
        # Hapus hasil seleksi yang terkait dengan kegiatan
        HasilSeleksi.query.filter_by(event_id=event_id).delete(synchronize_session=False)
        
        # Hapus data dari tabel tb_participant_kegiatan (relasi peserta-kegiatan)
        db.session.execute(
            tb_participant_kegiatan.delete().where(tb_participant_kegiatan.c.kegiatan_id == event_id)
        )
        
        # Hapus perbandingan AHP yang terkait
        PairwiseComparison.query.filter_by(event_id=event_id).delete(synchronize_session=False)
        
        # Hapus hasil AHP yang terkait
        AHPResults.query.filter_by(event_id=event_id).delete(synchronize_session=False)
        
        # Hapus arsip seleksi yang terkait
        ArsipSeleksi.query.filter_by(event_id=event_id).delete(synchronize_session=False)
        
        # Update participants yang memiliki kegiatan_id ini menjadi NULL
        Participants.query.filter_by(kegiatan_id=event_id).update({'kegiatan_id': None}, synchronize_session=False)
        
        # Hapus penugasan penilai
        event.evaluators = []

        # Hapus akan cascade otomatis ke Kuota dan Criteria karena cascade="all, delete-orphan"
        db.session.delete(event)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': f'Konfigurasi "{event_name}" berhasil dihapus'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/delete_config:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API Delete Banyak Konfigurasi Seleksi
@app.route('/api/delete_config_bulk', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def delete_config_bulk():
    try:
        data = request.get_json(force=True)
        event_ids = data.get('event_ids', [])
        
        if not event_ids or not isinstance(event_ids, list):
            return jsonify({'status': 'error', 'message': 'Tidak ada ID konfigurasi yang dipilih'}), 400
        
        # Validasi bahwa semua event_id adalah integer
        try:
            event_ids = [int(eid) for eid in event_ids]
        except (ValueError, TypeError):
            return jsonify({'status': 'error', 'message': 'Format ID konfigurasi tidak valid'}), 400
        
        if len(event_ids) == 0:
            return jsonify({'status': 'error', 'message': 'Tidak ada konfigurasi yang dipilih'}), 400
        
        # Query semua event yang akan dihapus
        events = Event.query.filter(Event.id_kegiatan.in_(event_ids)).all()
        
        if not events:
            return jsonify({'status': 'error', 'message': 'Konfigurasi yang dipilih tidak ditemukan'}), 404
        
        # Simpan nama-nama event untuk pesan sukses
        event_names = [event.nama_kegiatan for event in events]
        deleted_count = len(events)
        
        # Ambil semua kriteria ID dari kegiatan yang akan dihapus
        criteria_ids = []
        for event in events:
            criteria = Criteria.query.filter_by(event_id=event.id_kegiatan).all()
            criteria_ids.extend([c.id_kriteria for c in criteria])
        
        # Hapus penilaian yang terkait dengan kriteria tersebut
        if criteria_ids:
            Penilaian.query.filter(Penilaian.id_kriteria.in_(criteria_ids)).delete(synchronize_session=False)
            
            # Hapus himpunan kriteria yang terkait dengan kriteria tersebut
            HimpunanKriteria.query.filter(HimpunanKriteria.id_kriteria.in_(criteria_ids)).delete(synchronize_session=False)
        
        # Hapus hasil seleksi yang terkait dengan kegiatan
        HasilSeleksi.query.filter(HasilSeleksi.event_id.in_(event_ids)).delete(synchronize_session=False)
        
        # Hapus data dari tabel tb_participant_kegiatan (relasi peserta-kegiatan)
        for eid in event_ids:
            db.session.execute(
                tb_participant_kegiatan.delete().where(tb_participant_kegiatan.c.kegiatan_id == eid)
            )
        
        # Hapus perbandingan AHP yang terkait
        PairwiseComparison.query.filter(PairwiseComparison.event_id.in_(event_ids)).delete(synchronize_session=False)
        
        # Hapus hasil AHP yang terkait
        AHPResults.query.filter(AHPResults.event_id.in_(event_ids)).delete(synchronize_session=False)
        
        # Hapus arsip seleksi yang terkait
        ArsipSeleksi.query.filter(ArsipSeleksi.event_id.in_(event_ids)).delete(synchronize_session=False)
        
        # Update participants yang memiliki kegiatan_id ini menjadi NULL
        Participants.query.filter(Participants.kegiatan_id.in_(event_ids)).update({'kegiatan_id': None}, synchronize_session=False)
        
        # Hapus semua event (cascade akan menghapus Kuota dan Criteria secara otomatis)
        for event in events:
            # Hapus penugasan penilai
            event.evaluators = []
            db.session.delete(event)
        
        db.session.commit()
        
        message = f'{deleted_count} konfigurasi berhasil dihapus'
        if deleted_count == 1:
            message = f'Konfigurasi "{event_names[0]}" berhasil dihapus'
        elif deleted_count <= 3:
            message = f'{deleted_count} konfigurasi berhasil dihapus: {", ".join(event_names)}'
        
        return jsonify({'status': 'success', 'message': message}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/delete_config_bulk:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API Kegiatan 
@app.route('/api/kegiatan')
@login_required
@admin_required
def api_kegiatan():
    kegiatan = Event.query.all()
    result = [
        {
            "id": k.id_kegiatan,
            "nama_kegiatan": k.nama_kegiatan,
            "jenis_kegiatan": k.jenis_kegiatan,
            "waktu_pelaksanaan_dimulai": k.waktu_pelaksanaan_dimulai.strftime("%Y-%m-%d") if k.waktu_pelaksanaan_dimulai else None,
            "waktu_pelaksanaan_selesai": k.waktu_pelaksanaan_selesai.strftime("%Y-%m-%d") if k.waktu_pelaksanaan_selesai else None,
            "tempat_pelaksanaan": k.tempat_pelaksanaan,
            "skala_kegiatan": k.skala_kegiatan,
            "kwartir_penyelenggara": k.kwartir_penyelenggara,
            "mulai": k.mulai.strftime("%Y-%m-%d"),
            "selesai": k.selesai.strftime("%Y-%m-%d"),
        }
        for k in kegiatan
    ]
    return jsonify(result)

# API Kuota Kegiatan
@app.route('/api/kuota/<int:event_id>')
@login_required
@admin_required
def api_kuota(event_id):
    event = Event.query.get_or_404(event_id)
    if not event.kuota:
        return jsonify({"putra": 0, "putri": 0})

    # asumsi tabel Kuota punya kolom putra & putri
    kuota = event.kuota[0]
    return jsonify({
        "putra": kuota.putra,
        "putri": kuota.putri
    })

# API Data Peserta
@app.route("/api/peserta/<int:kegiatan_id>")
def get_peserta(kegiatan_id):
    peserta = Participants.query.filter_by(kegiatan_id=kegiatan_id).all()
    data = []
    for p in peserta:
        data.append({
            "nama_lengkap": p.nama_lengkap,
            "tanggal_lahir": str(p.tanggal_lahir),
            "jenis_kelamin": p.jenis_kelamin,
            "usia": p.usia,
            "alamat_tinggal": p.alamat_tinggal,
            "golongan": p.golongan,
            "tingkatan": p.tingkatan,
            "asal_gudep": p.asal_gudep,
            "asal_kwarran": p.asal_kwarran,
            "asal_kwarcab": p.asal_kwarcab,
            "asal_kwarda": p.asal_kwarda,
            "nomor_hp": p.nomor_hp,
            "email": p.email
        })
    return jsonify(data)

# API Search Peserta untuk Kelola Profil
@app.route("/api/peserta/search")
@login_required
@admin_required
def api_search_peserta():
    """API untuk mencari peserta berdasarkan email atau nama"""
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'success': False, 'message': 'Query tidak boleh kosong'}), 400
        
        # Cari peserta berdasarkan email atau nama
        peserta = Participants.query.filter(
            (Participants.email.ilike(f'%{query}%')) |
            (Participants.nama_lengkap.ilike(f'%{query}%'))
        ).first()
        
        if not peserta:
            return jsonify({'success': False, 'message': 'Peserta tidak ditemukan'}), 404
        
        # Format data peserta
        data = {
            'success': True,
            'peserta': {
                'id': peserta.id,
                'nama_lengkap': peserta.nama_lengkap or '',
                'tanggal_lahir': peserta.tanggal_lahir.strftime('%Y-%m-%d') if peserta.tanggal_lahir else '',
                'jenis_kelamin': peserta.jenis_kelamin or '',
                'usia': peserta.usia or '',
                'alamat_tinggal': peserta.alamat_tinggal or '',
                'golongan': peserta.golongan or '',
                'tingkatan': peserta.tingkatan or '',
                'asal_gudep': peserta.asal_gudep or '',
                'asal_kwarran': peserta.asal_kwarran or '',
                'asal_kwarcab': peserta.asal_kwarcab or '',
                'asal_kwarda': peserta.asal_kwarda or '',
                'nomor_hp': peserta.nomor_hp or '',
                'email': peserta.email or '',
                'foto': peserta.foto or 'img/default-user.png'
            }
        }
        
        return jsonify(data)
    except Exception as e:
        logging.error(f"Error in api_search_peserta: {e}")
        current_app.logger.exception('Error in api_search_peserta:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API List Peserta (gabungan users + participants)
@app.route("/api/peserta/list")
@login_required
@admin_required
def api_list_peserta():
    """API untuk mendapatkan semua data peserta (gabungan users dan participants)"""
    try:
        # Get optional filter parameters
        kegiatan_id = request.args.get('kegiatan_id', type=int)
        
        # Ambil semua user dengan level peserta
        users_peserta = Users.query.filter_by(level='peserta').all()
        
        peserta_data = []
        for user in users_peserta:
            # Cari data biodata dari tabel participants berdasarkan email
            biodata = Participants.query.filter_by(email=user.email).first()
            
            # Jika filter kegiatan_id diberikan, cek apakah peserta terdaftar di kegiatan tersebut
            if kegiatan_id and biodata:
                # Cek apakah peserta terdaftar di kegiatan ini
                is_registered = db.session.query(tb_participant_kegiatan).filter_by(
                    participant_id=biodata.id,
                    kegiatan_id=kegiatan_id
                ).first()
                
                if not is_registered:
                    continue  # Skip peserta yang tidak terdaftar di kegiatan ini
            
            # Get hasil seleksi untuk kegiatan tertentu atau latest
            if kegiatan_id:
                hasil_seleksi = HasilSeleksi.query.filter_by(
                    id_users=user.id,
                    event_id=kegiatan_id
                ).first()
            else:
                # Get latest or highest score
                hasil_seleksi = HasilSeleksi.query.filter_by(id_users=user.id).order_by(
                    HasilSeleksi.skor_akhir.desc()
                ).first()
            
            # Get registered activities for this participant
            registered_activities = []
            if biodata:
                activities = biodata.registered_activities.all()
                for activity in activities:
                    hasil_activity = HasilSeleksi.query.filter_by(
                        id_users=user.id,
                        event_id=activity.id_kegiatan
                    ).first()
                    
                    registered_activities.append({
                        'id': activity.id_kegiatan,
                        'nama': activity.nama_kegiatan,
                        'jenis': activity.jenis_kegiatan,
                        'skor': hasil_activity.skor_akhir if hasil_activity else None,
                        'ranking': hasil_activity.ranking if hasil_activity else None
                    })
            
            # Gabungkan data dari users dan participants
            peserta_item = {
                'id': user.id,
                'user_id': user.id,
                'participant_id': biodata.id if biodata else None,
                'username': user.username or '',
                'nama_lengkap': biodata.nama_lengkap if biodata and biodata.nama_lengkap else (user.nama_lengkap or ''),
                'email': user.email or '',
                'jenis_kelamin': biodata.jenis_kelamin if biodata and biodata.jenis_kelamin else (user.jenis_kelamin or ''),
                'usia': str(biodata.usia) if biodata and biodata.usia else (user.usia or '0'),
                'nomor_hp': biodata.nomor_hp if biodata and biodata.nomor_hp else (user.nomor_hp or ''),
                'foto': user.foto if user.foto and user.foto != 'img/default-user.png' else (biodata.foto if biodata and biodata.foto else 'img/default-user.png'),
                'status': user.status or 'aktif',
                'golongan': biodata.golongan if biodata else '',
                'tingkatan': biodata.tingkatan if biodata else '',
                'tanggal_lahir': biodata.tanggal_lahir.strftime('%Y-%m-%d') if biodata and biodata.tanggal_lahir else '',
                'alamat_tinggal': biodata.alamat_tinggal if biodata else '',
                'asal_gudep': biodata.asal_gudep if biodata else '',
                'asal_kwarran': biodata.asal_kwarran if biodata else '',
                'asal_kwarcab': biodata.asal_kwarcab if biodata else '',
                'asal_kwarda': biodata.asal_kwarda if biodata else '',
                'skor_akhir': hasil_seleksi.skor_akhir if hasil_seleksi else None,
                'ranking': hasil_seleksi.ranking if hasil_seleksi else None,
                'registered_activities': registered_activities
            }
            peserta_data.append(peserta_item)
        
        # Always return success with array, even if empty
        return jsonify({
            'success': True, 
            'peserta': peserta_data,
            'count': len(peserta_data)
        })
    except Exception as e:
        logging.error(f"Error in api_list_peserta: {e}")
        current_app.logger.exception('Error in api_list_peserta:')
        # Return error response in JSON format
        return jsonify({
            'success': False, 
            'message': str(e),
            'peserta': [],
            'count': 0
        }), 500

# API Add Peserta
@app.route("/api/peserta/add", methods=['POST'])
@login_required
@admin_required
def api_add_peserta():
    """API untuk menambah data peserta (users + participants)"""
    try:
        # Ambil data dari form
        nama_lengkap = request.form.get('nama_lengkap', '').strip()
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        jenis_kelamin = request.form.get('jenis_kelamin', '').strip()
        usia = request.form.get('usia', '0').strip()
        nomor_hp = request.form.get('nomor_hp', '').strip()
        status = request.form.get('status', 'aktif').strip()
        
        # Validasi required fields
        if not nama_lengkap or not username or not email:
            return jsonify({'success': False, 'message': 'Nama lengkap, username, dan email wajib diisi'}), 400
        
        # Cek apakah username sudah ada
        if Users.query.filter_by(username=username).first():
            return jsonify({'success': False, 'message': 'Username sudah digunakan'}), 400
        
        # Cek apakah email sudah ada
        if Users.query.filter_by(email=email).first():
            return jsonify({'success': False, 'message': 'Email sudah digunakan'}), 400
        
        # Hash password jika ada
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16) if password else ''
        
        # Buat user baru
        new_user = Users(
            username=username,
            password=hashed_password,
            nama_lengkap=nama_lengkap,
            email=email,
            jenis_kelamin=jenis_kelamin or 'laki-laki',
            usia=usia or '0',
            nomor_hp=nomor_hp,
            level='peserta',
            status=status,
            foto='img/default-user.png',
            login_method='manual'
        )
        db.session.add(new_user)
        db.session.flush()  # Untuk mendapatkan ID user
        
        # Buat data participants jika ada data tambahan
        golongan = request.form.get('golongan', '').strip()
        tingkatan = request.form.get('tingkatan', '').strip()
        tanggal_lahir = request.form.get('tanggal_lahir', '').strip()
        alamat_tinggal = request.form.get('alamat_tinggal', '').strip()
        asal_gudep = request.form.get('asal_gudep', '').strip()
        asal_kwarran = request.form.get('asal_kwarran', '').strip()
        asal_kwarcab = request.form.get('asal_kwarcab', '').strip()
        asal_kwarda = request.form.get('asal_kwarda', '').strip()
        
        if golongan or tingkatan or tanggal_lahir or alamat_tinggal:
            # Konversi usia ke integer jika ada
            usia_int = int(usia) if usia and usia.isdigit() else 0
            new_participant = Participants(
                nama_lengkap=nama_lengkap,
                email=email,
                jenis_kelamin=jenis_kelamin or 'laki-laki',
                usia=usia_int,
                nomor_hp=nomor_hp,
                tanggal_lahir=datetime.strptime(tanggal_lahir, '%Y-%m-%d').date() if tanggal_lahir else datetime.now().date(),
                alamat_tinggal=alamat_tinggal or '',
                golongan=golongan or 'siaga',
                tingkatan=tingkatan or 'siaga mula',
                asal_gudep=asal_gudep or '',
                asal_kwarran=asal_kwarran or '',
                asal_kwarcab=asal_kwarcab or '',
                asal_kwarda=asal_kwarda or '',
                foto='img/default-user.png',
                level='peserta'
            )
            db.session.add(new_participant)
        db.session.commit()
        log_activity(current_user.id, f'Menambah peserta baru: {username}')
        return jsonify({'success': True, 'message': 'Peserta berhasil ditambahkan'})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Data tidak valid: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_add_peserta: {e}")
        current_app.logger.exception('Error in api_add_peserta:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Edit Peserta
@app.route("/api/peserta/edit/<int:user_id>", methods=['POST'])
@login_required
@admin_required
def api_edit_peserta(user_id):
    """API untuk mengedit data peserta (users + participants)"""
    try:
        user = Users.query.get(user_id)
        if not user or user.level != 'peserta':
            return jsonify({'success': False, 'message': 'Peserta tidak ditemukan'}), 404
        
        # Update data user
        nama_lengkap = request.form.get('nama_lengkap', '').strip()
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip().lower()
        jenis_kelamin = request.form.get('jenis_kelamin', '').strip()
        usia = request.form.get('usia', '0').strip()
        nomor_hp = request.form.get('nomor_hp', '').strip()
        status = request.form.get('status', 'aktif').strip()
        
        if nama_lengkap:
            user.nama_lengkap = nama_lengkap
        if username and username != user.username:
            # Cek apakah username sudah digunakan
            if Users.query.filter(Users.username == username, Users.id != user_id).first():
                return jsonify({'success': False, 'message': 'Username sudah digunakan'}), 400
            user.username = username
        if email and email != user.email:
            # Cek apakah email sudah digunakan
            if Users.query.filter(Users.email == email, Users.id != user_id).first():
                return jsonify({'success': False, 'message': 'Email sudah digunakan'}), 400
            user.email = email
        if jenis_kelamin:
            user.jenis_kelamin = jenis_kelamin
        if usia:
            user.usia = usia
        if nomor_hp:
            user.nomor_hp = nomor_hp
        if status:
            user.status = status
        
        # Update atau buat data participants
        biodata = Participants.query.filter_by(email=user.email).first()
        
        golongan = request.form.get('golongan', '').strip()
        tingkatan = request.form.get('tingkatan', '').strip()
        tanggal_lahir = request.form.get('tanggal_lahir', '').strip()
        alamat_tinggal = request.form.get('alamat_tinggal', '').strip()
        asal_gudep = request.form.get('asal_gudep', '').strip()
        asal_kwarran = request.form.get('asal_kwarran', '').strip()
        asal_kwarcab = request.form.get('asal_kwarcab', '').strip()
        asal_kwarda = request.form.get('asal_kwarda', '').strip()
        
        if biodata:
            # Update existing biodata
            if nama_lengkap:
                biodata.nama_lengkap = nama_lengkap
            if jenis_kelamin:
                biodata.jenis_kelamin = jenis_kelamin
            if usia:
                biodata.usia = int(usia) if usia.isdigit() else 0
            if nomor_hp:
                biodata.nomor_hp = nomor_hp
            if email:
                biodata.email = email
            if tanggal_lahir:
                biodata.tanggal_lahir = datetime.strptime(tanggal_lahir, '%Y-%m-%d').date()
            if alamat_tinggal:
                biodata.alamat_tinggal = alamat_tinggal
            if golongan:
                biodata.golongan = golongan
            if tingkatan:
                biodata.tingkatan = tingkatan
            if asal_gudep:
                biodata.asal_gudep = asal_gudep
            if asal_kwarran:
                biodata.asal_kwarran = asal_kwarran
            if asal_kwarcab:
                biodata.asal_kwarcab = asal_kwarcab
            if asal_kwarda:
                biodata.asal_kwarda = asal_kwarda
        elif golongan or tingkatan or tanggal_lahir or alamat_tinggal:
            # Buat biodata baru jika ada data
            usia_int = int(usia) if usia and usia.isdigit() else 0
            new_biodata = Participants(
                nama_lengkap=nama_lengkap or user.nama_lengkap,
                email=email or user.email,
                jenis_kelamin=jenis_kelamin or user.jenis_kelamin,
                usia=usia_int,
                nomor_hp=nomor_hp or user.nomor_hp,
                tanggal_lahir=datetime.strptime(tanggal_lahir, '%Y-%m-%d').date() if tanggal_lahir else datetime.now().date(),
                alamat_tinggal=alamat_tinggal or '',
                golongan=golongan or 'siaga',
                tingkatan=tingkatan or 'siaga mula',
                asal_gudep=asal_gudep or '',
                asal_kwarran=asal_kwarran or '',
                asal_kwarcab=asal_kwarcab or '',
                asal_kwarda=asal_kwarda or '',
                foto=user.foto or 'img/default-user.png',
                level='peserta'
            )
            db.session.add(new_biodata)
        db.session.commit()
        log_activity(current_user.id, f'Mengupdate data peserta: {user.username}')
        return jsonify({'success': True, 'message': 'Data peserta berhasil diperbarui'})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Data tidak valid: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_edit_peserta: {e}")
        current_app.logger.exception('Error in api_edit_peserta:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Delete Peserta
@app.route("/api/peserta/delete/<int:user_id>", methods=['POST'])
@login_required
@admin_required
def api_delete_peserta(user_id):
    """API untuk menghapus data peserta (users + participants)"""
    try:
        user = Users.query.get(user_id)
        if not user or user.level != 'peserta':
            return jsonify({'success': False, 'message': 'Peserta tidak ditemukan'}), 404
        
        username = user.username
        email = user.email
        
        # Hapus data participants jika ada
        biodata = Participants.query.filter_by(email=email).first()
        if biodata:
            db.session.delete(biodata)
        
        # Hapus user
        db.session.delete(user)
        db.session.commit()
        
        log_activity(current_user.id, f'Menghapus peserta: {username}')
        return jsonify({'success': True, 'message': 'Peserta berhasil dihapus'})
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_delete_peserta: {e}")
        current_app.logger.exception('Error in api_delete_peserta:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Get Detail Peserta
@app.route("/api/peserta/detail/<int:user_id>")
@login_required
@admin_required
def api_detail_peserta(user_id):
    """API untuk mendapatkan detail data peserta"""
    try:
        user = Users.query.get(user_id)
        if not user or user.level != 'peserta':
            return jsonify({'success': False, 'message': 'Peserta tidak ditemukan'}), 404
        
        # Cari data biodata dari tabel participants berdasarkan email
        biodata = Participants.query.filter_by(email=user.email).first()
        
        # Get registered activities
        registered_activities = []
        if biodata:
            activities = biodata.registered_activities.all()
            for activity in activities:
                # Get hasil seleksi if exists
                hasil = HasilSeleksi.query.filter_by(
                    id_users=user.id,
                    event_id=activity.id_kegiatan
                ).first()
                
                registered_activities.append({
                    'id': activity.id_kegiatan,
                    'nama': activity.nama_kegiatan,
                    'jenis': activity.jenis_kegiatan,
                    'skor': hasil.skor_akhir if hasil else None,
                    'ranking': hasil.ranking if hasil else None
                })
        
        # Gabungkan data dari users dan participants
        peserta_data = {
            'id': user.id,
            'user_id': user.id,
            'participant_id': biodata.id if biodata else None,
            'username': user.username or '',
            'nama_lengkap': biodata.nama_lengkap if biodata and biodata.nama_lengkap else (user.nama_lengkap or ''),
            'email': user.email or '',
            'jenis_kelamin': biodata.jenis_kelamin if biodata and biodata.jenis_kelamin else (user.jenis_kelamin or ''),
            'usia': str(biodata.usia) if biodata and biodata.usia else (user.usia or '0'),
            'nomor_hp': biodata.nomor_hp if biodata and biodata.nomor_hp else (user.nomor_hp or ''),
            'foto': user.foto if user.foto and user.foto != 'img/default-user.png' else (biodata.foto if biodata and biodata.foto else 'img/default-user.png'),
            'status': user.status or 'aktif',
            'golongan': biodata.golongan if biodata else '',
            'tingkatan': biodata.tingkatan if biodata else '',
            'tanggal_lahir': biodata.tanggal_lahir.strftime('%Y-%m-%d') if biodata and biodata.tanggal_lahir else '',
            'alamat_tinggal': biodata.alamat_tinggal if biodata else '',
            'asal_gudep': biodata.asal_gudep if biodata else '',
            'asal_kwarran': biodata.asal_kwarran if biodata else '',
            'asal_kwarcab': biodata.asal_kwarcab if biodata else '',
            'asal_kwarda': biodata.asal_kwarda if biodata else '',
            'registered_activities': registered_activities
        }
        
        return jsonify({'success': True, 'peserta': peserta_data})
    except Exception as e:
        logging.error(f"Error in api_detail_peserta: {e}")
        current_app.logger.exception('Error in api_detail_peserta:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Get Statistics Peserta
@app.route("/api/peserta/statistics")
@login_required
@admin_required
def api_peserta_statistics():
    """API untuk mendapatkan statistik peserta untuk chart"""
    try:
        # Count by status
        total = Users.query.filter_by(level='peserta').count()
        aktif = Users.query.filter_by(level='peserta', status='aktif').count()
        nonaktif = Users.query.filter_by(level='peserta', status='non-aktif').count()
        
        # Count by gender (from Users table first, then Participants)
        laki_laki = Users.query.filter_by(level='peserta', jenis_kelamin='laki-laki').count()
        perempuan = Users.query.filter_by(level='peserta', jenis_kelamin='perempuan').count()
        
        # Also count from Participants table
        peserta_laki = Participants.query.filter_by(jenis_kelamin='laki-laki').join(
            Users, Participants.email == Users.email
        ).filter(Users.level == 'peserta').count()
        
        peserta_perempuan = Participants.query.filter_by(jenis_kelamin='perempuan').join(
            Users, Participants.email == Users.email
        ).filter(Users.level == 'peserta').count()
        
        # Combine counts
        laki_laki = max(laki_laki, peserta_laki)
        perempuan = max(perempuan, peserta_perempuan)
        
        # Count by golongan
        golongan_stats = db.session.query(
            Participants.golongan,
            db.func.count(Participants.id).label('count')
        ).join(
            Users, Participants.email == Users.email
        ).filter(
            Users.level == 'peserta'
        ).group_by(Participants.golongan).all()
        
        golongan_data = {golongan: count for golongan, count in golongan_stats if golongan}
        
        # Get average scores
        avg_score = db.session.query(
            db.func.avg(HasilSeleksi.skor_akhir)
        ).join(
            Users, HasilSeleksi.id_users == Users.id
        ).filter(
            Users.level == 'peserta'
        ).scalar() or 0
        
        return jsonify({
            'success': True,
            'statistics': {
                'status': {
                    'total': total,
                    'aktif': aktif,
                    'nonaktif': nonaktif
                },
                'gender': {
                    'laki_laki': laki_laki,
                    'perempuan': perempuan
                },
                'golongan': golongan_data,
                'average_score': float(avg_score)
            }
        })
    except Exception as e:
        logging.error(f"Error in api_peserta_statistics: {e}")
        current_app.logger.exception('Error in api_peserta_statistics:')
        return jsonify({'success': False, 'message': str(e)}), 500

# Tambah Kegiatan
@app.route('/admin/tambah_seleksi', methods=['GET', 'POST'])
@login_required
@admin_required
def tambah_seleksi():
    if request.method == 'POST':
        nama = request.form['nama_kegiatan']
        jenis = request.form['jenis_kegiatan']
        waktu_dimulai = request.form.get('waktu_pelaksanaan_dimulai', request.form.get('waktu_pelaksanaan', ''))
        waktu_selesai = request.form.get('waktu_pelaksanaan_selesai', waktu_dimulai)
        tempat = request.form['tempat_pelaksanaan']
        skala = request.form['skala_kegiatan']
        kwartir = request.form['kwartir_penyelenggara']
        
        # New Fields
        tanggal_tes = request.form.get('tanggal_tes')
        tempat_tes = request.form.get('tempat_tes')
        evaluator_ids = request.form.getlist('evaluators')

        new_event = Event(
            nama_kegiatan=nama,
            jenis_kegiatan=jenis,
            waktu_pelaksanaan_dimulai=datetime.strptime(waktu_dimulai, '%Y-%m-%d').date() if waktu_dimulai else datetime.utcnow().date(),
            waktu_pelaksanaan_selesai=datetime.strptime(waktu_selesai, '%Y-%m-%d').date() if waktu_selesai else datetime.utcnow().date(),
            tempat_pelaksanaan=tempat,
            skala_kegiatan=skala,
            kwartir_penyelenggara=kwartir,
            mulai=datetime.utcnow().date(),
            selesai=datetime.utcnow().date(),
            tanggal_tes=tanggal_tes if tanggal_tes else None,
            tempat_tes=tempat_tes
        )
        
        # Assign Evaluators
        if evaluator_ids:
            evaluators = Users.query.filter(Users.id.in_(evaluator_ids)).all()
            new_event.evaluators = evaluators
            
        db.session.add(new_event)
        db.session.commit()
        
        # Buat notifikasi untuk admin setelah commit berhasil
        notification_message = f"Kegiatan baru dibuat: {nama} ({jenis})"
        logging.info(f"[NOTIFICATION] Attempting to create notification: {notification_message}")
        
        try:
            create_notification_to_all_admins(notification_message)
            logging.info(f"[NOTIFICATION] Successfully created notifications for event: {nama}")
        except Exception as e:
            # Log error tapi jangan gagalkan proses utama
            logging.error(f"[NOTIFICATION] Failed to create notification for new event '{nama}': {e}")
            import traceback
            logging.error(f"[NOTIFICATION] Traceback: {traceback.format_exc()}")
            if hasattr(current_app, 'logger'):
                current_app.logger.exception('Error creating notification for new event:')

        flash('Kegiatan berhasil ditambahkan!', 'success')
        return redirect(url_for('admin_manajemen_seleksi'))
    
    # Get all evaluators
    evaluators = Users.query.filter_by(level='penilai').all()
    return render_template("tambah_kegiatan.html", evaluators=evaluators)

# Edit Kegiatan
@app.route('/admin/edit_kegiatan/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_kegiatan(id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    event = Event.query.get_or_404(id)
    if request.method == 'POST':
        event.nama_kegiatan = request.form['nama_kegiatan']
        event.jenis_kegiatan = request.form['jenis_kegiatan']
        if 'waktu_pelaksanaan_dimulai' in request.form:
            event.waktu_pelaksanaan_dimulai = datetime.strptime(request.form['waktu_pelaksanaan_dimulai'], '%Y-%m-%d').date()
        if 'waktu_pelaksanaan_selesai' in request.form:
            event.waktu_pelaksanaan_selesai = datetime.strptime(request.form['waktu_pelaksanaan_selesai'], '%Y-%m-%d').date()
        elif 'waktu_pelaksanaan' in request.form:
            # Fallback untuk backward compatibility
            waktu = datetime.strptime(request.form['waktu_pelaksanaan'], '%Y-%m-%d').date()
            event.waktu_pelaksanaan_dimulai = waktu
            event.waktu_pelaksanaan_selesai = waktu
        event.tempat_pelaksanaan = request.form['tempat_pelaksanaan']
        
        # Update Test Details
        tanggal_tes = request.form.get('tanggal_tes')
        if tanggal_tes:
            event.tanggal_tes = tanggal_tes
        event.tempat_tes = request.form.get('tempat_tes')
        
        # Update Evaluators
        evaluator_ids = request.form.getlist('evaluators')
        if evaluator_ids:
            evaluators = Users.query.filter(Users.id.in_(evaluator_ids)).all()
            event.evaluators = evaluators
        else:
            event.evaluators = [] 
        db.session.commit()
        flash(f"{t['event_updated']}", "success")
        return redirect(url_for('admin_manajemen_seleksi'))
    
    # Get all evaluators
    evaluators = Users.query.filter_by(level='penilai').all()
    return render_template("edit_kegiatan.html", event=event, evaluators=evaluators)

# Hapus Kegiatan
@app.route('/admin/hapus_kegiatan/<int:id>', methods=['GET'])
@login_required
@admin_required
def hapus_kegiatan(id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    event = Event.query.get_or_404(id)
    
    # Ambil semua kriteria ID dari kegiatan
    criteria = Criteria.query.filter_by(event_id=id).all()
    criteria_ids = [c.id_kriteria for c in criteria]
    
    # Hapus penilaian yang terkait dengan kriteria tersebut
    if criteria_ids:
        Penilaian.query.filter(Penilaian.id_kriteria.in_(criteria_ids)).delete(synchronize_session=False)
    
    # Hapus hasil seleksi yang terkait dengan kegiatan
    HasilSeleksi.query.filter_by(event_id=id).delete(synchronize_session=False)
    
    db.session.delete(event)
    db.session.commit()
    flash(f"{t['event_deleted']}", "danger")
    return redirect(url_for('admin_manajemen_seleksi'))

@app.route('/admin/detail_kegiatan/<int:id>')
@login_required
@admin_required
def detail_kegiatan(id):
    event = Event.query.get_or_404(id)
    return render_template("detail_kegiatan.html", event=event)

@app.route('/admin/kriteria')
@login_required
@admin_required
def admin_kriteria():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    return render_template('data_kriteria.html', sidebar_state=sidebar_state, user=users, time=time)
    
@app.route('/admin/pembobotan_kriteria')
@login_required
@admin_required
def admin_pembobotan_kriteria():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    
    # Ambil semua kegiatan untuk dropdown
    events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Ambil event_id dari query parameter jika ada
    selected_event_id = request.args.get('event_id', type=int)
    selected_event = None
    criteria_list = []
    pairwise_matrix = None
    ahp_results = None
    
    if selected_event_id:
        selected_event = Event.query.get(selected_event_id)
        if selected_event:
            criteria_list = Criteria.query.filter_by(event_id=selected_event_id).order_by(Criteria.id_kriteria).all()
            
            # Cek apakah sudah ada matriks perbandingan
            from app.fuzzy_ahp import get_pairwise_matrix_from_db
            import numpy as np
            criteria_ids = [c.id_kriteria for c in criteria_list]
            if criteria_ids:
                pairwise_matrix = get_pairwise_matrix_from_db(selected_event_id, criteria_ids)
            
            # Ambil hasil AHP jika ada
            ahp_results = AHPResults.query.filter_by(event_id=selected_event_id).first()
    
    return render_template(
        'pembobotan_kriteria.html', 
        sidebar_state=sidebar_state, 
        user=users, 
        time=time,
        events=events,
        selected_event=selected_event,
        criteria_list=criteria_list,
        pairwise_matrix=pairwise_matrix.tolist() if pairwise_matrix is not None else None,
        ahp_results=ahp_results
    )


# API untuk menyimpan matriks perbandingan berpasangan
@app.route('/api/save_pairwise_matrix/<int:event_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def save_pairwise_matrix(event_id):
    """API untuk menyimpan matriks perbandingan berpasangan"""
    try:
        data = request.get_json(force=True)
        matrix_data = data.get('matrix', [])
        
        if not matrix_data:
            return jsonify({'success': False, 'message': 'Matriks tidak boleh kosong'}), 400
        
        # Ambil kriteria
        criterias = Criteria.query.filter_by(event_id=event_id).order_by(Criteria.id_kriteria).all()
        if not criterias:
            return jsonify({'success': False, 'message': 'Tidak ada kriteria untuk kegiatan ini'}), 400
        
        criteria_ids = [c.id_kriteria for c in criterias]
        n = len(criterias)
        
        # Validasi ukuran matriks
        if len(matrix_data) != n or any(len(row) != n for row in matrix_data):
            return jsonify({'success': False, 'message': f'Ukuran matriks harus {n}x{n}'}), 400
        
        # Konversi ke numpy array
        import numpy as np
        matrix = np.array(matrix_data, dtype=float)
        
        # Validasi nilai (harus 1-9 atau kebalikannya)
        for i in range(n):
            for j in range(n):
                if i == j:
                    if matrix[i, j] != 1.0:
                        return jsonify({'success': False, 'message': f'Diagonal harus 1.0 (baris {i+1}, kolom {j+1})'}), 400
                else:
                    val = matrix[i, j]
                    if val < 1/9 or val > 9:
                        return jsonify({'success': False, 'message': f'Nilai harus antara 1/9 sampai 9 (baris {i+1}, kolom {j+1})'}), 400
        
        # Simpan matriks
        from app.fuzzy_ahp import save_pairwise_matrix
        success, message = save_pairwise_matrix(event_id, criteria_ids, matrix)
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message}), 400
            
    except Exception as e:
        logging.error(f"Error saving pairwise matrix: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


# API untuk menghitung bobot AHP
@app.route('/api/calculate_ahp/<int:event_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def calculate_ahp(event_id):
    """API untuk menghitung bobot menggunakan AHP"""
    try:
        data = request.get_json(force=True)
        use_fuzzy = data.get('use_fuzzy', True)
        
        from app.fuzzy_ahp import calculate_ahp_weights, calculate_fuzzy_ahp_weights
        
        if use_fuzzy:
            success, message, results = calculate_fuzzy_ahp_weights(event_id)
        else:
            success, message, results = calculate_ahp_weights(event_id)
        
        if success:
            # Ambil hasil AHP yang sudah disimpan
            ahp_result = AHPResults.query.filter_by(event_id=event_id).first()
            result_data = {
                'success': True,
                'message': message,
                'results': results
            }
            
            if ahp_result:
                result_data['ahp_result'] = {
                    'lambda_max': float(ahp_result.lambda_max) if ahp_result.lambda_max else None,
                    'ci': float(ahp_result.ci) if ahp_result.ci else None,
                    'cr': float(ahp_result.cr) if ahp_result.cr else None,
                    'is_consistent': ahp_result.is_consistent,
                    'weights_json': ahp_result.weights_json
                }
            
            return jsonify(result_data)
        else:
            return jsonify({'success': False, 'message': message}), 400
            
    except Exception as e:
        logging.error(f"Error calculating AHP: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# Route untuk Laporan & Arsip Seleksi
@app.route('/admin/laporan_arsip_seleksi')
@login_required
@admin_required
def admin_laporan_arsip_seleksi():
    """Halaman laporan dan arsip seleksi"""
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    # Ambil semua kegiatan
    events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Ambil semua arsip
    arsip_list = ArsipSeleksi.query.order_by(ArsipSeleksi.tanggal_arsip.desc()).all()
    
    # Format data arsip
    arsip_data = []
    for arsip in arsip_list:
        arsip_data.append({
            'id': arsip.id_arsip,
            'event_id': arsip.event_id,
            'nama_kegiatan': arsip.event.nama_kegiatan if arsip.event else 'N/A',
            'nama_arsip': arsip.nama_arsip,
            'deskripsi': arsip.deskripsi,
            'file_path': arsip.file_path,
            'file_type': arsip.file_type,
            'tanggal_arsip': arsip.tanggal_arsip.strftime('%d %b %Y') if arsip.tanggal_arsip else '',
            'pembuat': arsip.pembuat.nama_lengkap if arsip.pembuat else 'System',
            'status': arsip.status
        })
    
    return render_template(
        'manajemen_seleksi.html',
        kegiatan_list=events,
        kegiatan_data=[],
        arsip_list=arsip_data,
        sidebar_state=sidebar_state
    )

# API untuk generate laporan Excel
@app.route('/api/generate_laporan_excel/<int:event_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def generate_laporan_excel(event_id):
    """Generate laporan Excel untuk hasil seleksi"""
    try:
        event = Event.query.get_or_404(event_id)
        
        # Ambil hasil seleksi
        hasil_seleksi = db.session.query(
            HasilSeleksi,
            Users,
            Participants
        ).join(
            Users, HasilSeleksi.id_users == Users.id
        ).outerjoin(
            Participants, Users.email == Participants.email
        ).filter(
            HasilSeleksi.event_id == event_id
        ).order_by(
            HasilSeleksi.ranking.asc()
        ).all()
        
        if not hasil_seleksi:
            return jsonify({'success': False, 'message': 'Tidak ada data hasil seleksi'}), 400
        
        # Buat workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Hasil Seleksi"
        
        # Header
        headers = ['No', 'Peringkat', 'Nama Lengkap', 'Email', 'Golongan', 'Tingkatan', 'Asal Gudep', 'Skor Akhir']
        ws.append(headers)
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Data
        for idx, (hasil, user, participant) in enumerate(hasil_seleksi, start=1):
            ws.append([
                idx,
                hasil.ranking,
                user.nama_lengkap or '',
                user.email or '',
                participant.golongan if participant else '',
                participant.tingkatan if participant else '',
                participant.asal_gudep if participant else '',
                round(hasil.skor_akhir, 2)
            ])
        
        # Auto width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Simpan ke arsip
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"laporan_seleksi_{event.nama_kegiatan.replace(' ', '_')}_{timestamp}.xlsx"
        
        # Simpan file ke folder static/uploads/reports
        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'reports')
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, filename)
        
        with open(file_path, 'wb') as f:
            f.write(output.getvalue())
        
        # Simpan ke database arsip
        arsip = ArsipSeleksi(
            event_id=event_id,
            nama_arsip=f"Laporan Excel - {event.nama_kegiatan}",
            deskripsi=f"Laporan hasil seleksi dalam format Excel untuk kegiatan {event.nama_kegiatan}",
            file_path=f"static/uploads/reports/{filename}",
            file_type='excel',
            dibuat_oleh=current_user.id,
            status='aktif'
        )
        db.session.add(arsip)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Laporan Excel berhasil dibuat dan diarsipkan',
            'file_path': f"/{file_path}",
            'arsip_id': arsip.id_arsip
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error generating Excel report: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# API untuk generate laporan PDF (menggunakan HTML to PDF atau reportlab)
@app.route('/admin/laporan/preview/<int:event_id>')
@login_required
def preview_laporan_seleksi(event_id):
    # Allow admin, penilai, and peserta to access
    if current_user.level == 'penilai':
        event = Event.query.get_or_404(event_id)
        if current_user not in event.evaluators:
            flash('Akses ditolak! Anda tidak memiliki akses ke kegiatan ini.', 'error')
            return redirect(url_for('index'))
    elif current_user.level == 'peserta':
        # Peserta hanya bisa akses event yang dia ikuti
        participant = Participants.query.filter_by(email=current_user.email).first()
        if not participant:
            flash('Akses ditolak! Data peserta tidak ditemukan.', 'error')
            return redirect(url_for('index'))
        hasil = HasilSeleksi.query.filter_by(id_users=current_user.id, event_id=event_id).first()
        if not hasil:
            flash('Akses ditolak! Anda tidak terdaftar di kegiatan ini.', 'error')
            return redirect(url_for('index'))
    elif current_user.level != 'admin':
        flash('Akses ditolak!', 'error')
        return redirect(url_for('index'))
    
    event = Event.query.get_or_404(event_id)
    
    # Ambil hasil seleksi (Logika sama dengan admin_hasil_seleksi)
    hasil_seleksi = db.session.query(
        HasilSeleksi, Users, Participants
    ).join(
        Users, HasilSeleksi.id_users == Users.id
    ).outerjoin(
        Participants, Users.email == Participants.email
    ).filter(
        HasilSeleksi.event_id == event_id
    ).order_by(
        HasilSeleksi.ranking.asc()
    ).all()
    
    # Format tanggal indonesia
    now = datetime.now()
    bulan_list = [
        'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
    ]
    tanggal_laporan_indo = f"{now.day} {bulan_list[now.month-1]} {now.year}"
    
    return render_template(
        'laporan_pdf_template.html',
        event=event,
        hasil_seleksi=hasil_seleksi,
        tanggal_laporan=now.strftime('%d-%m-%Y'),
        tanggal_laporan_indo=tanggal_laporan_indo
    )

@app.route('/admin/laporan/word/<int:event_id>')
@login_required
def export_laporan_word(event_id):
    # Allow admin, penilai, and peserta to access
    if current_user.level == 'penilai':
        event = Event.query.get_or_404(event_id)
        if current_user not in event.evaluators:
            flash('Akses ditolak! Anda tidak memiliki akses ke kegiatan ini.', 'error')
            return redirect(url_for('index'))
    elif current_user.level == 'peserta':
        # Peserta hanya bisa akses event yang dia ikuti
        participant = Participants.query.filter_by(email=current_user.email).first()
        if not participant:
            flash('Akses ditolak! Data peserta tidak ditemukan.', 'error')
            return redirect(url_for('index'))
        hasil = HasilSeleksi.query.filter_by(id_users=current_user.id, event_id=event_id).first()
        if not hasil:
            flash('Akses ditolak! Anda tidak terdaftar di kegiatan ini.', 'error')
            return redirect(url_for('index'))
    elif current_user.level != 'admin':
        flash('Akses ditolak!', 'error')
        return redirect(url_for('index'))
    
    event = Event.query.get_or_404(event_id)
    
    # Render template yang sama
    # Logic data sama
    hasil_seleksi = db.session.query(
        HasilSeleksi, Users, Participants
    ).join(
        Users, HasilSeleksi.id_users == Users.id
    ).outerjoin(
        Participants, Users.email == Participants.email
    ).filter(
        HasilSeleksi.event_id == event_id
    ).order_by(
        HasilSeleksi.ranking.asc()
    ).all()
    
    now = datetime.now()
    bulan_list = [
        'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
    ]
    tanggal_laporan_indo = f"{now.day} {bulan_list[now.month-1]} {now.year}"
    
    html_content = render_template(
        'laporan_pdf_template.html',
        event=event,
        hasil_seleksi=hasil_seleksi,
        tanggal_laporan=now.strftime('%d-%m-%Y'),
        tanggal_laporan_indo=tanggal_laporan_indo
    )
    
    # Embed gambar logo sebagai Base64 data URI agar tampil di Word
    # Word tidak bisa resolve URL relatif dari server
    logo_files = {
        'images/Logo-Gerakan-Pramuka.png': 'Logo-Gerakan-Pramuka.png',
        'images/Logo-WOSM.png': 'Logo-WOSM.png'
    }
    
    for static_path, filename in logo_files.items():
        image_full_path = os.path.join(app.root_path, 'static', static_path.replace('/', os.sep))
        if os.path.exists(image_full_path):
            try:
                with open(image_full_path, 'rb') as img_file:
                    img_data = base64.b64encode(img_file.read()).decode('utf-8')
                # Ganti URL relatif dengan data URI base64
                relative_url = url_for('static', filename=static_path)
                data_uri = f'data:image/png;base64,{img_data}'
                html_content = html_content.replace(relative_url, data_uri)
            except Exception as e:
                logging.warning(f"Gagal embed logo {filename} ke Word: {str(e)}")
    
    # Return sebagai file Word (MIME type HTML tetapi extension doc trick)
    response = make_response(html_content)
    response.headers["Content-Type"] = "application/msword"
    response.headers["Content-Disposition"] = f"attachment; filename=Laporan_Hasil_Seleksi_{event.nama_kegiatan.replace(' ', '_')}.doc"
    return response

@app.route('/api/generate_laporan_pdf/<int:event_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def generate_laporan_pdf(event_id):
    """Generate laporan PDF untuk hasil seleksi"""
    try:
        event = Event.query.get_or_404(event_id)
        
        # Ambil hasil seleksi
        hasil_seleksi = db.session.query(
            HasilSeleksi,
            Users,
            Participants
        ).join(
            Users, HasilSeleksi.id_users == Users.id
        ).outerjoin(
            Participants, Users.email == Participants.email
        ).filter(
            HasilSeleksi.event_id == event_id
        ).order_by(
            HasilSeleksi.ranking.asc()
        ).all()
        
        if not hasil_seleksi:
            return jsonify({'success': False, 'message': 'Tidak ada data hasil seleksi'}), 400
        
        # Generate HTML untuk PDF (bisa menggunakan weasyprint atau pdfkit)
        # Untuk sekarang, kita simpan sebagai HTML dan bisa dikonversi nanti
        # Update: Menggunakan template baru
        now = datetime.now()
        bulan_list = [
            'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember'
        ]
        tanggal_laporan_indo = f"{now.day} {bulan_list[now.month-1]} {now.year}"
        
        html_content = render_template(
            'laporan_pdf_template.html',
            event=event,
            hasil_seleksi=hasil_seleksi,
            tanggal_laporan=now.strftime('%d-%m-%Y'),
            tanggal_laporan_indo=tanggal_laporan_indo
        )
        
        # Simpan HTML (atau konversi ke PDF jika ada library)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"laporan_seleksi_{event.nama_kegiatan.replace(' ', '_')}_{timestamp}.html"
        
        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'reports')
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, filename)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        # Simpan ke database arsip
        arsip = ArsipSeleksi(
            event_id=event_id,
            nama_arsip=f"Laporan PDF - {event.nama_kegiatan}",
            deskripsi=f"Laporan hasil seleksi dalam format PDF untuk kegiatan {event.nama_kegiatan}",
            file_path=f"static/uploads/reports/{filename}",
            file_type='pdf',
            dibuat_oleh=current_user.id,
            status='aktif'
        )
        db.session.add(arsip)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Laporan PDF berhasil dibuat dan diarsipkan',
            'file_path': f"/{file_path}",
            'arsip_id': arsip.id_arsip
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error generating PDF report: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# API untuk download file arsip
@app.route('/api/download_arsip/<int:arsip_id>')
@login_required
@admin_required
def download_arsip(arsip_id):
    """Download file arsip"""
    try:
        arsip = ArsipSeleksi.query.get_or_404(arsip_id)
        
        if not arsip.file_path:
            flash('File tidak ditemukan', 'error')
            return redirect(url_for('admin_manajemen_seleksi'))
        
        file_path = os.path.join(app.root_path, arsip.file_path)
        
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server', 'error')
            return redirect(url_for('admin_manajemen_seleksi'))
        
        return send_file(file_path, as_attachment=True, download_name=arsip.nama_arsip)
        
    except Exception as e:
        logging.error(f"Error downloading archive: {str(e)}")
        flash('Error saat mengunduh file', 'error')
        return redirect(url_for('admin_manajemen_seleksi'))

# API untuk hapus arsip
@app.route('/api/hapus_arsip/<int:arsip_id>', methods=['DELETE', 'POST'])
@login_required
@admin_required
@csrf.exempt
def hapus_arsip(arsip_id):
    """Hapus arsip seleksi"""
    try:
        arsip = ArsipSeleksi.query.get_or_404(arsip_id)
        
        # Hapus file jika ada
        if arsip.file_path:
            file_path = os.path.join(app.root_path, arsip.file_path)
            if os.path.exists(file_path):
                os.remove(file_path)
        
        db.session.delete(arsip)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Arsip berhasil dihapus'})
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error deleting archive: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/admin/peserta')
@login_required
@admin_required
def admin_peserta():
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    # Get statistics
    total_peserta = Users.query.filter_by(level='peserta').count()
    
    # Count by status
    peserta_aktif = Users.query.filter_by(level='peserta', status='aktif').count()
    peserta_nonaktif = Users.query.filter_by(level='peserta', status='non-aktif').count()
    
    # Get all events for filter
    events = Event.query.all()
    
    return render_template(
        'data_peserta.html', 
        sidebar_state=sidebar_state, 
        total_peserta=total_peserta,
        peserta_aktif=peserta_aktif,
        peserta_nonaktif=peserta_nonaktif,
        events=events,
        time=time
    )

# Route untuk halaman detail peserta
@app.route('/admin/peserta/detail/<int:user_id>')
@login_required
@admin_required
def detail_peserta(user_id):
    """Halaman detail peserta dengan sidebar"""
    try:
        # Get user data
        user = Users.query.get(user_id)
        if not user or user.level != 'peserta':
            flash("Peserta tidak ditemukan", "error")
            return redirect(url_for('admin_peserta'))
        
        # Get participant biodata
        biodata = Participants.query.filter_by(email=user.email).first()
        
        # Get registered activities
        registered_activities = []
        if biodata:
            activities = biodata.registered_activities.all()
            for activity in activities:
                hasil_activity = HasilSeleksi.query.filter_by(
                    id_users=user.id,
                    event_id=activity.id_kegiatan
                ).first()
                
                registered_activities.append({
                    'id': activity.id_kegiatan,
                    'nama': activity.nama_kegiatan,
                    'jenis': activity.jenis_kegiatan,
                    'skor': hasil_activity.skor_akhir if hasil_activity else None,
                    'ranking': hasil_activity.ranking if hasil_activity else None
                })
        
        sidebar_state = current_user.sidebar_state or 'expanded'
        
        return render_template(
            'detail_peserta.html',
            user=user,
            biodata=biodata,
            registered_activities=registered_activities,
            sidebar_state=sidebar_state
        )
    except Exception as e:
        logging.error(f"Error in detail_peserta: {e}")
        flash("Terjadi kesalahan saat memuat data peserta", "error")
        return redirect(url_for('admin_peserta'))

# Route untuk cetak kartu peserta
@app.route('/admin/peserta/kartu/<int:user_id>')
@login_required
@admin_required
def cetak_kartu_peserta(user_id):
    """Halaman untuk mencetak kartu peserta"""
    try:
        # Get user data
        user = Users.query.get(user_id)
        if not user or user.level != 'peserta':
            flash("Peserta tidak ditemukan", "error")
            return redirect(url_for('admin_peserta'))
        
        # Get participant biodata
        biodata = Participants.query.filter_by(email=user.email).first()
        
        # Get registered activities
        registered_activities = []
        if biodata:
            activities = biodata.registered_activities.all()
            for activity in activities:
                hasil_activity = HasilSeleksi.query.filter_by(
                    id_users=user.id,
                    event_id=activity.id_kegiatan
                ).first()
                
                registered_activities.append({
                    'id': activity.id_kegiatan,
                    'nama': activity.nama_kegiatan,
                    'jenis': activity.jenis_kegiatan,
                    'skor': hasil_activity.skor_akhir if hasil_activity else None,
                    'ranking': hasil_activity.ranking if hasil_activity else None
                })
        
        return render_template(
            'kartu_peserta.html',
            user=user,
            biodata=biodata,
            registered_activities=registered_activities
        )
    except Exception as e:
        logging.error(f"Error in cetak_kartu_peserta: {e}")
        flash("Terjadi kesalahan saat memuat data peserta", "error")
        return redirect(url_for('admin_peserta'))
    
# Route untuk halaman tambah peserta ke kegiatan
@app.route('/admin/peserta/tambah-kegiatan')
@login_required
@admin_required
def tambah_peserta_kegiatan():
    """Halaman untuk menambahkan peserta ke kegiatan"""
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    # Get all events
    events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Get all participants (users with level peserta)
    users_peserta = Users.query.filter_by(level='peserta').all()
    
    # Get all participants biodata
    participants = Participants.query.all()
    
    return render_template(
        'tambah_peserta_kegiatan.html',
        sidebar_state=sidebar_state,
        events=events,
        users_peserta=users_peserta,
        participants=participants
    )

# API untuk menambahkan peserta ke kegiatan
@app.route('/api/peserta/tambah-kegiatan', methods=['POST'])
@login_required
@admin_required
def api_tambah_peserta_kegiatan():
    """API untuk menambahkan peserta ke kegiatan"""
    try:
        data = request.get_json()
        participant_id = data.get('participant_id')
        kegiatan_id = data.get('kegiatan_id')
        
        if not participant_id or not kegiatan_id:
            return jsonify({
                'success': False,
                'message': 'Participant ID dan Kegiatan ID harus diisi'
            }), 400
        
        # Cek apakah peserta sudah terdaftar di kegiatan ini
        existing = db.session.query(tb_participant_kegiatan).filter_by(
            participant_id=participant_id,
            kegiatan_id=kegiatan_id
        ).first()
        
        if existing:
            return jsonify({
                'success': False,
                'message': 'Peserta sudah terdaftar di kegiatan ini'
            }), 400
        
        # Cek apakah participant dan event ada
        participant = Participants.query.get(participant_id)
        event = Event.query.get(kegiatan_id)
        
        if not participant:
            return jsonify({
                'success': False,
                'message': 'Peserta tidak ditemukan'
            }), 404
        
        if not event:
            return jsonify({
                'success': False,
                'message': 'Kegiatan tidak ditemukan'
            }), 404
        
        # Tambahkan peserta ke kegiatan
        db.session.execute(
            tb_participant_kegiatan.insert().values(
                participant_id=participant_id,
                kegiatan_id=kegiatan_id,
                tanggal_daftar=datetime.now()
            )
        )
        db.session.commit()
        
        log_activity(current_user.id, f'Menambahkan peserta {participant.nama_lengkap} ke kegiatan {event.nama_kegiatan}')
        
        return jsonify({
            'success': True,
            'message': f'Peserta berhasil ditambahkan ke kegiatan {event.nama_kegiatan}'
        })
        
    except IntegrityError as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': 'Peserta sudah terdaftar di kegiatan ini'
        }), 400
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_tambah_peserta_kegiatan: {e}")
        current_app.logger.exception('Error in api_tambah_peserta_kegiatan:')
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# API untuk menambahkan multiple peserta ke kegiatan
@app.route('/api/peserta/tambah-kegiatan-bulk', methods=['POST'])
@login_required
@admin_required
def api_tambah_peserta_kegiatan_bulk():
    """API untuk menambahkan multiple peserta ke kegiatan sekaligus"""
    try:
        data = request.get_json()
        participant_ids = data.get('participant_ids', [])
        kegiatan_id = data.get('kegiatan_id')
        
        if not participant_ids or not kegiatan_id:
            return jsonify({
                'success': False,
                'message': 'Participant IDs dan Kegiatan ID harus diisi'
            }), 400
        
        # Cek apakah event ada
        event = Event.query.get(kegiatan_id)
        if not event:
            return jsonify({
                'success': False,
                'message': 'Kegiatan tidak ditemukan'
            }), 404
        
        added_count = 0
        skipped_count = 0
        errors = []
        
        for participant_id in participant_ids:
            try:
                # Cek apakah ini user_id atau participant_id
                # Coba cari di Participants dulu
                participant = Participants.query.get(participant_id)
                
                # Jika tidak ada participant, cari user dan buat biodata minimal
                if not participant:
                    user = Users.query.get(participant_id)
                    if not user or user.level != 'peserta':
                        errors.append(f'User ID {participant_id} tidak ditemukan atau bukan peserta')
                        continue
                    
                    # Cek apakah sudah ada participant dengan email yang sama
                    participant = Participants.query.filter_by(email=user.email).first()
                    
                    # Jika belum ada, buat biodata minimal
                    if not participant:
                        participant = Participants(
                            nama_lengkap=user.nama_lengkap or user.username or 'Peserta',
                            email=user.email,
                            jenis_kelamin=user.jenis_kelamin if user.jenis_kelamin else 'laki-laki',
                            usia=int(user.usia) if user.usia and str(user.usia).isdigit() else 0,
                            nomor_hp=user.nomor_hp or '',
                            tanggal_lahir=date.today(),
                            alamat_tinggal='',
                            golongan='siaga',
                            tingkatan='siaga mula',
                            asal_gudep='',
                            asal_kwarran='',
                            asal_kwarcab='',
                            asal_kwarda='',
                            foto=user.foto if user.foto and user.foto != 'img/default-user.png' else 'img/default-user.png',
                            level='peserta'
                        )
                        db.session.add(participant)
                        db.session.flush()  # Untuk mendapatkan ID
                
                # Cek apakah peserta sudah terdaftar di kegiatan ini
                existing = db.session.query(tb_participant_kegiatan).filter_by(
                    participant_id=participant.id,
                    kegiatan_id=kegiatan_id
                ).first()
                
                if existing:
                    skipped_count += 1
                    continue
                
                # Tambahkan peserta ke kegiatan
                db.session.execute(
                    tb_participant_kegiatan.insert().values(
                        participant_id=participant.id,
                        kegiatan_id=kegiatan_id,
                        tanggal_daftar=datetime.now()
                    )
                )
                added_count += 1
                
            except Exception as e:
                errors.append(f'Error untuk peserta ID {participant_id}: {str(e)}')
                continue
        
        db.session.commit()
        
        log_activity(current_user.id, f'Menambahkan {added_count} peserta ke kegiatan {event.nama_kegiatan}')
        
        # Buat notifikasi untuk admin
        if added_count > 0:
            create_notification_to_all_admins(
                f"{added_count} peserta ditambahkan ke kegiatan: {event.nama_kegiatan}"
            )
        
        message = f'Berhasil menambahkan {added_count} peserta'
        if skipped_count > 0:
            message += f', {skipped_count} peserta sudah terdaftar'
        if errors:
            message += f', {len(errors)} error'
        
        return jsonify({
            'success': True,
            'message': message,
            'added': added_count,
            'skipped': skipped_count,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_tambah_peserta_kegiatan_bulk: {e}")
        current_app.logger.exception('Error in api_tambah_peserta_kegiatan_bulk:')
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    
@app.route('/admin/hasil_seleksi')
@login_required
@admin_required
def admin_hasil_seleksi():
    # Get all events
    all_events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Get selected event from query parameter
    selected_event_id = request.args.get('event_id', type=int)
    selected_event = None
    results = []
    
    if selected_event_id:
        selected_event = Event.query.get(selected_event_id)
        if selected_event:
            from app.fuzzy_ahp import calculate_spk
            
            # Calculate SPK for this event
            success, msg = calculate_spk(selected_event_id)
            if not success:
                logging.warning(f"Gagal hitung SPK untuk event {selected_event_id}: {msg}")
            else:
                # Buat notifikasi untuk admin saat hasil seleksi selesai
                create_notification_to_all_admins(
                    f"Hasil seleksi selesai dihitung untuk kegiatan: {selected_event.nama_kegiatan}"
                )
            
            # Fetch results for this event only
            hasil_seleksi = db.session.query(
                HasilSeleksi,
                Users,
                Participants
            ).join(
                Users, HasilSeleksi.id_users == Users.id
            ).outerjoin(
                Participants, Users.email == Participants.email
            ).filter(
                HasilSeleksi.event_id == selected_event_id
            ).order_by(
                HasilSeleksi.ranking.asc()
            ).all()
            
            # Build results list
            for hasil, user, participant in hasil_seleksi:
                results.append({
                    'hasil': hasil,
                    'user': user,
                    'participant': participant
                })
        else:
            flash("Kegiatan tidak ditemukan.", "error")
            selected_event = None

    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template(
        'admin/hasil_penilaian.html',
        assigned_events=all_events,
        selected_event=selected_event,
        results=results,
        sidebar_state=sidebar_state,
        show_back_button=False  # Tidak tampilkan tombol kembali karena dibuka dari sidebar
    )

# Manajemen Berita
@app.route('/admin/manajemen_berita')
@login_required
@admin_required
def admin_manajemen_berita():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    
    query = News.query
    
    # 🔍 Filter judul
    if search:
        query = query.filter(News.title.ilike(f"%{search}%"))

    # 🏷️ Filter status
    if status in ['published', 'draft', 'archived']:
        query = query.filter(News.status == status)

    # Ambil semua berita
    news_list_raw = query.order_by(News.created_at.desc()).all()

    news_list = []
    for news in news_list_raw:
        news_list.append({
            "id_news": news.id_news,
            "title": news.title,
            "content": news.content,
            "status": news.status,
            "created_at": news.created_at.strftime("%d-%m-%Y"),
            "author": {
                "nama_lengkap": news.author.nama_lengkap
            }
        })
    total_news = len(news_list_raw)
    published_news = News.query.filter_by(status='published').count()
    draft_news = News.query.filter_by(status='draft').count()

    # Ambil berita terakhir (updated_at fallback ke created_at)
    last_news = (
        News.query
        .order_by(
            db.func.coalesce(News.updated_at, News.created_at).desc()
        )
        .first()
    )
    last_update = (
        (last_news.updated_at or last_news.created_at).strftime("%d-%m-%Y")
        if last_news else '-'
    )
    return render_template('news_management.html', news_list=news_list, total_news=total_news, published_news=published_news, draft_news=draft_news, last_update=last_update, sidebar_state=sidebar_state, user=current_user, search=search, status=status, time=time)

# API Pagination Berita
@app.route("/admin/api/berita")
@login_required
@admin_required
def api_berita():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 6, type=int)

    pagination = (
        News.query
        .order_by(News.created_at.desc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    data = []
    for news in pagination.items:
        data.append({
            "id_news": news.id_news,
            "title": news.title,
            "status": news.status,
            "created_at": news.created_at.strftime("%d-%m-%Y"),
            "author": {
                "nama_lengkap": news.author.nama_lengkap
            }
        })

    return jsonify({
        "data": data,
        "pagination": {
            "page": pagination.page,
            "total_pages": pagination.pages,
            "total_items": pagination.total,
            "per_page": pagination.per_page,
        }
    })


# Halaman Tambah Berita (Baru)
@app.route('/admin/news/add', methods=['GET'])
@login_required
@admin_required
def admin_add_news_page():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
        
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('admin_add_news.html', sidebar_state=sidebar_state, user=current_user)

# Tambah Berita
@app.route('/admin/news/create', methods=['POST'])
@login_required
@admin_required
def admin_create_news():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    title = request.form['title']
    content = request.form['content']
    status = request.form['status']
    
    file = request.files.get('thumbnail')
    thumbnail_path = "images/default-news.jpg"

    if file and file.filename:
        filename = secure_filename(file.filename)
        upload_dir = os.path.join(app.static_folder, 'uploads/news')
        os.makedirs(upload_dir, exist_ok=True)

        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)

        thumbnail_path = f"uploads/news/{filename}"

    
    # Validasi dasar
    if not title or not content or not status:
        flash(t['all_fields_required_news'], 'error')
        return redirect(url_for('admin_manajemen_berita'))
    
    base_slug = slugify(title)
    slug = base_slug
    
    # Cegah slug duplikat
    counter = 1
    while News.query.filter_by(slug=slug).first():
        slug = f"{base_slug}-{counter}"
        counter += 1

    # Excerpt otomatis
    excerpt = content[:200] + '...' if len(content) > 200 else content
    news = News(
        title=title,
        slug=slug,
        content=content,
        excerpt=excerpt,
        status=status,
        author_id=current_user.id,
        thumbnail=thumbnail_path
    )

    if status == 'published':
        news.published_at = datetime.utcnow()
    db.session.add(news)
    db.session.commit()
    flash(t['news_added'], 'success')
    return redirect(url_for('admin_manajemen_berita'))

# Edit Berita
@app.route('/admin/news/edit/<int:id_news>', methods=['POST'])
@login_required
@admin_required
def admin_edit_news(id_news):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    news = News.query.get_or_404(id_news)
    news.title = request.form['title']
    news.content = request.form['content']
    news.status = request.form['status']
    news.slug = slugify(news.title)
    news.excerpt = news.content[:200] + '...'
    if news.status == 'published' and not news.published_at:
        news.published_at = datetime.utcnow()
    db.session.commit()
    flash(t['news_updated'], 'success')
    return redirect(url_for('admin_manajemen_berita'))

# Hapus Berita
@app.route('/admin/news/delete/<int:id_news>', methods=['POST'])
@login_required
@admin_required
def admin_delete_news(id_news):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    news = News.query.get_or_404(id_news)
    db.session.delete(news)
    db.session.commit()

    flash(t['news_deleted'], 'success')
    return redirect(url_for('admin_manajemen_berita'))

# Komentar Berita (Admin)
@app.route('/news/<int:news_id>/comment', methods=['POST'])
@login_required
@admin_required
def post_comment_admin(news_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'admin':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))

    data = request.get_json()
    content = data.get('content', '').strip()
    parent_id = data.get('parent_id')

    if not content:
        return jsonify({'error': 'Komentar tidak boleh kosong'}), 400

    comment = Comment(
        news_id=news_id,
        user_id=current_user.id,
        parent_id=parent_id,
        content=content
    )
    db.session.add(comment)
    db.session.commit()
    return jsonify({'message': 'Komentar berhasil ditambahkan'}), 201

@app.route('/author/<int:author_id>')
def author_news(author_id):
    author = Users.query.get_or_404(author_id)
    news_list = News.query.filter_by(author_id=author.id, status='published').all()
    return render_template('author_news.html', author=author, news_list=news_list)

# Detail Berita
@app.route('/news/<slug>')
def news_detail(slug):
    news = News.query.filter_by(
        slug=slug,
        status='published'
    ).first_or_404()

    # Ambil komentar utama (bukan reply)
    comment_count = Comment.query.filter(
        Comment.news_id == news.id_news,
        Comment.parent_id.is_(None),
        Comment.is_deleted.is_(False),
        Comment.is_approved.is_(True)
    ).count()
    return render_template('news_detail.html', news=news, comment_count=comment_count)

@app.route('/news/<slug>/comments')
def get_comments(slug):
    page = request.args.get('page', 1, type=int)
    per_page = 5

    news = News.query.filter_by(slug=slug).first_or_404()

    pagination = Comment.query.filter(
        Comment.news_id == news.id_news,
        Comment.parent_id.is_(None),
        Comment.is_deleted.is_(False),
        Comment.is_approved.is_(True)
    ).order_by(Comment.created_at.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    return jsonify({"total": pagination.total, "comments": [comment.to_dict() for comment in pagination.items], "has_next": pagination.has_next})

@app.route('/news/comment/<int:comment_id>/replies')
def get_comment_replies(comment_id):
    replies = Comment.query.filter_by(
        parent_id=comment_id,
        is_deleted=False,
        is_approved=True
    ).order_by(Comment.created_at.asc()).all()
    return jsonify({"replies": [r.to_dict() for r in replies]})

# Post Komentar Berita (support AJAX)
@app.route('/news/<slug>/comment', methods=['POST'])
# @login_required  <-- Removed to allow anonymous comments
def post_comment_user(slug):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    news = News.query.filter_by(
        slug=slug,
        status='published'
    ).first_or_404()

    # Ambil data komentar
    if request.is_json:
        data = request.get_json()
        content = data.get('content', '').strip()
        parent_id = data.get('parent_id')
    else:
        content = request.form.get('content', '').strip()
        parent_id = request.form.get('parent_id')

    if not content:
        if request.is_json:
            return jsonify({'success': False, 'message': t['comment_empty']}), 400
        else:
            flash(t['comment_empty'], 'error')
            return redirect(url_for('news_detail', slug=slug))

    # Buat komentar baru
    user_id = current_user.id if current_user.is_authenticated else None
    
    comment = Comment(
        news_id=news.id_news,
        user_id=user_id,
        content=content,
        parent_id=parent_id if parent_id else None
    )
    db.session.add(comment)
    db.session.commit()

    total_comments = Comment.query.filter_by(
        news_id=news.id_news,
        parent_id=None,
        is_deleted=False,
        is_approved=True
    ).count()

    if request.is_json:
        return jsonify({'success': True, 'comment': comment.to_dict(), 'total_comments': total_comments})

    # Redirect normal untuk submit form biasa
    flash(t['comment_posted'], 'success')
    return redirect(url_for('news_detail', slug=slug))

# Like / Unlike Komentar Berita
@app.route('/news/comment/<int:id>/like', methods=['POST'])
@login_required
def like_comment(id):
    comment = Comment.query.get_or_404(id)

    existing_like = CommentLike.query.filter_by(
        comment_id=comment.id,
        user_id=current_user.id
    ).first()

    if existing_like:
        # 💔 UNLIKE
        db.session.delete(existing_like)
        is_liked = False
    else:
        # ❤️ LIKE
        new_like = CommentLike(
            comment_id=comment.id,
            user_id=current_user.id
        )
        db.session.add(new_like)
        is_liked = True

    # 🔁 UPDATE COUNTER
    comment.likes = comment.likes_rel.count()
    db.session.commit()
    return jsonify({"success": True, "likes": comment.likes, "is_liked": is_liked})

# Hapus Komentar Berita
@app.route('/comment/<int:id>/delete', methods=['POST'])
@login_required
def delete_comment(id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    comment = Comment.query.get_or_404(id)

    if comment.user_id != current_user.id:
        return {"success": False, "message": t["unauthorized"]}, 403

    comment.is_deleted = True
    replies = Comment.query.filter_by(parent_id=id).all()
    for reply in replies:
        reply.is_deleted = True
    db.session.commit()
    return {"success": True, "message": t["commentDeleted"]}

# Edit Komentar Berita
@app.route('/comment/<int:id>/edit', methods=['POST'])
@login_required
def edit_comment(id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    comment = Comment.query.get_or_404(id)

    # Cek hak akses
    if comment.user_id != current_user.id:
        return {"success": False, "message": t["unauthorized-2"]}, 403

    data = request.get_json()
    content = data.get("content", "").strip()
    if not content:
        return {"success": False, "message": t["emptyContent"]}, 400

    # Update komentar
    comment.content = content
    db.session.commit()
    return {"success": True, "message": t["commentUpdated"], "comment": {"id": comment.id, "content": comment.content, "user": {"nama_lengkap": comment.user.nama_lengkap, "foto": comment.user.foto}, "likes": comment.likes, "is_owner": True, "is_liked": False, "reply_count": len(comment.replies), "parent_id": comment.parent_id}}

@app.route('/admin/notifikasi')
@login_required
@admin_required
def admin_notifikasi():
    # Get notifications for admin user
    notifications = Notification.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Notification.id.desc()
    ).all()
    
    # Count unread notifications
    unread_count = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).count()
    
    # Debug logging
    logging.info(f"Admin notifications page - User ID: {current_user.id}, Total notifications: {len(notifications)}, Unread: {unread_count}")
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template(
        'notifikasi.html',
        notifications=notifications,
        unread_count=unread_count,
        sidebar_state=sidebar_state
    )

# API Mark Notification as Read
@app.route('/api/notifikasi/mark-read/<int:notification_id>', methods=['POST'])
@login_required
def api_mark_notification_read(notification_id):
    try:
        notification = Notification.query.filter_by(
            id=notification_id,
            user_id=current_user.id
        ).first()
        
        if not notification:
            return jsonify({'success': False, 'message': 'Notifikasi tidak ditemukan'}), 404
        
        notification.is_read = True
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Notifikasi ditandai sebagai dibaca'})
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_mark_notification_read: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# API Mark All Notifications as Read
@app.route('/api/notifikasi/mark-all-read', methods=['POST'])
@login_required
def api_mark_all_notifications_read():
    try:
        updated = Notification.query.filter_by(
            user_id=current_user.id,
            is_read=False
        ).update({'is_read': True})
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{updated} notifikasi ditandai sebagai dibaca'
        })
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_mark_all_notifications_read: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# API Delete Notification
@app.route('/api/notifikasi/delete/<int:notification_id>', methods=['DELETE'])
@login_required
def api_delete_notification(notification_id):
    try:
        notification = Notification.query.filter_by(
            id=notification_id,
            user_id=current_user.id
        ).first()
        
        if not notification:
            return jsonify({'success': False, 'message': 'Notifikasi tidak ditemukan'}), 404
        
        db.session.delete(notification)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Notifikasi berhasil dihapus'})
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in api_delete_notification: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/admin/log_aktivitas')
@login_required
@admin_required
def admin_log_aktivitas():
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    # Get query parameters for filtering
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '', type=str)
    role_filter = request.args.get('role', '', type=str)
    date_filter = request.args.get('date', '', type=str)
    
    # Build query
    query = LogAktivitas.query.join(Users).order_by(LogAktivitas.timestamp.desc())
    
    # Apply filters
    if search:
        query = query.filter(
            db.or_(
                Users.username.ilike(f'%{search}%'),
                Users.nama_lengkap.ilike(f'%{search}%'),
                LogAktivitas.aktivitas.ilike(f'%{search}%')
            )
        )
    
    if role_filter:
        query = query.filter(Users.level == role_filter)
    
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter(db.cast(LogAktivitas.timestamp, db.Date) == filter_date)
        except ValueError:
            pass
    
    # Pagination
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    logs = pagination.items
    
    return render_template(
        'log_aktivity.html',
        sidebar_state=sidebar_state,
        logs=logs,
        pagination=pagination,
        search=search,
        role_filter=role_filter,
        date_filter=date_filter,
        current_page=page
    )

@app.route('/api/log_aktivitas/detail/<int:log_id>')
@login_required
@admin_required
def api_log_detail(log_id):
    try:
        log = LogAktivitas.query.get_or_404(log_id)
        return jsonify({
            'success': True,
            'log': {
                'id_log': log.id_log,
                'user_name': log.user.nama_lengkap or log.user.username,
                'user_role': log.user.level,
                'aktivitas': log.aktivitas,
                'ip_address': log.ip_address,
                'user_agent': log.user_agent,
                'timestamp': log.timestamp.strftime('%d/%m/%Y %H:%M:%S') if log.timestamp else '-'
            }
        })
    except Exception as e:
        current_app.logger.exception('Error in api_log_detail:')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/log_aktivitas/delete/<int:log_id>', methods=['DELETE'])
@login_required
@admin_required
def api_log_delete(log_id):
    try:
        log = LogAktivitas.query.get_or_404(log_id)
        db.session.delete(log)
        db.session.commit()
        log_activity(current_user.id, f'Menghapus log aktivitas ID: {log_id}')
        return jsonify({'success': True, 'message': 'Log aktivitas berhasil dihapus'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in api_log_delete:')
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/log_aktivitas/export')
@login_required
@admin_required
def admin_log_export():
    try:
        export_format = request.args.get('export', 'csv')
        search = request.args.get('search', '', type=str)
        role_filter = request.args.get('role', '', type=str)
        date_filter = request.args.get('date', '', type=str)
        
        # Build query (same as main route)
        query = LogAktivitas.query.join(Users).order_by(LogAktivitas.timestamp.desc())
        
        if search:
            query = query.filter(
                db.or_(
                    Users.username.ilike(f'%{search}%'),
                    Users.nama_lengkap.ilike(f'%{search}%'),
                    LogAktivitas.aktivitas.ilike(f'%{search}%')
                )
            )
        
        if role_filter:
            query = query.filter(Users.level == role_filter)
        
        if date_filter:
            try:
                filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                query = query.filter(db.cast(LogAktivitas.timestamp, db.Date) == filter_date)
            except ValueError:
                pass
        
        logs = query.all()
        
        if export_format == 'csv':
            # Export CSV
            output = io.StringIO()
            output.write('User,Role,Aktivitas,IP Address,User Agent,Tanggal & Waktu\n')
            
            for log in logs:
                user_name = log.user.nama_lengkap or log.user.username
                role = log.user.level
                aktivitas = log.aktivitas.replace('"', '""')  # Escape quotes
                ip = log.ip_address or '-'
                ua = (log.user_agent or '-').replace('"', '""')
                timestamp = log.timestamp.strftime('%d/%m/%Y %H:%M:%S') if log.timestamp else '-'
                
                output.write(f'"{user_name}","{role}","{aktivitas}","{ip}","{ua}","{timestamp}"\n')
            
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=log_aktivitas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
            )
        else:
            # Export PDF - Generate HTML and convert to PDF using weasyprint or return HTML
            # For now, return HTML that can be printed as PDF
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Log Aktivitas</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #333; }}
                    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #4CAF50; color: white; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                </style>
            </head>
            <body>
                <h1>Log Aktivitas</h1>
                <p>Tanggal Export: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
                <table>
                    <thead>
                        <tr>
                            <th>User</th>
                            <th>Role</th>
                            <th>Aktivitas</th>
                            <th>IP Address</th>
                            <th>Tanggal & Waktu</th>
                        </tr>
                    </thead>
                    <tbody>
            """
            
            for log in logs:
                user_name = log.user.nama_lengkap or log.user.username
                role = log.user.level
                aktivitas = log.aktivitas.replace('<', '&lt;').replace('>', '&gt;')
                ip = log.ip_address or '-'
                timestamp = log.timestamp.strftime('%d/%m/%Y %H:%M:%S') if log.timestamp else '-'
                html_content += f"""
                        <tr>
                            <td>{user_name}</td>
                            <td>{role}</td>
                            <td>{aktivitas}</td>
                            <td>{ip}</td>
                            <td>{timestamp}</td>
                        </tr>
                """
            
            html_content += """
                    </tbody>
                </table>
            </body>
            </html>
            """
            
            return Response(
                html_content,
                mimetype='text/html',
                headers={'Content-Disposition': f'attachment; filename=log_aktivitas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html'}
            )
    except Exception as e:
        current_app.logger.exception('Error in admin_log_export:')
        flash('Error exporting logs: ' + str(e), 'error')
        return redirect(url_for('admin_log_aktivitas'))

@app.route('/admin/settings')
@login_required
@admin_required
def admin_settings():
    sidebar_state = current_user.sidebar_state or 'expanded'
    users = Users.query.count()
    
    # Ambil semua pengaturan dari database
    settings = Settings.query.all()
    settings_dict = {s.key: s.value for s in settings}
    
    # Default values jika belum ada di database
    email_settings = {
        'mail_server': settings_dict.get('mail_server') or os.getenv("MAIL_SERVER", "smtp.gmail.com"),
        'mail_port': settings_dict.get('mail_port') or os.getenv("MAIL_PORT", "465"),
        'mail_use_tls': settings_dict.get('mail_use_tls', 'false'),
        'mail_use_ssl': settings_dict.get('mail_use_ssl', 'true'),
        'mail_username': settings_dict.get('mail_username') or os.getenv("MAIL_USERNAME", ""),
        'mail_password': "",  # Jangan tampilkan password
        'mail_enabled': settings_dict.get('mail_enabled', 'true')
    }
    
    sms_settings = {
        'twilio_account_sid': settings_dict.get('twilio_account_sid') or os.getenv("TWILIO_ACCOUNT_SID", ""),
        'twilio_auth_token': "",  # Jangan tampilkan token
        'twilio_whatsapp_from': settings_dict.get('twilio_whatsapp_from') or os.getenv("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886"),
        'sms_enabled': settings_dict.get('sms_enabled', 'true')
    }
    
    # Format logo_path untuk ditampilkan di template
    logo_path_raw = settings_dict.get('logo_path') or 'img/logo.png'
    if logo_path_raw.startswith('uploads/'):
        logo_path_display = f'/static/{logo_path_raw}'
    elif logo_path_raw.startswith('/static/'):
        logo_path_display = logo_path_raw
    elif not logo_path_raw.startswith('/'):
        logo_path_display = f'/static/{logo_path_raw}'
    else:
        logo_path_display = logo_path_raw
    
    app_settings = {
        'app_name': settings_dict.get('app_name') or 'SPK Pramuka',
        'app_description': settings_dict.get('app_description') or 'Sistem Pendukung Keputusan untuk Seleksi Pramuka',
        'app_version': settings_dict.get('app_version') or '1.0.0',
        'organization_name': settings_dict.get('organization_name') or 'Kwartir',
        'organization_address': settings_dict.get('organization_address') or '',
        'organization_phone': settings_dict.get('organization_phone') or '',
        'organization_email': settings_dict.get('organization_email') or '',
        'logo_path': logo_path_display,
        'default_language': settings_dict.get('default_language') or 'id'
    }
    
    return render_template('settings.html', 
                         sidebar_state=sidebar_state, 
                         user=users, 
                         time=time,
                         email_settings=email_settings,
                         sms_settings=sms_settings,
                         app_settings=app_settings)
    
@app.route('/penilai/dashboard')
@login_required
def penilai_dashboard():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))

    # Ambil semua kegiatan yang aktif
    events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Tambahkan flag is_assigned untuk setiap event
    for event in events:
        event.is_assigned = current_user in event.evaluators
    
    # Hitung total peserta (dari tabel participants)
    total_peserta = Participants.query.count()
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/dashboard.html', events=events, total_peserta=total_peserta, sidebar_state=sidebar_state)

@app.route('/penilai/penilaian')
@login_required
def penilai_penilaian():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))

    # Ambil semua kegiatan yang aktif
    events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Tambahkan flag is_assigned untuk setiap event
    for event in events:
        event.is_assigned = current_user in event.evaluators
        event.jumlah_peserta = event.registered_participants.count()
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/penilaian.html', events=events, sidebar_state=sidebar_state)

@app.route('/penilai/event/<int:event_id>/participants')
@login_required
def penilai_event_participants(event_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    event = Event.query.get_or_404(event_id)
    
    # Check if evaluator is assigned to this event
    if current_user not in event.evaluators:
        flash("Anda tidak ditugaskan untuk menilai kegiatan ini.", "error")
        return redirect(url_for('penilai_dashboard'))
    
    participants = event.registered_participants.all()
    
    # Cek status penilaian untuk setiap peserta oleh penilai ini
    for p in participants:
        # Cari user ID dari tabel Users berdasarkan email peserta
        user_peserta = Users.query.filter_by(email=p.email).first()
        if user_peserta:
            # Cek apakah sudah ada nilai dari penilai ini untuk peserta ini di EVENT INI
            # Join dengan Criteria untuk memastikan penilaian terkait dengan event yang sedang dilihat
            existing_score = db.session.query(Penilaian).join(
                Criteria, Penilaian.id_kriteria == Criteria.id_kriteria
            ).filter(
                Penilaian.id_users == user_peserta.id,
                Penilaian.evaluator_id == current_user.id,
                Criteria.event_id == event_id
            ).first()
            p.is_graded = True if existing_score else False
            p.user_id_for_link = user_peserta.id # Use temp attribute for link
        else:
            p.is_graded = False
            p.user_id_for_link = 0 # Fallback
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/list_peserta.html', event=event, participants=participants, sidebar_state=sidebar_state)

@app.route('/penilai/event/<int:event_id>/grade/<int:participant_id>', methods=['GET', 'POST'])
@login_required
def penilai_input_score(event_id, participant_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    event = Event.query.get_or_404(event_id)
    participant_user = Users.query.get_or_404(participant_id)
    participant_biodata = Participants.query.filter_by(email=participant_user.email).first()
    
    # If no biodata exists, create a temporary object with user data
    if not participant_biodata:
        # Create a simple object to hold the necessary data
        class ParticipantData:
            def __init__(self, user):
                self.id = user.id
                self.nama_lengkap = user.nama_lengkap or user.username
                self.email = user.email
                self.asal_gudep = ''
                self.golongan = 'N/A'
                self.tingkatan = 'N/A'
                self.usia = user.usia or '0'
                self.foto = user.foto or 'img/default-user.png'
        
        participant_biodata = ParticipantData(participant_user)
    
    # Ambil kriteria untuk event ini
    # Filter berdasarkan penugasan
    user_assigned_criteria = [c for c in current_user.assigned_criteria if c.event_id == event_id]
    
    if user_assigned_criteria:
        criterias = user_assigned_criteria
    else:
        # Fallback: jika tidak ada assignment spesifik (legacy), tampilkan semua
        criterias = Criteria.query.filter_by(event_id=event_id).all()
    
    # Ambil himpunan kriteria untuk dropdown
    for c in criterias:
        c.himpunan = HimpunanKriteria.query.filter_by(id_kriteria=c.id_kriteria).all()
        
    # Ambil nilai yang sudah ada (jika edit)
    existing_scores = {}
    scores_query = Penilaian.query.filter_by(
        id_users=participant_id,
        evaluator_id=current_user.id
    ).all()
    for s in scores_query:
        existing_scores[s.id_kriteria] = s.nilai

    if request.method == 'POST':
        try:
            # DEBUG LOGGING
            with open('debug_scores.log', 'a') as f:
                f.write(f"\n--- SAVING SCORES ---\n")
                f.write(f"Participant ID (from route): {participant_id}\n")
                f.write(f"Participant User ID: {participant_user.id}\n")
                f.write(f"Evaluator ID: {current_user.id}\n")
                f.write(f"Event ID: {event_id}\n")
            
            for criteria in criterias:
                score_val = request.form.get(f'score_{criteria.id_kriteria}')
                if score_val:
                    # Cek apakah update atau insert
                    penilaian = Penilaian.query.filter_by(
                        id_users=participant_id,
                        evaluator_id=current_user.id,
                        id_kriteria=criteria.id_kriteria
                    ).first()
                    if penilaian:
                        penilaian.nilai = float(score_val)
                    else:
                        penilaian = Penilaian(id_users=participant_id, evaluator_id=current_user.id, id_kriteria=criteria.id_kriteria, nilai=float(score_val))
                        db.session.add(penilaian)
                    
                    # DEBUG: Log what we're saving
                    with open('debug_scores.log', 'a') as f:
                        f.write(f"Saving: Criteria {criteria.id_kriteria}, Score {score_val}, id_users={participant_id}\n")
            db.session.commit()
            flash(f"{t['score_saved_success']}", "success")
            return redirect(url_for('penilai_event_participants', event_id=event_id))
        except Exception as e:
            db.session.rollback()
            logging.error(f"Error saving score: {e}")
            flash(f"{t['score_save_error']}", "danger")
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/form_penilaian.html', event=event, participant=participant_biodata, participant_user=participant_user, criterias=criterias, existing_scores=existing_scores, sidebar_state=sidebar_state)

@app.route('/penilai/event/<int:event_id>/view/<int:participant_id>')
@login_required
def penilai_view_score(event_id, participant_id):
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    event = Event.query.get_or_404(event_id)
    participant_user = Users.query.get_or_404(participant_id)
    participant_biodata = Participants.query.filter_by(email=participant_user.email).first()
    
    # If no biodata exists, create a temporary object with user data
    if not participant_biodata:
        class ParticipantData:
            def __init__(self, user):
                self.id = user.id
                self.nama_lengkap = user.nama_lengkap or user.username
                self.email = user.email
                self.asal_gudep = ''
                self.golongan = 'N/A'
                self.tingkatan = 'N/A'
                self.usia = user.usia or '0'
                self.foto = user.foto or 'img/default-user.png'
        
        participant_biodata = ParticipantData(participant_user)
    
    # Ambil SEMUA kriteria untuk event ini
    all_criterias = Criteria.query.filter_by(event_id=event_id).all()
    
    # Identifikasi kriteria yang ditugaskan ke user ini
    assigned_criteria_ids = [c.id_kriteria for c in current_user.assigned_criteria if c.event_id == event_id]
    
    # Ambil himpunan kriteria untuk dropdown
    for c in all_criterias:
        c.himpunan = HimpunanKriteria.query.filter_by(id_kriteria=c.id_kriteria).all()
        
    # Ambil SEMUA nilai yang sudah ada untuk peserta ini (dari penilai manapun)
    existing_scores = {}
    scores_query = Penilaian.query.filter_by(
        id_users=participant_id
    ).all()
    
    # Mapping nilai: Prioritaskan nilai dari current_user jika ada, jika tidak pakai nilai orang lain
    # (Dalam sistem ideal, mungkin kita ingin menampilkan siapa yang menilai, tapi untuk sekarang kita ambil nilai 'terbaru' atau 'milik sendiri')
    for s in scores_query:
        # Jika belum ada di map, masukkan
        if s.id_kriteria not in existing_scores:
            existing_scores[s.id_kriteria] = s.nilai
        # Jika sudah ada, tapi ini punya current_user, timpa (karena kita ingin lihat nilai kita sendiri jika ada)
        elif s.evaluator_id == current_user.id:
            existing_scores[s.id_kriteria] = s.nilai
            
    # DEBUG LOGGING TO FILE
    with open('debug_scores.log', 'a') as f:
        f.write(f"\n--- DEBUG: Viewing scores for Event {event_id}, Participant {participant_id} ---\n")
        f.write(f"Found {len(all_criterias)} criteria\n")
        f.write(f"Found {len(scores_query)} raw scores\n")
        f.write(f"Existing Scores Map: {existing_scores}\n")
        for c in all_criterias:
            f.write(f"Criteria {c.id_kriteria} ({c.nama_kriteria}) - Score: {existing_scores.get(c.id_kriteria)}\n")

    sidebar_state = current_user.sidebar_state or 'expanded'

    return render_template('penilai/view_penilaian.html', event=event, participant=participant_biodata, participant_user=participant_user, criterias=all_criterias, assigned_criteria_ids=assigned_criteria_ids, existing_scores=existing_scores, sidebar_state=sidebar_state)

@app.route('/penilai/biodata', methods=['GET', 'POST'])
@login_required
def penilai_biodata():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            nama_lengkap = request.form.get('nama_lengkap', '').strip()
            usia = request.form.get('usia', '0').strip()
            jenis_kelamin = request.form.get('jenis_kelamin', '').strip()
            nomor_hp = request.form.get('nomor_hp', '').strip()
            
            # Update Users table
            current_user.nama_lengkap = nama_lengkap
            current_user.usia = usia
            current_user.jenis_kelamin = jenis_kelamin
            current_user.nomor_hp = nomor_hp
            
            # Handle photo upload if any
            if 'foto' in request.files:
                file = request.files['foto']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Rename file to avoid conflict
                    ext = filename.rsplit('.', 1)[1].lower()
                    new_filename = f"{current_user.username}_{int(time.time())}.{ext}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], new_filename))
                    current_user.foto = f"img/{new_filename}"
            
            db.session.commit()
            flash(f"{t['profile_update_success']}", "success")
            return redirect(url_for('penilai_biodata'))
            
        except Exception as e:
            db.session.rollback()
            logging.error(f"Error updating evaluator data: {e}")
            flash(f"{t['profile_update_error']}", "danger")
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/biodata.html', sidebar_state=sidebar_state, user=current_user)

@app.route('/penilai/hasil-penilaian')
@login_required
def penilai_hasil_penilaian():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    # Get all assigned events
    assigned_events = Event.query.filter(Event.evaluators.any(id=current_user.id)).all()
    
    # Get selected event from query parameter
    selected_event_id = request.args.get('event_id', type=int)
    selected_event = None
    results = []
    
    if selected_event_id:
        # Verify the event is assigned to this evaluator
        selected_event = Event.query.get(selected_event_id)
        if selected_event and selected_event in assigned_events:
            from app.fuzzy_ahp import calculate_spk
            
            # Calculate SPK for this event
            success, msg = calculate_spk(selected_event_id)
            if not success:
                logging.warning(f"Gagal hitung SPK untuk event {selected_event_id}: {msg}")
            else:
                # Buat notifikasi untuk admin saat hasil seleksi selesai
                create_notification_to_all_admins(
                    f"Hasil seleksi selesai dihitung untuk kegiatan: {selected_event.nama_kegiatan}"
                )
            
            # Fetch results for this event only
            hasil_seleksi = db.session.query(
                HasilSeleksi,
                Users,
                Participants
            ).join(
                Users, HasilSeleksi.id_users == Users.id
            ).outerjoin(
                Participants, Users.email == Participants.email
            ).filter(
                HasilSeleksi.event_id == selected_event_id
            ).order_by(
                HasilSeleksi.ranking.asc()
            ).all()
            
            # Build results list
            for hasil, user, participant in hasil_seleksi:
                results.append({
                    'hasil': hasil,
                    'user': user,
                    'participant': participant
                })
        else:
            flash(f"{t['event_not_found_or_no_access']}", "error")
            selected_event = None
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/hasil_penilaian.html', assigned_events=assigned_events, selected_event=selected_event, results=results, sidebar_state=sidebar_state)

@app.route('/penilai/detail-nilai/<int:user_id>/<int:event_id>')
@login_required
def penilai_detail_nilai(user_id, event_id):
    from app.ahp_calculator import AHPCalculator, FuzzyAHPCalculator, TFN_SCALE, RI_TABLE, get_tfn_reciprocal
    from app.fuzzy_ahp import get_pairwise_matrix_from_db
    import numpy as np
    
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    # Verify event is assigned to this evaluator
    event = Event.query.get_or_404(event_id)
    if current_user not in event.evaluators:
        flash(f"{t['event_access_denied']}", "error")
        return redirect(url_for('penilai_hasil_penilaian'))
    
    # Get participant info
    user = Users.query.get_or_404(user_id)
    participant = Participants.query.filter_by(email=user.email).first()
    
    # Get final result
    hasil_seleksi = HasilSeleksi.query.filter_by(
        id_users=user_id,
        event_id=event_id
    ).first()
    
    # Get all criteria for this event (ordered)
    criterias = Criteria.query.filter_by(event_id=event_id).order_by(Criteria.id_kriteria).all()
    criteria_names = [c.nama_kriteria for c in criterias]
    criteria_ids = [c.id_kriteria for c in criterias]
    n = len(criterias)
    
    # ==========================================
    # LANGKAH 1: Penyusunan Matriks Perbandingan Berpasangan (Crisp)
    # ==========================================
    
    # Helper function to format TFN values as fractions
    def format_tfn_val(v):
        if abs(v - round(v)) < 0.01:  # It's essentially an integer
            return f"{int(round(v))}"
        # Check for common reciprocals 1/2 to 1/9
        for x in range(2, 10):
            if abs(v - 1/x) < 0.02:
                return f"1/{x}"
        # Check for values like 3/2, 5/2, 7/2, 9/2
        for num in [3, 5, 7, 9]:
            if abs(v - num/2) < 0.02:
                return f"{num}/2"
        # Check for 2/3
        if abs(v - 2/3) < 0.02:
            return f"2/3"
        return f"{v:.2f}"

    def format_tfn_tuple(tfn_tuple):
        l, m, u = tfn_tuple
        return f"({format_tfn_val(l)}, {format_tfn_val(m)}, {format_tfn_val(u)})"

    tfn_scale_table = []
    for intensity in range(1, 10):
        tfn = TFN_SCALE.get(intensity, (1, 1, 1))
        reciprocal = get_tfn_reciprocal(tfn)
        tfn_scale_table.append({
            'intensity': intensity,
            'tfn': format_tfn_tuple(tfn),
            'reciprocal': format_tfn_tuple(reciprocal)
        })
    
    # Get pairwise comparison matrix from database
    pairwise_matrix = get_pairwise_matrix_from_db(event_id, criteria_ids)
    pairwise_data = None
    fuzzy_pairwise_data = None
    use_generated_matrix = False
    fuzzy_ahp_calc = None
    ahp_calc = None
    
    # Jika tidak ada matriks di database tapi ada bobot kriteria, generate matriks dari bobot
    if pairwise_matrix is None and n > 0:
        total_bobot_check = sum(c.bobot for c in criterias)
        if total_bobot_check > 0:
            # Generate matriks perbandingan dari rasio bobot
            pairwise_matrix = np.ones((n, n))
            for i in range(n):
                for j in range(n):
                    if i != j:
                        wi = criterias[i].bobot if criterias[i].bobot > 0 else 0.001
                        wj = criterias[j].bobot if criterias[j].bobot > 0 else 0.001
                        ratio = wi / wj
                        if ratio >= 1:
                            pairwise_matrix[i, j] = min(9, max(1, ratio))
                        else:
                            pairwise_matrix[i, j] = max(1/9, ratio)
            use_generated_matrix = True
    
    if pairwise_matrix is not None and n > 0:
        pairwise_data = pairwise_matrix.tolist()
        
        # Convert to fuzzy (TFN) matrix with formatted strings
        fuzzy_ahp_calc = FuzzyAHPCalculator(criteria_names)
        fuzzy_ahp_calc.set_fuzzy_pairwise_matrix(pairwise_matrix)
        fuzzy_pairwise_data = []
        for i in range(n):
            row = []
            for j in range(n):
                tfn_tuple = tuple(fuzzy_ahp_calc.fuzzy_pairwise_matrix[i, j])
                row.append(format_tfn_tuple(tfn_tuple))
            fuzzy_pairwise_data.append(row)
    
    # ==========================================
    # LANGKAH 2: Perhitungan Vector Eigen
    # ==========================================
    eigenvector_data = None
    lambda_max = None
    
    if pairwise_matrix is not None and n > 0:
        ahp_calc = AHPCalculator(criteria_names)
        ahp_calc.set_pairwise_matrix(pairwise_matrix)
        eigenvector = ahp_calc.calculate_eigenvector()
        lambda_max = ahp_calc.calculate_lambda_max()
        
        eigenvector_data = []
        for i, name in enumerate(criteria_names):
            eigenvector_data.append({
                'criteria': name,
                'value': float(eigenvector[i])
            })
    
    # ==========================================
    # LANGKAH 3: Uji Konsistensi Matriks
    # ==========================================
    ci = None
    cr = None
    is_consistent = False
    ri_value = RI_TABLE.get(n, 1.58) if n > 0 else 0
    
    if pairwise_matrix is not None and n > 1 and ahp_calc is not None:
        ci, cr, is_consistent = ahp_calc.check_consistency()
    
    # ==========================================
    # LANGKAH 4: Sintesis Fuzzy (Fuzzy Synthetic Extent)
    # ==========================================
    fuzzy_synthetic_extent = None
    row_sums_data = None
    total_fuzzy_sum = None
    
    if pairwise_matrix is not None and n > 0 and fuzzy_ahp_calc is not None:
        synthetic_extents = fuzzy_ahp_calc.calculate_fuzzy_synthetic_extent()
        
        # Get row sums for display
        row_sums_data = []
        total_l, total_m, total_u = 0, 0, 0
        for i in range(n):
            l_sum, m_sum, u_sum = 0, 0, 0
            for j in range(n):
                l, m, u = fuzzy_ahp_calc.fuzzy_pairwise_matrix[i, j]
                l_sum += l
                m_sum += m
                u_sum += u
            row_sums_data.append({
                'criteria': criteria_names[i],
                'l': l_sum,
                'm': m_sum,
                'u': u_sum
            })
            total_l += l_sum
            total_m += m_sum
            total_u += u_sum
        
        total_fuzzy_sum = {'l': total_l, 'm': total_m, 'u': total_u}
        
        fuzzy_synthetic_extent = []
        for i, name in enumerate(criteria_names):
            si = synthetic_extents[i]
            fuzzy_synthetic_extent.append({
                'criteria': name,
                'l': si[0],
                'm': si[1],
                'u': si[2]
            })
    
    # ==========================================
    # LANGKAH 5: Perbandingan Probabilitas V(M2 >= M1)
    # ==========================================
    probability_matrix = None
    d_prime_values = None
    
    if pairwise_matrix is not None and n > 1 and fuzzy_ahp_calc is not None:
        synthetic_extents = fuzzy_ahp_calc.fuzzy_synthetic_extent
        
        # Build probability comparison matrix
        probability_matrix = []
        for i in range(n):
            row = []
            for j in range(n):
                if i == j:
                    row.append('-')
                else:
                    prob = fuzzy_ahp_calc.compare_fuzzy_probability(synthetic_extents[j], synthetic_extents[i])
                    row.append(round(prob, 4))
            probability_matrix.append(row)
        
        # Calculate d'(Ai) = min V(Si >= Sk) for all k != i
        d_prime_values = []
        for i in range(n):
            min_prob = float('inf')
            for j in range(n):
                if i != j:
                    prob = fuzzy_ahp_calc.compare_fuzzy_probability(synthetic_extents[j], synthetic_extents[i])
                    min_prob = min(min_prob, prob)
            d_prime_values.append({
                'criteria': criteria_names[i],
                'value': round(min_prob, 4) if min_prob != float('inf') else 0
            })
    
    # ==========================================
    # LANGKAH 6: Normalisasi & Perhitungan Bobot Global
    # ==========================================
    normalized_weights = None
    
    if pairwise_matrix is not None and n > 0 and fuzzy_ahp_calc is not None:
        weights = fuzzy_ahp_calc.calculate_fuzzy_weights()
        normalized_weights = []
        for name in criteria_names:
            normalized_weights.append({
                'criteria': name,
                'weight': round(weights.get(name, 0), 4)
            })
    
    # ==========================================
    # Calculate total weight and breakdown per participant
    # ==========================================
    # Use Fuzzy AHP weights from Step 6 if available, otherwise fall back to database weights
    fuzzy_ahp_weight_map = {}
    if normalized_weights:
        for item in normalized_weights:
            fuzzy_ahp_weight_map[item['criteria']] = item['weight']
    
    # Fallback: use database weights if Fuzzy AHP weights not available
    if not fuzzy_ahp_weight_map:
        total_bobot = sum(c.bobot for c in criterias)
        for criteria in criterias:
            fuzzy_ahp_weight_map[criteria.nama_kriteria] = (criteria.bobot / total_bobot) if total_bobot > 0 else 0
    
    calculation_details = []
    fuzzy_total_l = 0
    fuzzy_total_m = 0
    fuzzy_total_u = 0
    
    for criteria in criterias:
        # Use Fuzzy AHP weight (consistent with Step 6 Bobot Global)
        weight = fuzzy_ahp_weight_map.get(criteria.nama_kriteria, 0)
        
        # Get average score from all evaluators
        avg_score = db.session.query(db.func.avg(Penilaian.nilai)).filter_by(
            id_users=user_id,
            id_kriteria=criteria.id_kriteria
        ).scalar()
        
        if avg_score is not None:
            score = float(avg_score)
            
            # Fuzzification logic (same as fuzzy_ahp.py)
            if score <= 5:  # Likert scale
                if score <= 1:
                    l, m, u = 1, 1, 2
                elif score <= 2:
                    l, m, u = 1, 2, 3
                elif score <= 3:
                    l, m, u = 2, 3, 4
                elif score <= 4:
                    l, m, u = 3, 4, 5
                else:
                    l, m, u = 4, 5, 5
            else:  # 0-100 scale
                l = max(0, score - 5)
                m = score
                u = min(100, score + 5)
            
            # Weighted fuzzy values
            weighted_l = l * weight
            weighted_m = m * weight
            weighted_u = u * weight
            
            # Accumulate totals
            fuzzy_total_l += weighted_l
            fuzzy_total_m += weighted_m
            fuzzy_total_u += weighted_u
            
            calculation_details.append({
                'criteria': criteria,
                'weight': weight,
                'raw_score': score,
                'fuzzy_l': l,
                'fuzzy_m': m,
                'fuzzy_u': u,
                'weighted_l': weighted_l,
                'weighted_m': weighted_m,
                'weighted_u': weighted_u
            })
    
    # Final defuzzified score
    final_score = (fuzzy_total_l + fuzzy_total_m + fuzzy_total_u) / 3 if calculation_details else 0
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/detail_nilai.html', 
        user=user, 
        participant=participant, 
        event=event, 
        hasil_seleksi=hasil_seleksi, 
        calculation_details=calculation_details, 
        fuzzy_total_l=fuzzy_total_l, 
        fuzzy_total_m=fuzzy_total_m, 
        fuzzy_total_u=fuzzy_total_u, 
        final_score=final_score, 
        sidebar_state=sidebar_state,
        # Fuzzy AHP Step Data
        criteria_names=criteria_names,
        tfn_scale_table=tfn_scale_table,
        pairwise_data=pairwise_data,
        fuzzy_pairwise_data=fuzzy_pairwise_data,
        use_generated_matrix=use_generated_matrix,
        eigenvector_data=eigenvector_data,
        lambda_max=lambda_max,
        ci=ci,
        cr=cr,
        is_consistent=is_consistent,
        ri_value=ri_value,
        ri_table=RI_TABLE,
        row_sums_data=row_sums_data,
        total_fuzzy_sum=total_fuzzy_sum,
        fuzzy_synthetic_extent=fuzzy_synthetic_extent,
        probability_matrix=probability_matrix,
        d_prime_values=d_prime_values,
        normalized_weights=normalized_weights
    )

@app.route('/admin/hasil-penilaian')
@login_required
@admin_required
def admin_hasil_penilaian():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    # Get all events
    all_events = Event.query.order_by(Event.waktu_pelaksanaan_dimulai.desc()).all()
    
    # Get selected event from query parameter
    selected_event_id = request.args.get('event_id', type=int)
    selected_event = None
    results = []
    
    if selected_event_id:
        selected_event = Event.query.get(selected_event_id)
        if selected_event:
            from app.fuzzy_ahp import calculate_spk
            
            # Calculate SPK for this event
            success, msg = calculate_spk(selected_event_id)
            if not success:
                logging.warning(f"Gagal hitung SPK untuk event {selected_event_id}: {msg}")
            else:
                # Buat notifikasi untuk admin saat hasil seleksi selesai
                create_notification_to_all_admins(
                    f"Hasil seleksi selesai dihitung untuk kegiatan: {selected_event.nama_kegiatan}"
                )
            
            # Fetch results for this event only
            hasil_seleksi = db.session.query(
                HasilSeleksi,
                Users,
                Participants
            ).join(
                Users, HasilSeleksi.id_users == Users.id
            ).outerjoin(
                Participants, Users.email == Participants.email
            ).filter(
                HasilSeleksi.event_id == selected_event_id
            ).order_by(
                HasilSeleksi.ranking.asc()
            ).all()
            
            # Build results list
            for hasil, user, participant in hasil_seleksi:
                results.append({
                    'hasil': hasil,
                    'user': user,
                    'participant': participant
                })
        else:
            flash(f"{t['event_not_found']}", "error")
            selected_event = None
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('admin/hasil_penilaian.html', assigned_events=all_events, selected_event=selected_event, results=results, sidebar_state=sidebar_state, show_back_button=True)  # Tampilkan tombol kembali karena dibuka dari manajemen seleksi)

@app.route('/admin/hasil-penilaian/rekap/<int:event_id>')
@login_required
@admin_required
def admin_rekap_nilai_fuzzy(event_id):
    from app.ahp_calculator import AHPCalculator, FuzzyAHPCalculator, TFN_SCALE, RI_TABLE, get_tfn_reciprocal
    from app.fuzzy_ahp import get_pairwise_matrix_from_db, fuzzify_score
    import numpy as np

    
    # Get Event
    event = Event.query.get_or_404(event_id)
    
    # Get Criteria
    criterias = Criteria.query.filter_by(event_id=event_id).order_by(Criteria.id_kriteria).all()
    # Normalize weights
    total_bobot = sum(c.bobot for c in criterias)
    criteria_weights = {c.id_kriteria: (c.bobot / total_bobot if total_bobot > 0 else 0) for c in criterias}
    
    # Get Quota
    kuota = Kuota.query.filter_by(event_id=event_id).first()
    total_kuota = (kuota.putra + kuota.putri) if kuota else 0
    if total_kuota == 0:
        total_kuota = 9999 # Default all if no quota set

    # Get Criteria Names and IDs for Fuzzy Calculation
    criteria_names = [c.nama_kriteria for c in criterias]
    criteria_ids = [c.id_kriteria for c in criterias]
    n = len(criterias)

    # ==========================================
    # LANGKAH 1: Fuzzifikasi Matriks Perbandingan Berpasangan
    # ==========================================
    
    # Helper function to format TFN values as fractions
    def format_tfn_val(v):
        if abs(v - round(v)) < 0.01:  # It's essentially an integer
            return f"{int(round(v))}"
        # Check for common reciprocals 1/2 to 1/9
        for x in range(2, 10):
            if abs(v - 1/x) < 0.02:
                return f"1/{x}"
        # Check for values like 3/2, 5/2, 7/2, 9/2
        for num in [3, 5, 7, 9]:
            if abs(v - num/2) < 0.02:
                return f"{num}/2"
        # Check for 2/3
        if abs(v - 2/3) < 0.02:
            return f"2/3"
        return f"{v:.2f}"

    def format_tfn_tuple(tfn_tuple):
        l, m, u = tfn_tuple
        return f"({format_tfn_val(l)}, {format_tfn_val(m)}, {format_tfn_val(u)})"

    tfn_scale_table = []
    for intensity in range(1, 10):
        tfn = TFN_SCALE.get(intensity, (1, 1, 1))
        reciprocal = get_tfn_reciprocal(tfn)
        tfn_scale_table.append({
            'intensity': intensity,
            'tfn': format_tfn_tuple(tfn),
            'reciprocal': format_tfn_tuple(reciprocal)
        })

    # Get pairwise comparison matrix from database
    pairwise_matrix = get_pairwise_matrix_from_db(event_id, criteria_ids)
    pairwise_data = None
    fuzzy_pairwise_data = None
    use_generated_matrix = False

    # Jika tidak ada matriks di database tapi ada bobot kriteria, generate matriks dari bobot
    if pairwise_matrix is None and n > 0:
        total_bobot_check = sum(c.bobot for c in criterias)
        if total_bobot_check > 0:
            # Generate matriks perbandingan dari rasio bobot
            pairwise_matrix = np.ones((n, n))
            for i in range(n):
                for j in range(n):
                    if i != j:
                        # Rasio bobot sebagai nilai perbandingan
                        wi = criterias[i].bobot if criterias[i].bobot > 0 else 0.001
                        wj = criterias[j].bobot if criterias[j].bobot > 0 else 0.001
                        ratio = wi / wj
                        # Batasi ke skala 1-9
                        if ratio >= 1:
                            pairwise_matrix[i, j] = min(9, max(1, ratio))
                        else:
                            pairwise_matrix[i, j] = max(1/9, ratio)
            use_generated_matrix = True

    if pairwise_matrix is not None and n > 0:
        pairwise_data = pairwise_matrix.tolist()
        
        # Convert to fuzzy (TFN) matrix with formatted strings
        fuzzy_ahp_calc = FuzzyAHPCalculator(criteria_names)
        fuzzy_ahp_calc.set_fuzzy_pairwise_matrix(pairwise_matrix)
        fuzzy_pairwise_data = []

        for i in range(n):
            row = []
            for j in range(n):
                tfn_tuple = tuple(fuzzy_ahp_calc.fuzzy_pairwise_matrix[i, j])
                row.append(format_tfn_tuple(tfn_tuple))
            fuzzy_pairwise_data.append(row)

    # ==========================================
    # LANGKAH 2: Perhitungan Vector Eigen
    # ==========================================
    eigenvector_data = None
    lambda_max = None

    if pairwise_matrix is not None and n > 0:
        ahp_calc = AHPCalculator(criteria_names)
        ahp_calc.set_pairwise_matrix(pairwise_matrix)
        eigenvector = ahp_calc.calculate_eigenvector()
        lambda_max = ahp_calc.calculate_lambda_max()
        
        eigenvector_data = []
        for i, name in enumerate(criteria_names):
            eigenvector_data.append({
                'criteria': name,
                'value': float(eigenvector[i])
            })

    # ==========================================
    # LANGKAH 3: Uji Konsistensi Matriks
    # ==========================================
    ci = None
    cr = None
    is_consistent = False
    ri_value = RI_TABLE.get(n, 1.58) if n > 0 else 0

    if pairwise_matrix is not None and n > 1:
        ci, cr, is_consistent = ahp_calc.check_consistency()

    # ==========================================
    # LANGKAH 4: Sintesis Fuzzy (Fuzzy Synthetic Extent)
    # ==========================================
    fuzzy_synthetic_extent = None
    row_sums_data = None
    total_fuzzy_sum = None

    if pairwise_matrix is not None and n > 0:
        synthetic_extents = fuzzy_ahp_calc.calculate_fuzzy_synthetic_extent()
        
        # Get row sums for display
        row_sums_data = []
        total_l, total_m, total_u = 0, 0, 0
        for i in range(n):
            l_sum, m_sum, u_sum = 0, 0, 0
            for j in range(n):
                l, m, u = fuzzy_ahp_calc.fuzzy_pairwise_matrix[i, j]
                l_sum += l
                m_sum += m
                u_sum += u
            row_sums_data.append({
                'criteria': criteria_names[i],
                'l': l_sum,
                'm': m_sum,
                'u': u_sum
            })
            total_l += l_sum
            total_m += m_sum
            total_u += u_sum
        
        total_fuzzy_sum = {'l': total_l, 'm': total_m, 'u': total_u}
        
        fuzzy_synthetic_extent = []
        for i, name in enumerate(criteria_names):
            si = synthetic_extents[i]
            fuzzy_synthetic_extent.append({
                'criteria': name,
                'l': si[0],
                'm': si[1],
                'u': si[2]
            })

    # ==========================================
    # LANGKAH 5: Perbandingan Probabilitas V(M2 >= M1)
    # ==========================================
    probability_matrix = None
    d_prime_values = None

    if pairwise_matrix is not None and n > 1:
        synthetic_extents = fuzzy_ahp_calc.fuzzy_synthetic_extent
        
        # Build probability comparison matrix
        probability_matrix = []
        for i in range(n):
            row = []
            for j in range(n):
                if i == j:
                    row.append('-')
                else:
                    # V(Si >= Sj)
                    prob = fuzzy_ahp_calc.compare_fuzzy_probability(synthetic_extents[j], synthetic_extents[i])
                    row.append(round(prob, 4))
            probability_matrix.append(row)
        
        # Calculate d'(Ai) = min V(Si >= Sk) for all k != i
        d_prime_values = []
        for i in range(n):
            min_prob = float('inf')
            for j in range(n):
                if i != j:
                    prob = fuzzy_ahp_calc.compare_fuzzy_probability(synthetic_extents[j], synthetic_extents[i])
                    min_prob = min(min_prob, prob)
            d_prime_values.append({
                'criteria': criteria_names[i],
                'value': round(min_prob, 4) if min_prob != float('inf') else 0
            })

    # ==========================================
    # LANGKAH 6: Normalisasi & Perhitungan Bobot Global
    # ==========================================
    normalized_weights = None

    if pairwise_matrix is not None and n > 0:
        weights = fuzzy_ahp_calc.calculate_fuzzy_weights()
        normalized_weights = []
        for name in criteria_names:
            normalized_weights.append({
                'criteria': name,
                'weight': round(weights.get(name, 0), 4)
            })
        
        # Update criteria_weights to use Fuzzy AHP weights instead of database weights
        fuzzy_weight_map = {name: round(weights.get(name, 0), 4) for name in criteria_names}
        criteria_weights = {}
        for c in criterias:
            criteria_weights[c.id_kriteria] = fuzzy_weight_map.get(c.nama_kriteria, 0)
    
    # Get Results (Users)
    hasil_seleksi = db.session.query(
        HasilSeleksi, Users, Participants
    ).join(
        Users, HasilSeleksi.id_users == Users.id
    ).outerjoin(
        Participants, Users.email == Participants.email
    ).filter(
        HasilSeleksi.event_id == event_id
    ).order_by(
        HasilSeleksi.ranking.asc()
    ).all()
    
    rekap_data = []
    
    for hasil, user, participant in hasil_seleksi:
        row = {
            'nama': user.nama_lengkap,
            'user_id': user.id,
            'foto': user.foto if user.foto else 'img/default-user.png',
            'asal_gudep': participant.asal_gudep if participant else '-',
            'criteria_values': {},
            'total_score': hasil.skor_akhir,
            'rank': hasil.ranking,
            'cluster': 1 if hasil.ranking <= total_kuota else 2,
            'status': 'Berhak' if hasil.ranking <= total_kuota else 'Tidak Berhak',
            'calculation_details': [],
            'fuzzy_total_l': 0,
            'fuzzy_total_m': 0,
            'fuzzy_total_u': 0,
            'final_score': 0
        }
        
        fuzzy_total_l = 0
        fuzzy_total_m = 0
        fuzzy_total_u = 0
        
        # Calculate per-criteria weighted score
        for c in criterias:
             # Get average score from all evaluators
            avg_score = db.session.query(db.func.avg(Penilaian.nilai)).filter_by(
                id_users=user.id,
                id_kriteria=c.id_kriteria
            ).scalar()
            
            val = 0.0
            weight = criteria_weights.get(c.id_kriteria, 0)
            
            if avg_score is not None:
                score = float(avg_score)
                # Fuzzify
                l, m, u = fuzzify_score(score)
                
                # Weighted fuzzy values
                weighted_l = l * weight
                weighted_m = m * weight
                weighted_u = u * weight
                
                # Accumulate totals
                fuzzy_total_l += weighted_l
                fuzzy_total_m += weighted_m
                fuzzy_total_u += weighted_u
                
                val = weight * ((l + m + u) / 3.0)
                
                row['calculation_details'].append({
                    'criteria_name': c.nama_kriteria,
                    'criteria_aspek': c.aspek if hasattr(c, 'aspek') else '',
                    'weight': weight,
                    'raw_score': score,
                    'fuzzy_l': l,
                    'fuzzy_m': m,
                    'fuzzy_u': u,
                    'weighted_l': weighted_l,
                    'weighted_m': weighted_m,
                    'weighted_u': weighted_u
                })
            else:
                row['calculation_details'].append({
                    'criteria_name': c.nama_kriteria,
                    'criteria_aspek': c.aspek if hasattr(c, 'aspek') else '',
                    'weight': weight,
                    'raw_score': 0,
                    'fuzzy_l': 0,
                    'fuzzy_m': 0,
                    'fuzzy_u': 0,
                    'weighted_l': 0,
                    'weighted_m': 0,
                    'weighted_u': 0
                })
            
            row['criteria_values'][c.id_kriteria] = val
        
        row['fuzzy_total_l'] = fuzzy_total_l
        row['fuzzy_total_m'] = fuzzy_total_m
        row['fuzzy_total_u'] = fuzzy_total_u
        row['final_score'] = (fuzzy_total_l + fuzzy_total_m + fuzzy_total_u) / 3 if row['calculation_details'] else 0
            
        rekap_data.append(row)
        
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template(
        'admin/rekap_nilai_fuzzy.html',
        event=event,
        criteria_headers=criterias,
        rekap_data=rekap_data,
        sidebar_state=sidebar_state,
        # Fuzzy AHP Step Data
        criteria_names=criteria_names,
        tfn_scale_table=tfn_scale_table,
        pairwise_data=pairwise_data,
        fuzzy_pairwise_data=fuzzy_pairwise_data,
        use_generated_matrix=use_generated_matrix,
        eigenvector_data=eigenvector_data,
        lambda_max=lambda_max,
        ci=ci,
        cr=cr,
        is_consistent=is_consistent,
        ri_value=ri_value,
        ri_table=RI_TABLE,
        row_sums_data=row_sums_data,
        total_fuzzy_sum=total_fuzzy_sum,
        fuzzy_synthetic_extent=fuzzy_synthetic_extent,
        probability_matrix=probability_matrix,
        d_prime_values=d_prime_values,
        normalized_weights=normalized_weights
    )

@app.route('/admin/detail-nilai/<int:user_id>/<int:event_id>')
@login_required
@admin_required
def admin_detail_nilai(user_id, event_id):
    from app.ahp_calculator import AHPCalculator, FuzzyAHPCalculator, TFN_SCALE, RI_TABLE, get_tfn_reciprocal
    from app.fuzzy_ahp import get_pairwise_matrix_from_db
    import numpy as np
    import json
    
    # Get event
    event = Event.query.get_or_404(event_id)
    
    # Get participant info
    user = Users.query.get_or_404(user_id)
    participant = Participants.query.filter_by(email=user.email).first()
    
    # Get final result
    hasil_seleksi = HasilSeleksi.query.filter_by(
        id_users=user_id,
        event_id=event_id
    ).first()
    
    # Get all criteria for this event (ordered)
    criterias = Criteria.query.filter_by(event_id=event_id).order_by(Criteria.id_kriteria).all()
    criteria_names = [c.nama_kriteria for c in criterias]
    criteria_ids = [c.id_kriteria for c in criterias]
    n = len(criterias)
    
    # ==========================================
    # LANGKAH 1: Fuzzifikasi Matriks Perbandingan Berpasangan
    # ==========================================
    
    # Helper function to format TFN values as fractions
    def format_tfn_val(v):
        if abs(v - round(v)) < 0.01:  # It's essentially an integer
            return f"{int(round(v))}"
        # Check for common reciprocals 1/2 to 1/9
        for x in range(2, 10):
            if abs(v - 1/x) < 0.02:
                return f"1/{x}"
        # Check for values like 3/2, 5/2, 7/2, 9/2
        for num in [3, 5, 7, 9]:
            if abs(v - num/2) < 0.02:
                return f"{num}/2"
        # Check for 2/3
        if abs(v - 2/3) < 0.02:
            return f"2/3"
        return f"{v:.2f}"

    def format_tfn_tuple(tfn_tuple):
        l, m, u = tfn_tuple
        return f"({format_tfn_val(l)}, {format_tfn_val(m)}, {format_tfn_val(u)})"

    tfn_scale_table = []
    for intensity in range(1, 10):
        tfn = TFN_SCALE.get(intensity, (1, 1, 1))
        reciprocal = get_tfn_reciprocal(tfn)
        tfn_scale_table.append({
            'intensity': intensity,
            'tfn': format_tfn_tuple(tfn),
            'reciprocal': format_tfn_tuple(reciprocal)
        })
    
    # Get pairwise comparison matrix from database
    pairwise_matrix = get_pairwise_matrix_from_db(event_id, criteria_ids)
    pairwise_data = None
    fuzzy_pairwise_data = None
    use_generated_matrix = False
    
    # Jika tidak ada matriks di database tapi ada bobot kriteria, generate matriks dari bobot
    if pairwise_matrix is None and n > 0:
        total_bobot_check = sum(c.bobot for c in criterias)
        if total_bobot_check > 0:
            # Generate matriks perbandingan dari rasio bobot
            pairwise_matrix = np.ones((n, n))
            for i in range(n):
                for j in range(n):
                    if i != j:
                        # Rasio bobot sebagai nilai perbandingan
                        wi = criterias[i].bobot if criterias[i].bobot > 0 else 0.001
                        wj = criterias[j].bobot if criterias[j].bobot > 0 else 0.001
                        ratio = wi / wj
                        # Batasi ke skala 1-9
                        if ratio >= 1:
                            pairwise_matrix[i, j] = min(9, max(1, ratio))
                        else:
                            pairwise_matrix[i, j] = max(1/9, ratio)
            use_generated_matrix = True
    
    if pairwise_matrix is not None and n > 0:
        pairwise_data = pairwise_matrix.tolist()
        
        # Convert to fuzzy (TFN) matrix with formatted strings
        fuzzy_ahp_calc = FuzzyAHPCalculator(criteria_names)
        fuzzy_ahp_calc.set_fuzzy_pairwise_matrix(pairwise_matrix)
        fuzzy_pairwise_data = []
        for i in range(n):
            row = []
            for j in range(n):
                tfn_tuple = tuple(fuzzy_ahp_calc.fuzzy_pairwise_matrix[i, j])
                row.append(format_tfn_tuple(tfn_tuple))
            fuzzy_pairwise_data.append(row)
    
    # ==========================================
    # LANGKAH 2: Perhitungan Vector Eigen
    # ==========================================
    eigenvector_data = None
    lambda_max = None
    
    if pairwise_matrix is not None and n > 0:
        ahp_calc = AHPCalculator(criteria_names)
        ahp_calc.set_pairwise_matrix(pairwise_matrix)
        eigenvector = ahp_calc.calculate_eigenvector()
        lambda_max = ahp_calc.calculate_lambda_max()
        
        eigenvector_data = []
        for i, name in enumerate(criteria_names):
            eigenvector_data.append({
                'criteria': name,
                'value': float(eigenvector[i])
            })
    
    # ==========================================
    # LANGKAH 3: Uji Konsistensi Matriks
    # ==========================================
    ci = None
    cr = None
    is_consistent = False
    ri_value = RI_TABLE.get(n, 1.58) if n > 0 else 0
    
    if pairwise_matrix is not None and n > 1:
        ci, cr, is_consistent = ahp_calc.check_consistency()
    
    # ==========================================
    # LANGKAH 4: Sintesis Fuzzy (Fuzzy Synthetic Extent)
    # ==========================================
    fuzzy_synthetic_extent = None
    row_sums_data = None
    total_fuzzy_sum = None
    
    if pairwise_matrix is not None and n > 0:
        synthetic_extents = fuzzy_ahp_calc.calculate_fuzzy_synthetic_extent()
        
        # Get row sums for display
        row_sums_data = []
        total_l, total_m, total_u = 0, 0, 0
        for i in range(n):
            l_sum, m_sum, u_sum = 0, 0, 0
            for j in range(n):
                l, m, u = fuzzy_ahp_calc.fuzzy_pairwise_matrix[i, j]
                l_sum += l
                m_sum += m
                u_sum += u
            row_sums_data.append({
                'criteria': criteria_names[i],
                'l': l_sum,
                'm': m_sum,
                'u': u_sum
            })
            total_l += l_sum
            total_m += m_sum
            total_u += u_sum
        
        total_fuzzy_sum = {'l': total_l, 'm': total_m, 'u': total_u}
        
        fuzzy_synthetic_extent = []
        for i, name in enumerate(criteria_names):
            si = synthetic_extents[i]
            fuzzy_synthetic_extent.append({
                'criteria': name,
                'l': si[0],
                'm': si[1],
                'u': si[2]
            })
    
    # ==========================================
    # LANGKAH 5: Perbandingan Probabilitas V(M2 >= M1)
    # ==========================================
    probability_matrix = None
    d_prime_values = None
    
    if pairwise_matrix is not None and n > 1:
        synthetic_extents = fuzzy_ahp_calc.fuzzy_synthetic_extent
        
        # Build probability comparison matrix
        probability_matrix = []
        for i in range(n):
            row = []
            for j in range(n):
                if i == j:
                    row.append('-')
                else:
                    # V(Si >= Sj)
                    prob = fuzzy_ahp_calc.compare_fuzzy_probability(synthetic_extents[j], synthetic_extents[i])
                    row.append(round(prob, 4))
            probability_matrix.append(row)
        
        # Calculate d'(Ai) = min V(Si >= Sk) for all k != i
        d_prime_values = []
        for i in range(n):
            min_prob = float('inf')
            for j in range(n):
                if i != j:
                    prob = fuzzy_ahp_calc.compare_fuzzy_probability(synthetic_extents[j], synthetic_extents[i])
                    min_prob = min(min_prob, prob)
            d_prime_values.append({
                'criteria': criteria_names[i],
                'value': round(min_prob, 4) if min_prob != float('inf') else 0
            })
    
    # ==========================================
    # LANGKAH 6: Normalisasi & Perhitungan Bobot Global
    # ==========================================
    normalized_weights = None
    
    if pairwise_matrix is not None and n > 0:
        weights = fuzzy_ahp_calc.calculate_fuzzy_weights()
        normalized_weights = []
        for name in criteria_names:
            normalized_weights.append({
                'criteria': name,
                'weight': round(weights.get(name, 0), 4)
            })
    
    # ==========================================
    # Calculate total weight and breakdown per participant
    # ==========================================
    # Use Fuzzy AHP weights from Step 6 if available, otherwise fall back to database weights
    fuzzy_ahp_weight_map = {}
    if normalized_weights:
        for item in normalized_weights:
            fuzzy_ahp_weight_map[item['criteria']] = item['weight']
    
    # Fallback: use database weights if Fuzzy AHP weights not available
    if not fuzzy_ahp_weight_map:
        total_bobot = sum(c.bobot for c in criterias)
        for criteria in criterias:
            fuzzy_ahp_weight_map[criteria.nama_kriteria] = (criteria.bobot / total_bobot) if total_bobot > 0 else 0
    
    calculation_details = []
    fuzzy_total_l = 0
    fuzzy_total_m = 0
    fuzzy_total_u = 0
    
    for criteria in criterias:
        # Use Fuzzy AHP weight (consistent with Step 6 Bobot Global)
        weight = fuzzy_ahp_weight_map.get(criteria.nama_kriteria, 0)
        
        # Get average score from all evaluators
        avg_score = db.session.query(db.func.avg(Penilaian.nilai)).filter_by(
            id_users=user_id,
            id_kriteria=criteria.id_kriteria
        ).scalar()
        
        if avg_score is not None:
            score = float(avg_score)
            
            # Fuzzification logic (same as fuzzy_ahp.py)
            if score <= 5:  # Likert scale
                if score <= 1:
                    l, m, u = 1, 1, 2
                elif score <= 2:
                    l, m, u = 1, 2, 3
                elif score <= 3:
                    l, m, u = 2, 3, 4
                elif score <= 4:
                    l, m, u = 3, 4, 5
                else:
                    l, m, u = 4, 5, 5
            else:  # 0-100 scale
                l = max(0, score - 5)
                m = score
                u = min(100, score + 5)
            
            # Weighted fuzzy values
            weighted_l = l * weight
            weighted_m = m * weight
            weighted_u = u * weight
            
            # Accumulate totals
            fuzzy_total_l += weighted_l
            fuzzy_total_m += weighted_m
            fuzzy_total_u += weighted_u
            
            calculation_details.append({
                'criteria': criteria,
                'weight': weight,
                'raw_score': score,
                'fuzzy_l': l,
                'fuzzy_m': m,
                'fuzzy_u': u,
                'weighted_l': weighted_l,
                'weighted_m': weighted_m,
                'weighted_u': weighted_u
            })
    
    # Final defuzzified score
    final_score = (fuzzy_total_l + fuzzy_total_m + fuzzy_total_u) / 3 if calculation_details else 0
    sidebar_state = current_user.sidebar_state or 'expanded'
    
    return render_template('admin/detail_nilai.html', 
        user=user, 
        participant=participant, 
        event=event, 
        hasil_seleksi=hasil_seleksi, 
        calculation_details=calculation_details, 
        fuzzy_total_l=fuzzy_total_l, 
        fuzzy_total_m=fuzzy_total_m, 
        fuzzy_total_u=fuzzy_total_u, 
        final_score=final_score, 
        sidebar_state=sidebar_state,
        # Fuzzy AHP Step Data
        criteria_names=criteria_names,
        tfn_scale_table=tfn_scale_table,
        pairwise_data=pairwise_data,
        fuzzy_pairwise_data=fuzzy_pairwise_data,
        use_generated_matrix=use_generated_matrix,
        eigenvector_data=eigenvector_data,
        lambda_max=lambda_max,
        ci=ci,
        cr=cr,
        is_consistent=is_consistent,
        ri_value=ri_value,
        ri_table=RI_TABLE,
        row_sums_data=row_sums_data,
        total_fuzzy_sum=total_fuzzy_sum,
        fuzzy_synthetic_extent=fuzzy_synthetic_extent,
        probability_matrix=probability_matrix,
        d_prime_values=d_prime_values,
        normalized_weights=normalized_weights
    )

@app.route('/penilai/hasil-seleksi')
@login_required
def penilai_hasil_seleksi():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    # Get all assigned events
    assigned_events = Event.query.filter(Event.evaluators.any(id=current_user.id)).all()
    
    # Get selected event from query parameter
    selected_event_id = request.args.get('event_id', type=int)
    selected_event = None
    results = []
    
    if selected_event_id:
        # Verify the event is assigned to this evaluator
        selected_event = Event.query.get(selected_event_id)
        if selected_event and selected_event in assigned_events:
            # Fetch results for this event only
            # Join with Participants to get gender (for quota check if needed in template)
            hasil_seleksi_query = db.session.query(
                HasilSeleksi,
                Users,
                Participants
            ).join(
                Users, HasilSeleksi.id_users == Users.id
            ).outerjoin(
                Participants, Users.email == Participants.email
            ).filter(
                HasilSeleksi.event_id == selected_event_id
            ).order_by(
                HasilSeleksi.ranking.asc()
            ).all()
            
            # Process results to include passing status logic explicitly if needed
            # Although template can do it, it's good to have it ready.
            # Using simple query logic for now.
            
            kuota = Kuota.query.filter_by(event_id=selected_event_id).first()
            
            for hasil, user, participant in hasil_seleksi_query:
                results.append({
                    'hasil': hasil,
                    'user': user,
                    'participant': participant
                })
        else:
            flash(f"{t['event_not_found_or_no_access']}", "error")
            selected_event = None
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/hasil_seleksi.html', assigned_events=assigned_events, selected_event=selected_event, results=results, sidebar_state=sidebar_state)

@app.route('/penilai/notifikasi')
@login_required
def penilai_notifikasi():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    
    # Get notifications for this evaluator
    notifications = Notification.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Notification.created_at.desc()
    ).all()
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('penilai/notifikasi.html', notifications=notifications, sidebar_state=sidebar_state)

@app.route('/penilai/settings')
@login_required
def penilai_settings():
    sidebar_state = current_user.sidebar_state or 'expanded'
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'penilai':
        flash(f"{t['evaluator_access_denied']}", "error")
        return redirect(url_for('index'))
    return render_template('settings.html', sidebar_state=sidebar_state, time=time)

@app.route('/peserta/dashboard')
@login_required
def peserta_dashboard():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    """Dashboard for participants showing scores and rankings for all registered activities"""
    
    if current_user.level != 'peserta':
        flash(f"{t['participant_access_denied']}", "error")
        return redirect(url_for('index'))
    
    # Get current user's participant record
    # Cek biodata
    participant = Participants.query.filter_by(email=current_user.email).first()
    
    # Flash warning jika biodata belum ada
    if not participant:
        msg = 'Biodata Anda belum terdaftar. Silakan lengkapi biodata terlebih dahulu.' if lang == 'id' else 'Your biodata is not registered yet. Please complete your biodata first.'
        # Cek apakah pesan sudah ada di flashed messages untuk menghindari duplikasi
        # Namun karena kita tidak bisa peek message dengan mudah, kita flash saja. 
        # Frontend sebaiknya handle deduplikasi atau kita asumsikan ini page load baru.
        flash(msg, 'warning')
        
    # Get all registered activities for this participant
    registered_activities = []
    if participant:
        registered_activities = Event.query.join(
            tb_participant_kegiatan,
            Event.id_kegiatan == tb_participant_kegiatan.c.kegiatan_id
        ).filter(
            tb_participant_kegiatan.c.participant_id == participant.id
        ).all()
    
    # Calculate scores for each activity
    activity_scores = []
    for event in registered_activities:
        # Get all criteria for this event
        criteria_list = Criteria.query.filter_by(event_id=event.id_kegiatan).all()
        
        # Calculate total score
        total_score = 0
        has_scores = False
        
        for criterion in criteria_list:
            penilaian = Penilaian.query.filter_by(
                id_users=current_user.id,
                id_kriteria=criterion.id_kriteria
            ).first()
            
            if penilaian:
                # Calculate weighted score
                weighted_score = penilaian.nilai * (criterion.bobot / 100)
                total_score += weighted_score
                has_scores = True
        
        # Get ranking from HasilSeleksi table
        hasil = HasilSeleksi.query.filter_by(
            id_users=current_user.id,
            event_id=event.id_kegiatan
        ).first()
        
        activity_scores.append({
            'event': event,
            'final_score': round(total_score, 2) if has_scores else None,
            'ranking': hasil.ranking if hasil else None,
            'has_scores': has_scores
        })
    
    # Check if any selection period has ended
    is_selection_ended = any(
        event.selesai and event.selesai < date.today()
        for event in registered_activities
    )
    
    # Determine status
    status_seleksi = 'Terdaftar' if registered_activities else 'Belum ada status'
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('peserta/dashboard.html', biodata=participant, registered_activities=registered_activities, activity_scores=activity_scores, is_selection_ended=is_selection_ended, status_seleksi=status_seleksi, user=current_user, sidebar_state=sidebar_state, today=date.today())

@app.route('/peserta/notifikasi')
@login_required
def peserta_notifikasi():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'peserta':
        flash(f"{t['participant_access_denied']}", "error")
        return redirect(url_for('index'))
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.id.desc()).all()
    return render_template('peserta/notifikasi.html', notifications=notifications, sidebar_state=sidebar_state, user=current_user)

@app.route('/peserta/settings')
@login_required
def peserta_settings():
    sidebar_state = current_user.sidebar_state or 'expanded'
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'peserta':
        flash(f"{t['participant_access_denied']}", "error")
        return redirect(url_for('index'))
    return render_template('settings.html', sidebar_state=sidebar_state, time=time)

@app.route('/peserta/hasil_seleksi')
@login_required
def peserta_hasil_seleksi():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'peserta':
        flash(f"{t['participant_access_denied']}", "error")
        return redirect(url_for('index'))
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    biodata = Participants.query.filter_by(email=current_user.email).first()
    
    results_data = []
    
    if biodata:
        # Get all registered activities via the many-to-many relationship
        # Assuming biodata.registered_activities is the relationship to Event
        registered_activities = biodata.registered_activities.all()
        
        for event in registered_activities:
            # Check for existing result
            hasil = HasilSeleksi.query.filter_by(
                id_users=current_user.id,
                event_id=event.id_kegiatan
            ).first()
            
            status_text = "Dalam Proses"
            temp_score = 0
            has_temp_score = False
            
            if hasil:
                    status_text = "Selesai"
            else:
                # If no final result, calculate temporary score from Penilaian
                criteria_list = Criteria.query.filter_by(event_id=event.id_kegiatan).all()
                current_score = 0
                count_rated = 0
                
                # Check directly in Penilaian table
                # We need to sum (nilai * bobot) / 100 or similar based on formula
                # Using simple weighted sum for display
                # Note: This is an approximation if the final formula is Fuzzy AHP
                # But good enough for "Temporary Score"
                
                # Retrieve all ratings for this user and event criterias
                if criteria_list:
                    criteria_ids = [c.id_kriteria for c in criteria_list]
                    ratings = Penilaian.query.filter(
                        Penilaian.id_users == current_user.id,
                        Penilaian.id_kriteria.in_(criteria_ids)
                    ).all()
                    rating_map = {r.id_kriteria: r.nilai for r in ratings}
                    total_bobot = sum(c.bobot for c in criteria_list)
                    if ratings:
                         has_temp_score = True
                         for c in criteria_list:
                             if c.id_kriteria in rating_map:
                                 # Normalize weight usually happens in calculation, 
                                 # here we assume simple weighted sum: value * (bobot/total_bobot)
                                 # or just value * bobot if bobot is percentage.
                                 # Let's align with dashboard logic: weighted_score = nilai * (bobot / 100)
                                 # Assuming bobot is 0-100.
                                 if total_bobot > 0:
                                     val = rating_map[c.id_kriteria]
                                     # Simple weighted average 
                                     # (value * weight) / total_weight
                                     # This keeps result in same scale as value (e.g. 1-100)
                                     current_score += val * (c.bobot / total_bobot)
                    temp_score = current_score
            results_data.append({
                'event': event,
                'hasil': hasil,
                'status_text': status_text,
                'temp_score': temp_score,
                'has_temp_score': has_temp_score
            })

    return render_template('peserta/hasil_seleksi.html', results_data=results_data, biodata=biodata, sidebar_state=sidebar_state, user=current_user)

@app.route('/peserta/hasil_seleksi/<int:event_id>')
@login_required
def peserta_detail_nilai(event_id):
    """Halaman detail nilai semua peserta untuk kegiatan tertentu"""
    if current_user.level != 'peserta':
        flash("Anda tidak memiliki akses ke halaman ini.", "error")
        return redirect(url_for('index'))
    
    # Verify that the current user is registered for this event
    biodata = Participants.query.filter_by(email=current_user.email).first()
    if not biodata:
        flash("Data peserta tidak ditemukan.", "error")
        return redirect(url_for('peserta_hasil_seleksi'))
    
    # Check if user is registered for this event
    event = Event.query.get_or_404(event_id)
    if event not in biodata.registered_activities.all():
        flash("Anda tidak terdaftar untuk kegiatan ini.", "error")
        return redirect(url_for('peserta_hasil_seleksi'))
    
    # Get all results for this event, ordered by ranking
    hasil_seleksi_query = db.session.query(
        HasilSeleksi,
        Users,
        Participants
    ).join(
        Users, HasilSeleksi.id_users == Users.id
    ).outerjoin(
        Participants, Users.email == Participants.email
    ).filter(
        HasilSeleksi.event_id == event_id
    ).order_by(
        HasilSeleksi.ranking.asc()
    ).all()
    
    # Process results
    results = []
    kuota = Kuota.query.filter_by(event_id=event_id).first()
    
    for hasil, user, participant in hasil_seleksi_query:
        results.append({
            'hasil': hasil,
            'user': user,
            'participant': participant
        })
    
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template(
        'peserta/detail_nilai.html',
        event=event,
        results=results,
        kuota=kuota,
        sidebar_state=sidebar_state,
        user=current_user,
        biodata=biodata
    )

@app.route('/peserta/data', methods=['GET', 'POST'])
@login_required
def peserta_data():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    if current_user.level != 'peserta':
        flash(f"{t['participant_access_denied']}", "error")
        return redirect(url_for('index'))

    # Ambil data biodata
    biodata = Participants.query.filter_by(email=current_user.email).first()
    
    if request.method == 'POST':
        try:
            nama_lengkap = request.form.get('nama_lengkap', '').strip()
            tanggal_lahir = request.form.get('tanggal_lahir', '').strip()
            alamat_tinggal = request.form.get('alamat_tinggal', '').strip()
            golongan = request.form.get('golongan', '').strip()
            tingkatan = request.form.get('tingkatan', '').strip()
            asal_gudep = request.form.get('asal_gudep', '').strip()
            asal_kwarran = request.form.get('asal_kwarran', '').strip()
            asal_kwarcab = request.form.get('asal_kwarcab', '').strip()
            asal_kwarda = request.form.get('asal_kwarda', '').strip()
            usia = request.form.get('usia', '0').strip()
            jenis_kelamin = request.form.get('jenis_kelamin', '').strip()
            nomor_hp = request.form.get('nomor_hp', '').strip()
            
            # Update Users table
            current_user.nama_lengkap = nama_lengkap
            current_user.usia = usia
            current_user.jenis_kelamin = jenis_kelamin
            current_user.nomor_hp = nomor_hp
            
            # Handle photo upload if any
            if 'foto' in request.files:
                file = request.files['foto']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Rename file to avoid conflict
                    ext = filename.rsplit('.', 1)[1].lower()
                    new_filename = f"{current_user.username}_{int(time.time())}.{ext}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], new_filename))
                    current_user.foto = f"uploads/{new_filename}"
                    if biodata:
                        biodata.foto = f"uploads/{new_filename}"

            if not biodata:
                # Create new biodata
                biodata = Participants(
                    nama_lengkap=nama_lengkap,
                    email=current_user.email,
                    tanggal_lahir=datetime.strptime(tanggal_lahir, '%Y-%m-%d').date() if tanggal_lahir else datetime.now().date(),
                    alamat_tinggal=alamat_tinggal,
                    golongan=golongan,
                    tingkatan=tingkatan,
                    asal_gudep=asal_gudep,
                    asal_kwarran=asal_kwarran,
                    asal_kwarcab=asal_kwarcab,
                    asal_kwarda=asal_kwarda,
                    usia=int(usia) if usia.isdigit() else 0,
                    jenis_kelamin=jenis_kelamin,
                    nomor_hp=nomor_hp,
                    foto=current_user.foto,
                    level='peserta'
                )
                db.session.add(biodata)
            else:
                # Update existing biodata
                biodata.nama_lengkap = nama_lengkap
                if tanggal_lahir:
                    biodata.tanggal_lahir = datetime.strptime(tanggal_lahir, '%Y-%m-%d').date()
                biodata.alamat_tinggal = alamat_tinggal
                biodata.golongan = golongan
                biodata.tingkatan = tingkatan
                biodata.asal_gudep = asal_gudep
                biodata.asal_kwarran = asal_kwarran
                biodata.asal_kwarcab = asal_kwarcab
                biodata.asal_kwarda = asal_kwarda
                biodata.usia = int(usia) if usia.isdigit() else 0
                biodata.jenis_kelamin = jenis_kelamin
                biodata.nomor_hp = nomor_hp
            
            db.session.commit()
            flash(f"{t['participant_updated']}", "success")
            return redirect(url_for('peserta_data'))
            
        except Exception as e:
            db.session.rollback()
            logging.error(f"Error updating participant data: {e}")
            flash(f"{t['save_error']}", "danger")

    # Check completeness
    is_complete = False
    missing_fields = []
    if biodata:
        required_fields = [
            'nama_lengkap', 'tanggal_lahir', 'alamat_tinggal', 'golongan', 
            'tingkatan', 'asal_gudep', 'asal_kwarran', 'asal_kwarcab', 
            'asal_kwarda', 'usia', 'jenis_kelamin', 'nomor_hp'
        ]
        for field in required_fields:
            val = getattr(biodata, field)
            if not val or val == '' or val == 0:
                missing_fields.append(field)
        
        if not missing_fields:
            is_complete = True
    else:
        missing_fields = ['All Data']
    sidebar_state = current_user.sidebar_state or 'expanded'
    return render_template('peserta/data_peserta.html', biodata=biodata, user=current_user, sidebar_state=sidebar_state, is_complete=is_complete, missing_fields=missing_fields)

# API untuk mendapatkan kegiatan/seleksi yang tersedia (sedang dibuka)
@app.route('/api/kegiatan_tersedia')
@login_required
def api_kegiatan_tersedia():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    """Mengembalikan daftar kegiatan yang sedang membuka seleksi (tanggal sekarang antara mulai dan selesai)"""
    try:
        if current_user.level != 'peserta':
            return jsonify({'status': 'error', 'message': t['access_denied']}), 403
        
        today = datetime.utcnow().date()
        
        # Ambil kegiatan yang sedang membuka seleksi (tanggal sekarang antara mulai dan selesai)
        kegiatan_list = Event.query.filter(
            Event.selesai >= today
        ).order_by(Event.mulai.asc()).all()
        
        # Ambil biodata peserta untuk cek apakah sudah terdaftar
        biodata = Participants.query.filter_by(email=current_user.email).first()
        peserta_kegiatan_ids = []
        if biodata:
            # Get all registered activities from many-to-many relationship
            peserta_kegiatan_ids = [k.id_kegiatan for k in biodata.registered_activities.all()]
        
        result = []
        for kegiatan in kegiatan_list:
            kuota = Kuota.query.filter_by(event_id=kegiatan.id_kegiatan).first()
            
            # Hitung jumlah peserta yang sudah terdaftar
            peserta_terdaftar = kegiatan.registered_participants.count()
            peserta_putra = kegiatan.registered_participants.filter(Participants.jenis_kelamin == 'laki-laki').count()
            peserta_putri = kegiatan.registered_participants.filter(Participants.jenis_kelamin == 'perempuan').count()
            
            # Cek apakah peserta sudah terdaftar di kegiatan ini
            sudah_terdaftar = kegiatan.id_kegiatan in peserta_kegiatan_ids
            
            # Cek apakah sudah dinilai
            is_graded = False
            if sudah_terdaftar:
                # Get criteria IDs for this event
                criteria_ids = [c.id_kriteria for c in Criteria.query.filter_by(event_id=kegiatan.id_kegiatan).all()]
                if criteria_ids:
                    # Check if any score exists
                    score_exists = Penilaian.query.filter(
                        Penilaian.id_users == current_user.id,
                        Penilaian.id_kriteria.in_(criteria_ids)
                    ).first()
                    if score_exists:
                        is_graded = True

            result.append({
                'id_kegiatan': kegiatan.id_kegiatan,
                'nama_kegiatan': kegiatan.nama_kegiatan,
                'jenis_kegiatan': kegiatan.jenis_kegiatan,
                'skala_kegiatan': kegiatan.skala_kegiatan,
                'kwartir_penyelenggara': kegiatan.kwartir_penyelenggara,
                'tempat_pelaksanaan': kegiatan.tempat_pelaksanaan,
                'waktu_pelaksanaan_dimulai': kegiatan.waktu_pelaksanaan_dimulai.strftime('%Y-%m-%d') if kegiatan.waktu_pelaksanaan_dimulai else None,
                'waktu_pelaksanaan_selesai': kegiatan.waktu_pelaksanaan_selesai.strftime('%Y-%m-%d') if kegiatan.waktu_pelaksanaan_selesai else None,
                'periode_seleksi_mulai': kegiatan.mulai.strftime('%Y-%m-%d') if kegiatan.mulai else None,
                'periode_seleksi_selesai': kegiatan.selesai.strftime('%Y-%m-%d') if kegiatan.selesai else None,
                'kuota_putra': kuota.putra if kuota else 0,
                'kuota_putri': kuota.putri if kuota else 0,
                'peserta_terdaftar': peserta_terdaftar,
                'peserta_putra_terdaftar': peserta_putra,
                'peserta_putri_terdaftar': peserta_putri,
                'sisa_kuota_putra': (kuota.putra if kuota else 0) - peserta_putra,
                'sisa_kuota_putri': (kuota.putri if kuota else 0) - peserta_putri,
                'sudah_terdaftar': sudah_terdaftar,
                'is_graded': is_graded,
                'status': 'Terdaftar' if sudah_terdaftar else 'Tersedia'
            })
        return jsonify({'status': 'success', 'data': result}), 200
    except Exception as e:
        current_app.logger.exception('Error in /api/kegiatan_tersedia:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API untuk peserta mendaftar/bergabung ke seleksi
@app.route('/api/daftar_seleksi', methods=['POST'])
@login_required
@csrf.exempt
def api_daftar_seleksi():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    """Endpoint untuk peserta mendaftar ke seleksi kegiatan"""
    try:
        if current_user.level != 'peserta':
            return jsonify({'status': 'error', 'message': t['access_denied']}), 403
        
        data = request.get_json(force=True)
        kegiatan_id = data.get('kegiatan_id')
        if not kegiatan_id:
            return jsonify({'status': 'error', 'message': t['id_event_not_found']}), 400
        
        # Cek apakah kegiatan ada dan sedang membuka seleksi
        kegiatan = Event.query.get(kegiatan_id)
        if not kegiatan:
            return jsonify({'status': 'error', 'message': t['event_not_found']}), 404
        today = datetime.utcnow().date()
        if today < kegiatan.mulai or today > kegiatan.selesai:
            return jsonify({
                'status': 'error', 
                'message': t['registration_closed']
            }), 400
        
        # Cek apakah peserta sudah punya biodata
        biodata = Participants.query.filter_by(email=current_user.email).first()
        if not biodata:
            return jsonify({
                'status': 'error', 
                'message': t['biodata_not_registered']
            }), 400
            
        # Cek kesesuaian golongan
        # Pastikan string tidak None, lalu lowercase untuk perbandingan case-insensitive
        participant_golongan = (biodata.golongan or "").lower().strip()
        event_golongan = (kegiatan.jenis_kegiatan or "").lower().strip()
        
        # Logika: Golongan peserta harus ada di dalam jenis kegiatan
        # Contoh: "penegak" ada di dalam "penegak" atau "penegak dan pandega"
        if participant_golongan not in event_golongan:
             return jsonify({
                'status': 'error', 
                'message': f"Maaf, golongan Anda ({biodata.golongan}) tidak sesuai dengan golongan kegiatan ({kegiatan.jenis_kegiatan})."
            }), 400
        
        # Cek apakah sudah terdaftar di kegiatan yang sama
        if kegiatan in biodata.registered_activities.all():
            return jsonify({
                'status': 'error', 
                'message': t['already_registered_for_event']
            }), 400
        
        # Cek kuota
        kuota = Kuota.query.filter_by(event_id=kegiatan_id).first()
        if kuota:
            peserta_putra = kegiatan.registered_participants.filter(Participants.jenis_kelamin == 'laki-laki').count()
            peserta_putri = kegiatan.registered_participants.filter(Participants.jenis_kelamin == 'perempuan').count()
            if biodata.jenis_kelamin == 'laki-laki' and peserta_putra >= kuota.putra:
                return jsonify({
                    'status': 'error', 
                    'message': t['male_quota_full']
                }), 400
            elif biodata.jenis_kelamin == 'perempuan' and peserta_putri >= kuota.putri:
                return jsonify({
                    'status': 'error', 
                    'message': t['female_quota_full']
                }), 400
        
        # Daftarkan peserta ke kegiatan menggunakan many-to-many relationship
        biodata.registered_activities.append(kegiatan)
        db.session.commit()
        
        # Log aktivitas
        log_activity(
            current_user.id,
            f"{t['register_event']}: {kegiatan.nama_kegiatan}"
        )
        
        # Buat notifikasi untuk admin
        create_notification_to_all_admins(
            f"Peserta {biodata.nama_lengkap or current_user.nama_lengkap} mendaftar ke seleksi: {kegiatan.nama_kegiatan}"
        )
        return jsonify({
            'status': 'success', 
            'message': f"{t['register_event_success']}: {kegiatan.nama_kegiatan}"
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/daftar_seleksi:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# API untuk peserta membatalkan pendaftaran seleksi
@app.route('/api/batal_daftar_seleksi', methods=['POST'])
@login_required
@csrf.exempt
def api_batal_daftar_seleksi():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    """Endpoint untuk peserta membatalkan pendaftaran ke seleksi kegiatan"""
    try:
        if current_user.level != 'peserta':
            return jsonify({'status': 'error', 'message': t['access_denied']}), 403
        
        data = request.get_json(force=True)
        kegiatan_id = data.get('kegiatan_id')
        
        if not kegiatan_id:
            return jsonify({'status': 'error', 'message': t['id_event_not_found']}), 400
        
        # Cek apakah kegiatan ada
        kegiatan = Event.query.get(kegiatan_id)
        if not kegiatan:
            return jsonify({'status': 'error', 'message': t['event_not_found']}), 404
        
        # Cek apakah peserta sudah punya biodata
        biodata = Participants.query.filter_by(email=current_user.email).first()
        if not biodata:
            return jsonify({
                'status': 'error', 
                'message': t['biodata_not_registered']
            }), 400
        
        # Cek apakah peserta terdaftar di kegiatan ini
        if kegiatan not in biodata.registered_activities.all():
            return jsonify({
                'status': 'error', 
                'message': f"{t['not_registered_for_event']}"
            }), 400
        
        # Cek apakah sudah ada hasil seleksi (jika sudah ada hasil seleksi, tidak bisa dibatalkan)
        hasil_seleksi = HasilSeleksi.query.filter_by(
            id_users=current_user.id, 
            event_id=kegiatan.id_kegiatan
        ).first()
        
        if hasil_seleksi:
            return jsonify({
                'status': 'error', 
                'message': f"{t['cannot_cancel_after_event_ended']}"
            }), 400
            
        # Cek apakah sudah ada penilaian (jika sudah dinilai, tidak bisa batal)
        criteria_ids = [c.id_kriteria for c in Criteria.query.filter_by(event_id=kegiatan.id_kegiatan).all()]
        if criteria_ids:
            score_exists = Penilaian.query.filter(
                Penilaian.id_users == current_user.id,
                Penilaian.id_kriteria.in_(criteria_ids)
            ).first()
            if score_exists:
                return jsonify({
                    'status': 'error', 
                    'message': "Tidak dapat membatalkan pendaftaran karena Anda sudah diberi nilai." if lang == 'id' else "Cannot cancel registration because you have been graded."
                }), 400
        
        # Batalkan pendaftaran (remove from many-to-many relationship)
        biodata.registered_activities.remove(kegiatan)
        db.session.commit()
        
        # Log aktivitas
        log_activity(
            current_user.id,
            f"{t['cancel_registration_event']}: {kegiatan.nama_kegiatan}"
        )
        
        return jsonify({
            'status': 'success', 
            'message': f"{t['cancel_registration_success']}: {kegiatan.nama_kegiatan}"
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/batal_daftar_seleksi:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/save_settings', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def save_settings():
    try:
        data = request.get_json()
        category = data.get('category')
        
        if not category:
            return jsonify({'status': 'error', 'message': 'Category is required'}), 400
        
        # Simpan setiap setting berdasarkan category
        if category == 'email':
            settings_to_save = {
                'mail_server': data.get('mail_server', ''),
                'mail_port': data.get('mail_port', '465'),
                'mail_use_tls': data.get('mail_use_tls', 'false'),
                'mail_use_ssl': data.get('mail_use_ssl', 'true'),
                'mail_username': data.get('mail_username', ''),
                'mail_enabled': data.get('mail_enabled', 'true')
            }
            # Hanya update password jika diisi
            if data.get('mail_password'):
                settings_to_save['mail_password'] = data.get('mail_password')
                
        elif category == 'sms':
            settings_to_save = {
                'twilio_account_sid': data.get('twilio_account_sid', ''),
                'twilio_whatsapp_from': data.get('twilio_whatsapp_from', 'whatsapp:+14155238886'),
                'sms_enabled': data.get('sms_enabled', 'true')
            }
            # Hanya update auth token jika diisi
            if data.get('twilio_auth_token'):
                settings_to_save['twilio_auth_token'] = data.get('twilio_auth_token')
                
        elif category == 'app':
            settings_to_save = {
                'app_name': data.get('app_name', ''),
                'app_description': data.get('app_description', ''),
                'app_version': data.get('app_version', ''),
                'organization_name': data.get('organization_name', ''),
                'organization_address': data.get('organization_address', ''),
                'organization_phone': data.get('organization_phone', ''),
                'organization_email': data.get('organization_email', ''),
                'default_language': data.get('default_language', 'id')
            }
        elif category == 'logo':
            settings_to_save = {
                'logo_path': data.get('logo_path', 'img/logo.png')
            }
        else:
            return jsonify({'status': 'error', 'message': 'Invalid category'}), 400
        
        # Simpan ke database
        for key, value in settings_to_save.items():
            setting = Settings.query.filter_by(key=key).first()
            if setting:
                setting.value = str(value)
                setting.updated_by = current_user.id
            else:
                setting = Settings(
                    key=key,
                    value=str(value),
                    category=category,
                    updated_by=current_user.id
                )
                db.session.add(setting)
        
        db.session.commit()
        
        # Update app config untuk email jika category email
        if category == 'email':
            app.config['MAIL_SERVER'] = settings_to_save.get('mail_server', app.config.get('MAIL_SERVER'))
            app.config['MAIL_PORT'] = int(settings_to_save.get('mail_port', app.config.get('MAIL_PORT', 465)))
            app.config['MAIL_USE_TLS'] = settings_to_save.get('mail_use_tls', 'false') == 'true'
            app.config['MAIL_USE_SSL'] = settings_to_save.get('mail_use_ssl', 'true') == 'true'
            app.config['MAIL_USERNAME'] = settings_to_save.get('mail_username', app.config.get('MAIL_USERNAME'))
            if 'mail_password' in settings_to_save:
                app.config['MAIL_PASSWORD'] = settings_to_save['mail_password']
        
        # Log aktivitas
        log_activity(
            current_user.id,
            f'Mengubah pengaturan {category}'
        )
        
        return jsonify({
            'status': 'success',
            'message': f'Pengaturan {category} berhasil disimpan'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/save_settings:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/test_email', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def test_email():
    try:
        data = request.get_json()
        test_email_address = data.get('email', current_user.email)
        
        # Ambil pengaturan email dari database
        mail_settings = {}
        settings = Settings.query.filter_by(category='email').all()
        for s in settings:
            mail_settings[s.key] = s.value
        
        # Gunakan pengaturan dari database atau fallback ke env/config
        mail_server = mail_settings.get('mail_server') or app.config.get('MAIL_SERVER')
        mail_port = int(mail_settings.get('mail_port') or app.config.get('MAIL_PORT', 465))
        mail_use_ssl = mail_settings.get('mail_use_ssl', 'true') == 'true'
        mail_use_tls = mail_settings.get('mail_use_tls', 'false') == 'true'
        mail_username = mail_settings.get('mail_username') or app.config.get('MAIL_USERNAME')
        mail_password = mail_settings.get('mail_password') or app.config.get('MAIL_PASSWORD')
        
        # Buat Mail instance sementara dengan konfigurasi baru
        from flask_mail import Mail
        test_mail = Mail()
        test_mail.init_app(app)
        
        # Update config sementara
        original_config = {
            'MAIL_SERVER': app.config.get('MAIL_SERVER'),
            'MAIL_PORT': app.config.get('MAIL_PORT'),
            'MAIL_USE_SSL': app.config.get('MAIL_USE_SSL'),
            'MAIL_USE_TLS': app.config.get('MAIL_USE_TLS'),
            'MAIL_USERNAME': app.config.get('MAIL_USERNAME'),
            'MAIL_PASSWORD': app.config.get('MAIL_PASSWORD')
        }
        
        app.config['MAIL_SERVER'] = mail_server
        app.config['MAIL_PORT'] = mail_port
        app.config['MAIL_USE_SSL'] = mail_use_ssl
        app.config['MAIL_USE_TLS'] = mail_use_tls
        app.config['MAIL_USERNAME'] = mail_username
        app.config['MAIL_PASSWORD'] = mail_password
        
        # Reinitialize mail dengan config baru
        mail.init_app(app)
        
        try:
            msg = Message(
                subject='Test Email - SPK Pramuka',
                recipients=[test_email_address],
                body='Ini adalah email test dari sistem SPK Pramuka. Jika Anda menerima email ini, berarti konfigurasi email sudah benar.',
                sender=mail_username
            )
            mail.send(msg)
            
            # Reinitialize mail dengan config asli
            mail.init_app(app)
            
            # Restore original config
            for key, value in original_config.items():
                app.config[key] = value
            
            return jsonify({
                'status': 'success',
                'message': f'Email test berhasil dikirim ke {test_email_address}'
            }), 200
        except Exception as e:
            # Restore original config
            for key, value in original_config.items():
                app.config[key] = value
            raise e
            
    except Exception as e:
        current_app.logger.exception('Error in /api/test_email:')
        return jsonify({'status': 'error', 'message': f'Gagal mengirim email test: {str(e)}'}), 500

@app.route('/api/upload_logo', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def upload_logo():
    try:
        if 'logo' not in request.files:
            return jsonify({'status': 'error', 'message': 'Tidak ada file yang diupload'}), 400
        
        file = request.files['logo']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'Tidak ada file yang dipilih'}), 400
        
        # Validasi file
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg'}
        if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'status': 'error', 'message': 'Format file tidak didukung. Gunakan PNG, JPG, JPEG, GIF, atau SVG'}), 400
        
        # Simpan file
        filename = secure_filename(file.filename)
        logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'logo')
        os.makedirs(logo_dir, exist_ok=True)
        
        # Generate unique filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"logo_{timestamp}_{filename}"
        filepath = os.path.join(logo_dir, filename)
        file.save(filepath)
        
        # Simpan path ke database
        logo_path = f"uploads/logo/{filename}"
        setting = Settings.query.filter_by(key='logo_path').first()
        if setting:
            # Hapus logo lama jika ada
            old_value = setting.value
            # Bersihkan path dari prefix /static/ jika ada
            if old_value.startswith('/static/'):
                old_value = old_value.replace('/static/', '')
            # Pastikan path relatif dari uploads/
            if old_value.startswith('uploads/'):
                old_value = old_value.replace('uploads/', '')
            old_path = os.path.join(app.config['UPLOAD_FOLDER'], old_value)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except:
                    pass
            setting.value = logo_path
        else:
            setting = Settings(key='logo_path', value=logo_path, category='logo', updated_by=current_user.id)
            db.session.add(setting)
        
        db.session.commit()
        
        # Log aktivitas
        log_activity(current_user.id, 'Mengupload logo aplikasi')
        
        # Kembalikan path dengan /static/ untuk ditampilkan di frontend
        logo_path_display = f'/static/{logo_path}'
        
        return jsonify({
            'status': 'success',
            'message': 'Logo berhasil diupload',
            'logo_path': logo_path_display
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/upload_logo:')
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================================
# API UNTUK LAPORAN & ARSIP SELEKSI
# =====================================================

# API Download Arsip Seleksi
@app.route('/api/download_arsip/<int:arsip_id>')
@login_required
@admin_required
def api_download_arsip(arsip_id):
    """Download file arsip seleksi"""
    try:
        arsip = ArsipSeleksi.query.get_or_404(arsip_id)
        
        if not arsip.file_path:
            return jsonify({'success': False, 'message': 'File tidak ditemukan'}), 404
        
        # Path file
        file_path = os.path.join(app.root_path, 'static', arsip.file_path)
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File tidak ditemukan di server'}), 404
        
        # Tentukan MIME type berdasarkan file_type
        if arsip.file_type == 'excel':
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif arsip.file_type == 'pdf':
            mimetype = 'application/pdf'
        else:
            mimetype = 'application/octet-stream'
        
        # Nama file untuk download
        filename = f"{arsip.nama_arsip}.{'xlsx' if arsip.file_type == 'excel' else 'pdf'}"
        
        return send_file(
            file_path,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        current_app.logger.exception('Error in /api/download_arsip:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Hapus Arsip Seleksi
@app.route('/api/hapus_arsip/<int:arsip_id>', methods=['DELETE'])
@login_required
@admin_required
@csrf.exempt
def api_hapus_arsip(arsip_id):
    """Hapus arsip seleksi"""
    try:
        arsip = ArsipSeleksi.query.get_or_404(arsip_id)
        nama_arsip = arsip.nama_arsip
        
        # Hapus file fisik jika ada
        if arsip.file_path:
            file_path = os.path.join(app.root_path, 'static', arsip.file_path)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    current_app.logger.warning(f'Gagal menghapus file arsip: {e}')
        
        # Hapus record dari database
        db.session.delete(arsip)
        db.session.commit()
        
        log_activity(current_user.id, f'Menghapus arsip seleksi: {nama_arsip}')
        
        return jsonify({'success': True, 'message': f'Arsip "{nama_arsip}" berhasil dihapus'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/hapus_arsip:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Generate Laporan Excel
@app.route('/api/generate_laporan_excel/<int:event_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def api_generate_laporan_excel(event_id):
    """Generate laporan hasil seleksi dalam format Excel"""
    try:
        event = Event.query.get_or_404(event_id)
        
        # Ambil data hasil seleksi
        hasil_list = HasilSeleksi.query.filter_by(event_id=event_id).order_by(HasilSeleksi.ranking.asc()).all()
        
        if not hasil_list:
            return jsonify({'success': False, 'message': 'Belum ada hasil seleksi untuk kegiatan ini'}), 400
        
        # Buat file Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            return jsonify({'success': False, 'message': 'Library openpyxl tidak tersedia. Silakan install dengan: pip install openpyxl'}), 500
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hasil Seleksi"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Judul
        ws.merge_cells('A1:F1')
        ws['A1'] = f"LAPORAN HASIL SELEKSI"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"{event.nama_kegiatan}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Tanggal: {datetime.now().strftime('%d %B %Y')}"
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Header tabel
        headers = ['Ranking', 'Nama Peserta', 'Email', 'Jenis Kelamin', 'Asal Gudep', 'Skor Akhir']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row_idx, hasil in enumerate(hasil_list, start=6):
            user = Users.query.get(hasil.id_users)
            participant = Participants.query.filter_by(email=user.email).first() if user else None
            
            data = [
                hasil.ranking,
                user.nama_lengkap if user else 'N/A',
                user.email if user else 'N/A',
                participant.jenis_kelamin if participant else (user.jenis_kelamin if user else 'N/A'),
                participant.asal_gudep if participant else 'N/A',
                round(hasil.skor_akhir, 4)
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [1, 6]:  # Ranking dan skor = center
                    cell.alignment = Alignment(horizontal="center")
        
        # Atur lebar kolom
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        
        # Simpan file
        reports_dir = os.path.join(app.root_path, 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"laporan_{event.id_kegiatan}_{timestamp}.xlsx"
        file_path = os.path.join(reports_dir, filename)
        wb.save(file_path)
        
        # Simpan ke database arsip
        arsip = ArsipSeleksi(
            event_id=event_id,
            nama_arsip=f"Laporan {event.nama_kegiatan} - {datetime.now().strftime('%d %b %Y')}",
            deskripsi=f"Laporan hasil seleksi kegiatan {event.nama_kegiatan}",
            file_path=f"reports/{filename}",
            file_type='excel',
            dibuat_oleh=current_user.id,
            status='aktif'
        )
        db.session.add(arsip)
        db.session.commit()
        
        log_activity(current_user.id, f'Generate laporan Excel untuk kegiatan: {event.nama_kegiatan}')
        
        return jsonify({
            'success': True, 
            'message': f'Laporan Excel berhasil di-generate untuk kegiatan "{event.nama_kegiatan}"'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/generate_laporan_excel:')
        return jsonify({'success': False, 'message': str(e)}), 500

# API Generate Laporan PDF
@app.route('/api/generate_laporan_pdf/<int:event_id>', methods=['POST'])
@login_required
@admin_required
@csrf.exempt
def api_generate_laporan_pdf(event_id):
    """Generate laporan hasil seleksi dalam format PDF"""
    try:
        event = Event.query.get_or_404(event_id)
        
        # Ambil data hasil seleksi
        hasil_list = HasilSeleksi.query.filter_by(event_id=event_id).order_by(HasilSeleksi.ranking.asc()).all()
        
        if not hasil_list:
            return jsonify({'success': False, 'message': 'Belum ada hasil seleksi untuk kegiatan ini'}), 400
        
        # Buat file PDF
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            return jsonify({'success': False, 'message': 'Library reportlab tidak tersedia. Silakan install dengan: pip install reportlab'}), 500
        
        # Simpan file
        reports_dir = os.path.join(app.root_path, 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"laporan_{event.id_kegiatan}_{timestamp}.pdf"
        file_path = os.path.join(reports_dir, filename)
        
        # Buat dokumen PDF
        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,  # Center
            spaceAfter=12
        )
        subtitle_style = ParagraphStyle(
            'CustomSubTitle',
            parent=styles['Heading2'],
            fontSize=12,
            alignment=1,
            spaceAfter=20
        )
        
        # Judul
        elements.append(Paragraph("LAPORAN HASIL SELEKSI", title_style))
        elements.append(Paragraph(f"{event.nama_kegiatan}", subtitle_style))
        elements.append(Paragraph(f"Tanggal: {datetime.now().strftime('%d %B %Y')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Data tabel
        table_data = [['Ranking', 'Nama Peserta', 'Email', 'Jenis Kelamin', 'Asal Gudep', 'Skor Akhir']]
        
        for hasil in hasil_list:
            user = Users.query.get(hasil.id_users)
            participant = Participants.query.filter_by(email=user.email).first() if user else None
            
            table_data.append([
                str(hasil.ranking),
                user.nama_lengkap if user else 'N/A',
                user.email if user else 'N/A',
                participant.jenis_kelamin if participant else (user.jenis_kelamin if user else 'N/A'),
                participant.asal_gudep if participant else 'N/A',
                str(round(hasil.skor_akhir, 4))
            ])
        
        # Buat tabel
        table = Table(table_data, colWidths=[0.7*inch, 2*inch, 2.5*inch, 1.2*inch, 2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Simpan ke database arsip
        arsip = ArsipSeleksi(
            event_id=event_id,
            nama_arsip=f"Laporan {event.nama_kegiatan} - {datetime.now().strftime('%d %b %Y')}",
            deskripsi=f"Laporan hasil seleksi kegiatan {event.nama_kegiatan}",
            file_path=f"reports/{filename}",
            file_type='pdf',
            dibuat_oleh=current_user.id,
            status='aktif'
        )
        db.session.add(arsip)
        db.session.commit()
        
        log_activity(current_user.id, f'Generate laporan PDF untuk kegiatan: {event.nama_kegiatan}')
        
        return jsonify({
            'success': True, 
            'message': f'Laporan PDF berhasil di-generate untuk kegiatan "{event.nama_kegiatan}"'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Error in /api/generate_laporan_pdf:')
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/logout/')
def logout():
    lang = session.get('lang', 'id')
    t = TRANSLATIONS.get(lang, TRANSLATIONS['id'])
    
    session.clear()
    session.pop('username', None)
    flash(f"{t['logged_out']}", "info")
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
 